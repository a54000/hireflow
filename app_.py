import os, re, hashlib, csv, io, sqlite3, json, smtplib, secrets, zipfile, traceback
import xml.etree.ElementTree as ET
import requests
from flask import send_from_directory
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv
from flask import (Flask, render_template, jsonify, request,
                   send_file, session, redirect, url_for)
from datetime import datetime, date, timedelta
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.exceptions import HTTPException, RequestEntityTooLarge
from werkzeug.middleware.proxy_fix import ProxyFix
from functools import wraps
from authlib.integrations.flask_client import OAuth
from email.mime.text import MIMEText
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request as GoogleAuthRequest
from googleapiclient.discovery import build as google_api_build

import base64
from urllib.parse import urlparse
from ats_schema import ensure_ats_pipeline_schema
from ats_pipeline import MATCH_PIPELINE_VERSION, parse_jd as parse_jd_structured, parse_resume as parse_resume_structured, run_hybrid_match, versioned_text_hash
from embedding_engine import deserialize_embedding, serialize_embedding
from services.match_analysis import build_match_dashboard
from services.match_pdf import build_match_pdf
from services.screening_pdf import build_screening_questions_pdf
from skill_aliases import canonical_skill, skill_aliases_for


import sys
sys.path.insert(0, '/Users/surindersingh/Library/Python/3.9/lib/python3.9/site-packages')
print(f"Python: {sys.executable}", flush=True)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_OK = True
    print("openpyxl loaded!", flush=True)
except ImportError:
    XLSX_OK = False
    print("openpyxl NOT loaded", flush=True)

try:
    import cloudinary, cloudinary.uploader
    CLOUD_OK = True
except ImportError:
    CLOUD_OK = False

app = Flask(__name__)
#app = Flask(__name__, template_folder="/Users/surindersingh/Downloads/ats6-dev/templates")
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)

load_dotenv()
oauth = OAuth(app)

google = oauth.register(
    name='google',
    client_id=os.getenv("GOOGLE_CLIENT_ID"),
    client_secret=os.getenv("GOOGLE_CLIENT_SECRET"),
    authorize_url='https://accounts.google.com/o/oauth2/v2/auth',
    access_token_url='https://oauth2.googleapis.com/token',
    api_base_url='https://www.googleapis.com/oauth2/v3/',
    jwks_uri='https://www.googleapis.com/oauth2/v3/certs',
    client_kwargs={
        'scope': 'openid email profile https://www.googleapis.com/auth/gmail.send'
    }
)

app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
app.config["PREFERRED_URL_SCHEME"] = os.getenv("PREFERRED_URL_SCHEME", "https")
SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    raise RuntimeError("SECRET_KEY is required. Set it in your environment or .env file.")
app.secret_key = SECRET_KEY

@app.errorhandler(RequestEntityTooLarge)
def handle_file_too_large(error):
    if request.path.startswith("/api/"):
        return jsonify({
            "error": "Uploaded files are too large. Please keep the total upload under 50 MB and try again."
        }), 413
    return "Uploaded files are too large. Please keep the total upload under 50 MB and try again.", 413

@app.errorhandler(500)
def handle_internal_error(error):
    print("Unhandled server error:", error, flush=True)
    traceback.print_exc()
    if request.path.startswith("/api/"):
        return jsonify({
            "error": "Server error while processing this request. Please check the ATS terminal/logs for details."
        }), 500
    return "Server error while processing this request. Please check the ATS terminal/logs for details.", 500

@app.errorhandler(Exception)
def handle_unexpected_error(error):
    if isinstance(error, HTTPException):
        if request.path.startswith("/api/"):
            return jsonify({"error": error.description or error.name}), error.code
        return error
    print("Unhandled exception:", error, flush=True)
    traceback.print_exc()
    if request.path.startswith("/api/"):
        return jsonify({"error": str(error) or "Server error while processing this request."}), 500
    return "Server error while processing this request. Please check the ATS terminal/logs for details.", 500

BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
DB_PATH        = os.path.join(BASE_DIR, "ats.db")
print("DB PATH:", DB_PATH)
CLOUDINARY_URL = os.getenv("CLOUDINARY_URL", "")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "")

# Gmail config for weekly summary
GMAIL_USER     = os.getenv("GMAIL_USER", "")       # your gmail address
GMAIL_APP_PASS = os.getenv("GMAIL_APP_PASS", "")   # gmail app password
ADMIN_EMAIL    = os.getenv("ADMIN_EMAIL", "")       # where to send the summary

# CV Parsing config
CV_PARSE_MODE = os.getenv("CV_PARSE_MODE", "free")  # "free" or "ai"
CV_AI_PROVIDER = os.getenv("CV_AI_PROVIDER", "")   # openai, gemini, claude
CV_AI_API_KEY = os.getenv("CV_AI_API_KEY", "")

# Candidate statuses
STATUSES = ["New", "Shortlisted", "Screening Pending", "Screen Rejected", 
            "Interviewed", "HM Rejected", "Offered", "Joined", "Dropped", "Duplicate", "OnHold"]

# Submission statuses
SUBMISSION_STATUSES = ["Pending", "Reviewing", "Interview Scheduled", "Selected", "Rejected", "Withdrawn"]

# ── DB ────────────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    print("Connecting DB...")
    print("DB PATH:", os.path.abspath(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def hash_password(password):
    return generate_password_hash(password, method="pbkdf2:sha256", salt_length=16)

def set_login_session_from_app_user(app_user):
    team_member = None
    if app_user["team_member_id"]:
        conn = get_db()
        team_member = conn.execute("SELECT * FROM team_members WHERE id=?", (app_user["team_member_id"],)).fetchone()
        conn.close()
    display_name = (team_member["name"] if team_member else app_user["username"]) or app_user["username"]
    email = (team_member["email"] if team_member else "") or ""
    role = (team_member["role"] if team_member else "") or ""
    can_bulk_upload = bool(
        app_user["is_admin"] or
        app_user["is_bulk_admin"] or
        (team_member and (team_member["can_bulk_upload"] or role in {"Bulk Admin", "Admin"}))
    )
    session["logged_in"] = True
    session["user_id"] = app_user["id"]
    session["app_user_id"] = app_user["id"]
    session["team_member_id"] = app_user["team_member_id"]
    session["username"] = display_name
    session["recruiter_name"] = display_name
    session["email"] = email
    session["recruiter_email"] = email
    session["is_admin"] = 1 if app_user["is_admin"] else 0
    session["can_bulk_upload"] = 1 if can_bulk_upload else 0

def non_admin_candidate_owner_clause(current_session, alias="c"):
    if not current_session or current_session.get("is_admin"):
        return "", []
    prefix = f"{alias}." if alias else ""
    team_member_id = current_session.get("team_member_id")
    recruiter_email = (current_session.get("recruiter_email") or "").strip().lower()
    if team_member_id:
        return f" AND {prefix}sourcer_id=?", [team_member_id]
    if recruiter_email:
        return f" AND lower({prefix}recruiter_email)=?", [recruiter_email]
    return " AND 1=0", []

def system_email(to_addr, subject, body):
    if not to_addr:
        return {"error": "No recipient configured"}
    if not GMAIL_USER or not GMAIL_APP_PASS:
        return {"error": "Email not configured"}
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = GMAIL_USER
    msg["To"] = to_addr
    msg.attach(MIMEText(body, "plain"))
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
        s.login(GMAIL_USER, GMAIL_APP_PASS)
        s.sendmail(GMAIL_USER, to_addr, msg.as_string())
    return {"ok": True}

def admin_approval_emails():
    conn = get_db()
    rows = conn.execute("""
        SELECT DISTINCT COALESCE(t.email, '') AS email
        FROM app_users u
        LEFT JOIN team_members t ON t.id = u.team_member_id
        WHERE u.is_admin=1 AND u.is_active=1
    """).fetchall()
    team_admins = conn.execute("""
        SELECT DISTINCT email
        FROM team_members
        WHERE email IS NOT NULL
          AND email != ''
          AND LOWER(COALESCE(role,'')) IN ('admin', 'bulk admin')
    """).fetchall()
    conn.close()
    emails = [r["email"] for r in rows if r["email"]]
    emails.extend([r["email"] for r in team_admins if r["email"]])
    if ADMIN_EMAIL:
        emails.append(ADMIN_EMAIL)
    return sorted(set(emails))

def normalize_login_name(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())

def resolve_user_for_password_reset(conn, identifier):
    identifier = (identifier or "").strip().lower()
    user = conn.execute("""
        SELECT u.*, t.email AS team_email
        FROM app_users u
        LEFT JOIN team_members t ON t.id = u.team_member_id
        WHERE LOWER(u.username)=LOWER(?) OR LOWER(t.email)=LOWER(?)
    """, (identifier, identifier)).fetchone()
    if user:
        return user
    team = conn.execute("""
        SELECT *
        FROM team_members
        WHERE LOWER(email)=LOWER(?) OR LOWER(name)=LOWER(?)
    """, (identifier, identifier)).fetchone()
    if not team:
        return None
    normalized_name = normalize_login_name(team["name"])
    return conn.execute("""
        SELECT u.*, ? AS team_email
        FROM app_users u
        WHERE LOWER(u.username)=LOWER(?)
           OR LOWER(u.username)=LOWER(?)
    """, (team["email"], team["name"], normalized_name)).fetchone()

def reset_recipients_for_user(user):
    recipients = []
    if user and user["team_email"]:
        recipients.append(user["team_email"])
    recipients.extend(admin_approval_emails())
    return sorted(set([email for email in recipients if email]))

def ensure_auth_workflow_schema(conn):
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS password_reset_tokens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        token TEXT UNIQUE NOT NULL,
        expires_at TEXT NOT NULL,
        used INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        FOREIGN KEY(user_id) REFERENCES app_users(id)
    );
    CREATE TABLE IF NOT EXISTS user_registration_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        email TEXT NOT NULL,
        username TEXT NOT NULL,
        password_hash TEXT NOT NULL,
        token TEXT UNIQUE NOT NULL,
        status TEXT DEFAULT 'pending',
        approved_by INTEGER,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        approved_at TEXT
    );
    """)

def init_db():
    conn = get_db()
    ensure_ats_pipeline_schema(conn)
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS job_details (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        candidate_id INTEGER NOT NULL,
        title TEXT,
        location TEXT,
        salary TEXT,
        department TEXT,
        job_id TEXT,
        notes TEXT,
        tags TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now','localtime')),
        updated_at TEXT DEFAULT (datetime('now','localtime')),
        FOREIGN KEY(candidate_id) REFERENCES candidates(id)
    );
    CREATE TABLE IF NOT EXISTS standard_roles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        domain TEXT NOT NULL,
        role_name TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        UNIQUE(domain, role_name)
    );
    CREATE TABLE IF NOT EXISTS standard_skills (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        skill_name TEXT UNIQUE,
        category TEXT,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS clients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_name TEXT UNIQUE,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS pipelines (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        role_name TEXT UNIQUE,
        status_list TEXT,
        is_default INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS jobs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        job_id TEXT UNIQUE,
        title TEXT NOT NULL,
        description TEXT,
        client_name TEXT,
        status TEXT DEFAULT 'Open',
        created_at TEXT DEFAULT (datetime('now','localtime')),
        updated_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS requirements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        description TEXT,
        client_name TEXT,
        jd_filename TEXT,
        jd_url TEXT,
        jd_public_id TEXT,
        assigned_sourcer_id INTEGER,
        assigned_recruiter_id INTEGER,
        daily_target INTEGER DEFAULT 3,
        status TEXT DEFAULT 'Open',
        created_by TEXT,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        updated_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS requirement_checks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        requirement_id INTEGER NOT NULL,
        check_name TEXT NOT NULL,
        check_description TEXT,
        check_type TEXT DEFAULT 'boolean',
        pass_criteria TEXT,
        is_mandatory INTEGER DEFAULT 1,
        sort_order INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS candidates (
        id                 INTEGER PRIMARY KEY AUTOINCREMENT,
        upload_batch       TEXT,
        recruiter_name     TEXT,
        recruiter_email    TEXT,
        role_name          TEXT,
        education       TEXT,
        candidate_name     TEXT,
        email_addr         TEXT,
        phone              TEXT,
        current_company    TEXT,
        current_role       TEXT,
        experience_years   TEXT,
        key_skills         TEXT,
        notice_period      TEXT,
        current_salary     TEXT,
        expected_salary    TEXT,
        current_location   TEXT,
        preferred_location TEXT,
        remarks            TEXT,
        cv_filename        TEXT,
        cv_url             TEXT,
        cv_public_id       TEXT,
        cv_summary         TEXT,
        status             TEXT DEFAULT 'New',
        tags               TEXT DEFAULT '',
        is_duplicate       INTEGER DEFAULT 0,
        duplicate_of       INTEGER,
        missing_info       TEXT,
        job_id            TEXT,
        sourcer_id         INTEGER,
        requirement_id     INTEGER,
        created_at         TEXT DEFAULT (datetime('now','localtime')),
        updated_at         TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS requirement_submissions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        candidate_id INTEGER NOT NULL,
        requirement_id INTEGER NOT NULL,
        sourcer_id INTEGER,
        submitted_by TEXT,
        status TEXT DEFAULT 'Submitted',
        notes TEXT,
        recruiter_feedback TEXT,
        feedback_by INTEGER,
        submitted_at TEXT DEFAULT (datetime('now','localtime')),
        updated_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS submission_checks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        submission_id INTEGER NOT NULL,
        check_id INTEGER NOT NULL,
        check_name TEXT,
        passed INTEGER DEFAULT 0,
        notes TEXT
    );
    CREATE TABLE IF NOT EXISTS team_members (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        name      TEXT,
        email     TEXT UNIQUE,
        phone     TEXT,
        role      TEXT,
        can_bulk_upload INTEGER DEFAULT 0,
        is_fixed  INTEGER DEFAULT 0,
        added_at  TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS app_users (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        username     TEXT UNIQUE NOT NULL,
        password     TEXT NOT NULL,
        team_member_id INTEGER,
        is_admin     INTEGER DEFAULT 0,
        is_bulk_admin INTEGER DEFAULT 0,
        is_active    INTEGER DEFAULT 1,
        created_at   TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS password_reset_tokens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        token TEXT UNIQUE NOT NULL,
        expires_at TEXT NOT NULL,
        used INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        FOREIGN KEY(user_id) REFERENCES app_users(id)
    );
    CREATE TABLE IF NOT EXISTS user_registration_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        email TEXT NOT NULL,
        username TEXT NOT NULL,
        password_hash TEXT NOT NULL,
        token TEXT UNIQUE NOT NULL,
        status TEXT DEFAULT 'pending',
        approved_by INTEGER,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        approved_at TEXT
    );
    CREATE TABLE IF NOT EXISTS alerts (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        alert_type      TEXT,
        message         TEXT,
        candidate_id    INTEGER,
        recruiter_email TEXT,
        is_read         INTEGER DEFAULT 0,
        created_at      TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS upload_log (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        batch_id         TEXT,
        recruiter_name   TEXT,
        recruiter_email  TEXT,
        filename         TEXT,
        candidates_added INTEGER DEFAULT 0,
        duplicates_found INTEGER DEFAULT 0,
        missing_count    INTEGER DEFAULT 0,
        uploaded_at      TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS saved_searches (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        name       TEXT NOT NULL,
        filters    TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS submissions (
        id                 INTEGER PRIMARY KEY AUTOINCREMENT,
        candidate_id       INTEGER NOT NULL,
        recruiter_name     TEXT,
        recruiter_email    TEXT,
        hm_name            TEXT NOT NULL,
        hm_email           TEXT,
        hm_company         TEXT,
        role_submitted     TEXT,
        status             TEXT DEFAULT 'Pending',
        submitted_at       TEXT DEFAULT (datetime('now','localtime')),
        updated_at         TEXT DEFAULT (datetime('now','localtime')),
        notes              TEXT,
        FOREIGN KEY (candidate_id) REFERENCES candidates(id)
    );
    CREATE TABLE IF NOT EXISTS hiring_managers (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        name        TEXT NOT NULL,
        email       TEXT,
        company     TEXT,
        created_at  TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS pipelines (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        role_name   TEXT NOT NULL,
        status_list TEXT NOT NULL,
        is_default  INTEGER DEFAULT 0,
        created_at  TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS automation_rules (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        rule_name    TEXT NOT NULL,
        trigger_type TEXT NOT NULL,
        trigger_value TEXT,
        action_type  TEXT NOT NULL,
        action_config TEXT,
        is_active    INTEGER DEFAULT 1,
        created_at   TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS interviews (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        candidate_id   INTEGER NOT NULL,
        interviewer_name TEXT,
        interviewer_email TEXT,
        scheduled_at  TEXT,
        duration_mins INTEGER DEFAULT 60,
        location      TEXT,
        meeting_link  TEXT,
        status        TEXT DEFAULT 'Scheduled',
        notes         TEXT,
        created_at    TEXT DEFAULT (datetime('now','localtime')),
        FOREIGN KEY (candidate_id) REFERENCES candidates(id)
    );
    CREATE TABLE IF NOT EXISTS approvals (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        candidate_id    INTEGER NOT NULL,
        requestor_name  TEXT,
        requestor_email TEXT,
        approval_type   TEXT,
        status          TEXT DEFAULT 'Pending',
        comments       TEXT,
        reviewed_by    TEXT,
        reviewed_at    TEXT,
        created_at     TEXT DEFAULT (datetime('now','localtime')),
        FOREIGN KEY (candidate_id) REFERENCES candidates(id)
    );
    CREATE TABLE IF NOT EXISTS email_templates (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        name        TEXT NOT NULL UNIQUE,
        subject     TEXT NOT NULL,
        body        TEXT NOT NULL,
        trigger_event TEXT,
        is_default  INTEGER DEFAULT 0,
        created_at  TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS email_log (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        candidate_id   INTEGER,
        template_name  TEXT,
        recipient     TEXT NOT NULL,
        subject       TEXT,
        body          TEXT,
        status        TEXT DEFAULT 'Sent',
        error_msg     TEXT,
        sent_at       TEXT DEFAULT (datetime('now','localtime')),
        FOREIGN KEY (candidate_id) REFERENCES candidates(id)
    );
    CREATE TABLE IF NOT EXISTS skills (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        raw_value TEXT,
        canonical_value TEXT NOT NULL UNIQUE,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS skill_aliases (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        skill_id INTEGER NOT NULL,
        alias TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        UNIQUE(skill_id, alias),
        FOREIGN KEY(skill_id) REFERENCES skills(id)
    );
    CREATE TABLE IF NOT EXISTS candidate_skills (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        candidate_id INTEGER,
        resume_hash TEXT,
        raw_value TEXT,
        canonical_value TEXT NOT NULL,
        confidence REAL DEFAULT 0,
        skill_type TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS candidate_roles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        candidate_id INTEGER,
        resume_hash TEXT,
        raw_value TEXT,
        canonical_value TEXT NOT NULL,
        confidence REAL DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS candidate_domains (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        candidate_id INTEGER,
        resume_hash TEXT,
        raw_value TEXT,
        canonical_value TEXT NOT NULL,
        confidence REAL DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS jd_requirements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        jd_hash TEXT UNIQUE,
        role_title TEXT,
        parsed_json TEXT,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        updated_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS match_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        jd_hash TEXT NOT NULL,
        resume_hash TEXT NOT NULL,
        final_score INTEGER,
        structured_score INTEGER,
        semantic_score INTEGER,
        hard_filter_score INTEGER,
        result_json TEXT,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        UNIQUE(jd_hash, resume_hash)
    );
    CREATE TABLE IF NOT EXISTS parsed_resume_cache (
        resume_hash TEXT PRIMARY KEY,
        parsed_json TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        updated_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS normalized_skill_cache (
        raw_value TEXT PRIMARY KEY,
        canonical_value TEXT NOT NULL,
        aliases_json TEXT,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS embedding_cache (
        text_hash TEXT NOT NULL,
        model TEXT NOT NULL,
        embedding_json TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        PRIMARY KEY(text_hash, model)
    );
    """)
    candidate_columns = {
        row["name"] for row in conn.execute("PRAGMA table_info(candidates)").fetchall()
    }
    if "sourcer_id" not in candidate_columns:
        conn.execute("ALTER TABLE candidates ADD COLUMN sourcer_id INTEGER")
    if "requirement_id" not in candidate_columns:
        conn.execute("ALTER TABLE candidates ADD COLUMN requirement_id INTEGER")
    requirement_columns = {
        row["name"] for row in conn.execute("PRAGMA table_info(requirements)").fetchall()
    }
    for col in ["jd_filename", "jd_url", "jd_public_id"]:
        if col not in requirement_columns:
            conn.execute(f"ALTER TABLE requirements ADD COLUMN {col} TEXT")
    team_columns = {
        row["name"] for row in conn.execute("PRAGMA table_info(team_members)").fetchall()
    }
    if "can_bulk_upload" not in team_columns:
        conn.execute("ALTER TABLE team_members ADD COLUMN can_bulk_upload INTEGER DEFAULT 0")
    user_columns = {
        row["name"] for row in conn.execute("PRAGMA table_info(app_users)").fetchall()
    }
    if "is_bulk_admin" not in user_columns:
        conn.execute("ALTER TABLE app_users ADD COLUMN is_bulk_admin INTEGER DEFAULT 0")
    ensure_ats_pipeline_schema(conn)

    # Insert default pipeline if not exists
    conn.execute("INSERT OR IGNORE INTO pipelines (role_name,status_list,is_default) VALUES ('Default',?,1)",
                 (json.dumps(["New","Shortlisted","Feedback Pending","Offered","On Hold","Joined","Rejected"]),))
    # Insert default email templates
    defaults = [
        ('welcome', 'Application received for {role_name}', 
         'Hi {candidate_name},\n\nThank you for applying for {role_name}. We have received your profile and will review it shortly.\n\nBest regards,\n{recruiter_name}',
         'new_candidate', 1),
        ('status_update', 'Update on your {role_name} application',
         'Hi {candidate_name},\n\nYour application for {role_name} is now marked as {status}.\n\nBest regards,\n{recruiter_name}',
         'status_change', 1),
        ('interview_invite', 'Interview invitation for {role_name}',
         'Hi {candidate_name},\n\nCongratulations! You have been shortlisted for an interview for {role_name}.\n\nInterview Details:\nDate: {scheduled_at}\nInterviewer: {interviewer_name}\nLocation: {location}\n\nBest regards,\n{recruiter_name}',
         'interview_scheduled', 1)
    ]
    for name, subject, body, trigger, is_def in defaults:
        conn.execute("INSERT OR IGNORE INTO email_templates (name,subject,body,trigger_event,is_default) VALUES (?,?,?,?,?)",
                     (name, subject, body, trigger, is_def))
        conn.execute("UPDATE email_templates SET subject=?, body=?, trigger_event=?, is_default=? WHERE name=? AND is_default=1",
                     (subject, body, trigger, is_def, name))
    # Insert standard roles if not exists
    standard_roles = [
        ('IT', 'Program Manager'), ('IT', 'Software Engineer'), ('IT', 'Senior Software Engineer'),
        ('IT', 'Tech Lead'), ('IT', 'Data Analyst'), ('IT', 'DevOps Engineer'),
        ('IT', 'QA Engineer'), ('IT', 'UI/UX Designer'), ('IT', 'System Administrator'),
        ('Manufacturing', 'Plant Head'), ('Manufacturing', 'Production Manager'),
        ('Manufacturing', 'Quality Manager'), ('Manufacturing', 'Supply Chain Manager'),
        ('Automobile', 'Design Engineer'), ('Automobile', 'Production Engineer'),
        ('Automobile', 'Quality Engineer'), ('Automobile', 'R&D Engineer'),
        ('Finance', 'Accountant'), ('Finance', 'Financial Analyst'),
        ('Finance', 'Tax Specialist'), ('Finance', 'Auditor'),
        ('HR', 'Recruiter'), ('HR', 'HR Manager'), ('HR', 'Training Coordinator'),
        ('Sales', 'Sales Manager'), ('Sales', 'Business Developer'),
        ('Sales', 'Account Manager'), ('Marketing', 'Marketing Manager'),
        ('Marketing', 'Content Writer'), ('Marketing', 'SEO Specialist'),
        ('Operations', 'Operations Manager'), ('Operations', 'Logistics Manager'),
        ('Legal', 'Legal Counsel'), ('Legal', 'Compliance Officer'),
    ]
    for domain, role in standard_roles:
        conn.execute("INSERT OR IGNORE INTO standard_roles (domain, role_name) VALUES (?,?)", (domain, role))
    # Insert standard skills if not exists
    standard_skills = [
        ('Python', 'Programming'), ('JavaScript', 'Programming'), ('Java', 'Programming'), ('C++', 'Programming'),
        ('React', 'Frontend'), ('Angular', 'Frontend'), ('Vue.js', 'Frontend'), ('HTML', 'Frontend'), ('CSS', 'Frontend'),
        ('Node.js', 'Backend'), ('Django', 'Backend'), ('Spring', 'Backend'), ('Flask', 'Backend'),
        ('SQL', 'Database'), ('MongoDB', 'Database'), ('PostgreSQL', 'Database'), ('MySQL', 'Database'),
        ('AWS', 'Cloud'), ('Azure', 'Cloud'), ('GCP', 'Cloud'), ('Docker', 'DevOps'), ('Kubernetes', 'DevOps'),
        ('Git', 'Tools'), ('Jira', 'Tools'), (' Jenkins', 'CI/CD'),
        ('Machine Learning', 'AI/ML'), ('Data Analysis', 'AI/ML'), ('NLP', 'AI/ML'),
        ('Excel', 'Office'), ('PowerPoint', 'Office'), ('Word', 'Office'),
        ('Communication', 'Soft Skills'), ('Teamwork', 'Soft Skills'), ('Leadership', 'Soft Skills'),
    ]
    for skill, category in standard_skills:
        conn.execute("INSERT OR IGNORE INTO standard_skills (skill_name, category) VALUES (?,?)", (skill, category))
    # Insert standard clients if not exists
    standard_clients = [
        'Diamanti', 'EY', 'KPMG', 'Taggd-Birla Opus', 'Taggd-HPE', 'Taggd-Hyundai',
        'Taggd-M&M', 'Taggd-Neosoft', 'Taggd-Pidilite', 'Taggd-Siemens', 'Taggd-TCPL', 'e-Zest'
    ]
    for client_name in standard_clients:
        conn.execute("INSERT OR IGNORE INTO clients (client_name) VALUES (?)", (client_name,))
    # Insert standard requirements if not exists
    standard_reqs = [
        ('Python Developer', 'Tech Corp', 'Open'),
        ('Java Developer', 'Tech Corp', 'Open'),
        ('Frontend Developer', 'Tech Corp', 'Open'),
        ('Full Stack Developer', 'ABC Corp', 'Open'),
        ('React Developer', 'XYZ Ltd', 'Open'),
        ('Node.js Developer', 'Tech Corp', 'Open'),
        ('Data Analyst', 'Data Inc', 'Open'),
        ('DevOps Engineer', 'Cloud Co', 'Open'),
        ('QA Engineer', 'Tech Corp', 'Open'),
        ('Project Manager', 'PM Solutions', 'Open'),
        ('UI/UX Designer', 'Design Studio', 'In Progress'),
        ('Machine Learning Engineer', 'AI Labs', 'Open'),
    ]
    default_checks = [
        "Technical Skills", "Years of Relevant Experience", "Within Given Budget",
        "Notice Period", "Location", "Non-Poachable Employee", "Updated CV"
    ]
    for title, client, status in standard_reqs:
        # Check if requirements already exists
        existing = conn.execute("SELECT id FROM requirements WHERE title=? AND client_name=?", (title, client)).fetchone()
        if not existing:
            cursor = conn.execute("INSERT INTO requirements (title, client_name, status) VALUES (?,?,?)", (title, client, status))
            rid = cursor.lastrowid
            # Add default checks for new requirement
            for i, check_name in enumerate(default_checks):
                conn.execute("""INSERT INTO requirement_checks 
                    (requirement_id,check_name,check_description,check_type,pass_criteria,sort_order)
                    VALUES (?,?,?,?,?,?)""",
                    (rid, check_name, "", "boolean", "Yes", i))
    conn.commit()
    conn.close()

# ── Auth ──────────────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            # API routes return 401, page routes redirect
            if request.path.startswith("/api/"):
                return jsonify({"error": "Unauthorised"}), 401
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated

def has_bulk_upload_access():
    if session.get("is_admin") or session.get("can_bulk_upload"):
        return True
    tid = session.get("team_member_id")
    if not tid:
        return False
    conn = get_db()
    row = conn.execute("SELECT can_bulk_upload, role FROM team_members WHERE id=?", (tid,)).fetchone()
    conn.close()
    allowed = bool(row and (row["can_bulk_upload"] or row["role"] == "Bulk Admin" or row["role"] == "Admin"))
    if allowed:
        session["can_bulk_upload"] = 1
    return allowed

@app.route("/login", methods=["GET"])
def login_page():
    if session.get("logged_in"):
        #return "THIS IS NEW LOGIN PAGE"
        return redirect(url_for("index"))
    return render_template("login.html")

@app.route("/login", methods=["POST"])
def login_password():
    username = (request.form.get("username") or "").strip().lower()
    password = request.form.get("password") or ""
    if not username or not password:
        return render_template("login.html", error="Enter username and password."), 400
    conn = get_db()
    user = conn.execute("SELECT * FROM app_users WHERE LOWER(username)=LOWER(?)", (username,)).fetchone()
    conn.close()
    if not user or not user["is_active"] or not check_password_hash(user["password"], password):
        return render_template("login.html", error="Invalid username or password."), 401
    set_login_session_from_app_user(user)
    return redirect(url_for("index"))

@app.route("/forgot_password", methods=["POST"])
def forgot_password():
    username = (request.form.get("username") or "").strip().lower()
    conn = get_db()
    ensure_auth_workflow_schema(conn)
    user = resolve_user_for_password_reset(conn, username)
    email_results = []
    if user and user["is_active"]:
        token = secrets.token_urlsafe(32)
        conn.execute(
            "INSERT INTO password_reset_tokens (user_id, token, expires_at) VALUES (?,?,datetime('now','+2 hours'))",
            (user["id"], token)
        )
        conn.commit()
        reset_url = url_for("reset_password", token=token, _external=True)
        recipients = reset_recipients_for_user(user)
        for recipient in recipients:
            try:
                result = system_email(
                    recipient,
                    "Reset your HR Guru ATS password",
                    f"Use this link to reset your password. It expires in 2 hours:\n\n{reset_url}"
                )
                email_results.append({"recipient": recipient, **result})
            except Exception as e:
                print("Password reset email error:", e)
                email_results.append({"recipient": recipient, "error": str(e)})
    conn.close()
    if user and user["is_active"] and not email_results:
        return render_template("login.html", message="Reset link was created, but no admin/user email is configured. Please set ADMIN_EMAIL or link the user to a team email.")
    if any(item.get("ok") for item in email_results):
        return render_template("login.html", message="Password reset email has been sent.")
    if email_results:
        return render_template("login.html", message="Reset link was created, but email sending failed. Please check GMAIL_USER and GMAIL_APP_PASS.")
    return render_template("login.html", message="If the account exists, a password reset email has been sent.")

@app.route("/reset_password/<token>", methods=["GET", "POST"])
def reset_password(token):
    if session.get("logged_in"):
        session.clear()
    conn = get_db()
    ensure_auth_workflow_schema(conn)
    row = conn.execute("""
        SELECT * FROM password_reset_tokens
        WHERE token=? AND used=0 AND datetime(expires_at) >= datetime('now')
    """, (token,)).fetchone()
    if not row:
        conn.close()
        return render_template("login.html", error="This reset link is invalid or expired."), 400
    if request.method == "POST":
        password = request.form.get("password") or ""
        if len(password) < 8:
            conn.close()
            return render_template("login.html", reset_token=token, error="Password must be at least 8 characters."), 400
        new_hash = hash_password(password)
        conn.execute("UPDATE app_users SET password=? WHERE id=?", (new_hash, row["user_id"]))
        conn.execute("UPDATE password_reset_tokens SET used=1 WHERE user_id=?", (row["user_id"],))
        conn.commit()
        updated = conn.execute("SELECT password FROM app_users WHERE id=?", (row["user_id"],)).fetchone()
        if not updated or not check_password_hash(updated["password"], password):
            conn.close()
            return render_template("login.html", reset_token=token, error="Password update could not be verified. Please try again."), 500
        conn.close()
        return render_template("login.html", message="Password reset successfully. You can sign in now.")
    conn.close()
    return render_template("login.html", reset_token=token)

@app.route("/register", methods=["POST"])
def register_user_request():
    name = (request.form.get("name") or "").strip()
    email = (request.form.get("email") or "").strip().lower()
    username = (request.form.get("username") or "").strip().lower()
    password = request.form.get("password") or ""
    if not email or not username or len(password) < 8:
        return render_template("login.html", register_error="Email, username, and an 8+ character password are required."), 400
    token = secrets.token_urlsafe(32)
    conn = get_db()
    ensure_auth_workflow_schema(conn)
    existing = conn.execute("SELECT id FROM app_users WHERE LOWER(username)=LOWER(?)", (username,)).fetchone()
    if existing:
        conn.close()
        return render_template("login.html", register_error="That username is already taken."), 400
    conn.execute(
        """INSERT INTO user_registration_requests
           (name, email, username, password_hash, token)
           VALUES (?,?,?,?,?)""",
        (name, email, username, hash_password(password), token)
    )
    conn.commit()
    conn.close()
    approve_url = url_for("approve_registration", token=token, _external=True)
    body = (
        f"New HR Guru ATS registration request:\n\n"
        f"Name: {name or '-'}\nEmail: {email}\nUsername: {username}\n\n"
        f"Admin approval link:\n{approve_url}\n\n"
        "You must be logged in as an admin to approve this request."
    )
    for admin_email in admin_approval_emails():
        try:
            system_email(admin_email, "Approve new HR Guru ATS user", body)
        except Exception as e:
            print("Registration approval email error:", e)
    return render_template("login.html", message="Registration request submitted. An admin must approve it before you can sign in.")

@app.route("/register/approve/<token>")
@login_required
def approve_registration(token):
    if not session.get("is_admin"):
        return "Admin approval required.", 403
    conn = get_db()
    ensure_auth_workflow_schema(conn)
    req = conn.execute("SELECT * FROM user_registration_requests WHERE token=? AND status='pending'", (token,)).fetchone()
    if not req:
        conn.close()
        return "Registration request is invalid or already processed.", 400
    member = conn.execute("SELECT id FROM team_members WHERE LOWER(email)=LOWER(?)", (req["email"],)).fetchone()
    team_member_id = member["id"] if member else None
    if not team_member_id:
        team_member_id = conn.execute(
            "INSERT INTO team_members (name,email,role) VALUES (?,?,?)",
            (req["name"] or req["username"], req["email"], "User")
        ).lastrowid
    conn.execute(
        """INSERT INTO app_users (username,password,team_member_id,is_admin,is_bulk_admin,is_active)
           VALUES (?,?,?,?,?,1)""",
        (req["username"], req["password_hash"], team_member_id, 0, 0)
    )
    conn.execute(
        "UPDATE user_registration_requests SET status='approved', approved_by=?, approved_at=datetime('now','localtime') WHERE id=?",
        (session.get("user_id"), req["id"])
    )
    conn.commit()
    conn.close()
    try:
        system_email(req["email"], "Your HR Guru ATS account is approved", "Your account has been approved. You can now sign in with the username and password you registered.")
    except Exception as e:
        print("Registration approved email error:", e)
    return redirect(url_for("index"))

@app.route("/login/google")
def login_google():
    public_base_url = (os.getenv("PUBLIC_BASE_URL") or "").strip().rstrip("/")
    request_host = (request.host or "").lower()
    is_local_request = (
        request_host.startswith("localhost")
        or request_host.startswith("127.0.0.1")
        or request_host.startswith("0.0.0.0")
    )
    if is_local_request:
        public_base_url = ""
    if not public_base_url:
        for header_name in ("X-Forwarded-Host", "Host"):
            host = (request.headers.get(header_name) or "").split(",")[0].strip()
            if host.lower().startswith("hireflow.hrgp.in"):
                public_base_url = "https://hireflow.hrgp.in"
                break
    if not public_base_url:
        for header_name in ("Referer", "Origin"):
            parsed = urlparse(request.headers.get(header_name) or "")
            if parsed.netloc.lower().startswith("hireflow.hrgp.in"):
                public_base_url = "https://hireflow.hrgp.in"
                break
    redirect_uri = (os.getenv("GOOGLE_REDIRECT_URI") or "").strip()
    if is_local_request:
        redirect_uri = ""
    if public_base_url and (
        not redirect_uri
        or "localhost" in redirect_uri
        or "127.0.0.1" in redirect_uri
        or redirect_uri.startswith("http://")
    ):
        redirect_uri = f"{public_base_url}/login/callback"
    if not redirect_uri:
        if public_base_url:
            redirect_uri = f"{public_base_url}/login/callback"
        elif request.host.lower().startswith("hireflow.hrgp.in"):
            redirect_uri = "https://hireflow.hrgp.in/login/callback"
        else:
            redirect_uri = url_for("google_callback", _external=True).replace("/auth/google/callback", "/login/callback")
    if redirect_uri.startswith("http://hireflow.hrgp.in"):
        redirect_uri = redirect_uri.replace("http://hireflow.hrgp.in", "https://hireflow.hrgp.in", 1)
    print("Google OAuth redirect_uri:", redirect_uri, flush=True)
    try:
        return google.authorize_redirect(redirect_uri, access_type="offline", prompt="consent")
    except Exception as e:
        print("Google login start error:", e, flush=True)
        traceback.print_exc()
        return render_template(
            "login.html",
            message="Google login could not start. Please check internet access from the ATS server and try again."
        ), 503

@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"ok": True})

@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory('uploads', filename, as_attachment=True)

@app.route("/login/callback")
@app.route("/auth/google/callback")
def google_callback():
    try:
        token = google.authorize_access_token()
        session["google_token"] = token

        #resp = google.get("userinfo")
        resp = google.get("https://www.googleapis.com/oauth2/v3/userinfo")
        user_info = resp.json()
    except Exception as e:
        print("Google login callback error:", e, flush=True)
        traceback.print_exc()
        return render_template(
            "login.html",
            message="Google login could not be completed. Please reconnect and try again."
        ), 503

    email = user_info["email"].lower()
    username = email.split("@")[0]
    full_name = user_info.get("name", username.title())

    print("Google userinfo:", email)
    

    conn = get_db()
    user = conn.execute(
    "SELECT * FROM team_members WHERE LOWER(TRIM(email)) = LOWER(TRIM(?))",
    (email,)
    ).fetchone()


    if not user:
        conn.close()
        return "Access not provisioned. Contact Admin.", 403
        
    session["logged_in"] = True
    session["user_id"] = user["id"]
    session["team_member_id"] = user["id"]   # important
    session["username"] = user["name"]
    session["recruiter_name"] = user["name"]
    session["email"] = user["email"]
    session["recruiter_email"] = user["email"]
    session["is_admin"] = 1 if user["role"] == "Admin" else 0
    session["can_bulk_upload"] = 1 if session["is_admin"] or user["can_bulk_upload"] or user["role"] == "Bulk Admin" else 0

    return redirect(url_for("index"))

# ── Column map ────────────────────────────────────────────────────────────────
COL_MAP = {
    "candidate name":"candidate_name","name":"candidate_name","candidate":"candidate_name",
    "email":"email_addr","email id":"email_addr","email address":"email_addr","email-id":"email_addr",
    "phone":"phone","mobile":"phone","contact":"phone","contact no":"phone","contact number":"phone",
    "current company":"current_company","company":"current_company","employer":"current_company",
    "organisation":"current_company","organization":"current_company","current org":"current_company",
    "current role":"current_role","designation":"current_role","title":"current_role",
    "position":"current_role","current designation":"current_role",
    "experience":"experience_years","exp":"experience_years","total exp":"experience_years",
    "years of exp":"experience_years","experience (years)":"experience_years",
    "yrs":"experience_years","total experience":"experience_years","total experience (in yrs)":"experience_years",
    "skills":"key_skills","key skills":"key_skills","skill set":"key_skills",
    "tech skills":"key_skills","technical skills":"key_skills",
    "notice":"notice_period","notice period":"notice_period","np":"notice_period",
    "current salary":"current_salary","current ctc":"current_salary","ctc":"current_salary","salary":"current_salary",
    "expected salary":"expected_salary","expected ctc":"expected_salary","ectc":"expected_salary","exp salary":"expected_salary",
    "location":"current_location","current location":"current_location","city":"current_location",
    "preferred location":"preferred_location","pref location":"preferred_location","preferred city":"preferred_location",
    "job location":"preferred_location",
    "remarks":"remarks","notes":"remarks","comments":"remarks","feedback":"remarks",
    "role":"role_name","job role":"role_name","applied for":"role_name",
    "position applied":"role_name","job title":"role_name","position name":"role_name",
    "education":"education",
}
ALL_FIELDS = ["candidate_name","email_addr","phone","current_company","current_role",
              "experience_years","key_skills","notice_period","current_salary",
              "expected_salary","current_location","preferred_location","remarks",
              "role_name","education"]

def norm_key(k): return COL_MAP.get(str(k).strip().lower(), None)
def empty_row(): return {f: "" for f in ALL_FIELDS}

# ── Parsers ───────────────────────────────────────────────────────────────────
def parse_xlsx(file_bytes, role_override=""):
    if not XLSX_OK: return [], "openpyxl not installed"
    wb   = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    rows = []
    for ws in wb.worksheets:
        data = list(ws.iter_rows(values_only=True))
        if not data: continue
        hdr_idx = 0
        for i, row in enumerate(data):
            if len([c for c in row if c is not None and str(c).strip()]) >= 3:
                hdr_idx = i; break
        headers = [norm_key(str(c).strip() if c else "") for c in data[hdr_idx]]
        for dr in data[hdr_idx+1:]:
            if not any(c is not None and str(c).strip() for c in dr): continue
            row = empty_row()
            for i, cell in enumerate(dr):
                if i >= len(headers): break
                key = headers[i]
                val = str(cell).strip() if cell is not None else ""
                if key and key in ALL_FIELDS: row[key] = val
            if role_override and not row["role_name"]: row["role_name"] = role_override
            rows.append(row)
    return rows, None

def parse_csv(file_bytes, role_override=""):
    rows   = []
    text   = file_bytes.decode("utf-8-sig", errors="replace")
    # Check for tabs first
    if '\t' in text:
        reader = csv.DictReader(io.StringIO(text), dialect=csv.excel_tab)
    else:
        reader = csv.DictReader(io.StringIO(text))
    for dr in reader:
        row = empty_row()
        for k, v in dr.items():
            key = norm_key(k)
            if key and key in ALL_FIELDS: row[key] = (v or "").strip()
        if role_override and not row["role_name"]: row["role_name"] = role_override
        rows.append(row)
    return rows, None

# ── CV upload ─────────────────────────────────────────────────────────────────
def upload_cv_cloud(file_storage, batch_id):
    fname = secure_filename(file_storage.filename)
    ext   = os.path.splitext(fname)[1].lower()
    if ext not in (".pdf", ".doc", ".docx"): return fname, None, None
    if not CLOUD_OK or not CLOUDINARY_URL:   return fname, None, None
    cloudinary.config(cloudinary_url=CLOUDINARY_URL)
    
    result = cloudinary.uploader.upload(
        file_storage, folder="hrguru_cvs",
        public_id=f"{batch_id}_{fname}",
        resource_type="raw", use_filename=True, unique_filename=False)
    return fname, result.get("secure_url"), result.get("public_id")

def upload_cv(file_storage, batch_id):
    fname = secure_filename(file_storage.filename)
    ext = os.path.splitext(fname)[1].lower()

    if ext not in (".pdf", ".doc", ".docx"):
        return fname, None, None

    upload_folder = os.path.join(app.root_path, "uploads")
    os.makedirs(upload_folder, exist_ok=True)

    saved_name = f"{batch_id}_{fname}"
    file_path = os.path.join(upload_folder, saved_name)

    file_storage.save(file_path)

    file_url = f"/uploads/{saved_name}"

    return fname, file_url, saved_name

def match_cv(name, cv_files):
    """
    Match a CV file to a candidate by name.

    Strategy:
    1. Strip the filename down to lowercase alpha-only characters (removes
       underscores, digits, dots, prefixes like 'Naukri', 'IR_' etc.)
       e.g. 'IR_RahulSPawar.pdf'   → 'irrahulspawar'
            'Naukri_DineshM4y.pdf' → 'naukridineshmy'
    2. Check if the candidate's FIRST NAME appears as a substring in that
       cleaned string — this is the primary signal.
    3. Bonus points if the last name also appears.
    4. Minimum requirement: first name must match (len >= 3).
    """
    if not cv_files: return None, None, None, False
    name_parts = (name or "").strip().split()
    if not name_parts: return None, None, None, False

    first_name = name_parts[0].lower()
    last_name  = name_parts[-1].lower() if len(name_parts) > 1 else ""

    # First name must be meaningful (skip very short names)
    if len(first_name) < 3: return None, None, None, False

    best_score, best = 0, (None, None, None)
    for orig, url, pub_id in cv_files:
        # Clean filename: keep only lowercase letters for substring search
        fname_alpha = re.sub(r'[^a-z]', '', orig.lower())
        fname_lower = orig.lower()   # also keep original for exact-word check

        score = 0

        # Primary: first name as substring in alpha-stripped filename
        if first_name in fname_alpha:
            score += 10
            # Bonus: last name also present
            if last_name and len(last_name) >= 3 and last_name in fname_alpha:
                score += 5

        # Fallback: check original filename with separators
        # e.g. 'rahul_pawar.pdf' — first name as a word token
        elif re.search(r'(?<![a-z])' + re.escape(first_name) + r'(?![a-z])', fname_lower):
            score += 8

        if score > best_score:
            best_score, best = score, (orig, url, pub_id)

    if best_score > 0: return *best, True
    return None, None, None, False

def unmatched_cvs(cv_files, matched_ids):
    return [(o, u, p) for o, u, p in cv_files if p not in matched_ids]

def parse_huggingface(text, api_key=""):
    """Use HuggingFace Inference API for NER-based CV parsing"""
    HF_API_URL = "https://api-inference.huggingface.co/models/dslim/bert-base-NER"
    
    headers = {"Authorization": f"Bearer {api_key}"} if api_key else {}
    
    result = {"candidate_name":"","current_company":"","current_role":"",
              "experience_years":"","email_addr":"","phone":"",
              "key_skills":"","current_location":""}
    
    try:
        response = requests.post(HF_API_URL, headers=headers, json={"inputs": text[:1500]}, timeout=30)
        if response.status_code == 200:
            entities = response.json()
            current_entity = None
            current_text = []
            
            for entity in entities:
                if isinstance(entity, dict):
                    word = entity.get("word", "")
                    entity_type = entity.get("entity_group", "")
                    
                    if entity_type == current_entity:
                        current_text.append(word)
                    else:
                        if current_entity == "PER" and current_text:
                            result["candidate_name"] = " ".join(current_text)
                        elif current_entity == "ORG" and current_text:
                            result["current_company"] = " ".join(current_text)
                        current_entity = entity_type
                        current_text = [word]
            
            if current_entity == "PER" and current_text:
                result["candidate_name"] = " ".join(current_text)
            elif current_entity == "ORG" and current_text:
                result["current_company"] = " ".join(current_text)
        
        if not api_key:
            result["_hf_limited"] = True
    except Exception as e:
        result["_hf_error"] = str(e)
        result["_hf_limited"] = True
    
    return result

def extract_cv_text(cv_path):
    ext = os.path.splitext(cv_path)[1].lower()
    if ext == ".txt":
        try:
            with open(cv_path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        except Exception:
            return ""
    if ext == ".docx":
        try:
            with zipfile.ZipFile(cv_path) as z:
                xml = z.read("word/document.xml")
            root = ET.fromstring(xml)
            ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
            lines = []
            for para in root.findall(".//w:p", ns):
                parts = [t.text for t in para.findall(".//w:t", ns) if t.text]
                line = re.sub(r"\s+", " ", "".join(parts)).strip()
                if line:
                    lines.append(line)
            return "\n".join(lines)
        except Exception:
            return ""
    if ext == ".doc":
        try:
            data = open(cv_path, "rb").read()
            chunks = []
            current = []
            for idx in range(0, len(data) - 1, 2):
                code = data[idx] + (data[idx + 1] << 8)
                if code in (9, 10, 13) or 32 <= code <= 126:
                    current.append(chr(code))
                else:
                    if len(current) >= 4:
                        chunks.append("".join(current))
                    current = []
            if len(current) >= 4:
                chunks.append("".join(current))
            lines = []
            seen = set()
            for chunk in chunks:
                for part in re.split(r"[\r\n]+", chunk):
                    line = re.sub(r"\s+", " ", part).strip()
                    if len(line) < 3:
                        continue
                    if not re.search(r"[A-Za-z]", line):
                        continue
                    key = line.lower()
                    if key in seen:
                        continue
                    seen.add(key)
                    lines.append(line)
            for idx, line in enumerate(lines):
                if line.strip().lower() in {"resume", "curriculum vitae", "cv"}:
                    lines = lines[idx:]
                    break
            return "\n".join(lines)
        except Exception:
            return ""
    try:
        import pymupdf
        doc = pymupdf.open(cv_path)
        text = ""
        for page in doc:
            blocks = page.get_text("blocks") or []
            if blocks:
                sorted_blocks = sorted(blocks, key=lambda b: (round(float(b[1]) / 4) * 4, float(b[0])))
                page_text = "\n".join(re.sub(r"\s*\n\s*", "\n", str(b[4] or "").strip()) for b in sorted_blocks if str(b[4] or "").strip())
                text += page_text + "\n"
            else:
                text += page.get_text() + "\n"
        doc.close()
        return text
    except Exception:
        return ""

def keyword_set(text):
    stop = {
        "and","or","the","a","an","to","for","of","in","on","with","by","as","is","are",
        "be","this","that","from","at","it","we","you","your","our","job","role","candidate",
        "experience","years","year","skills","skill","responsibilities","requirements"
    }
    words = re.findall(r"[a-zA-Z][a-zA-Z0-9+#.-]{2,}", (text or "").lower())
    return {w.strip(".-") for w in words if w not in stop and len(w.strip(".-")) > 2}

SKILL_ALIASES = {
    "ml": "Machine Learning",
    "machine learning": "Machine Learning",
    "ai": "Artificial Intelligence",
    "artificial intelligence": "Artificial Intelligence",
    "pyspark": "Apache Spark",
    "spark": "Apache Spark",
    "apache spark": "Apache Spark",
    "js": "JavaScript",
    "javascript": "JavaScript",
    "node": "Node.js",
    "node.js": "Node.js",
    "react": "React",
    "react.js": "React",
    "angular": "Angular",
    "vue": "Vue.js",
    "python": "Python",
    "java": "Java",
    "spring boot": "Spring Boot",
    "spring": "Spring",
    "sql": "SQL",
    "mysql": "MySQL",
    "postgresql": "PostgreSQL",
    "mongodb": "MongoDB",
    "aws": "AWS",
    "amazon web services": "AWS",
    "azure": "Microsoft Azure",
    "gcp": "Google Cloud Platform",
    "google cloud": "Google Cloud Platform",
    "docker": "Docker",
    "kubernetes": "Kubernetes",
    "k8s": "Kubernetes",
    "terraform": "Terraform",
    "jenkins": "Jenkins",
    "git": "Git",
    "jira": "Jira",
    "selenium": "Selenium",
    "cypress": "Cypress",
    "playwright": "Playwright",
    "rest api": "REST API",
    "restful": "REST API",
    "graphql": "GraphQL",
    "microservices": "Microservices",
    "linux": "Linux",
    "unix": "Unix",
    "tableau": "Tableau",
    "power bi": "Power BI",
    "excel": "Microsoft Excel",
    "tensorflow": "TensorFlow",
    "pytorch": "PyTorch",
    "scikit-learn": "Scikit-learn",
    "pandas": "Pandas",
    "numpy": "NumPy",
    "snowflake": "Snowflake",
    "databricks": "Databricks",
    "airflow": "Apache Airflow",
    "apache airflow": "Apache Airflow",
    "etl": "ETL",
    "ci/cd": "CI/CD",
    "devops": "DevOps",
    "html": "HTML",
    "css": "CSS",
}

ROLE_CATEGORIES = {
    "Data Engineering": ["data engineer", "etl", "big data", "spark", "snowflake", "databricks"],
    "Data Science": ["data scientist", "machine learning", "ml engineer", "ai engineer"],
    "Software Engineering": ["software engineer", "developer", "backend", "frontend", "full stack", "java", "python"],
    "Quality Assurance": ["qa", "quality", "test engineer", "automation testing", "selenium"],
    "DevOps": ["devops", "sre", "site reliability", "cloud engineer", "kubernetes"],
    "Product Management": ["product manager", "product owner"],
    "Business Analysis": ["business analyst", "system analyst"],
    "IT Support": ["it support", "desktop support", "linux admin", "system administrator"],
    "Cybersecurity": ["security", "cyber", "penetration", "soc", "iam"],
}

DOMAIN_HINTS = {
    "Banking / Financial Services": ["bank", "finance", "fintech", "payment", "trading", "loan", "insurance"],
    "Healthcare": ["healthcare", "clinical", "hospital", "patient", "medical"],
    "Retail / E-commerce": ["retail", "ecommerce", "e-commerce", "marketplace"],
    "Telecom": ["telecom", "network operator", "5g", "lte"],
    "SaaS": ["saas", "subscription", "multi-tenant"],
    "Manufacturing": ["manufacturing", "plant", "supply chain", "erp"],
}

def normalize_skill(skill):
    cleaned = re.sub(r"\s+", " ", str(skill or "").strip(" ,.;:|/\\")).lower()
    if not cleaned:
        return ""
    return SKILL_ALIASES.get(cleaned, cleaned.title())

def unique_list(items, limit=None):
    seen, out = set(), []
    for item in items or []:
        val = str(item or "").strip()
        if not val:
            continue
        key = val.lower()
        if key not in seen:
            seen.add(key)
            out.append(val)
    return out[:limit] if limit else out

def extract_known_skills(text):
    text_l = (text or "").lower()
    found = []
    for raw, normalized in sorted(SKILL_ALIASES.items(), key=lambda x: len(x[0]), reverse=True):
        pattern = r"(?<![a-z0-9])" + re.escape(raw.lower()) + r"(?![a-z0-9])"
        if re.search(pattern, text_l):
            found.append(normalized)
    return unique_list(found)

def infer_role_category(text):
    text_l = (text or "").lower()
    best = ("", 0)
    for category, hints in ROLE_CATEGORIES.items():
        score = sum(1 for hint in hints if hint in text_l)
        if score > best[1]:
            best = (category, score)
    return best[0]

def infer_domain(text):
    text_l = (text or "").lower()
    for domain, hints in DOMAIN_HINTS.items():
        if any(h in text_l for h in hints):
            return domain
    return ""

def extract_year_range(text):
    text_l = (text or "").lower()
    ranges = re.findall(r"(\d{1,2})\s*(?:-|to)\s*(\d{1,2})\s*\+?\s*(?:years?|yrs?)", text_l)
    if ranges:
        a, b = ranges[0]
        return {"min_years": int(a), "max_years": int(b)}
    singles = re.findall(r"(\d{1,2})\s*\+?\s*(?:years?|yrs?)", text_l)
    if singles:
        val = int(singles[0])
        return {"min_years": val, "max_years": 0}
    return {"min_years": 0, "max_years": 0}

def extract_role_title(text):
    lines = [re.sub(r"\s+", " ", l).strip() for l in (text or "").splitlines() if l.strip()]
    label_patterns = [
        r"(?:job\s*title|role|position|designation)\s*[:\-]\s*(.+)",
        r"(?:hiring\s+for|opening\s+for)\s*[:\-]?\s*(.+)",
    ]
    for line in lines[:30]:
        for pattern in label_patterns:
            m = re.search(pattern, line, re.I)
            if m:
                return clean_cv_value(m.group(1), 80).title()
    for line in lines[:12]:
        if len(line) <= 90 and infer_role_category(line):
            return clean_cv_value(line, 80).title()
    return ""

def extract_location(text):
    m = re.search(r"(?:location|job\s+location|work\s+location)\s*[:\-]\s*([A-Za-z0-9, /.-]{2,80})", text or "", re.I)
    if m:
        return clean_cv_value(m.group(1), 80)
    cities = ['Bangalore', 'Bengaluru', 'Mumbai', 'Delhi', 'Hyderabad', 'Chennai', 'Pune', 'Noida', 'Gurgaon', 'Gurugram', 'Kolkata', 'Ahmedabad', 'Remote']
    text_l = (text or "").lower()
    found = [c for c in cities if c.lower() in text_l]
    return ", ".join(unique_list(found))

def extract_certifications(text):
    certs = []
    patterns = [
        r"\bAWS Certified [A-Za-z ]+",
        r"\bAzure [A-Za-z ]+",
        r"\bPMP\b",
        r"\bCSM\b",
        r"\bCISSP\b",
        r"\bCEH\b",
        r"\bISTQB\b",
        r"\bCCNA\b",
        r"\bITIL\b",
    ]
    for pattern in patterns:
        certs.extend(re.findall(pattern, text or "", re.I))
    return unique_list([c.upper() if len(c) <= 5 else c.strip() for c in certs], 10)

def extract_responsibilities(text):
    lines = [re.sub(r"\s+", " ", l).strip(" -•*\t") for l in (text or "").splitlines()]
    signals = ["responsible", "develop", "design", "manage", "build", "implement", "lead", "support", "analyze", "maintain", "create", "work with"]
    picked = []
    for line in lines:
        if 20 <= len(line) <= 180 and any(sig in line.lower() for sig in signals):
            picked.append(line)
    return unique_list(picked, 8)

def split_must_nice_skills(text):
    all_skills = extract_known_skills(text)
    must, nice = [], []
    must_markers = ["must", "required", "mandatory", "should have", "need", "hands-on", "strong experience"]
    nice_markers = ["preferred", "nice to have", "good to have", "plus", "advantage", "optional"]
    sentences = re.split(r"[\n.;]", text or "")
    for sentence in sentences:
        skills = extract_known_skills(sentence)
        sent_l = sentence.lower()
        if any(m in sent_l for m in nice_markers):
            nice.extend(skills)
        elif any(m in sent_l for m in must_markers):
            must.extend(skills)
    if not must:
        must = all_skills[:8]
    for skill in all_skills:
        if skill not in must:
            nice.append(skill)
    return unique_list(must, 12), unique_list(nice, 12)

def parse_jd_requirements(jd_text):
    role_title = extract_role_title(jd_text)
    category = infer_role_category((role_title or "") + "\n" + (jd_text or ""))
    must, nice = split_must_nice_skills(jd_text)
    text_l = (jd_text or "").lower()
    employment = ""
    if "contract" in text_l:
        employment = "Contract"
    elif "part time" in text_l or "part-time" in text_l:
        employment = "Part-time"
    elif "full time" in text_l or "full-time" in text_l or "permanent" in text_l:
        employment = "Full-time"
    seniority = ""
    if re.search(r"\b(lead|principal|staff|architect|manager)\b", text_l):
        seniority = "Lead/Senior"
    elif re.search(r"\b(senior|sr\.?)\b", text_l):
        seniority = "Senior"
    elif re.search(r"\b(junior|fresher|entry)\b", text_l):
        seniority = "Junior"
    return {
        "role_title": role_title,
        "role_category": category,
        "employment_type": employment,
        "experience_required": extract_year_range(jd_text),
        "must_have_skills": must,
        "nice_to_have_skills": nice,
        "tools_technologies": unique_list(extract_known_skills(jd_text), 18),
        "domain": infer_domain(jd_text),
        "sub_domain": "",
        "responsibilities": extract_responsibilities(jd_text),
        "seniority_level": seniority,
        "certifications_required": extract_certifications(jd_text),
        "location": extract_location(jd_text),
        "keywords_expanded": unique_list(must + nice + ([category] if category else []), 25)
    }

def parse_experience_years(value):
    if isinstance(value, (int, float)):
        return float(value)
    m = re.search(r"(\d+(?:\.\d+)?)", str(value or ""))
    return float(m.group(1)) if m else 0

def extract_education_items(text):
    items = re.findall(r"\b(?:B\.?Tech|B\.?E\.?|B\.?Sc|M\.?Tech|M\.?Sc|MBA|MCA|BCA|Ph\.?D|Bachelor(?:'s)?|Master(?:'s)?)\b", text or "", re.I)
    return unique_list([i.upper().replace("BACHELOR'S", "Bachelor").replace("MASTER'S", "Master") for i in items], 8)

def parse_candidate_profile(cv_text, parsed_cv=None):
    parsed = parsed_cv or {}
    skills = extract_known_skills(cv_text)
    parsed_skills = [normalize_skill(s) for s in re.split(r"[,;/|]", parsed.get("key_skills", "")) if s.strip()]
    primary = unique_list(parsed_skills + skills, 15)
    current_role = parsed.get("current_role", "")
    role_history = []
    if current_role or parsed.get("current_company"):
        role_history.append({
            "title": current_role,
            "company": parsed.get("current_company", ""),
            "duration_years": parse_experience_years(parsed.get("experience_years", "")),
            "responsibilities": extract_responsibilities(cv_text)[:5],
            "skills_used": primary[:10]
        })
    total_exp = parse_experience_years(parsed.get("experience_years", ""))
    stability = 60 if role_history else 0
    if total_exp >= 3 and role_history:
        stability = 75
    red_flags = []
    if not current_role:
        red_flags.append("Current role not clearly stated")
    if not total_exp:
        red_flags.append("Total experience not clearly stated")
    return {
        "candidate_name": parsed.get("candidate_name", ""),
        "total_experience_years": total_exp,
        "current_role": current_role,
        "role_history": role_history,
        "primary_skills": primary[:10],
        "secondary_skills": primary[10:20],
        "tools_technologies": primary[:18],
        "domain_experience": unique_list([infer_domain(cv_text)] if infer_domain(cv_text) else []),
        "certifications": extract_certifications(cv_text),
        "education": unique_list([parsed.get("education", "")] + extract_education_items(cv_text), 8),
        "location": parsed.get("current_location", "") or extract_location(cv_text),
        "career_stability_score": stability,
        "red_flags": red_flags,
        "project_complexity_indicators": extract_project_complexity(cv_text),
        "ownership_signals": extract_ownership_signals(cv_text)
    }

def extract_project_complexity(text):
    indicators = []
    hints = ["microservices", "distributed", "large scale", "high availability", "migration", "architecture", "pipeline", "automation", "performance", "cloud"]
    for hint in hints:
        if hint in (text or "").lower():
            indicators.append(hint.title())
    return unique_list(indicators, 8)

def extract_ownership_signals(text):
    signals = []
    for phrase in ["led", "owned", "managed", "designed", "architected", "delivered", "mentored", "implemented"]:
        if re.search(r"\b" + re.escape(phrase) + r"\b", text or "", re.I):
            signals.append(phrase.title())
    return unique_list(signals, 8)

def skill_match_score(required, available):
    available_norm = {normalize_skill(s).lower() for s in available or []}
    matched, missing = [], []
    for skill in required or []:
        normalized = normalize_skill(skill)
        key = normalized.lower()
        if key in available_norm:
            matched.append(normalized)
        else:
            missing.append(normalized)
    return unique_list(matched), unique_list(missing)

def score_candidate_fit(jd, cv):
    required = jd.get("must_have_skills", [])
    available = (cv.get("primary_skills", []) or []) + (cv.get("secondary_skills", []) or []) + (cv.get("tools_technologies", []) or [])
    matched, missing = skill_match_score(required, available)
    must_score = round((len(matched) / len(required)) * 100) if required else 0
    jd_category = (jd.get("role_category") or "").lower()
    cv_role_text = (cv.get("current_role") or "") + " " + " ".join(cv.get("primary_skills", []) or [])
    cv_category = infer_role_category(cv_role_text).lower()
    role_score = 85 if jd_category and jd_category == cv_category else 35 if jd_category and cv_category else 50
    req_exp = jd.get("experience_required") or {}
    min_exp, max_exp = req_exp.get("min_years", 0) or 0, req_exp.get("max_years", 0) or 0
    cand_exp = cv.get("total_experience_years", 0) or 0
    if min_exp and cand_exp < min_exp:
        exp_score = max(0, round((cand_exp / min_exp) * 70))
    elif max_exp and cand_exp > max_exp + 3:
        exp_score = 70
    elif min_exp or max_exp:
        exp_score = 90
    else:
        exp_score = 50
    jd_domain = (jd.get("domain") or "").lower()
    cv_domains = [str(d).lower() for d in cv.get("domain_experience", []) or []]
    domain_score = 85 if jd_domain and jd_domain in cv_domains else 50 if not jd_domain else 25
    nice = jd.get("nice_to_have_skills", []) or []
    nice_matched, _ = skill_match_score(nice, available)
    secondary_score = round((len(nice_matched) / len(nice)) * 100) if nice else 50
    stability_score = int(cv.get("career_stability_score", 0) or 0)
    context_score = 70 if cv.get("project_complexity_indicators") or cv.get("ownership_signals") else 45
    penalties = []
    if missing:
        penalties.append({"reason": "Missing must-have skills: " + ", ".join(missing[:6]), "impact": min(45, len(missing) * 12)})
    if jd_category and cv_category and jd_category != cv_category:
        penalties.append({"reason": "Role category mismatch", "impact": 30})
    if min_exp and cand_exp < min_exp:
        penalties.append({"reason": f"Experience below required minimum of {min_exp} years", "impact": 20})
    if cv.get("red_flags"):
        penalties.append({"reason": "CV has missing or unclear core details", "impact": 10})
    weighted = (
        must_score * 0.34 + role_score * 0.20 + exp_score * 0.15 + domain_score * 0.10 +
        secondary_score * 0.08 + stability_score * 0.06 + context_score * 0.07
    )
    final_score = max(0, min(100, round(weighted - sum(p["impact"] for p in penalties) * 0.35)))
    if final_score >= 80:
        verdict = "Strong Match"
    elif final_score >= 65:
        verdict = "Moderate Match"
    elif final_score >= 45:
        verdict = "Weak Match"
    else:
        verdict = "Reject / Not Recommended"
    strengths = []
    if matched:
        strengths.append("Matches must-have skills: " + ", ".join(matched[:6]))
    if cv.get("ownership_signals"):
        strengths.append("Shows ownership signals: " + ", ".join(cv.get("ownership_signals", [])[:4]))
    concerns = []
    if missing:
        concerns.append("Missing must-have skills: " + ", ".join(missing[:6]))
    if role_score < 50:
        concerns.append("Current role does not align well with the JD role category")
    if exp_score < 70:
        concerns.append("Experience is below the JD requirement")
    return {
        "final_score": final_score,
        "verdict": verdict,
        "score_breakdown": {
            "must_have_skills": must_score,
            "role_relevance": role_score,
            "experience_fit": exp_score,
            "domain_fit": domain_score,
            "secondary_skills": secondary_score,
            "stability": stability_score,
            "contextual_intelligence": context_score
        },
        "matched_must_have_skills": matched,
        "missing_must_have_skills": missing,
        "strengths": strengths[:8],
        "concerns": concerns[:8],
        "red_flags": cv.get("red_flags", []),
        "penalties_applied": penalties,
        "explanation_summary": f"Candidate scored {final_score}/100. {verdict}. " + ("; ".join(concerns[:2]) if concerns else "Core hiring signals are reasonably aligned.")
    }

def build_jd_match_response(jd_json, cv_json, score_json, mode):
    score = int(score_json.get("final_score", 0) or 0)
    strengths = score_json.get("strengths", []) or []
    gaps = (score_json.get("concerns", []) or []) + [p.get("reason", "") for p in score_json.get("penalties_applied", []) if p.get("reason")]
    return {
        "score": score,
        "verdict": score_json.get("verdict", ""),
        "summary": score_json.get("explanation_summary", ""),
        "strengths": strengths[:8],
        "gaps": unique_list(gaps, 8),
        "mode": mode,
        "jd_json": jd_json,
        "cv_json": cv_json,
        "score_json": score_json
    }

def heuristic_jd_match(jd_text, cv_text, parsed_cv=None):
    jd_json = parse_jd_requirements(jd_text)
    cv_json = parse_candidate_profile(cv_text, parsed_cv)
    score_json = score_candidate_fit(jd_json, cv_json)
    return build_jd_match_response(jd_json, cv_json, score_json, "structured_keyword")

def build_candidate_summary_from_cv(cv_path, cv_text=""):
    parsed = {}
    try:
        parsed = parse_cv(cv_path) or {}
    except Exception as e:
        print("CV summary parse failed:", e)
    skills = parsed.get("key_skills", "")
    if isinstance(skills, str):
        skills_list = [s.strip() for s in skills.split(",") if s.strip()]
    else:
        skills_list = skills or []
    relevant = parsed.get("cv_summary", "")
    if not relevant and cv_text:
        lines = [re.sub(r"\s+", " ", line).strip() for line in cv_text.splitlines() if line.strip()]
        relevant = " ".join(lines[:4])[:700]
    return {
        "current_designation": parsed.get("current_role", ""),
        "key_skills": skills_list[:18],
        "total_experience": parsed.get("experience_years", ""),
        "education": parsed.get("education", ""),
        "current_company": parsed.get("current_company", ""),
        "relevant_details": relevant
    }

def ai_jd_match(jd_text, cv_text, candidate_name=""):
    # Deprecated: scoring is now owned by deterministic Python modules.
    # LLM extraction is handled by ats_pipeline.ollama_extract and never returns scores.
    return None

def save_uploaded_analysis_file(file_storage, prefix):
    if not file_storage or not file_storage.filename:
        raise ValueError("Missing file")
    ext = os.path.splitext(file_storage.filename)[1].lower()
    if ext not in [".pdf", ".doc", ".docx", ".txt"]:
        raise ValueError("Upload PDF, DOC, DOCX, or TXT files only")
    upload_folder = os.path.join(app.root_path, "uploads")
    os.makedirs(upload_folder, exist_ok=True)
    safe_name = secure_filename(file_storage.filename)
    saved_name = f"{prefix}_{int(datetime.now().timestamp())}_{safe_name}"
    path = os.path.join(upload_folder, saved_name)
    file_storage.save(path)
    return path, safe_name

def get_cached_embedding(text_hash_value, model):
    conn = get_db()
    ensure_ats_pipeline_schema(conn)
    row = conn.execute(
        "SELECT embedding_json FROM embedding_cache WHERE text_hash=? AND model=?",
        (text_hash_value, model)
    ).fetchone()
    conn.close()
    return deserialize_embedding(row["embedding_json"]) if row else None

def set_cached_embedding(text_hash_value, model, embedding):
    conn = get_db()
    ensure_ats_pipeline_schema(conn)
    conn.execute(
        "INSERT OR REPLACE INTO embedding_cache (text_hash, model, embedding_json) VALUES (?,?,?)",
        (text_hash_value, model, serialize_embedding(embedding))
    )
    conn.commit()
    conn.close()

def get_cached_match_result(jd_hash, resume_hash):
    conn = get_db()
    ensure_ats_pipeline_schema(conn)
    row = conn.execute(
        "SELECT result_json FROM match_results WHERE jd_hash=? AND resume_hash=?",
        (jd_hash, resume_hash)
    ).fetchone()
    conn.close()
    if not row:
        return None
    try:
        return json.loads(row["result_json"])
    except Exception:
        return None

def get_cached_parsed_resume(resume_hash):
    conn = get_db()
    ensure_ats_pipeline_schema(conn)
    row = conn.execute(
        "SELECT parsed_json FROM parsed_resume_cache WHERE resume_hash=?",
        (resume_hash,)
    ).fetchone()
    conn.close()
    if not row:
        return None
    try:
        parsed = json.loads(row["parsed_json"] or "{}")
        return parsed if isinstance(parsed, dict) else None
    except Exception:
        return None

def set_cached_parsed_resume(resume_hash, parsed_candidate):
    if not parsed_candidate:
        return
    conn = get_db()
    ensure_ats_pipeline_schema(conn)
    conn.execute(
        """INSERT OR REPLACE INTO parsed_resume_cache
           (resume_hash, parsed_json, updated_at)
           VALUES (?,?,datetime('now','localtime'))""",
        (resume_hash, json.dumps(parsed_candidate))
    )
    conn.commit()
    conn.close()

def get_cached_parsed_jd(jd_hash):
    conn = get_db()
    ensure_ats_pipeline_schema(conn)
    row = conn.execute(
        "SELECT parsed_json FROM jd_requirements WHERE jd_hash=?",
        (jd_hash,)
    ).fetchone()
    conn.close()
    if not row:
        return None
    try:
        parsed = json.loads(row["parsed_json"] or "{}")
        return parsed if isinstance(parsed, dict) else None
    except Exception:
        return None

def set_cached_parsed_jd(jd_hash, parsed_jd):
    if not parsed_jd:
        return
    conn = get_db()
    ensure_ats_pipeline_schema(conn)
    conn.execute(
        """INSERT OR REPLACE INTO jd_requirements
           (jd_hash, role_title, parsed_json, updated_at)
           VALUES (?,?,?,datetime('now','localtime'))""",
        (jd_hash, parsed_jd.get("role_title", ""), json.dumps(parsed_jd))
    )
    conn.commit()
    conn.close()

def persist_skill_aliases(conn, canonical_value):
    canonical_value = canonical_skill(canonical_value)
    if not canonical_value:
        return
    cursor = conn.execute(
        "INSERT OR IGNORE INTO skills (raw_value, canonical_value) VALUES (?,?)",
        (canonical_value, canonical_value)
    )
    row = conn.execute("SELECT id FROM skills WHERE canonical_value=?", (canonical_value,)).fetchone()
    if not row:
        return
    skill_id = row["id"]
    aliases = skill_aliases_for(canonical_value)
    conn.execute(
        "INSERT OR REPLACE INTO normalized_skill_cache (raw_value, canonical_value, aliases_json) VALUES (?,?,?)",
        (canonical_value.lower(), canonical_value, json.dumps(aliases))
    )
    for alias in aliases:
        conn.execute(
            "INSERT OR IGNORE INTO skill_aliases (skill_id, alias) VALUES (?,?)",
            (skill_id, alias)
        )

def persist_match_artifacts(jd_hash, resume_hash, result):
    conn = get_db()
    ensure_ats_pipeline_schema(conn)
    parsed_jd = result.get("parsed_jd") or result.get("jd_json") or {}
    parsed_candidate = result.get("parsed_candidate") or result.get("cv_json") or {}
    conn.execute(
        """INSERT OR REPLACE INTO jd_requirements
           (jd_hash, role_title, parsed_json, updated_at)
           VALUES (?,?,?,datetime('now','localtime'))""",
        (jd_hash, parsed_jd.get("role_title", ""), json.dumps(parsed_jd))
    )
    conn.execute(
        """INSERT OR REPLACE INTO parsed_resume_cache
           (resume_hash, parsed_json, updated_at)
           VALUES (?,?,datetime('now','localtime'))""",
        (resume_hash, json.dumps(parsed_candidate))
    )
    conn.execute("DELETE FROM match_results WHERE jd_hash=? AND resume_hash=?", (jd_hash, resume_hash))
    conn.execute(
        """INSERT INTO match_results
           (jd_hash, resume_hash, final_score, structured_score, semantic_score, hard_filter_score, result_json)
           VALUES (?,?,?,?,?,?,?)""",
        (
            jd_hash,
            resume_hash,
            int(result.get("final_score", 0) or 0),
            int(result.get("structured_score", 0) or 0),
            int(result.get("semantic_score", 0) or 0),
            int(result.get("hard_filter_score", 0) or 0),
            json.dumps(result)
        )
    )
    conn.execute("DELETE FROM candidate_skills WHERE resume_hash=?", (resume_hash,))
    for item in parsed_candidate.get("skill_confidence_scores", []) or []:
        canonical = canonical_skill(item.get("skill"))
        persist_skill_aliases(conn, canonical)
        skill_type = "production" if canonical in (parsed_candidate.get("production_skills") or []) else "exposure"
        conn.execute(
            """INSERT INTO candidate_skills
               (resume_hash, raw_value, canonical_value, confidence, skill_type)
               VALUES (?,?,?,?,?)""",
            (resume_hash, item.get("skill", ""), canonical, float(item.get("confidence", 0) or 0), skill_type)
        )
    conn.execute("DELETE FROM candidate_roles WHERE resume_hash=?", (resume_hash,))
    for role in parsed_candidate.get("normalized_roles", []) or []:
        conn.execute(
            """INSERT INTO candidate_roles (resume_hash, raw_value, canonical_value, confidence)
               VALUES (?,?,?,?)""",
            (resume_hash, role, role, 0.8)
        )
    conn.execute("DELETE FROM candidate_domains WHERE resume_hash=?", (resume_hash,))
    for item in parsed_candidate.get("domain_confidence_scores", []) or []:
        conn.execute(
            """INSERT INTO candidate_domains (resume_hash, raw_value, canonical_value, confidence)
               VALUES (?,?,?,?)""",
            (resume_hash, item.get("domain", ""), item.get("domain", ""), float(item.get("confidence", 0) or 0))
        )
    conn.commit()
    conn.close()

def clean_cv_value(value, limit=80):
    value = re.sub(r"\s+", " ", value or "").strip(" -:|,\t\r\n")
    value = re.sub(r"\b(email|phone|mobile|contact|location|address)\b.*$", "", value, flags=re.I).strip(" -:|,")
    return value[:limit]

EMAIL_DOMAIN_WORDS = {
    "gmail", "yahoo", "hotmail", "outlook", "icloud", "proton", "zoho", "rediffmail",
    "live", "msn", "aol", "mail", "com", "co", "in", "net", "org", "edu"
}

def normalize_person_name(value):
    raw = str(value or "").strip()
    if not raw:
        return ""
    email_match = re.search(r"[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}", raw)
    if email_match and raw.strip() == email_match.group(0):
        raw = email_match.group(0).split("@", 1)[0]
    raw = re.sub(r"^\s*(?:name|candidate\s+name|full\s+name)\s*[:\-]\s*", "", raw, flags=re.I)
    raw = re.sub(r"\b(?:gmail|yahoo|hotmail|outlook|icloud|proton|zoho|rediffmail|live|msn|aol|mail)\s*(?:\.|\s+)?(?:com|co\.in|in|net|org|edu)\b.*$", "", raw, flags=re.I)
    raw = re.sub(r"[@_]+", " ", raw)
    raw = re.sub(r"\.+", " ", raw)
    raw = re.sub(r"\b(?:email|e-mail|mail|phone|mobile|contact)\b.*$", "", raw, flags=re.I)
    raw = re.sub(r"[^A-Za-z'\-\s]", " ", raw)
    words = [w for w in re.findall(r"[A-Za-z][A-Za-z'-]*", raw) if w.lower() not in EMAIL_DOMAIN_WORDS]
    if len(words) < 2 or len(words) > 4:
        return ""
    return " ".join(words).title()

def looks_like_person_name(line):
    if not line or any(ch in line for ch in "@:/\\|()[]{}"):
        return False
    lower = line.lower()
    blocked = [
        "resume", "curriculum", "vitae", "profile", "summary", "experience", "education",
        "skill", "skills", "core skill", "developer", "engineer", "manager", "analyst", "consultant", "specialist",
        "email", "phone", "mobile", "contact", "address", "linkedin", "github",
        "gmail", "yahoo", "hotmail", "outlook", "icloud", "proton", "zoho", "rediffmail"
    ]
    if any(word in lower for word in blocked):
        return False
    words = re.findall(r"[A-Za-z][A-Za-z'-]*", line.replace(".", " "))
    if len(words) < 2 or len(words) > 4:
        return False
    if any(w.lower() in EMAIL_DOMAIN_WORDS for w in words):
        return False
    return sum(1 for w in words if w[:1].isupper()) >= min(2, len(words))

def clean_company_name(value):
    value = clean_cv_value(value, 70)
    value = re.sub(r"\s+", " ", value).strip(" -:|,\t\r\n")
    if not value:
        return ""
    lower = value.lower()
    blocked = {
        "that", "would", "utilise", "organization", "organisation", "professional",
        "technology", "driven", "career", "objective", "summary"
    }
    words = re.findall(r"[A-Za-z0-9&.'-]+", value)
    if not words or words[0].lower() in blocked:
        return ""
    if len(words) == 1 and words[0].lower() in blocked:
        return ""
    return " ".join(words[:8]).title()

def clean_company_display(value):
    value = re.sub(r"\s*\([^)]*\).*$", "", str(value or ""))
    value = re.sub(r"(?i)eicher\s*ve", "Eicher VE", value)
    company = clean_company_name(value)
    company = re.sub(r"\bVe\b", "VE", company)
    company = re.sub(r"\bPvt\b", "Pvt", company)
    company = re.sub(r"\bLtd\b", "Ltd", company)
    return company

def split_company_role_line(line):
    value = clean_cv_value(line, 120)
    value = re.sub(r"^\s*\d+[\).]?\s*", "", value).strip()
    match = re.match(r"(.+?)\s+[–—-]\s+(.+)$", value)
    if not match:
        return "", ""
    left, right = match.group(1), match.group(2)
    date_words = r"\b(current|present|20\d{2}|19\d{2}|jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec|january|february|march|april|june|july|august|september|october|november|december)\b"
    if re.search(date_words, left, re.I) or re.search(r"^\s*(current|present|till\s+date|to\s+date)\s*$", right, re.I):
        return "", ""
    company = clean_company_display(left)
    role = clean_role_title(right)
    return company, role

def split_company_date_line(line):
    value = clean_cv_value(line, 120)
    match = re.match(
        r"(.+?)\s*[–—-]\s*(?:\(?[A-Za-z ]+\)?\s*)?(?:(?:jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*\s+)?(?:19|20)\d{2}\b.*(?:current|present|till\s+working|till\s+date|to\s+date|working)?",
        value,
        re.I
    )
    if not match:
        return ""
    return clean_company_display(match.group(1))

def infer_presently_working_company_role(lines):
    text = "\n".join(lines[:100])
    flat = re.sub(r"\s+", " ", text)
    match = re.search(
        r"presently\s+working.*?(customer\s+service\s+manager\s*&?\s*senior\s+eos\s+executive).*?\bin\s+(.+?)(?:\s+for\s+|\s+from\s+|\s+my\s+\d|\n|$)",
        flat,
        re.I
    )
    if match:
        return clean_company_display(match.group(2)), clean_role_title(match.group(1))
    match = re.search(r"presently\s+working.*?\bin\s+([A-Za-z0-9&()., \-]+?\b(?:ltd|limited|pvt|private|motors|vehicles)\b[^.,\n]*)", flat, re.I)
    if match:
        return clean_company_display(match.group(1)), ""
    return "", ""

def looks_like_role_title_line(line):
    role = clean_role_title(line)
    if not role:
        return False
    return bool(re.search(r"\b(engineer|developer|manager|analyst|consultant|specialist|associate|administrator|officer|executive|lead|architect|designer|director|head|service|support|quality|technical)\b", role, re.I))

def infer_current_company_role_from_experience(lines):
    if not lines:
        return "", ""
    company, role = infer_presently_working_company_role(lines)
    if company:
        return company, role
    heading_idx = -1
    for idx, line in enumerate(lines[:100]):
        lower = line.lower()
        if any(key in lower for key in ["work experience", "professional experience", "employment history", "career history", "employment"]):
            heading_idx = idx
            break
    if heading_idx < 0:
        for idx in range(0, min(len(lines) - 1, 80)):
            role_line = lines[idx]
            company = split_company_date_line(lines[idx + 1])
            nearby = " ".join(lines[idx + 1:idx + 4]).lower()
            if company and looks_like_role_title_line(role_line) and re.search(r"\b(current|present|till\s+working|till\s+date|to\s+date|working)\b", nearby):
                return company, clean_role_title(role_line)
        return "", ""
    stop_words = {"education", "academic", "skills", "certification", "certifications", "projects", "personal details"}
    for idx in range(heading_idx + 1, min(len(lines), heading_idx + 50)):
        lower = lines[idx].strip().lower()
        if lower in stop_words:
            break
        company, role = split_company_role_line(lines[idx])
        if not company:
            continue
        nearby = " ".join(lines[idx:idx + 4]).lower()
        if re.search(r"\b(current|present|till\s+date|to\s+date)\b", nearby):
            return company, role
    return "", ""

def infer_current_company_from_experience(lines):
    if not lines:
        return ""
    split_company, _ = infer_current_company_role_from_experience(lines)
    if split_company:
        return split_company
    heading_idx = -1
    for idx, line in enumerate(lines[:80]):
        lower = line.lower()
        if any(key in lower for key in ["work experience", "professional experience", "employment history", "career history", "employment"]):
            heading_idx = idx
            break
    if heading_idx < 0:
        return ""
    stop_words = {
        "career objective", "professional synopsis", "education", "academic", "skills",
        "certification", "certifications", "projects", "personal details"
    }
    role_words = {
        "associate", "consultant", "analyst", "engineer", "developer", "manager",
        "accountant", "auditor", "executive", "lead", "architect", "specialist"
    }
    for idx in range(heading_idx + 1, min(len(lines), heading_idx + 45)):
        line = clean_cv_value(lines[idx], 90)
        lower = line.lower()
        if not line:
            continue
        if lower in stop_words:
            break
        if re.search(r"\b(current|present|till\s+date|to\s+date)\b", lower):
            for back in range(idx - 1, heading_idx, -1):
                candidate = clean_cv_value(lines[back], 90)
                cand_lower = candidate.lower()
                if not candidate or re.search(r"\b(current|present|20\d{2}|19\d{2}|jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b", cand_lower):
                    continue
                if any(word in cand_lower for word in role_words):
                    continue
                if len(candidate.split()) <= 6:
                    cleaned = clean_company_name(candidate)
                    if cleaned:
                        return cleaned
    return ""

def clean_role_title(value):
    value = clean_cv_value(value, 90)
    value = re.sub(r"\s*&\s*", " & ", value)
    value = re.sub(r"\s+", " ", value).strip(" -:|,\t\r\n")
    if not value:
        return ""
    lower = value.lower()
    blocked_fragments = [
        "professional and technology", "career objective", "would utilise", "growth of the organisation",
        "challenging position", "looking for", "seeking", "aspiring"
    ]
    if any(fragment in lower for fragment in blocked_fragments):
        return ""
    words = re.findall(r"[A-Za-z0-9&/.'+-]+", value)
    if len(words) < 1 or len(words) > 9:
        return ""
    role = " ".join(words).title()
    role = re.sub(r"\bEos\b", "EOS", role)
    return role

def infer_current_role_from_experience(lines):
    if not lines:
        return ""
    _, split_role = infer_current_company_role_from_experience(lines)
    if split_role:
        return split_role
    heading_idx = -1
    for idx, line in enumerate(lines[:80]):
        lower = line.lower()
        if any(key in lower for key in ["work experience", "professional experience", "employment history", "career history", "employment"]):
            heading_idx = idx
            break
    if heading_idx < 0:
        return ""
    role_words = {
        "associate", "consultant", "analyst", "engineer", "developer", "manager",
        "accountant", "auditor", "executive", "lead", "architect", "specialist",
        "administrator", "director", "officer"
    }
    for idx in range(heading_idx + 1, min(len(lines), heading_idx + 45)):
        line = clean_cv_value(lines[idx], 90)
        lower = line.lower()
        if re.search(r"\b(current|present|till\s+date|to\s+date)\b", lower):
            for forward in range(idx + 1, min(len(lines), idx + 5)):
                candidate = clean_cv_value(lines[forward], 90).strip(",")
                cand_lower = candidate.lower()
                if not candidate:
                    continue
                if re.search(r"\b(20\d{2}|19\d{2}|jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec|current|present)\b", cand_lower):
                    continue
                if any(word in cand_lower for word in role_words):
                    cleaned = clean_role_title(candidate)
                    if cleaned:
                        return cleaned
            for back in range(idx - 1, heading_idx, -1):
                candidate = clean_cv_value(lines[back], 90).strip(",")
                cand_lower = candidate.lower()
                if any(word in cand_lower for word in role_words):
                    cleaned = clean_role_title(candidate)
                    if cleaned:
                        return cleaned
    return ""

def clean_header_role(value):
    role = clean_role_title(value)
    if not role:
        return ""
    if not re.search(r"\b(engineer|developer|manager|analyst|consultant|specialist|associate|administrator|officer|executive|lead|architect|designer|director|head|service)\b", role, re.I):
        return ""
    return role

def extract_pdf_header_fields(cv_path):
    fields = {}
    if os.path.splitext(cv_path or "")[1].lower() != ".pdf":
        return fields
    try:
        import pymupdf
        doc = pymupdf.open(cv_path)
        if not doc:
            return fields
        page = doc[0]
        blocks = []
        for block in page.get_text("blocks") or []:
            x0, y0, x1, y1, text = block[:5]
            if y0 <= min(page.rect.height * 0.28, 180) and str(text or "").strip():
                for line in str(text or "").splitlines():
                    cleaned = re.sub(r"\s+", " ", line).strip()
                    if cleaned:
                        blocks.append((float(y0), float(x0), cleaned))
        doc.close()
    except Exception:
        return fields
    lines = [item[2] for item in sorted(blocks, key=lambda b: (round(b[0] / 4) * 4, b[1]))]
    header_text = "\n".join(lines)
    email_match = re.search(r"(?<![\w.-])[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}(?![\w.-])", header_text)
    if email_match:
        fields["email_addr"] = email_match.group(0).strip().lower()
    phone_match = re.search(r"(?:\+?91[\s\-]?)?[6-9]\d{4}[\s\-]?\d{5}|\+?\d[\d\s\-]{9,}\d", header_text)
    if phone_match:
        phone = re.sub(r"[^\d+]", "", phone_match.group(0))
        fields["phone"] = phone[-10:] if phone.startswith("91") and len(phone) > 10 else phone
    exp_match = re.search(r"(\d+(?:\.\d+)?)\s*\+?\s*(?:years?|yrs?)\s+(?:of\s+)?(?:experience|exp)", header_text, re.I)
    if exp_match:
        years = float(exp_match.group(1))
        if years <= 40:
            fields["experience_years"] = f"{years:g} years"
    for line in lines[:8]:
        candidate = normalize_person_name(line)
        if candidate and looks_like_person_name(candidate):
            fields["candidate_name"] = candidate
            break
    if fields.get("candidate_name"):
        name_idx = next((i for i, line in enumerate(lines[:8]) if normalize_person_name(line) == fields["candidate_name"]), -1)
        for line in lines[name_idx + 1:name_idx + 6]:
            if "@" in line or re.search(r"\b(contact|phone|mobile|email|experience)\b", line, re.I):
                continue
            role = clean_header_role(line)
            if role:
                fields["current_role"] = role
                break
    cities = ['Bangalore', 'Bengaluru', 'Mumbai', 'Delhi', 'Hyderabad', 'Chennai', 'Pune', 'Noida', 'Gurgaon', 'Gurugram', 'Kolkata', 'Ahmedabad']
    text_l = header_text.lower()
    for city in cities:
        if city.lower() in text_l:
            fields["current_location"] = city
            break
    return fields

def clean_skill_phrase(value):
    value = re.sub(r"^[\s\-*•\d.)]+", "", str(value or ""))
    value = re.sub(r"\s+", " ", value).strip(" .;:-|")
    value = re.sub(r"^(?:core\s+)?(?:skills?|competenc(?:y|ies)|expertise|areas?\s+of\s+expertise)\s*[:\-]\s*", "", value, flags=re.I)
    if not value or len(value) < 3:
        return ""
    phrase = value.title() if value.isupper() else value
    sentence_patterns = [
        (r"ability to handle all kinds of automobile customers.*", "Customer Complaint Handling"),
        (r".*\bcomplaints?\b.*\bsolution\b.*customer satisfaction.*", "Customer Complaint Handling"),
        (r".*\bautomobile(?:s)?\b.*\bparts\b.*\brepair\b.*", "Automobile Parts & Repair Knowledge"),
        (r".*\bcommunication\b.*\borganizational skills\b.*", "Communication & Organization Skills"),
        (r".*\btaking charge of work\b.*", "Accountability & Responsibility Handling"),
        (r".*\bverbal\b.*\bcommunication\b.*\breporting\b.*", "Communication & Reporting"),
        (r".*\bhandling projects\b.*", "Project Handling"),
        (r".*\bms word\b.*\bms excel\b.*\bms outlook\b.*", "MS Word, MS Excel, MS Outlook"),
        (r".*\bfi[- ]mm integration\b.*", "FI-MM Integration"),
        (r".*\bmonth end\b.*", "Month-End Closing"),
        (r".*\byear[- ]end\b.*", "Year-End Closing"),
        (r".*\bopening\b.*\bclosing posting\b.*", "Opening & Closing Posting Periods"),
        (r".*\bconfiguration\b.*\bfico\b.*", "SAP FICO Configuration"),
        (r".*\binvoice processing\b.*", "Invoice Processing"),
        (r".*\bpayment runs?\b.*", "Payment Runs"),
    ]
    for pattern, replacement in sentence_patterns:
        if re.match(pattern, phrase, re.I):
            return replacement
    if len(value) > 80:
        return ""
    lower = value.lower()
    blocked = [
        "page ", "resume", "curriculum vitae", "professional experience", "work experience",
        "education", "personal details", "contact number", "email", "references",
        "willing to work on challenging assignments"
    ]
    if any(fragment in lower for fragment in blocked):
        return ""
    if re.search(r"\b(19|20)\d{2}\b", value) or re.search(r"@|www\.|https?://", value, re.I):
        return ""
    words = re.findall(r"[A-Za-z][A-Za-z0-9+#.-]*", value)
    allowed_single = {"SAP", "FICO", "Excel", "GST", "TDS", "Python", "Java", "React", "Angular", "Docker", "Kubernetes", "SQL"}
    if len(words) == 1 and words[0].upper() not in {x.upper() for x in allowed_single}:
        return ""
    return phrase

def extract_skill_section_phrases(lines, limit=12):
    headings = [
        "core skill", "core skills", "key skill", "key skills", "technical skills",
        "functional skills", "sap functional skills", "skills", "skill set",
        "areas of expertise", "area of expertise", "competencies", "core competencies",
        "professional skills", "domain skills"
    ]
    stops = [
        "professional experience", "work experience", "employment history", "career history",
        "education", "academic", "certification", "certifications", "projects",
        "personal details", "declaration", "languages known", "hobbies", "responsibilities"
    ]
    found = []
    for idx, line in enumerate(lines[:160]):
        normalized = re.sub(r"[^a-z0-9 ]+", " ", line.lower()).strip()
        if normalized not in headings:
            continue
        items = []
        current = ""
        for item in lines[idx + 1:min(len(lines), idx + 35)]:
            item_norm = re.sub(r"[^a-z0-9 ]+", " ", item.lower()).strip()
            if item_norm in stops or item_norm in headings or looks_like_role_title_line(item) or split_company_date_line(item):
                break
            if re.fullmatch(r"[\s\-*•]+", item):
                if current:
                    items.append(current)
                current = ""
                continue
            if current and re.match(r"^[a-z,;&(). ]+$", item.strip()):
                current += " " + item.strip()
            else:
                if current:
                    items.append(current)
                current = item.strip()
        if current:
            items.append(current)
        for item in items:
            whole_skill = clean_skill_phrase(item)
            if whole_skill and whole_skill != item.strip() and whole_skill.lower() not in {x.lower() for x in found}:
                found.append(whole_skill)
                if len(found) >= limit:
                    return found
                continue
            parts = re.split(r"[,;|]+", item)
            for part in parts:
                skill = clean_skill_phrase(part)
                if skill and skill.lower() not in {x.lower() for x in found}:
                    found.append(skill)
                    if len(found) >= limit:
                        return found
    return found

def extract_domain_skill_phrases(text, limit=12):
    phrases = [
        "Field Breakdown Analysis", "Warranty Handling", "Service Operations", "Dealer Management",
        "Customer Complaint Handling", "Root Cause Analysis", "Diesel Engines", "Engine Diagnostics",
        "Preventive Maintenance", "Technical Support", "After Sales Service", "Team Handling",
        "Vendor Management", "Client Relationship Management", "Business Development", "Channel Sales",
        "Key Account Management", "Lead Generation", "Negotiation", "Recruitment", "Talent Acquisition",
        "Sourcing", "Screening", "Interview Coordination", "Payroll", "Employee Relations",
        "Accounts Payable", "Accounts Receivable", "General Ledger", "GST", "TDS", "Bank Reconciliation",
        "Financial Reporting", "SAP FICO", "Month End Closing", "Invoice Processing", "Payment Runs",
        "Customer Satisfaction Index", "Post Service Feedback", "Retrofitments", "AMC", "Technical Issues",
        "E-Learning Monitoring", "Value Added Services", "KMPL Checkup", "Onsite Services",
        "AutoCAD", "Siebel", "SAP", "CRM DMS", "Vehicle Management", "Service Advisor",
        "Job Card Opening", "Vehicle Delivery", "Breakdown Service"
    ]
    text_l = (text or "").lower()
    found = []
    for phrase in phrases:
        pattern = r"(?<![a-z0-9])" + re.escape(phrase.lower()) + r"(?![a-z0-9])"
        if re.search(pattern, text_l):
            found.append(phrase)
            if len(found) >= limit:
                break
    return found

def extract_resume_skills(text, lines, skills_db, limit=12):
    section_skills = extract_skill_section_phrases(lines, limit)
    domain_skills = extract_domain_skill_phrases(text, limit)
    combined = unique_list(section_skills + domain_skills, limit)
    if combined:
        return ", ".join(combined)

    found_skills = set()
    text_lower = (text or "").lower()
    for category, skills in skills_db.items():
        for skill in skills:
            pattern = r"(?<![a-z0-9])" + re.escape(skill.lower()) + r"(?![a-z0-9])"
            if re.search(pattern, text_lower):
                display = skill.replace('r programming', 'R').replace('c programming', 'C')
                found_skills.add(display.title())
    if found_skills:
        sorted_skills = sorted(found_skills, key=lambda x: len(x), reverse=True)
        return ", ".join(sorted_skills[:limit])
    return ""

def parse_cv(cv_path, use_ai=False, ai_provider="", ai_api_key=""):
    """
    CV parsing with HuggingFace NER (when available) + enhanced regex/heuristics fallback.
    Extracts: name, email, phone, company, role, experience, skills, location, education, notice period, salary, summary.
    """
    result = {"candidate_name":"","email_addr":"","phone":"",
              "current_company":"","current_role":"","experience_years":"",
              "key_skills":"","current_location":"","cv_summary":"",
              "education":"","notice_period":"","current_salary":"","expected_salary":""}
    try:
        text = extract_cv_text(cv_path)
        if not text.strip():
            if os.path.splitext(cv_path or "")[1].lower() == ".pdf":
                result["_parse_warning"] = "No selectable text was found in this PDF. It looks like a scanned/image resume, so ATS could not auto-read it. Please enter details manually or use OCR before upload."
            else:
                result["_parse_warning"] = "ATS could not read text from this CV file. Please check the file or enter details manually."
            return result
        
        import re
        text_lower = text.lower()
        
        # Step 1: Try HuggingFace NER (free tier - no API key needed)
        hf_result = parse_huggingface(text, ai_api_key or "")
        if not hf_result.get("_hf_limited") and not hf_result.get("_hf_error"):
            # HF worked! Use its results and fill remaining with heuristics
            if hf_result.get("candidate_name"):
                result["candidate_name"] = normalize_person_name(hf_result["candidate_name"])
            if hf_result.get("current_company"):
                result["current_company"] = hf_result["current_company"]
        else:
            # HF not available (rate limit or no API key) - use full heuristics
            pass
        
        # Step 2: Always run enhanced heuristics to fill gaps + extract other fields
        lines_list = [l.strip() for l in text.split('\n') if l.strip()]
        header_fields = extract_pdf_header_fields(cv_path)
        for field in ("candidate_name", "email_addr", "phone", "current_role", "experience_years", "current_location"):
            if header_fields.get(field):
                result[field] = header_fields[field]
        
        # ===== EMAIL =====
        if not result.get("email_addr"):
            email_match = re.search(r"(?<![\w.-])[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}(?![\w.-])", text)
            if email_match:
                result["email_addr"] = email_match.group(0).strip().lower()
        
        # ===== PHONE =====
        if not result.get("phone"):
            phone_patterns = [
                r"(?:\+?91[\s\-]?)?[6-9]\d{4}[\s\-]?\d{5}",  # Indian mobile
                r"\+?1[\s\-]?\(?\d{3}\)[\s\-]?\d{3}[\s\-]?\d{4}",  # US
                r"\+?\d[\d\s\-]{9,}\d",  # Generic international
                r"\(?\d{2,4}\)?[\s\-]?\d{3,4}[\s\-]?\d{3,4}",  # Landline
            ]
            for pattern in phone_patterns:
                match = re.search(pattern, text)
                if match:
                    phone = re.sub(r"[^\d+]", "", match.group(0))
                    result["phone"] = phone[-10:] if phone.startswith("91") and len(phone) > 10 else phone
                    break
        
        # ===== NAME =====
        if not result.get("candidate_name"):
            name_patterns = [
                r"(?:^|\n)\s*(?:name|candidate\s+name|full\s+name)\s*[:\-]\s*([A-Za-z][A-Za-z'.\- \t]{3,70})",
                r"(?:^|\n)\s*([A-Za-z][A-Za-z'.-]+[ \t]+[A-Za-z][A-Za-z'.-]+(?:[ \t]+[A-Za-z][A-Za-z'.-]+){0,2})\s*(?:\n|$)",
            ]
            for pattern in name_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    candidate = clean_cv_value(match.group(1), 60)
                    normalized = normalize_person_name(candidate)
                    if normalized and looks_like_person_name(normalized):
                        result["candidate_name"] = normalized
                        break

        if not result.get("candidate_name"):
            for line in lines_list[:18]:
                candidate = clean_cv_value(line, 60)
                normalized = normalize_person_name(candidate)
                if normalized and looks_like_person_name(normalized):
                    result["candidate_name"] = normalized
                    break
        
        # ===== COMPANY =====
        email_domains = ['gmail', 'yahoo', 'hotmail', 'outlook', 'aol', 'icloud', 'proton', 'zoho', 'mail', 'live', 'msn']
        def is_not_email(s):
            return not any(d in s.lower() for d in email_domains)
        experience_company = infer_current_company_from_experience(lines_list)
        if experience_company:
            result["current_company"] = experience_company
        
        company_suffixes = ['inc', 'llc', 'corp', 'pvt', 'ltd', 'llp', 'co', 'company', 'technologies', 'tech',
                          'solutions', 'systems', 'services', 'labs', 'studios', 'digital', 'software',
                          'consulting', 'group', 'international', 'global', 'innovations', 'partners',
                          'ventures', 'info', 'institute', 'bank', 'finance', 'enterprises', 'holdings',
                          'analytics', 'data', 'ai', 'learning']
        
        # Pattern 1: Explicit labels
        if not result.get("current_company"):
            label_patterns = [
                r"(?:current\s+company|current\s+employer|company|employer)[:\s-]*([A-Za-z0-9&.,' -]{2,60})",
                r"(?:works?\s+(?:at|with)|employed\s+(?:at|by)|associated\s+with)[:\s-]*([A-Za-z0-9&.,' -]{2,60})",
                r"(?:currently\s+(?:at|with|working\s+at))[:\s-]*([A-Za-z0-9&.,' -]{2,60})",
            ]
            for pattern in label_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match and is_not_email(match.group(1)):
                    result["current_company"] = clean_company_name(match.group(1))
                    if not result["current_company"]:
                        continue
                    break
        
        # Pattern 2: Look for company in experience entries
        if not result.get("current_company"):
            exp_section = False
            for i, line in enumerate(lines_list[:50]):
                lower = line.lower()
                if any(kw in lower for kw in ['experience', 'employment', 'work history', 'professional experience']):
                    exp_section = True
                    continue
                if exp_section and line.strip():
                    line_clean = line.strip()
                    if len(line_clean) > 3:
                        words_in_line = line_clean.split()
                        if len(words_in_line) >= 2:
                            for suffix in company_suffixes:
                                if suffix in lower:
                                    match = re.match(r"([A-Za-z][\w\s&.,]+)", line_clean)
                                    if match and is_not_email(line_clean):
                                        result["current_company"] = match.group(1).strip().title()[:45]
                                        break
                        if result.get("current_company"):
                            break
                        if len(line.split()) > 6:
                            break
        
        # Pattern 3: Company name with suffix
        if not result.get("current_company"):
            for line in lines_list[:30]:
                for suffix in company_suffixes:
                    if suffix in line.lower() and len(line) > 5 and len(line) < 60:
                        if is_not_email(line):
                            match = re.match(r"([A-Za-z][\w\s&.,]+)", line.strip())
                            if match:
                                result["current_company"] = match.group(1).strip().title()[:45]
                                break
                if result.get("current_company"):
                    break
        
        # ===== ROLE =====
        role_suffixes = ['engineer', 'developer', 'manager', 'analyst', 'designer', 'architect', 
                        'consultant', 'lead', 'specialist', 'administrator', 'director', 'head',
                        'associate', 'coordinator', 'officer', 'executive', 'president', 'vp']
        
        experience_role = infer_current_role_from_experience(lines_list)
        if experience_role:
            result["current_role"] = experience_role

        # Pattern 1: Explicit labels
        if not result.get("current_role"):
            role_patterns = [
                r"(?:current\s+role|current\s+designation|designation|job\s+title|role|title)[:\s-]*([A-Za-z][^\n,]{2,60})",
                r"(?:working\s+as)[:\s-]*([A-Za-z][^\n,]{2,60})",
            ]
            for pattern in role_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    result["current_role"] = clean_role_title(match.group(1))
                    if not result["current_role"]:
                        continue
                    break
        
        # Pattern 2: Look for role titles in experience section
        if not result.get("current_role"):
            exp_section = False
            for i, line in enumerate(lines_list[:60]):
                lower = line.lower()
                if any(kw in lower for kw in ['experience', 'employment']):
                    exp_section = True
                    continue
                if exp_section:
                    for suffix in role_suffixes:
                        if suffix in lower:
                            match = re.search(r"([A-Za-z][\w\s/\-]+(?:\s+(?:engineer|developer|manager|analyst|designer|architect|consultant|lead|specialist))?)", line, re.IGNORECASE)
                            if match:
                                result["current_role"] = clean_role_title(match.group(1))
                                break
                    if result.get("current_role"):
                        break
        
        # Pattern 3: Common title formats anywhere
        if not result.get("current_role"):
            title_patterns = [
                r"((?:Senior|Junior|Lead|Principal|Staff|Associate)\s+(?:Software|Frontend|Backend|Full\s*Stack|Data|ML|DevOps)\s+Engineer)",
                r"((?:Software|Full\s*Stack|Frontend|Backend)\s+Developer)",
                r"((?:Product|Project|Program|Technical|Engineering)\s+Manager)",
                r"((?:Business|System|Data|Business\s+Intelligence)\s+Analyst)",
                r"((?:DevOps|Cloud|Site\s*Reliability)\s+Engineer)",
                r"((?:UI|UX|Graphic)\s+Designer)",
                r"((?:Machine\s*Learning|ML)\s+Engineer)",
                r"((?:Data|Business)\s+Scientist)",
            ]
            for pattern in title_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    result["current_role"] = match.group(1).strip().title()[:55]
                    break
        
        # ===== SKILLS =====
        skills_db = {
            'languages': ['python', 'java', 'javascript', 'c++', 'c#', 'c programming', 'php', 'ruby', 'go', 'golang', 'rust', 'swift', 'kotlin', 'scala', 'r programming', 'matlab', 'perl', 'shell', 'bash'],
            'frameworks': ['react', 'angular', 'vue', 'node', 'node.js', 'express', 'django', 'flask', 'spring', 'spring boot', 'rails', 'laravel', 'next.js', 'gatsby', 'nuxt'],
            'databases': ['sql', 'mysql', 'postgresql', 'mongodb', 'oracle', 'redis', 'elasticsearch', 'cassandra', 'dynamodb', 'sqlite', 'sql server', 'plsql'],
            'cloud': ['aws', 'amazon web services', 'azure', 'google cloud', 'gcp', 'heroku', 'digitalocean', 'docker', 'kubernetes', 'k8s', 'terraform'],
            'tools': ['git', 'github', 'gitlab', 'jenkins', 'ci/cd', 'jira', 'agile', 'scrum', 'jira', 'confluence', 'maven', 'gradle', 'npm', 'yarn', 'webpack'],
            'data': ['machine learning', 'ml', 'ai', 'artificial intelligence', 'deep learning', 'tensorflow', 'pytorch', 'pandas', 'numpy', 'scikit-learn', 'data science', 'tableau', 'power bi', 'excel', 'statistics'],
            'web': ['html', 'css', 'rest api', 'graphql', 'json', 'xml', 'web services', 'sass', 'less', 'bootstrap', 'tailwind'],
            'other': ['linux', 'unix', 'windows server', 'networking', 'security', 'devops', 'microservices', 'api', 'backend', 'frontend', 'full stack'],
        }
        
        result["key_skills"] = extract_resume_skills(text, lines_list, skills_db, 12)
        
        # ===== EXPERIENCE =====
        exp_patterns = [
            r"(?:total|overall)\s+(\d+(?:\.\d+)?)\s*-\s*years?",
            r"(?:total|overall)\s+(?:experience\s+)?(?:of\s+)?(\d+(?:\.\d+)?)\s*\+?\s*years?",
            r"(?:experience|total\s+experience)[:\s]*(\d+(?:\.\d+)?)\s*\+?\s*years?",
            r"(\d+(?:\.\d+)?)\s*\+?\s*years?\s+(?:of\s+)?(?:experience|work)",
            r"(\d+(?:\.\d+)?)\s*years?\s+(?:in|of|working)",
            r"(\d+(?:\.\d+)?)\+?\s*yrs?\s+(?:exp|experience)",
        ]
        for pattern in exp_patterns:
            match = re.search(pattern, text_lower)
            if match:
                years = float(match.group(1))
                if years <= 40:
                    result["experience_years"] = f"{years:g} years"
                    break
        
        # Fallback: count months
        if not result.get("experience_years"):
            months = len(re.findall(r"\d+\s*(?:months?|yrs?|years?)", text_lower))
            if months > 0:
                total_years = max(1, months // 12)
                result["experience_years"] = f"{total_years} years"
        
        # ===== LOCATION =====
        indian_cities = ['bangalore', 'bengaluru', 'mumbai', 'delhi', 'hyderabad', 'chennai', 'pune', 
                         'kolkata', 'ahmedabad', 'noida', 'gurgaon', 'gurugram', 'chandigarh', 'kolkata',
                         'jaipur', 'surat', 'lucknow', 'indore', 'coimbatore', 'vizag', 'kochi']
        for city in indian_cities:
            if city in text_lower:
                result["current_location"] = city.title()
                break
        
        if not result.get("current_location"):
            loc_match = re.search(r"(?:location|based\s+in|residing\s+in|city)[:\s]*([A-Za-z][a-z]+)", text_lower)
            if loc_match:
                result["current_location"] = loc_match.group(1).title()
        
        # ===== EDUCATION =====
        edu_patterns = [
            r"(?:b\.?tech|b\.?e\.?|b\.?sc\.?|m\.?tech|m\.?sc\.?|mba|m\.?ca\.?|ph\.?d\.?|b\.?a\.?|m\.?com\.?)",
        ]
        for pattern in edu_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                result["education"] = match.group(0).upper()
                break
        
        # ===== NOTICE PERIOD =====
        notice_patterns = [
            r"(?:notice\s+period|np)[:\s]*(\d+)\s*(?:days?|months?)",
            r"(?:serving\s+notice|serving\s+np)[:\s]*(\d+)",
            r"can\s+join\s+(?:within|in)\s+(\d+)\s*(?:days?|months?)",
            r"(\d+)\s*(?:days?|months?)\s+notice",
        ]
        for pattern in notice_patterns:
            match = re.search(pattern, text_lower)
            if match:
                result["notice_period"] = match.group(0).strip()
                break
        
        # ===== SALARY =====
        ctc_patterns = [
            r"(?:current\s+ctc|present\s+ctc|existing\s+ctc|salary)[:\s]*₹?\s*(\d+(?:\.\d+)?)\s*(?:lpa|lakh|lacs?|inr)?",
            r"(?:ctc)[:\s]*₹?\s*(\d+(?:\.\d+)?)\s*(?:lpa|lakh|lacs?)",
        ]
        for pattern in ctc_patterns:
            match = re.search(pattern, text_lower)
            if match:
                try:
                    val = float(match.group(1))
                    if val < 1000:
                        val *= 100000
                    result["current_salary"] = f"₹{val/100000:.1f} LPA"
                except:
                    pass
                break
        
        exp_sal_patterns = [
            r"(?:expected\s+ctc|expected\s+salary|notice\s+salary)[:\s]*₹?\s*(\d+(?:\.\d+)?)\s*(?:lpa|lakh|lacs?)?",
        ]
        for pattern in exp_sal_patterns:
            match = re.search(pattern, text_lower)
            if match:
                try:
                    val = float(match.group(1))
                    if val < 1000:
                        val *= 100000
                    result["expected_salary"] = f"₹{val/100000:.1f} LPA"
                except:
                    pass
                break
        
        # ===== SUMMARY =====
        summary_parts = []
        
        line1_parts = []
        if result.get("current_role"):
            line1_parts.append(result["current_role"])
        if result.get("experience_years"):
            line1_parts.append(f"{result['experience_years']}")
        if result.get("current_company"):
            line1_parts.append(f"@ {result['current_company']}")
        if line1_parts:
            summary_parts.append(" • ".join(line1_parts))
        
        line2_parts = []
        if result.get("key_skills"):
            skills = result["key_skills"]
            if len(skills) > 55:
                skill_list = skills.split(", ")
                skills = ", ".join(skill_list[:6])
                if len(skill_list) > 6:
                    skills += f" +{len(skill_list)-6}"
            line2_parts.append(skills)
        if result.get("current_location"):
            line2_parts.append(f"📍 {result['current_location']}")
        if line2_parts:
            summary_parts.append(" • ".join(line2_parts))
        
        result["candidate_name"] = normalize_person_name(result.get("candidate_name", ""))
        result["cv_summary"] = " | ".join(summary_parts) if summary_parts else ""
        
        return result
    except Exception as e:
        return {"error": f"CV parsing failed: {str(e)}"}
def norm_phone(p): return re.sub(r'[^\d]', '', p or "")[-10:]

def check_dup(conn, row):
    phone = norm_phone(row.get("phone", ""))
    email = (row.get("email_addr", "") or "").strip().lower()
    if phone and len(phone) >= 8:
        r = conn.execute(
            "SELECT id FROM candidates WHERE replace(replace(replace(phone,'+',''),'-',''),' ','') LIKE ? AND is_duplicate=0 LIMIT 1",
            (f"%{phone}",)).fetchone()
        if r: return True, r[0], "phone"
    if email:
        r = conn.execute(
            "SELECT id FROM candidates WHERE lower(trim(email_addr))=? AND is_duplicate=0 LIMIT 1",
            (email,)).fetchone()
        if r: return True, r[0], "email"
    return False, None, None

def check_missing(row):
    m = []
    if not row.get("candidate_name"): m.append("name")
    if not row.get("phone") and not row.get("email_addr"): m.append("phone/email")
    elif not row.get("phone"):       m.append("phone")
    elif not row.get("email_addr"):  m.append("email")
    return m

# ── Upload handler ────────────────────────────────────────────────────────────
def process_upload(recruiter_name, recruiter_email, role_override, sourcer_id, excel_file, cv_file_list):
    batch_id = hashlib.md5(f"{recruiter_email}{datetime.now().isoformat()}".encode()).hexdigest()[:10]
    result   = {"added":0,"duplicates":0,"missing":0,"errors":[],"cv_warnings":[]}

    fname = excel_file.filename.lower()
    data  = excel_file.read()
    if   fname.endswith(".csv"):           rows, err = parse_csv(data, role_override)
    elif fname.endswith((".xlsx",".xls")): rows, err = parse_xlsx(data, role_override)
    else: return {"error": "Please upload a .xlsx or .csv file"}
    if err:   return {"error": err}
    if not rows: return {"error": "No rows found. Check the file has a header row with data below it."}

    saved_cvs = []
    for cv in cv_file_list:
        if cv and cv.filename:
            orig, url, pub_id = upload_cv(cv, batch_id)
            if orig: saved_cvs.append((orig, url, pub_id))

    conn = get_db()
    conn.execute("INSERT OR IGNORE INTO team_members (name,email,is_fixed) VALUES (?,?,0)",
                 (recruiter_name, recruiter_email.lower()))

    matched_ids = set()
    for row in rows:
        if not any(str(v).strip() for v in row.values()): continue
        cname = row.get("candidate_name", "").strip()
        cv_orig, cv_url, cv_pub, cv_ok = match_cv(cname, saved_cvs)
        if cv_pub: matched_ids.add(cv_pub)
        is_dup, dup_id, dup_why = check_dup(conn, row)
        missing = check_missing(row)

        cid = conn.execute("""INSERT INTO candidates
            (upload_batch,recruiter_name,recruiter_email,sourcer_id,role_name,candidate_name,
             email_addr,phone,current_company,current_role,experience_years,key_skills,
             notice_period,current_salary,expected_salary,current_location,
             preferred_location,remarks,cv_filename,cv_url,cv_public_id,cv_summary,
             status,tags,is_duplicate,duplicate_of,missing_info)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (batch_id,recruiter_name,recruiter_email.lower(),sourcer_id,
             row.get("role_name",""),cname,row.get("email_addr",""),row.get("phone",""),
             row.get("current_company",""),row.get("current_role",""),
             row.get("experience_years",""),row.get("key_skills",""),
             row.get("notice_period",""),row.get("current_salary",""),
             row.get("expected_salary",""),row.get("current_location",""),
             row.get("preferred_location",""),row.get("remarks",""),
             cv_orig,cv_url,cv_pub,row.get("cv_summary",""),"New","",
             1 if is_dup else 0,dup_id,
             ",".join(missing) if missing else None)).lastrowid
        conn.commit()

        if is_dup:
            conn.execute("INSERT INTO alerts (alert_type,message,candidate_id,recruiter_email) VALUES (?,?,?,?)",
                ("duplicate",f"Duplicate: {cname or '?'} matched via {dup_why} (original ID #{dup_id})",cid,recruiter_email))
            result["duplicates"] += 1
        if missing:
            conn.execute("INSERT INTO alerts (alert_type,message,candidate_id,recruiter_email) VALUES (?,?,?,?)",
                ("missing_info",f"Missing {', '.join(missing)} for {cname or 'unnamed'}",cid,recruiter_email))
            result["missing"] += 1
        result["added"] += 1

    for orig, url, pid in unmatched_cvs(saved_cvs, matched_ids):
        conn.execute("INSERT INTO alerts (alert_type,message,recruiter_email) VALUES (?,?,?)",
            ("cv_mismatch",
             f"CV '{orig}' by {recruiter_name} doesn't match any candidate in the Excel. Please re-upload with correct filename.",
             recruiter_email))
        result["cv_warnings"].append(orig)

    conn.execute("""INSERT INTO upload_log
        (batch_id,recruiter_name,recruiter_email,filename,candidates_added,duplicates_found,missing_count)
        VALUES (?,?,?,?,?,?,?)""",
        (batch_id,recruiter_name,recruiter_email.lower(),excel_file.filename,
         result["added"],result["duplicates"],result["missing"]))

    today = date.today().isoformat()
    submitted = {r[0] for r in conn.execute(
        "SELECT DISTINCT recruiter_email FROM candidates WHERE date(created_at)=?",(today,)).fetchall()}
    for member in conn.execute("SELECT name,email FROM team_members").fetchall():
        if member["email"] not in submitted:
            if not conn.execute(
                "SELECT 1 FROM alerts WHERE alert_type='no_submission' AND recruiter_email=? AND date(created_at)=?",
                (member["email"],today)).fetchone():
                conn.execute("INSERT INTO alerts (alert_type,message,recruiter_email) VALUES (?,?,?)",
                    ("no_submission",f"{member['name']} has not submitted any profiles today",member["email"]))
    conn.commit()
    conn.close()
    return result

# ── Weekly summary email ──────────────────────────────────────────────────────
def build_weekly_summary():
    conn  = get_db()
    since = (date.today() - timedelta(days=7)).isoformat()
    today = date.today().isoformat()

    total   = conn.execute("SELECT COUNT(*) FROM candidates WHERE date(created_at)>=? AND is_duplicate=0",(since,)).fetchone()[0]
    dups    = conn.execute("SELECT COUNT(*) FROM candidates WHERE date(created_at)>=? AND is_duplicate=1",(since,)).fetchone()[0]
    missing = conn.execute("SELECT COUNT(*) FROM candidates WHERE date(created_at)>=? AND missing_info IS NOT NULL",(since,)).fetchone()[0]

    # Per recruiter
    perf = conn.execute("""
        SELECT recruiter_name, COUNT(*) as total,
               SUM(CASE WHEN is_duplicate=0 THEN 1 ELSE 0 END) as unique_c
        FROM candidates WHERE date(created_at)>=?
        GROUP BY recruiter_email ORDER BY unique_c DESC""",(since,)).fetchall()

    # Status breakdown
    statuses = conn.execute("""
        SELECT status, COUNT(*) as cnt FROM candidates
        WHERE is_duplicate=0 GROUP BY status ORDER BY cnt DESC""").fetchall()

    # Who didn't submit this week
    all_team  = conn.execute("SELECT name,email FROM team_members WHERE is_fixed=1").fetchall()
    submitted = {r[0] for r in conn.execute(
        "SELECT DISTINCT recruiter_email FROM candidates WHERE date(created_at)>=?",(since,)).fetchall()}
    no_sub = [m["name"] for m in all_team if m["email"] not in submitted]

    conn.close()

    # Build HTML email
    perf_rows = "".join(f"<tr><td>{r['recruiter_name']}</td><td>{r['total']}</td><td>{r['unique_c']}</td></tr>" for r in perf)
    status_rows = "".join(f"<tr><td>{r['status']}</td><td>{r['cnt']}</td></tr>" for r in statuses)
    no_sub_txt = ", ".join(no_sub) if no_sub else "Everyone submitted this week ✓"

    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;background:#f9f9f9;padding:24px;border-radius:8px">
      <h2 style="color:#e8643a;margin-bottom:4px">HR Guru ATS — Weekly Summary</h2>
      <p style="color:#666;font-size:13px">Week ending {date.today().strftime('%d %B %Y')}</p>

      <div style="display:flex;gap:12px;margin:20px 0">
        <div style="background:#fff;border-radius:8px;padding:16px 20px;flex:1;text-align:center;border:1px solid #eee">
          <div style="font-size:28px;font-weight:800;color:#e8643a">{total}</div>
          <div style="font-size:12px;color:#888">Candidates Added</div>
        </div>
        <div style="background:#fff;border-radius:8px;padding:16px 20px;flex:1;text-align:center;border:1px solid #eee">
          <div style="font-size:28px;font-weight:800;color:#e83a3a">{dups}</div>
          <div style="font-size:12px;color:#888">Duplicates</div>
        </div>
        <div style="background:#fff;border-radius:8px;padding:16px 20px;flex:1;text-align:center;border:1px solid #eee">
          <div style="font-size:28px;font-weight:800;color:#e8c53a">{missing}</div>
          <div style="font-size:12px;color:#888">Missing Info</div>
        </div>
      </div>

      <h3 style="color:#333;font-size:14px;margin-bottom:8px">Recruiter Performance</h3>
      <table style="width:100%;border-collapse:collapse;font-size:13px;background:#fff;border-radius:8px;overflow:hidden;margin-bottom:20px">
        <thead><tr style="background:#1c2030;color:#e8643a">
          <th style="padding:8px 12px;text-align:left">Recruiter</th>
          <th style="padding:8px 12px;text-align:left">Submitted</th>
          <th style="padding:8px 12px;text-align:left">Unique</th>
        </tr></thead>
        <tbody>{perf_rows if perf_rows else '<tr><td colspan="3" style="padding:8px 12px;color:#888">No submissions this week</td></tr>'}</tbody>
      </table>

      <h3 style="color:#333;font-size:14px;margin-bottom:8px">Pipeline Status (All Time)</h3>
      <table style="width:100%;border-collapse:collapse;font-size:13px;background:#fff;border-radius:8px;overflow:hidden;margin-bottom:20px">
        <thead><tr style="background:#1c2030;color:#e8643a">
          <th style="padding:8px 12px;text-align:left">Status</th>
          <th style="padding:8px 12px;text-align:left">Count</th>
        </tr></thead>
        <tbody>{status_rows}</tbody>
      </table>

      <div style="background:#fff;border-radius:8px;padding:14px 16px;border:1px solid #eee;font-size:13px;margin-bottom:20px">
        <strong>Not submitted this week:</strong><br/>
        <span style="color:#e8c53a">{no_sub_txt}</span>
      </div>

      <p style="font-size:11px;color:#aaa;text-align:center">HR Guru ATS · Auto-generated weekly report</p>
    </div>"""
    return html

def send_weekly_email():
    if not GMAIL_USER or not GMAIL_APP_PASS or not ADMIN_EMAIL:
        return {"error": "Gmail not configured — set GMAIL_USER, GMAIL_APP_PASS, ADMIN_EMAIL in wsgi.py"}
    try:
        html = build_weekly_summary()
        msg  = MIMEMultipart("alternative")
        msg["Subject"] = f"HR Guru Weekly Summary — {date.today().strftime('%d %b %Y')}"
        msg["From"]    = GMAIL_USER
        msg["To"]      = ADMIN_EMAIL
        msg.attach(MIMEText(html, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(GMAIL_USER, GMAIL_APP_PASS)
            s.sendmail(GMAIL_USER, ADMIN_EMAIL, msg.as_string())
        return {"ok": True, "sent_to": ADMIN_EMAIL}
    except Exception as e:
        return {"error": str(e)}

# ── Template ──────────────────────────────────────────────────────────────────
def make_template():
    if not XLSX_OK: return None
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Candidates"
    headers = ["Candidate Name","Email","Phone","Current Company","Current Role",
               "Experience (Years)","Key Skills","Notice Period","Current Salary",
               "Expected Salary","Current Location","Preferred Location","Role","Remarks"]
    hfill = PatternFill("solid",fgColor="1C2030"); hfont = Font(bold=True,color="E8643A")
    for i,h in enumerate(headers,1):
        c = ws.cell(row=1,column=i,value=h); c.font=hfont; c.fill=hfill
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[c.column_letter].width = max(len(h)+4,18)
    hints = ["← Full name","Email address","10-digit mobile","Current employer","Job title",
             "e.g. 5 years","Comma separated","e.g. 30 days","e.g. 12 LPA","e.g. 18 LPA",
             "Current city","Preferred city","Role being screened","Any notes"]
    ifill = PatternFill("solid",fgColor="0D1017"); ifont = Font(italic=True,color="6B7494")
    for i,h in enumerate(hints,1):
        c = ws.cell(row=2,column=i,value=h); c.font=ifont; c.fill=ifill
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ── Export ────────────────────────────────────────────────────────────────────
EXPORT_COLS = ["candidate_name","email_addr","phone","current_company","current_role",
               "experience_years","key_skills","notice_period","current_salary",
               "expected_salary","current_location","preferred_location",
               "role_name","status","tags","recruiter_name","remarks","created_at"]
EXPORT_HDR  = ["Candidate Name","Email","Phone","Current Company","Current Role",
               "Experience","Key Skills","Notice Period","Current Salary","Expected Salary",
               "Current Location","Preferred Location","Role","Status","Tags",
               "Recruiter","Remarks","Date Added"]

def build_query(args, current_session=None):
    role=args.get("role",""); sender=args.get("sender",""); loc=args.get("location","")
    notice=args.get("notice",""); q=args.get("q",""); show_d=args.get("show_dups","1")
    status=args.get("status",""); tag=args.get("tag","")
    exp_min=args.get("exp_min",""); exp_max=args.get("exp_max","")
    sal_min=args.get("sal_min",""); sal_max=args.get("sal_max","")
    sort=args.get("sort","newest")
    req_id=args.get("requirement_id","")
    client=args.get("client","")
    sql="SELECT c.*, r.title as requirement_title, r.client_name as client_name FROM candidates c LEFT JOIN requirements r ON c.requirement_id = r.id WHERE 1=1"; p=[]
    
    # Filter by sourcer_id for non-admin users
    owner_sql, owner_params = non_admin_candidate_owner_clause(current_session, "c")
    sql += owner_sql
    p.extend(owner_params)
        
    if role:   sql+=" AND role_name=?";   p.append(role)
    if sender: sql+=" AND recruiter_email=?"; p.append(sender)
    if loc:    sql+=" AND (current_location LIKE ? OR preferred_location LIKE ?)"; p+=[f"%{loc}%"]*2
    if notice: sql+=" AND notice_period LIKE ?"; p.append(f"%{notice}%")
    if status: sql+=" AND c.status=?"; p.append(status)
    if tag:    sql+=" AND (',' || tags || ',') LIKE ?"; p.append(f"%,{tag},%")
    if req_id: sql+=" AND requirement_id=?"; p.append(req_id)
    if client: sql+=" AND r.client_name=?"; p.append(client)
    if show_d=="0": sql+=" AND is_duplicate=0"
    # Search by name, skills, company, role, email, phone
    if q:
        sql+=" AND (candidate_name LIKE ? OR key_skills LIKE ? OR current_company LIKE ? OR current_role LIKE ? OR email_addr LIKE ? OR phone LIKE ?)"
        p += [f"%{q}%"]*6
    # Skills filter
    skills=args.get("skills","")
    if skills:
        skill_list = [s.strip().lower() for s in skills.split(",") if s.strip()]
        for sk in skill_list:
            sql+=" AND LOWER(key_skills) LIKE ?"
            p.append(f"%{sk}%")
    # Experience range filter (e.g. "6-10")
    exp_range=args.get("exp_range","")
    if exp_range and '-' in exp_range:
        parts = exp_range.split('-')
        exp_min = parts[0].strip() if parts[0].strip() else None
        exp_max = parts[1].strip() if len(parts) > 1 and parts[1].strip() else None
        # Extract numeric from experience_years (handle "6 Years", "6 years", "6", etc)
        if exp_min:
            sql+=" AND CAST(replace(replace(replace(replace(lower(experience_years),'years',''),'year',''),' ',''),'yrs','') AS REAL) >= ?"
            p.append(float(exp_min))
        if exp_max:
            sql+=" AND CAST(replace(replace(replace(replace(lower(experience_years),'years',''),'year',''),' ',''),'yrs','') AS REAL) <= ?"
            p.append(float(exp_max))
    # Sorting
    if sort == "newest":
        sql += " ORDER BY created_at DESC"
    elif sort == "oldest":
        sql += " ORDER BY created_at ASC"
    elif sort == "name":
        sql += " ORDER BY candidate_name ASC"
    elif sort == "status":
        sql += " ORDER BY status ASC"
    else:
        sql += " ORDER BY created_at DESC"
    return sql, p

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
@login_required
def index(): return render_template("index.html")

@app.route("/guide")
def guide(): return render_template("guide.html")

@app.route("/requirements")
@login_required
def requirements_page(): return render_template("index.html", active_tab="requirements")

@app.route("/upload")
@login_required
def upload_page():
    if not has_bulk_upload_access():
        return redirect("/")
    return render_template("upload.html")

@app.route("/healthz")
def health(): return "ok"

@app.route("/api/me")
@login_required
def api_me():
    return jsonify({
        "user_id": session.get("user_id"),
        "team_member_id": session.get("team_member_id"),
        "username": session.get("username"),
        "recruiter_name": session.get("recruiter_name"),
        "email": session.get("email"),
        "recruiter_email": session.get("recruiter_email"),
        "is_admin": session.get("is_admin"),
        "can_bulk_upload": 1 if has_bulk_upload_access() else 0
    }
)

@app.route("/api/upload", methods=["POST"])
@login_required
def api_upload():
    if not has_bulk_upload_access():
        return jsonify({"error":"Bulk upload permission required"}),403
    sourcer_id = session.get("team_member_id")
    name  = request.form.get("recruiter_name","").strip() or session.get("recruiter_name","System")
    email = request.form.get("recruiter_email","").strip() or session.get("recruiter_email","")
    if not name or name == "System": name = session.get("recruiter_name","System")
    if not email: email = session.get("recruiter_email","system@hrguru.com")
    email = email.lower()
    role  = request.form.get("role_override","").strip()
    if "excel_file" not in request.files: return jsonify({"error":"No file attached"}),400
    return jsonify(process_upload(name,email,role,sourcer_id,
                   request.files["excel_file"],request.files.getlist("cv_files")))

@app.route("/api/template")
def api_template():
    buf = make_template()
    if not buf: return "openpyxl not installed",500
    return send_file(buf,download_name="candidate_template.xlsx",as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/candidates")
@login_required
def api_candidates():
    conn = get_db(); sql,p = build_query(request.args, session)
    rows = conn.execute(sql,p).fetchall(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/candidates/export")
@login_required
def export_candidates():
    fmt  = request.args.get("format","csv")
    ids  = request.args.get("ids","")
    conn = get_db()
    if ids:
        id_list = [int(x) for x in ids.split(",") if x.isdigit()]
        if id_list:
            sql = "SELECT c.*, r.title as requirement_title FROM candidates c LEFT JOIN requirements r ON c.requirement_id = r.id WHERE c.id IN (" + ",".join("?" * len(id_list)) + ")"
            p = id_list
            owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
            sql += owner_sql
            p.extend(owner_params)
        else:
            sql, p = build_query(request.args, session)
    else:
        sql, p = build_query(request.args, session)
    rows = [dict(r) for r in conn.execute(sql,p).fetchall()]; conn.close()

    if fmt=="xlsx" and XLSX_OK:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="Candidates"
        hfill=PatternFill("solid",fgColor="1C2030"); hfont=Font(bold=True,color="E8643A")
        for i,h in enumerate(EXPORT_HDR,1):
            c=ws.cell(row=1,column=i,value=h); c.font=hfont; c.fill=hfill
            ws.column_dimensions[c.column_letter].width=max(len(h)+4,16)
        for ri,row in enumerate(rows,2):
            for ci,col in enumerate(EXPORT_COLS,1):
                v=row.get(col,""); ws.cell(row=ri,column=ci,value=str(v) if v else "")
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf,download_name=f"candidates_{date.today()}.xlsx",as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    buf=io.StringIO(); writer=csv.writer(buf); writer.writerow(EXPORT_HDR)
    for row in rows: writer.writerow([str(row.get(c,"") or "") for c in EXPORT_COLS])
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode()),
                     download_name=f"candidates_{date.today()}.csv",
                     as_attachment=True,mimetype="text/csv")

@app.route("/api/reports")
@login_required
def api_reports():
    report_type = request.args.get("type", "summary")
    from_date = request.args.get("from_date", "")
    to_date = request.args.get("to_date", "")
    conn = get_db()
    
    date_filter = ""
    params = []
    if from_date:
        date_filter += " AND created_at >= ?"
        params.append(from_date)
    if to_date:
        date_filter += " AND created_at <= ?"
        params.append(to_date + " 23:59:59")
    
    user_filter = ""
    user_params = []
    if not session.get("is_admin") and session.get("recruiter_email"):
        user_filter = " AND recruiter_email = ?"
        user_params = [session["recruiter_email"].lower()]
    
    base_where = "1=1" + date_filter + user_filter
    
    if report_type == "summary":
        total = conn.execute("SELECT COUNT(*) FROM candidates WHERE " + base_where, params + user_params).fetchone()[0]
        by_status = conn.execute("SELECT status, COUNT(*) as cnt FROM candidates WHERE " + base_where + " GROUP BY status", params + user_params).fetchall()
        by_role = conn.execute("SELECT role_name, COUNT(*) as cnt FROM candidates WHERE role_name!=''" + date_filter + user_filter + " GROUP BY role_name ORDER BY cnt DESC LIMIT 10", params + user_params).fetchall()
        return jsonify({"total": total, "by_status": [dict(r) for r in by_status], "by_role": [dict(r) for r in by_role]})
    
    elif report_type == "daily_dashboard":
        if not session.get("is_admin"):
            return jsonify({"error":"Admin only"}),403
        today = date.today().isoformat()
        rows = conn.execute("""
            SELECT
                COALESCE(NULLIF(c.recruiter_name,''), tm.name, c.recruiter_email, 'Unassigned') as recruiter,
                COALESCE(NULLIF(r.title,''), NULLIF(c.role_name,''), 'No Requirement') as requirement,
                COUNT(c.id) as submissions
            FROM candidates c
            LEFT JOIN team_members tm ON c.sourcer_id=tm.id
            LEFT JOIN requirements r ON c.requirement_id=r.id
            WHERE date(c.created_at)=?
            GROUP BY recruiter, requirement
            ORDER BY recruiter ASC, submissions DESC
        """, (today,)).fetchall()
        totals = conn.execute("""
            SELECT
                COUNT(*) as submissions,
                COUNT(DISTINCT COALESCE(NULLIF(recruiter_email,''), sourcer_id)) as recruiters,
                COUNT(DISTINCT COALESCE(requirement_id, role_name)) as requirements
            FROM candidates
            WHERE date(created_at)=?
        """, (today,)).fetchone()
        return jsonify({
            "date": today,
            "totals": dict(totals),
            "rows": [dict(r) for r in rows]
        })
    
    elif report_type == "sourcer":
        rows = conn.execute("""SELECT tm.name as sourcer, COUNT(c.id) as submissions, 
            SUM(CASE WHEN c.status IN ('Shortlisted','Hired','Joined') THEN 1 ELSE 0 END) as success
            FROM candidates c LEFT JOIN team_members tm ON c.sourcer_id = tm.id
            WHERE c.sourcer_id IS NOT NULL""" + date_filter + user_filter + """ GROUP BY c.sourcer_id
            ORDER BY submissions DESC""", params + user_params).fetchall()
        return jsonify([dict(r) for r in rows])
    
    elif report_type == "status":
        rows = conn.execute("SELECT status, COUNT(*) as cnt FROM candidates WHERE " + base_where + " GROUP BY status ORDER BY cnt DESC", params + user_params).fetchall()
        return jsonify([{"status": r[0], "count": r[1]} for r in rows])
    
    elif report_type == "skills":
        all_skills = conn.execute("SELECT key_skills FROM candidates WHERE key_skills!=''" + date_filter + user_filter, params + user_params).fetchall()
        skill_counts = {}
        for row in all_skills:
            for s in row[0].split(","):
                s = s.strip().lower()
                if s:
                    skill_counts[s] = skill_counts.get(s, 0) + 1
        sorted_skills = sorted(skill_counts.items(), key=lambda x: x[1], reverse=True)[:20]
        return jsonify([{"skill": k, "count": v} for k, v in sorted_skills])
    
    elif report_type == "submissions":
        rows = conn.execute("""SELECT DATE(created_at) as date, COUNT(*) as submissions
            FROM candidates WHERE """ + base_where + """ GROUP BY DATE(created_at)
            ORDER BY date DESC LIMIT 30""", params + user_params).fetchall()
        return jsonify([{"date": str(r[0]), "count": r[1]} for r in rows])
    
    elif report_type == "status_detail":
        status = request.args.get("status", "New")
        q = [status]
        sql = "SELECT role_name, recruiter_name, COUNT(*) as cnt FROM candidates WHERE status=?"
        if from_date:
            sql += " AND created_at >= ?"
            q.append(from_date)
        if to_date:
            sql += " AND created_at <= ?"
            q.append(to_date + " 23:59:59")
        if not session.get("is_admin") and session.get("team_member_id"):
            sql += " AND sourcer_id = ?"
            q.append(session["team_member_id"])
        sql += " GROUP BY role_name, recruiter_name ORDER BY cnt DESC"
        rows = conn.execute(sql, q).fetchall()
        return jsonify([{"role_name": r[0] or "-", "recruiter_name": r[1] or "-", "cnt": r[2]} for r in rows])
    
    return jsonify({"error": "Unknown report type"})

@app.route("/api/filters")
@login_required
def api_filters():
    conn = get_db()
    roles   = [r[0] for r in conn.execute("SELECT DISTINCT role_name FROM candidates WHERE role_name!='' ORDER BY role_name").fetchall()]
    senders = [{"recruiter_email":r[0],"recruiter_name":r[1]} for r in
               conn.execute("SELECT DISTINCT recruiter_email,recruiter_name FROM candidates WHERE recruiter_email!='' ORDER BY recruiter_name").fetchall()]
    locs    = [r[0] for r in conn.execute("SELECT DISTINCT current_location FROM candidates WHERE current_location!='' ORDER BY current_location").fetchall()]
    hms     = [dict(r) for r in conn.execute("SELECT * FROM hiring_managers ORDER BY name").fetchall()]
    # All unique tags
    all_tags_raw = conn.execute("SELECT tags FROM candidates WHERE tags!=''").fetchall()
    tag_set = set()
    for row in all_tags_raw:
        for t in (row[0] or "").split(","):
            t = t.strip()
            if t: tag_set.add(t)
    # Pipelines query before closing
    pipelines = [dict(r) for r in conn.execute("SELECT * FROM pipelines ORDER BY role_name").fetchall()]
    for p in pipelines: p["status_list"] = json.loads(p["status_list"])
    conn.close()
    # Return all data
    return jsonify({"roles":roles,"senders":senders,"locations":locs,
                    "statuses":STATUSES,"tags":sorted(tag_set),
                    "submission_statuses":SUBMISSION_STATUSES,"hiring_managers":hms,
                    "pipelines":pipelines})

@app.route("/api/stats")
@login_required
def api_stats():
    conn = get_db()

    team_member_id = session.get("team_member_id")
    is_admin = session.get("is_admin")

    where_clause = ""
    params = ()

    # non-admin sees only own stats
    if not is_admin:
        where_clause = " AND sourcer_id = ?"
        params = (team_member_id,)

    total = conn.execute(
        "SELECT COUNT(*) FROM candidates WHERE is_duplicate=0" + where_clause,
        params
    ).fetchone()[0]

    dups = conn.execute(
        "SELECT COUNT(*) FROM candidates WHERE is_duplicate=1" + where_clause,
        params
    ).fetchone()[0]

    roles = conn.execute(
        "SELECT COUNT(DISTINCT role_name) FROM candidates WHERE is_duplicate=0" + where_clause,
        params
    ).fetchone()[0]

    unread = conn.execute(
        "SELECT COUNT(*) FROM alerts WHERE is_read=0"
    ).fetchone()[0]

    last = conn.execute(
        "SELECT uploaded_at FROM upload_log ORDER BY id DESC LIMIT 1"
    ).fetchone()

    status_counts = {
        r[0]: r[1] for r in conn.execute(
            "SELECT status, COUNT(*) FROM candidates WHERE is_duplicate=0" + where_clause + " GROUP BY status",
            params
        ).fetchall()
    }

    conn.close()

    return jsonify({
        "total": total,
        "duplicates": dups,
        "roles": roles,
        "unread_alerts": unread,
        "last_upload": last[0] if last else None,
        "status_counts": status_counts
    })
@app.route("/api/roles")
@login_required
def get_roles():
    conn = get_db()
    # Get standard roles grouped by domain
    std_roles = conn.execute("SELECT domain, role_name FROM standard_roles ORDER BY domain, role_name").fetchall()
    # Get dynamic roles from candidates
    if not session.get("is_admin") and session.get("team_member_id"):
        dyn_rows = conn.execute("SELECT DISTINCT current_role as role_name FROM candidates WHERE current_role != '' AND sourcer_id = ? ORDER BY current_role", (session["team_member_id"],)).fetchall()
    else:
        dyn_rows = conn.execute("SELECT DISTINCT current_role as role_name FROM candidates WHERE current_role != '' ORDER BY current_role").fetchall()
    conn.close()
    # Group standard roles by domain
    grouped = {}
    for r in std_roles:
        d = dict(r)
        domain = d.get('domain', 'Other')
        role = d.get('role_name', '')
        if domain not in grouped:
            grouped[domain] = []
        grouped[domain].append(role)
    return jsonify({"grouped": grouped, "dynamic": [dict(r) for r in dyn_rows]})

@app.route("/api/roles", methods=["POST"])
@login_required
def add_role():
    data = request.json or {}
    domain = data.get("domain", "").strip()
    role = data.get("role_name", "").strip()
    if not domain or not role:
        return jsonify({"error": "domain and role_name required"}), 400
    conn = get_db()
    conn.execute("INSERT OR IGNORE INTO standard_roles (domain, role_name) VALUES (?, ?)", (domain, role))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/statuses")
@login_required
def get_statuses():
    statuses = ["New", "Contacted", "Interview", "Offer", "Hired", "Rejected", "Shortlisted", "Feedback Pending", "On Hold", "Joined"]
    return jsonify([{"status": s} for s in statuses])

@app.route("/api/skills")
@login_required
def get_skills():
    conn = get_db()
    rows = conn.execute("SELECT id, skill_name, category FROM standard_skills ORDER BY category, skill_name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/clients")
@login_required
def get_clients():
    conn = get_db()
    rows = conn.execute("SELECT id, client_name FROM clients ORDER BY client_name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/clients", methods=["POST"])
@login_required
def add_client():
    data = request.json or {}
    client_name = data.get("client_name", "").strip()
    if not client_name:
        return jsonify({"error": "client_name required"}), 400
    conn = get_db()
    conn.execute("INSERT OR IGNORE INTO clients (client_name) VALUES (?)", (client_name,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/candidate/<int:cid>/tags", methods=["PATCH"])
@login_required
def update_tags(cid):
    tags = request.json.get("tags","")
    # Normalise: strip whitespace from each tag
    clean = ",".join(t.strip() for t in tags.split(",") if t.strip())
    conn  = get_db()
    conn.execute("UPDATE candidates SET tags=?, updated_at=datetime('now','localtime') WHERE id=?",
                 (clean, cid))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/alerts")
@login_required
def api_alerts():
    conn  = get_db()
    atype = request.args.get("type","")
    sql   = "SELECT * FROM alerts WHERE 1=1"; p=[]
    if atype: sql+=" AND alert_type=?"; p.append(atype)
    sql  += " ORDER BY created_at DESC LIMIT 150"
    rows  = conn.execute(sql,p).fetchall(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/alerts/read",methods=["POST"])
@login_required
def mark_all_read():
    conn=get_db(); conn.execute("UPDATE alerts SET is_read=1"); conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/alerts/<int:aid>/read",methods=["POST"])
@login_required
def mark_one_read(aid):
    conn=get_db(); conn.execute("UPDATE alerts SET is_read=1 WHERE id=?",(aid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# ── Saved searches ────────────────────────────────────────────────────────────
@app.route("/api/saved_searches", methods=["GET"])
@login_required
def get_saved_searches():
    conn = get_db()
    rows = conn.execute("SELECT * FROM saved_searches ORDER BY created_at DESC").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/saved_searches", methods=["POST"])
@login_required
def save_search():
    data = request.json
    name    = (data.get("name") or "").strip()
    filters = data.get("filters", {})
    if not name: return jsonify({"error":"Name required"}),400
    conn = get_db()
    sid = conn.execute("INSERT INTO saved_searches (name,filters) VALUES (?,?)",
                       (name, json.dumps(filters))).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":sid})

@app.route("/api/candidates/bulk_delete", methods=["POST"])
@login_required
def bulk_delete_candidates():
    data = request.json or {}
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"error": "no ids provided"}), 400
    
    conn = get_db()
    
    # Ownership check for non-admins
    if not session.get("is_admin"):
        tm_id = session.get("team_member_id")
        placeholders = ",".join(["?"]*len(ids))
        # Only allow deleting candidates owned by this user
        if tm_id:
            conn.execute(f"DELETE FROM candidates WHERE id IN ({placeholders}) AND sourcer_id=?", ids + [tm_id])
        else:
            conn.execute(f"DELETE FROM candidates WHERE id IN ({placeholders}) AND sourcer_id IS NULL", ids)
    else:
        placeholders = ",".join(["?"]*len(ids))
        conn.execute(f"DELETE FROM candidates WHERE id IN ({placeholders})", ids)
    
    conn.commit(); conn.close()
    return jsonify({"ok": True, "deleted": len(ids)})

# ── Job Details ─────────────────────────────────────────────────────────────
@app.route("/api/job_details", methods=["GET"])
@login_required
def get_job_details():
    cid = request.args.get("candidate_id") or request.args.get("cid")
    if not cid:
        return jsonify([])
    conn = get_db()
    rows = conn.execute("SELECT * FROM job_details WHERE candidate_id=? ORDER BY created_at DESC", (cid,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/job_details", methods=["POST"])
@login_required
def add_job_details():
    data = request.json or {}
    cid = data.get("candidate_id")
    if not cid:
        return jsonify({"error":"candidate_id required"}), 400
    conn = get_db()
    jd = conn.execute(
        "INSERT INTO job_details (candidate_id,title,location,salary,department,job_id,notes,tags) VALUES (?,?,?,?,?,?,?,?)",
        (cid, data.get("title",""), data.get("location",""), data.get("salary",""),
         data.get("department",""), data.get("job_id",""), data.get("notes",""), data.get("tags",""))
    ).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok": True, "id": jd})

@app.route("/api/saved_searches/<int:sid>", methods=["DELETE"])
@login_required
def delete_saved_search(sid):
    conn = get_db(); conn.execute("DELETE FROM saved_searches WHERE id=?",(sid,))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

# ── Reporting ─────────────────────────────────────────────────────────────────
@app.route("/api/reporting/recruiter_performance")
@login_required
def recruiter_performance():
    conn  = get_db()
    days  = int(request.args.get("days",30))
    since = (date.today()-timedelta(days=days)).isoformat()
    rows  = conn.execute("""
        SELECT recruiter_name,recruiter_email,COUNT(*) as total,
            SUM(CASE WHEN is_duplicate=0 THEN 1 ELSE 0 END) as unique_candidates,
            SUM(CASE WHEN is_duplicate=1 THEN 1 ELSE 0 END) as duplicates,
            COUNT(DISTINCT role_name) as roles_worked,
            COUNT(DISTINCT date(created_at)) as active_days
        FROM candidates WHERE created_at>=?
        GROUP BY recruiter_email ORDER BY unique_candidates DESC""",(since,)).fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/reporting/daily_trend")
@login_required
def daily_trend():
    conn  = get_db()
    days  = int(request.args.get("days",14))
    since = (date.today()-timedelta(days=days)).isoformat()
    rows  = conn.execute("""
        SELECT date(created_at) as day,COUNT(*) as total,
            SUM(CASE WHEN is_duplicate=0 THEN 1 ELSE 0 END) as unique_c,
            COUNT(DISTINCT recruiter_email) as active_recruiters
        FROM candidates WHERE created_at>=?
        GROUP BY date(created_at) ORDER BY day""",(since,)).fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/reporting/no_submission_today")
@login_required
def no_submission_today():
    conn  = get_db(); today=date.today().isoformat()
    submitted={r[0] for r in conn.execute(
        "SELECT DISTINCT recruiter_email FROM candidates WHERE date(created_at)=?",(today,)).fetchall()}
    missing=[{"name":r["name"],"email":r["email"]} for r in
             conn.execute("SELECT name,email FROM team_members").fetchall()
             if r["email"] not in submitted]
    conn.close(); return jsonify(missing)

@app.route("/api/reporting/send_weekly", methods=["POST"])
@login_required
def api_send_weekly():
    return jsonify(send_weekly_email())

@app.route("/api/upload_cv", methods=["POST"])
@login_required
def upload_cv_file():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files["file"]

    if not f.filename:
        return jsonify({"error": "Empty filename"}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in [".pdf", ".doc", ".docx"]:
        return jsonify({"error": "Invalid file type"}), 400

    try:
        upload_folder = os.path.join(app.root_path, "uploads")
        os.makedirs(upload_folder, exist_ok=True)

        safe_name = secure_filename(f.filename)
        saved_name = f"single_{int(datetime.now().timestamp())}_{safe_name}"
        file_path = os.path.join(upload_folder, saved_name)

        f.save(file_path)

        file_url = f"/uploads/{saved_name}"

        parsed = None
        try:
            parsed = parse_cv(file_path)
            if parsed:
                parsed["candidate_name"] = normalize_person_name(parsed.get("candidate_name", ""))
                parsed = {
                    "candidate_name": parsed.get("candidate_name", ""),
                    "email_addr": parsed.get("email_addr", ""),
                    "phone": parsed.get("phone", "")
                }
        except Exception as pe:
            print("CV Parse Error:", pe)

        return jsonify({
            "ok": True,
            "filename": safe_name,
            "url": file_url,
            "public_id": saved_name,
            "parsed": parsed or {}
        })

    except Exception as e:
        print("Upload Error:", str(e))
        return jsonify({"error": str(e)}), 500

@app.route("/api/candidate/<int:cid>/upload_cv", methods=["POST"])
@login_required
def upload_candidate_cv(cid):
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    if not session.get("is_admin"):
        conn = get_db()
        c = conn.execute("SELECT sourcer_id FROM candidates WHERE id=?", (cid,)).fetchone()
        conn.close()
        if not c or c["sourcer_id"] != session.get("team_member_id"):
            return jsonify({"error": "Permission denied"}), 403
    upload_response = upload_cv_file()
    if isinstance(upload_response, tuple):
        return upload_response
    data = upload_response.get_json()
    if not data or not data.get("ok"):
        return upload_response
    conn = get_db()
    conn.execute(
        """UPDATE candidates
           SET cv_filename=?, cv_url=?, cv_public_id=?, updated_at=datetime('now','localtime')
           WHERE id=?""",
        (data.get("filename", ""), data.get("url", ""), data.get("public_id", ""), cid)
    )
    conn.commit()
    conn.close()
    return jsonify(data)

@app.route("/api/jd_match", methods=["POST"])
@login_required
def api_jd_match():
    jd_text_input = (request.form.get("jd_text") or "").strip()
    custom_hard_filters = (request.form.get("custom_hard_filters") or "").strip()
    if not jd_text_input and "jd_file" not in request.files:
        return jsonify({"error": "Please upload a JD file or paste the JD text."}), 400
    cv_files = request.files.getlist("cv_files") or request.files.getlist("cv_file")
    cv_files = [file for file in cv_files if file and file.filename]
    if not cv_files:
        return jsonify({"error": "Please upload the candidate CV file."}), 400
    try:
        schema_conn = get_db()
        ensure_ats_pipeline_schema(schema_conn)
        schema_conn.commit()
        schema_conn.close()
        jd_path, jd_name = "", "Pasted JD"
        if "jd_file" in request.files and request.files["jd_file"].filename:
            jd_path, jd_name = save_uploaded_analysis_file(request.files["jd_file"], "jd")
        jd_text = jd_text_input or extract_cv_text(jd_path)
        if not jd_text.strip():
            return jsonify({"error": "Could not read text from the job description."}), 400
        jd_hash = versioned_text_hash(jd_text + "\nCUSTOM_HARD_FILTERS:\n" + custom_hard_filters)
        parsed_jd = get_cached_parsed_jd(jd_hash)
        if not parsed_jd:
            parsed_jd = parse_jd_structured(jd_text)
            set_cached_parsed_jd(jd_hash, parsed_jd)
        results = []
        errors = []
        for cv_file in cv_files:
            cv_path, cv_name = save_uploaded_analysis_file(cv_file, "match_cv")
            cv_text = extract_cv_text(cv_path)
            if not cv_text.strip():
                errors.append({"cv_filename": cv_name, "error": "Could not read text from this CV."})
                continue
            resume_hash = versioned_text_hash(cv_text)
            cached = get_cached_match_result(jd_hash, resume_hash)
            if cached:
                cached["dashboard"] = build_match_dashboard(cached)
                cached["candidate_summary"] = build_candidate_summary_from_cv(cv_path, cv_text)
                cached["jd_filename"] = jd_name
                cached["cv_filename"] = cv_name
                cached["cache_hit"] = True
                results.append(cached)
                continue
            parsed_candidate = get_cached_parsed_resume(resume_hash)
            parsed_cv = {} if parsed_candidate else (parse_cv(cv_path) or {})
            result = run_hybrid_match(
                jd_text,
                cv_text,
                parsed_cv,
                cache_get=get_cached_embedding,
                cache_set=set_cached_embedding,
                custom_hard_filters=custom_hard_filters,
                parsed_jd=parsed_jd,
                parsed_candidate=parsed_candidate
            )
            if not parsed_candidate:
                set_cached_parsed_resume(resume_hash, result.get("parsed_candidate") or result.get("cv_json") or {})
            result["dashboard"] = build_match_dashboard(result)
            persist_match_artifacts(jd_hash, resume_hash, result)
            result["candidate_summary"] = build_candidate_summary_from_cv(cv_path, cv_text)
            result["jd_filename"] = jd_name
            result["jd_source"] = "text" if jd_text_input else "file"
            result["cv_filename"] = cv_name
            result["jd_hash"] = jd_hash
            result["resume_hash"] = resume_hash
            result["pipeline_version"] = MATCH_PIPELINE_VERSION
            result["custom_hard_filters"] = custom_hard_filters
            result["cache_hit"] = False
            results.append(result)
        results.sort(key=lambda item: int(item.get("final_score", item.get("score", 0)) or 0), reverse=True)
        if len(cv_files) == 1 and results and not request.files.getlist("cv_files"):
            return jsonify(results[0])
        return jsonify({
            "batch": True,
            "jd_filename": jd_name,
            "jd_hash": jd_hash,
            "count": len(results),
            "errors": errors,
            "results": results,
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print("JD match error:", e)
        return jsonify({"error": str(e)}), 500

@app.route("/api/parse_jd", methods=["POST"])
@login_required
def api_parse_jd_structured():
    data = request.get_json(silent=True) or {}
    jd_text = (data.get("jd_text") or request.form.get("jd_text") or "").strip()
    if not jd_text:
        return jsonify({"error": "Please provide JD text."}), 400
    return jsonify(parse_jd_structured(jd_text))

@app.route("/api/parse_cv", methods=["POST"])
@login_required
def api_parse_cv_structured():
    data = request.get_json(silent=True) or {}
    cv_text = (data.get("cv_text") or request.form.get("cv_text") or "").strip()
    if not cv_text:
        return jsonify({"error": "Please provide CV text."}), 400
    return jsonify(parse_resume_structured(cv_text, {}))

@app.route("/api/match_jd_cv", methods=["POST"])
@login_required
def api_match_jd_cv_structured():
    data = request.get_json(silent=True) or {}
    jd_text = (data.get("jd_text") or "").strip()
    cv_text = (data.get("cv_text") or "").strip()
    if not jd_text or not cv_text:
        return jsonify({"error": "Please provide both jd_text and cv_text."}), 400
    custom_hard_filters = (data.get("custom_hard_filters") or "").strip()
    result = run_hybrid_match(
        jd_text,
        cv_text,
        {},
        cache_get=get_cached_embedding,
        cache_set=set_cached_embedding,
        custom_hard_filters=custom_hard_filters
    )
    result["dashboard"] = build_match_dashboard(result)
    return jsonify(result)

@app.route("/api/match/export_pdf", methods=["POST"])
@login_required
def export_match_pdf():
    try:
        analysis = request.get_json(silent=True) or {}
        if not analysis:
            return jsonify({"error": "No match analysis data received."}), 400
        dashboard = analysis.get("dashboard") or build_match_dashboard(analysis)
        analysis["dashboard"] = dashboard
        buf = build_match_pdf(analysis)
        candidate = (dashboard.get("candidate_snapshot") or {}).get("candidate_name") or "candidate"
        safe_candidate = re.sub(r"[^A-Za-z0-9_-]+", "_", candidate).strip("_") or "candidate"
        return send_file(
            buf,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"match_analysis_{safe_candidate}_{date.today()}.pdf"
        )
    except Exception as e:
        print("Match PDF export error:", e)
        return jsonify({"error": str(e)}), 500

@app.route("/api/match/screening_questions", methods=["POST"])
@login_required
def generate_match_screening_questions():
    data = request.get_json(silent=True) or {}
    if request.form:
        data = dict(request.form)
        if data.get("analysis_json"):
            try:
                data["analysis"] = json.loads(data.get("analysis_json") or "{}")
            except Exception:
                data["analysis"] = {}
    analysis = data.get("analysis") or {}
    parsed_jd = analysis.get("parsed_jd") or analysis.get("jd_json") or {}
    jd_text = (data.get("jd_text") or "").strip()
    if not jd_text and "jd_file" in request.files and request.files["jd_file"].filename:
        jd_path, _ = save_uploaded_analysis_file(request.files["jd_file"], "jd_screening")
        jd_text = extract_cv_text(jd_path)
    if not parsed_jd and jd_text:
        parsed_jd = parse_jd_structured(jd_text)
    must_skills = parsed_jd.get("must_have_skills") or [
        item.get("skill") for item in parsed_jd.get("must_have_skills_weighted", []) or [] if item.get("skill")
    ]
    must_skills = [skill for skill in must_skills if skill][:10]
    if not must_skills:
        return jsonify({"error": "No mandatory skills found in the JD."}), 400
    candidate = analysis.get("parsed_candidate") or analysis.get("cv_json") or {}
    prompt = (
        "Generate recruiter phone-screen questions only. Do not score the candidate. "
        "Use plain language and vary the question style. Avoid repeating the same opening phrase. "
        "Include a mix of practical scenario, depth-check, troubleshooting, prioritization, stakeholder, and trade-off questions. "
        "Return JSON with a questions array. Each item must have skill, question, expected_signal, and follow_up. "
        f"Mandatory skills: {', '.join(must_skills)}. "
        f"Candidate current role: {candidate.get('current_role','')}. "
        f"Candidate skills: {', '.join((candidate.get('normalized_skills') or [])[:20])}."
    )
    try:
        response = requests.post(
            f"{os.getenv('LLM_API_BASE', 'http://localhost:11434').rstrip('/')}/api/generate",
            json={"model": os.getenv("LLM_MODEL", "qwen2.5:7b"), "prompt": prompt, "format": "json", "stream": False},
            timeout=float(os.getenv("OLLAMA_QUESTION_TIMEOUT", "12"))
        )
        response.raise_for_status()
        raw = response.json().get("response") or "{}"
        parsed = json.loads(raw)
        questions = parsed.get("questions", []) if isinstance(parsed, dict) else []
    except Exception as e:
        question_patterns = [
            (
                "Walk me through how you would use {skill} to solve a real customer or platform problem.",
                "Look for a practical workflow, clear decision points, and real ownership.",
                "What trade-off did you consider, and why did you choose that path?"
            ),
            (
                "What are the first three things you would check when a {skill}-related implementation is not working as expected?",
                "Look for structured troubleshooting rather than generic familiarity.",
                "Which signal would tell you the issue is resolved?"
            ),
            (
                "How do you decide whether {skill} is truly required for a product requirement or just a nice-to-have?",
                "Look for prioritization, customer impact thinking, and technical judgment.",
                "How would you explain that decision to engineering and business stakeholders?"
            ),
            (
                "Tell me about a time you had to align engineering, product, and customers around {skill}.",
                "Look for stakeholder handling, requirement translation, and follow-through.",
                "What disagreement came up, and how did you resolve it?"
            ),
            (
                "What good looks like for {skill} in a production environment?",
                "Look for depth, operational awareness, risks, and measurable outcomes.",
                "Which metrics or checks would you track after launch?"
            ),
            (
                "If a candidate says they know {skill}, what evidence would convince you they have used it hands-on?",
                "Look for concrete artifacts, examples, and implementation details.",
                "What answer would make you doubt their depth?"
            ),
        ]
        questions = []
        for idx, skill in enumerate(must_skills[:10]):
            question, expected_signal, follow_up = question_patterns[idx % len(question_patterns)]
            questions.append({
                "skill": skill,
                "question": question.format(skill=skill),
                "expected_signal": expected_signal,
                "follow_up": follow_up
            })
        return jsonify({"ok": True, "questions": questions[:12], "skills": must_skills, "source": "fallback", "note": "Ollama is not reachable, so deterministic fallback questions were generated."})
    return jsonify({"ok": True, "questions": questions[:12], "skills": must_skills, "source": "llm"})

@app.route("/api/match/screening_questions_pdf", methods=["POST"])
@login_required
def export_screening_questions_pdf():
    try:
        payload = request.get_json(silent=True) or {}
        questions = payload.get("questions") or []
        if not questions:
            return jsonify({"error": "No screening questions received."}), 400
        buf = build_screening_questions_pdf(payload)
        return send_file(
            buf,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"interview_screening_questions_{date.today()}.pdf"
        )
    except Exception as e:
        print("Screening PDF export error:", e)
        return jsonify({"error": str(e)}), 500
        
@app.route("/api/team",methods=["GET"])
@login_required
def get_team():
    conn=get_db()
    rows=conn.execute("SELECT * FROM team_members ORDER BY name").fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/team",methods=["POST"])
@login_required
def add_team():
    d=request.json; conn=get_db()
    conn.execute("INSERT OR IGNORE INTO team_members (name,email,phone,role,can_bulk_upload,is_fixed) VALUES (?,?,?,?,?,1)",
                 (d["name"],(d["email"] or "").lower(),d.get("phone",""),d.get("role",""),
                  1 if d.get("can_bulk_upload") else 0))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/team/<int:tid>",methods=["PATCH"])
@login_required
def update_team_member(tid):
    if not session.get("is_admin"):
        return jsonify({"error":"Admin only"}),403
    d=request.json or {}
    allowed=["name","email","phone","role","can_bulk_upload"]
    sets=", ".join(f"{k}=?" for k in d if k in allowed)
    vals=[]
    for k,v in d.items():
        if k not in allowed:
            continue
        if k == "email":
            v = (v or "").lower()
        if k == "can_bulk_upload":
            v = 1 if v else 0
        vals.append(v)
    if not sets:
        return jsonify({"error":"nothing to update"}),400
    vals.append(tid)
    conn=get_db()
    conn.execute(f"UPDATE team_members SET {sets} WHERE id=?", vals)
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/team/bulk",methods=["POST"])
@login_required
def bulk_add_team():
    if 'file' not in request.files:
        return jsonify({"error":"No file attached"}),400
    f = request.files['file']
    data = f.read()
    fname = f.filename.lower()
    rows = []
    if fname.endswith('.csv'):
        import io as io_module
        reader = csv.DictReader(io_module.StringIO(data.decode('utf-8', errors='ignore')))
        for row in reader:
            rows.append(row)
    elif fname.endswith(('.xlsx','.xls')):
        try:
            import openpyxl
            import io as io_module
            wb = openpyxl.load_workbook(io_module.BytesIO(data))
            ws = wb.active
            headers = [h.value or h for h in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell for cell in row):
                    d = dict(zip(headers, row))
                    rows.append(d)
        except: pass
    if not rows:
        return jsonify({"error":"No valid data found in file"}),400
    conn=get_db()
    added = 0; skipped = 0
    for row in rows:
        name = (row.get("Name") or row.get("name") or row.get("Name ","") or "").strip()
        email = (row.get("Email") or row.get("email") or row.get("Email Id") or row.get("Email-Id") or row.get("Email Id ") or row.get("Email-id") or "").strip().lower()
        phone = (row.get("Phone") or row.get("phone") or row.get("Phone Number") or row.get("Phone number") or row.get("Mobile") or "").strip().replace(",", "")
        role = (row.get("Role") or row.get("role") or row.get("Designation") or row.get("Title") or "").strip()
        if name and email:
            try:
                conn.execute("INSERT INTO team_members (name,email,phone,role,is_fixed) VALUES (?,?,?,?,1)",
                             (name,email,phone,role))
                added += 1
            except:
                skipped += 1
        else:
            skipped += 1
    conn.commit(); conn.close()
    return jsonify({"ok":True,"added":added,"skipped":skipped})

@app.route("/api/team/<int:tid>",methods=["DELETE"])
@login_required
def del_team(tid):
    if not session.get("is_admin"):
        return jsonify({"error":"Admin only"}),403
    conn=get_db(); conn.execute("DELETE FROM team_members WHERE id=?",(tid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# ── User Management ───────────────────────────────────────────────────────────
@app.route("/api/users",methods=["GET"])
@login_required
def get_users():
    if not session.get("is_admin"):
        return jsonify({"error":"Admin only"}),403
    conn=get_db()
    rows=conn.execute("SELECT id,username,is_admin,is_bulk_admin,is_active,created_at FROM app_users ORDER BY created_at DESC").fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/users",methods=["POST"])
@login_required
def add_user():
    if not session.get("is_admin"):
        return jsonify({"error":"Admin only"}),403
    d=request.json
    username = (d.get("username") or "").strip().lower()
    password = (d.get("password") or "").strip()
    if not username or not password:
        return jsonify({"error":"Username and password required"}),400
    password_hash = hash_password(password)
    conn=get_db()
    try:
        conn.execute("INSERT INTO app_users (username,password,is_admin,is_bulk_admin) VALUES (?,?,?,?)",
                     (username,password_hash,1 if d.get("is_admin") else 0,1 if d.get("is_bulk_admin") else 0))
        conn.commit()
        conn.close()
        return jsonify({"ok":True})
    except:
        conn.close()
        return jsonify({"error":"Username already exists"}),400

@app.route("/api/users/<int:uid>",methods=["DELETE"])
@login_required
def del_user(uid):
    if not session.get("is_admin"):
        return jsonify({"error":"Admin only"}),403
    if uid == session.get("user_id"):
        return jsonify({"error":"Cannot delete yourself"}),400
    conn=get_db(); conn.execute("DELETE FROM app_users WHERE id=? AND is_admin=0",(uid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/users/create-from-team", methods=["POST"])
@login_required
def create_users_from_team():
    if not session.get("is_admin"):
        return jsonify({"error":"Admin only"}),403
    
    conn = get_db()
    
    # Get all team members without app users
    rows = conn.execute("""
        SELECT t.id, t.name, t.email, t.role 
        FROM team_members t 
        LEFT JOIN app_users u ON t.id = u.team_member_id 
        WHERE u.id IS NULL AND t.email IS NOT NULL AND t.email != ''
        ORDER BY t.name
    """).fetchall()
    
    created = []
    skipped = []
    
    for member in rows:
        name = member["name"]
        email = member["email"]
        role = member["role"]
        
        # Generate username from name
        username = name.lower().replace(" ", "").replace(".", "")
        
        temp_password = secrets.token_urlsafe(10)
        password_hash = hash_password(temp_password)
        
        # Reetu is admin
        is_admin = 1 if name.lower() == "reetu" else 0
        
        try:
            conn.execute("""
                INSERT INTO app_users (username, password, team_member_id, is_admin) 
                VALUES (?, ?, ?, ?)
            """, (username, password_hash, member["id"], is_admin))
            created.append({
                "name": name,
                "username": username,
                "email": email,
                "role": role,
                "is_admin": is_admin,
                "password": temp_password
            })
        except Exception as e:
            skipped.append({"name": name, "error": str(e)})
    
    conn.commit()
    conn.close()
    
    return jsonify({"ok": True, "created": created, "skipped": skipped, "total_created": len(created)})

@app.route("/api/users/<int:uid>",methods=["PATCH"])
@login_required
def update_user(uid):
    if not session.get("is_admin"):
        return jsonify({"error":"Admin only"}),403
    d=request.json
    conn=get_db()
    if d.get("password"):
        password_hash = hash_password(d["password"])
        conn.execute("UPDATE app_users SET password=? WHERE id=?",(password_hash,uid))
    if d.get("is_active") is not None:
        conn.execute("UPDATE app_users SET is_active=? WHERE id=?",(1 if d["is_active"] else 0,uid))
    if d.get("is_bulk_admin") is not None:
        conn.execute("UPDATE app_users SET is_bulk_admin=? WHERE id=?",(1 if d["is_bulk_admin"] else 0,uid))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# ── Candidate CRUD ────────────────────────────────────────────────────────────
@app.route("/api/candidate/<int:cid>",methods=["PATCH"])
@login_required
def update_candidate(cid):
    data=request.json
    allowed=["candidate_name","email_addr","phone","current_company","current_role",
             "experience_years","key_skills","notice_period","current_salary",
             "expected_salary","current_location","preferred_location","remarks","role_name","status","requirement_id"]
    
    # Ownership check for non-admins
    if not session.get("is_admin"):
        conn = get_db()
        c = conn.execute("SELECT sourcer_id FROM candidates WHERE id=?", (cid,)).fetchone()
        conn.close()
        if not c or c["sourcer_id"] != session.get("team_member_id"):
            return jsonify({"error": "Permission denied"}), 403
    elif data.get("sourcer_id"):
        allowed.append("sourcer_id")
        conn = get_db()
        member = conn.execute("SELECT name,email FROM team_members WHERE id=?", (data.get("sourcer_id"),)).fetchone()
        conn.close()
        if member:
            data["recruiter_name"] = member["name"]
            data["recruiter_email"] = (member["email"] or "").lower()
            allowed.extend(["recruiter_name", "recruiter_email"])
    
    sets=", ".join(f"{k}=?" for k in data if k in allowed)
    vals=[v for k,v in data.items() if k in allowed]+[cid]
    if not sets: return jsonify({"error":"nothing to update"}),400
    conn=get_db()
    conn.execute(f"UPDATE candidates SET {sets}, updated_at=datetime('now','localtime') WHERE id=?",vals)
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/candidate",methods=["POST"])
@login_required
def add_single_candidate():
    d=request.json or {}
    required_fields = {
        "requirement_id": "Requirement",
        "candidate_name": "Candidate name",
        "email_addr": "Email",
        "phone": "Phone",
        "current_company": "Current company",
        "current_role": "Current role",
        "experience_years": "Experience",
        "key_skills": "Skills",
        "notice_period": "Notice period",
        "current_salary": "Current salary",
        "expected_salary": "Expected salary",
        "current_location": "Current location",
        "preferred_location": "Preferred location",
        "cv_filename": "CV/Resume"
    }
    missing_required = [
        label for field, label in required_fields.items()
        if not str(d.get(field) or "").strip()
    ]
    if not (d.get("sourcer_id") or session.get("team_member_id")):
        missing_required.append("Recruiter")
    if missing_required:
        return jsonify({
            "ok": False,
            "error": "Missing required fields: " + ", ".join(missing_required),
            "missing_fields": missing_required
        }), 400
    
    # Auto-populate recruiter info from session if not provided
    recruiter_name = d.get("recruiter_name","") or session.get("recruiter_name","")
    recruiter_email = d.get("recruiter_email","") or session.get("recruiter_email","")
    
    if not recruiter_name:
        recruiter_name = "Unknown"
    if not recruiter_email:
        recruiter_email = "unknown@system.com"
    
    conn = get_db()
    is_dup, dup_id, dup_why = check_dup(conn, d)
    missing = check_missing(d)
    
    if is_dup:
        conn.close()
        return jsonify({"ok":False,"error":f"Duplicate candidate! Already exists as ID #{dup_id} (matched by {dup_why})"}), 400
    
    # CV fields
    cv_filename = d.get("cv_filename","")
    cv_url      = d.get("cv_url","")
    cv_public_id = d.get("cv_public_id","")
    sourcer_id = d.get("sourcer_id") or session.get("team_member_id")
    
    cid = conn.execute("""INSERT INTO candidates
        (upload_batch,recruiter_name,recruiter_email,role_name,candidate_name,
         email_addr,phone,current_company,current_role,experience_years,key_skills,
         notice_period,current_salary,expected_salary,current_location,
         preferred_location,remarks,cv_filename,cv_url,cv_public_id,cv_summary,
         status,tags,is_duplicate,duplicate_of,missing_info,job_id,sourcer_id,requirement_id)
         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        ("single_"+datetime.now().strftime("%Y%m%d%H%M%S"),
         recruiter_name,recruiter_email,d.get("role_name",""),d.get("candidate_name",""),
         d.get("email_addr",""),d.get("phone",""),
         d.get("current_company",""),d.get("current_role",""),
         d.get("experience_years",""),d.get("key_skills",""),
         d.get("notice_period",""),d.get("current_salary",""),
         d.get("expected_salary",""),d.get("current_location",""),
         d.get("preferred_location",""),d.get("remarks",""),
         cv_filename,cv_url,cv_public_id,d.get("cv_summary",""),
         d.get("status","New"),d.get("tags",""),
         1 if is_dup else 0,
         dup_id if is_dup else None,
         ",".join(missing) if missing else None,
         d.get("job_id",""),
         sourcer_id,
         d.get("requirement_id"))).lastrowid
    conn.commit()
    conn.commit()
    if is_dup:
        conn.execute("INSERT INTO alerts (alert_type,message,candidate_id,recruiter_email) VALUES (?,?,?,?)",
            ("duplicate",f"Duplicate: {d.get('candidate_name','?')} via {dup_why} (original ID #{dup_id})",cid,recruiter_email))
    if missing:
        conn.execute("INSERT INTO alerts (alert_type,message,candidate_id,recruiter_email) VALUES (?,?,?,?)",
            ("missing_info",f"Missing {','.join(missing)} for {d.get('candidate_name','unnamed')}",cid,recruiter_email))
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":cid})

@app.route("/api/candidate/<int:cid>/status", methods=["PATCH"])
@login_required
def update_status(cid):
    data   = request.json
    status = data.get("status","")
    if not status: return jsonify({"error":"Status required"}),400
    conn = get_db()
    c = conn.execute(
        "SELECT recruiter_email, role_name FROM candidates WHERE id=?",
        (cid,)
    ).fetchone()
    if not c:
        conn.close()
        return jsonify({"error": "Not found"}), 404

    if not session.get("is_admin"):
        recruiter_email = session.get("recruiter_email")
        if not recruiter_email or (c["recruiter_email"] or "").lower() != recruiter_email.lower():
            conn.close()
            return jsonify({"error": "Permission denied"}), 403
    
    # Get candidate role to find custom pipeline
    role_name = c["role_name"] or ""
    # Get valid statuses for this role
    pipe = conn.execute("SELECT status_list FROM pipelines WHERE role_name=? LIMIT 1",(role_name,)).fetchone()
    if not pipe:
        pipe = conn.execute("SELECT status_list FROM pipelines WHERE is_default=1 LIMIT 1").fetchone()
    if pipe:
        valid_statuses = json.loads(pipe[0])
        if status not in valid_statuses:
            conn.close()
            return jsonify({"error":f"Invalid status. Must be one of: {', '.join(valid_statuses)}"}),400
    conn.execute("UPDATE candidates SET status=?, updated_at=datetime('now','localtime') WHERE id=?",
                 (status, cid))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/candidate/<int:cid>",methods=["GET"])
@login_required
def get_candidate(cid):
    conn = get_db()

    if not session.get("is_admin"):
        recruiter_email = session.get("recruiter_email")
        row = conn.execute(
            """SELECT c.*, r.title as requirement_title, r.client_name as client_name
               FROM candidates c
               LEFT JOIN requirements r ON c.requirement_id = r.id
               WHERE c.id=? AND lower(c.recruiter_email)=?""",
            (cid, recruiter_email.lower())
        ).fetchone()
    else:
        row = conn.execute(
            """SELECT c.*, r.title as requirement_title, r.client_name as client_name
               FROM candidates c
               LEFT JOIN requirements r ON c.requirement_id = r.id
               WHERE c.id=?""",
            (cid,)
        ).fetchone()

    conn.close()

    if row:
        return jsonify(dict(row))

    return jsonify({"error":"Not found"}),404

@app.route("/api/candidate/<int:cid>",methods=["DELETE"])
@login_required
def del_candidate(cid):
    conn = get_db()
    c = conn.execute(
        "SELECT recruiter_email FROM candidates WHERE id=?",
        (cid,)
    ).fetchone()
    if not c:
        conn.close()
        return jsonify({"error": "Not found"}), 404

    if not session.get("is_admin"):
        recruiter_email = session.get("recruiter_email")
        if not recruiter_email or (c["recruiter_email"] or "").lower() != recruiter_email.lower():
            conn.close()
            return jsonify({"error": "Permission denied"}), 403

    conn.execute("DELETE FROM candidates WHERE id=?",(cid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# ── Hiring Managers ───────────────────────────────────────────────────────────
@app.route("/api/hiring_managers",methods=["GET"])
@login_required
def get_hiring_managers():
    conn=get_db()
    rows=conn.execute("SELECT * FROM hiring_managers ORDER BY name").fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/hiring_managers",methods=["POST"])
@login_required
def add_hiring_manager():
    d=request.json
    if not d.get("name"): return jsonify({"error":"Name required"}),400
    conn=get_db()
    hid = conn.execute("INSERT INTO hiring_managers (name,email,company) VALUES (?,?,?)",
                       (d["name"],d.get("email",""),d.get("company",""))).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":hid})

@app.route("/api/hiring_managers/<int:hid>",methods=["DELETE"])
@login_required
def del_hiring_manager(hid):
    conn=get_db(); conn.execute("DELETE FROM hiring_managers WHERE id=?",(hid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# ── Pipelines ──────────────────────────────────────────────────────────────
@app.route("/api/pipelines",methods=["GET"])
@login_required
def get_pipelines():
    conn=get_db()
    rows=conn.execute("SELECT * FROM pipelines ORDER BY role_name").fetchall()
    result=[dict(r) for r in rows]
    for r in result: r["status_list"]=json.loads(r["status_list"])
    conn.close(); return jsonify(result)

@app.route("/api/pipelines",methods=["POST"])
@login_required
def add_pipeline():
    d=request.json
    if not d.get("role_name") or not d.get("status_list"):
        return jsonify({"error":"role_name and status_list required"}),400
    conn=get_db()
    pid=conn.execute("INSERT INTO pipelines (role_name,status_list) VALUES (?,?)",
                     (d["role_name"],json.dumps(d["status_list"]))).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":pid})

@app.route("/api/pipelines/<int:pid>",methods=["PATCH"])
@login_required
def update_pipeline(pid):
    d=request.json
    if "status_list" in d:
        conn=get_db()
        conn.execute("UPDATE pipelines SET status_list=? WHERE id=?",(json.dumps(d["status_list"]),pid))
        conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/pipelines/<int:pid>",methods=["DELETE"])
@login_required
def del_pipeline(pid):
    conn=get_db(); conn.execute("DELETE FROM pipelines WHERE id=?",(pid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/pipelines/for_role")
@login_required
def get_pipeline_for_role():
    role=request.args.get("role","")
    conn=get_db()
    row=conn.execute("SELECT status_list FROM pipelines WHERE role_name=? LIMIT 1",(role,)).fetchone()
    if not row:
        row=conn.execute("SELECT status_list FROM pipelines WHERE is_default=1 LIMIT 1").fetchone()
    conn.close()
    if row: return jsonify(json.loads(row[0]))
    return jsonify(["New","Shortlisted","Feedback Pending","Offered","On Hold","Joined","Rejected"])

# ── Jobs ─────────────────────────────────────────────────────────────────────
@app.route("/api/jobs",methods=["GET"])
@login_required
def get_jobs():
    conn=get_db()
    rows=conn.execute("SELECT * FROM jobs ORDER BY created_at DESC").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/jobs",methods=["POST"])
@login_required
def create_job():
    d=request.json
    if not d.get("title"):
        return jsonify({"error":"title required"}),400
    conn=get_db()
    # Generate job_id if not provided
    job_id = d.get("job_id") or f"JOB-{conn.execute('SELECT COUNT(*) FROM jobs').fetchone()[0] + 1:04d}"
    pid=conn.execute("INSERT INTO jobs (job_id,title,description,client_name,status) VALUES (?,?,?,?,?)",
                     (job_id,d["title"],d.get("description",""),d.get("client_name",""),d.get("status","Open"))).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":pid})

@app.route("/api/jobs/<int:jid>",methods=["PATCH"])
@login_required
def update_job(jid):
    d=request.json
    conn=get_db()
    if "title" in d: conn.execute("UPDATE jobs SET title=? WHERE id=?",(d["title"],jid))
    if "description" in d: conn.execute("UPDATE jobs SET description=? WHERE id=?",(d["description"],jid))
    if "client_name" in d: conn.execute("UPDATE jobs SET client_name=? WHERE id=?",(d["client_name"],jid))
    if "status" in d: conn.execute("UPDATE jobs SET status=? WHERE id=?",(d["status"],jid))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/jobs/<int:jid>",methods=["DELETE"])
@login_required
def del_job(jid):
    conn=get_db(); conn.execute("DELETE FROM jobs WHERE id=?",(jid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# ── Requirements ───────────────────────────────────────────────────────────
@app.route("/api/requirements",methods=["GET"])
@login_required
def get_requirements():
    conn=get_db()
    rows=conn.execute("""
        SELECT r.*, 
               ts.name as sourcer_name, 
               tr.name as recruiter_name,
               (SELECT COUNT(*) FROM candidates c WHERE c.requirement_id=r.id) as submissions
        FROM requirements r 
        LEFT JOIN team_members ts ON r.assigned_sourcer_id=ts.id
        LEFT JOIN team_members tr ON r.assigned_recruiter_id=tr.id
        ORDER BY r.created_at DESC
    """).fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/requirements/search", methods=["GET"])
@login_required
def search_requirements():
    q = (request.args.get("q") or "").strip().lower()
    if len(q) < 2:
        return jsonify([])
    try:
        limit = max(1, min(int(request.args.get("limit") or 25), 50))
    except (TypeError, ValueError):
        limit = 25
    like = f"%{q}%"
    conn = get_db()
    rows = conn.execute("""
        SELECT r.id, r.title, r.client_name, COALESCE(NULLIF(TRIM(r.status),''),'New') AS status
        FROM requirements r
        WHERE COALESCE(NULLIF(TRIM(r.status),''),'New') IN ('New','Open','In Progress')
          AND (
            lower(COALESCE(r.title,'')) LIKE ?
            OR lower(COALESCE(r.client_name,'')) LIKE ?
            OR CAST(r.id AS TEXT)=?
          )
        ORDER BY
          CASE COALESCE(NULLIF(TRIM(r.status),''),'New')
            WHEN 'Open' THEN 1
            WHEN 'In Progress' THEN 2
            WHEN 'New' THEN 3
            ELSE 4
          END,
          r.created_at DESC
        LIMIT ?
    """, (like, like, q, limit)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/requirements",methods=["POST"])
@login_required
def add_requirement():
    d=request.json
    conn=get_db()
    title = (d.get("title") or "").strip()
    client_name = (d.get("client_name") or "").strip()
    if not title:
        conn.close()
        return jsonify({"ok":False,"error":"Requirement title is required"}),400
    duplicate = conn.execute(
        "SELECT id FROM requirements WHERE lower(trim(title))=lower(trim(?)) AND lower(trim(COALESCE(client_name,'')))=lower(trim(?)) LIMIT 1",
        (title, client_name)
    ).fetchone()
    if duplicate:
        conn.close()
        return jsonify({"ok":False,"error":"A requirement with this title and client already exists."}),409
    cursor=conn.execute("""INSERT INTO requirements 
        (title,description,client_name,assigned_sourcer_id,assigned_recruiter_id,daily_target,status,created_by)
        VALUES (?,?,?,?,?,?,?,?)""",
        (title,d.get("description"),client_name,d.get("assigned_sourcer_id"),
         d.get("assigned_recruiter_id"),d.get("daily_target",3),d.get("status","Open"),
         session.get("username")))
    rid=cursor.lastrowid
    
    # Auto-add default screening checks if none provided
    checks = d.get("checks",[])
    if not checks:
        default_checks = [
            "Technical Skills", "Years of Relevant Experience", "Within Given Budget",
            "Notice Period", "Location", "Non-Poachable Employee", "Updated CV"
        ]
        for i, check_name in enumerate(default_checks):
            conn.execute("""INSERT INTO requirement_checks 
                (requirement_id,check_name,check_description,check_type,pass_criteria,sort_order)
                VALUES (?,?,?,?,?,?)""",
                (rid, check_name, "", "boolean", "Yes", i))
    else:
        for i, check in enumerate(checks):
            check_name = check.get("check_name") or check.get("name")
            if not check_name:
                continue
            conn.execute("""INSERT INTO requirement_checks 
                (requirement_id,check_name,check_description,check_type,pass_criteria,sort_order)
                VALUES (?,?,?,?,?,?)""",
                (rid, check_name, check.get("description"), check.get("check_type") or check.get("type","boolean"), 
                 check.get("pass_criteria"), check.get("sort_order", i)))
    
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":rid})

@app.route("/api/requirements/<int:rid>",methods=["GET"])
@login_required
def get_requirement(rid):
    conn=get_db()
    row=conn.execute("SELECT * FROM requirements WHERE id=?",(rid,)).fetchone()
    conn.close()
    if row:
        return jsonify(dict(row))
    return jsonify({"error":"Not found"}),404

@app.route("/api/requirements/<int:rid>",methods=["PATCH"])
@login_required
def update_requirement(rid):
    d=request.json
    conn=get_db()
    current = conn.execute("SELECT title, client_name FROM requirements WHERE id=?", (rid,)).fetchone()
    if not current:
        conn.close()
        return jsonify({"ok":False,"error":"Requirement not found"}),404
    new_title = (d.get("title", current["title"]) or "").strip()
    new_client = (d.get("client_name", current["client_name"]) or "").strip()
    duplicate = conn.execute(
        """SELECT id FROM requirements
           WHERE id<>? AND lower(trim(title))=lower(trim(?))
           AND lower(trim(COALESCE(client_name,'')))=lower(trim(?)) LIMIT 1""",
        (rid, new_title, new_client)
    ).fetchone()
    if duplicate:
        conn.close()
        return jsonify({"ok":False,"error":"A requirement with this title and client already exists."}),409
    if "title" in d: conn.execute("UPDATE requirements SET title=? WHERE id=?",(new_title,rid))
    if "description" in d: conn.execute("UPDATE requirements SET description=? WHERE id=?",(d["description"],rid))
    if "client_name" in d: conn.execute("UPDATE requirements SET client_name=? WHERE id=?",(new_client,rid))
    if "location" in d: conn.execute("UPDATE requirements SET location=? WHERE id=?",(d["location"],rid))
    if "remote" in d: conn.execute("UPDATE requirements SET remote=? WHERE id=?",(1 if d["remote"] else 0,rid))
    if "assigned_sourcer_id" in d: conn.execute("UPDATE requirements SET assigned_sourcer_id=? WHERE id=?",(d["assigned_sourcer_id"],rid))
    if "assigned_recruiter_id" in d: conn.execute("UPDATE requirements SET assigned_recruiter_id=? WHERE id=?",(d["assigned_recruiter_id"],rid))
    if "daily_target" in d: conn.execute("UPDATE requirements SET daily_target=? WHERE id=?",(d["daily_target"],rid))
    if "status" in d: conn.execute("UPDATE requirements SET status=? WHERE id=?",(d["status"],rid))
    conn.execute("UPDATE requirements SET updated_at=datetime('now','localtime') WHERE id=?",(rid,))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/requirements/<int:rid>/upload_jd", methods=["POST"])
@login_required
def upload_requirement_jd(rid):
    if "file" not in request.files:
        return jsonify({"error": "No JD file uploaded"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "Empty filename"}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in [".pdf", ".doc", ".docx", ".txt"]:
        return jsonify({"error": "Invalid JD file type"}), 400
    conn = get_db()
    exists = conn.execute("SELECT id FROM requirements WHERE id=?", (rid,)).fetchone()
    if not exists:
        conn.close()
        return jsonify({"error": "Requirement not found"}), 404
    upload_folder = os.path.join(app.root_path, "uploads")
    os.makedirs(upload_folder, exist_ok=True)
    safe_name = secure_filename(f.filename)
    saved_name = f"requirement_jd_{rid}_{int(datetime.now().timestamp())}_{safe_name}"
    file_path = os.path.join(upload_folder, saved_name)
    f.save(file_path)
    file_url = f"/uploads/{saved_name}"
    conn.execute(
        """UPDATE requirements
           SET jd_filename=?, jd_url=?, jd_public_id=?, updated_at=datetime('now','localtime')
           WHERE id=?""",
        (safe_name, file_url, saved_name, rid)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "filename": safe_name, "url": file_url, "public_id": saved_name})

@app.route("/api/requirements/<int:rid>",methods=["DELETE"])
@login_required
def del_requirement(rid):
    conn=get_db()
    conn.execute("UPDATE candidates SET requirement_id=NULL WHERE requirement_id=?",(rid,))
    conn.execute("DELETE FROM requirement_submissions WHERE requirement_id=?",(rid,))
    conn.execute("DELETE FROM requirement_checks WHERE requirement_id=?",(rid,))
    conn.execute("DELETE FROM requirements WHERE id=?",(rid,))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/requirements/<int:rid>/checks",methods=["GET"])
@login_required
def get_requirement_checks(rid):
    conn=get_db()
    rows=conn.execute("SELECT * FROM requirement_checks WHERE requirement_id=? ORDER BY sort_order",(rid,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/requirements/<int:rid>/checks",methods=["POST"])
@login_required
def add_requirement_check(rid):
    d=request.json
    conn=get_db()
    cursor=conn.execute("""INSERT INTO requirement_checks 
        (requirement_id,check_name,check_description,check_type,pass_criteria,is_mandatory,sort_order)
        VALUES (?,?,?,?,?,?,?)""",
        (rid,d.get("check_name"),d.get("check_description"),d.get("check_type","boolean"),
         d.get("pass_criteria"),1 if d.get("is_mandatory",True) else 0,
         d.get("sort_order",0)))
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":cursor.lastrowid})

@app.route("/api/requirements/<int:rid>/checks/<int:cid>",methods=["DELETE"])
@login_required
def del_requirement_check(rid,cid):
    conn=get_db()
    conn.execute("DELETE FROM requirement_checks WHERE id=?",(cid,))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

# ── Requirement Submissions ─────────────────────────────────────────────────────
@app.route("/api/submissions",methods=["GET"])
@login_required
def get_submissions():
    conn=get_db()
    rows=conn.execute("""
        SELECT rs.*, r.title as requirement_title, r.client_name,
               c.candidate_name, c.email_addr, c.phone, c.current_company, c.current_role
        FROM requirement_submissions rs
        LEFT JOIN requirements r ON rs.requirement_id=r.id
        LEFT JOIN candidates c ON rs.candidate_id=c.id
        ORDER BY rs.submitted_at DESC
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/submissions",methods=["POST"])
@login_required
def add_submission():
    d=request.json
    conn=get_db()
    
    # Create or get candidate
    candidate_id = d.get("candidate_id")
    if not candidate_id:
        # Create new candidate from data
        cursor = conn.execute("""INSERT INTO candidates 
            (candidate_name,email_addr,phone,current_company,current_role,experience_years,key_skills,notice_period,current_salary,expected_salary,current_location,requirement_id,recruiter_name,recruiter_email)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (d.get("candidate_name"),d.get("email_addr"),d.get("phone"),d.get("current_company"),
             d.get("current_role"),d.get("experience_years"),d.get("key_skills"),d.get("notice_period"),
             d.get("current_salary"),d.get("expected_salary"),d.get("current_location"),
             d.get("requirement_id"),session.get("recruiter_name"),session.get("recruiter_email")))
        candidate_id = cursor.lastrowid
    
    # Create submission
    cursor = conn.execute("""INSERT INTO requirement_submissions
        (candidate_id,requirement_id,sourcer_id,submitted_by,status,notes)
        VALUES (?,?,?,?,?,?)""",
        (candidate_id, d.get("requirement_id"), d.get("sourcer_id"), 
         session.get("username"), 'Submitted', d.get("notes","")))
    submission_id = cursor.lastrowid
    
    # Add check results
    checks = d.get("checks",[])
    for check in checks:
        conn.execute("""INSERT INTO submission_checks 
            (submission_id,check_id,check_name,passed,notes)
            VALUES (?,?,?,?,?)""",
            (submission_id, check.get("check_id"), check.get("check_name"), 
             1 if check.get("passed") else 0, check.get("notes","")))
    
    conn.commit()
    
    # Get requirement to update submission count
    req = conn.execute("SELECT * FROM requirements WHERE id=?", (d.get("requirement_id"),)).fetchone()
    conn.close()
    
    return jsonify({"ok":True,"id":submission_id,"candidate_id":candidate_id})

@app.route("/api/submissions/<int:sid>",methods=["PATCH"])
@login_required
def update_submission(sid):
    d=request.json
    conn=get_db()
    if "status" in d:
        conn.execute("UPDATE requirement_submissions SET status=?, updated_at=datetime('now','localtime') WHERE id=?",(d["status"],sid))
    if "recruiter_feedback" in d:
        conn.execute("UPDATE requirement_submissions SET recruiter_feedback=?, feedback_by=?, updated_at=datetime('now','localtime') WHERE id=?",
            (d["recruiter_feedback"], session.get("user_id"), sid))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/submissions/<int:sid>/checks",methods=["GET"])
@login_required
def get_submission_checks(sid):
    conn=get_db()
    rows=conn.execute("SELECT * FROM submission_checks WHERE submission_id=?",(sid,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])
    return jsonify({"ok":True})

# ── Automation Rules ──────────────────────────────────────────────────────
@app.route("/api/automation_rules",methods=["GET"])
@login_required
def get_automation_rules():
    conn=get_db()
    rows=conn.execute("SELECT * FROM automation_rules WHERE is_active=1 ORDER BY id").fetchall()
    result=[dict(r) for r in rows]
    for r in result: r["action_config"]=json.loads(r["action_config"] or "{}")
    conn.close(); return jsonify(result)

@app.route("/api/automation_rules",methods=["POST"])
@login_required
def add_automation_rule():
    d=request.json
    if not d.get("rule_name") or not d.get("trigger_type") or not d.get("action_type"):
        return jsonify({"error":"rule_name, trigger_type, action_type required"}),400
    conn=get_db()
    rid=conn.execute("INSERT INTO automation_rules (rule_name,trigger_type,trigger_value,action_type,action_config) VALUES (?,?,?,?,?)",
                     (d["rule_name"],d["trigger_type"],d.get("trigger_value",""),
                      d["action_type"],json.dumps(d.get("action_config",{})))).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":rid})

@app.route("/api/automation_rules/<int:rid>",methods=["DELETE"])
@login_required
def del_automation_rule(rid):
    conn=get_db(); conn.execute("DELETE FROM automation_rules WHERE id=?",(rid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# ── Interviews ────────────────────────────────────────────────────────────
@app.route("/api/interviews",methods=["GET"])
@login_required
def get_interviews():
    conn=get_db()
    rows=conn.execute("""SELECT i.*, c.candidate_name, c.role_name
                         FROM interviews i
                         JOIN candidates c ON i.candidate_id=c.id
                         ORDER BY scheduled_at DESC""").fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/interviews",methods=["POST"])
@login_required
def add_interview():
    d=request.json
    if not d.get("candidate_id") or not d.get("scheduled_at"):
        return jsonify({"error":"candidate_id and scheduled_at required"}),400
    conn=get_db()
    iid=conn.execute("""INSERT INTO interviews
        (candidate_id,interviewer_name,interviewer_email,scheduled_at,duration_mins,location,meeting_link,notes)
        VALUES (?,?,?,?,?,?,?,?)""",
        (d["candidate_id"],d.get("interviewer_name",""),d.get("interviewer_email",""),
         d["scheduled_at"],d.get("duration_mins",60),d.get("location",""),
         d.get("meeting_link",""),d.get("notes",""))).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":iid})

@app.route("/api/interviews/<int:iid>",methods=["PATCH"])
@login_required
def update_interview(iid):
    d=request.json
    allowed=["interviewer_name","interviewer_email","scheduled_at","duration_mins","location","meeting_link","status","notes"]
    sets=", ".join(f"{k}=?" for k in d if k in allowed)
    vals=[v for k,v in d.items() if k in allowed]+[iid]
    if not sets: return jsonify({"error":"nothing to update"}),400
    conn=get_db()
    conn.execute(f"UPDATE interviews SET {sets} WHERE id=?",vals)
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/interviews/<int:iid>",methods=["DELETE"])
@login_required
def del_interview(iid):
    conn=get_db(); conn.execute("DELETE FROM interviews WHERE id=?",(iid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# ── Approvals ─────────────────────────────────────────────────────────────
@app.route("/api/approvals",methods=["GET"])
@login_required
def get_approvals():
    conn=get_db()
    rows=conn.execute("""SELECT a.*, c.candidate_name, c.role_name
                         FROM approvals a
                         JOIN candidates c ON a.candidate_id=c.id
                         ORDER BY a.created_at DESC""").fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/approvals",methods=["POST"])
@login_required
def add_approval():
    d=request.json
    if not d.get("candidate_id") or not d.get("approval_type"):
        return jsonify({"error":"candidate_id and approval_type required"}),400
    conn=get_db()
    aid=conn.execute("""INSERT INTO approvals
        (candidate_id,requester_name,requester_email,approval_type,comments)
        VALUES (?,?,?,?,?)""",
        (d["candidate_id"],d.get("requester_name",""),d.get("requester_email",""),
         d["approval_type"],d.get("comments",""))).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":aid})

@app.route("/api/approvals/<int:aid>",methods=["PATCH"])
@login_required
def update_approval(aid):
    d=request.json
    conn=get_db()
    if "status" in d:
        conn.execute("UPDATE approvals SET status=?, reviewed_by=?, reviewed_at=datetime('now','localtime') WHERE id=?",
                     (d["status"],d.get("reviewed_by",""),aid))
    if "comments" in d:
        conn.execute("UPDATE approvals SET comments=? WHERE id=?",(d["comments"],aid))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/approvals/<int:aid>",methods=["DELETE"])
@login_required
def del_approval(aid):
    conn=get_db(); conn.execute("DELETE FROM approvals WHERE id=?",(aid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})


@app.route("/api/email_templates",methods=["POST"])
@login_required
def add_email_template():
    d=request.json
    if not d.get("name") or not d.get("subject") or not d.get("body"):
        return jsonify({"error":"name, subject, body required"}),400
    conn=get_db()
    tid=conn.execute("INSERT INTO email_templates (name,subject,body,trigger_event) VALUES (?,?,?,?)",
                     (d["name"],d["subject"],d["body"],d.get("trigger_event",""))).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":tid})

@app.route("/api/email_templates/<int:tid>",methods=["PATCH"])
@login_required
def update_email_template(tid):
    d=request.json
    allowed=["name","subject","body","trigger_event"]
    sets=", ".join(f"{k}=?" for k in d if k in allowed)
    vals=[v for k,v in d.items() if k in allowed]+[tid]
    if not sets: return jsonify({"error":"nothing to update"}),400
    conn=get_db()
    conn.execute(f"UPDATE email_templates SET {sets} WHERE id=?",vals)
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/email_templates/<int:tid>",methods=["DELETE"])
@login_required
def del_email_template(tid):
    conn=get_db(); conn.execute("DELETE FROM email_templates WHERE id=?",(tid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# ── Send Email ──────────────────────────────────────────────────────
def unwrap_template_braces(text, variables):
    output = text or ""
    for value in variables.values():
        clean = str(value or "").strip()
        if not clean:
            continue
        output = re.sub(r"\{\s*" + re.escape(clean) + r"\s*\}", clean, output)
    return output

def send_google_oauth_email(to_addr, subject, body):
    token = session.get("google_token") or {}
    access_token = token.get("access_token")
    if not access_token:
        return {"error": "Please log in with Google again before sending email."}
    creds = Credentials(
        token=access_token,
        refresh_token=token.get("refresh_token"),
        token_uri=token.get("token_uri") or "https://oauth2.googleapis.com/token",
        client_id=os.getenv("GOOGLE_CLIENT_ID"),
        client_secret=os.getenv("GOOGLE_CLIENT_SECRET"),
        scopes=["openid", "email", "profile", "https://www.googleapis.com/auth/gmail.send"]
    )
    if creds.expired and creds.refresh_token:
        creds.refresh(GoogleAuthRequest())
        token["access_token"] = creds.token
        session["google_token"] = token
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = session.get("email") or "me"
    msg["To"] = to_addr
    msg.attach(MIMEText(body, "plain"))
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
    service = google_api_build("gmail", "v1", credentials=creds, cache_discovery=False)
    sent = service.users().messages().send(userId="me", body={"raw": raw}).execute()
    return {"ok": True, "sent_to": to_addr, "message_id": sent.get("id"), "sender": session.get("email")}

def send_candidate_email(candidate_id, template_name, extra_vars={}):
    """Send email to candidate using a template"""
    conn=get_db()
    c=conn.execute("""SELECT c.*, r.client_name as client_name
                      FROM candidates c
                      LEFT JOIN requirements r ON c.requirement_id = r.id
                      WHERE c.id=?""",(candidate_id,)).fetchone()
    if not c: conn.close(); return {"error":"Candidate not found"}
    t=conn.execute("SELECT * FROM email_templates WHERE name=?",(template_name,)).fetchone()
    if not t: conn.close(); return {"error":"Template not found"}
    # Replace variables in subject and body
    variables={"candidate_name":c["candidate_name"] or "",
               "role_name":c["role_name"] or "",
               "role":c["role_name"] or c["current_role"] or "",
               "current_role":c["current_role"] or "",
               "status":c["status"] or "",
               "current_company":c["current_company"] or "",
               "company":c["client_name"] or c["current_company"] or "",
               "client_name":c["client_name"] or "",
               "email_addr":c["email_addr"] or "",
               "phone":c["phone"] or "",
               "recruiter_name":c["recruiter_name"] or "",
               **extra_vars}
    subject=t["subject"]
    body=t["body"]
    for k,v in variables.items():
        subject=subject.replace("{{"+k+"}}",str(v))
        body=body.replace("{{"+k+"}}",str(v))
        subject=subject.replace("{"+k+"}",str(v))
        body=body.replace("{"+k+"}",str(v))
    subject=unwrap_template_braces(subject, variables)
    body=unwrap_template_braces(body, variables)
    recipient=c["email_addr"]
    if not recipient: conn.close(); return {"error":"No email address"}
    try:
        if session.get("google_token"):
            sent_result = send_google_oauth_email(recipient, subject, body)
            if not sent_result.get("ok"):
                conn.close()
                return sent_result
        else:
            if not GMAIL_USER or not GMAIL_APP_PASS:
                conn.close()
                return {"error":"Please log in with Google again before sending email."}
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"] = GMAIL_USER
            msg["To"] = recipient
            msg.attach(MIMEText(body,"plain"))
            with smtplib.SMTP_SSL("smtp.gmail.com",465) as s:
                s.login(GMAIL_USER,GMAIL_APP_PASS)
                s.sendmail(GMAIL_USER,recipient,msg.as_string())
        # Log the email
        conn.execute("INSERT INTO email_log (candidate_id,template_name,recipient,subject,body,status) VALUES (?,?,?,?,?,'Sent')",
                     (candidate_id,template_name,recipient,subject,body))
        conn.commit(); conn.close()
        return {"ok":True,"sent_to":recipient}
    except Exception as e:
        conn.execute("INSERT INTO email_log (candidate_id,template_name,recipient,subject,body,status,error_msg) VALUES (?,?,?,?,?,'Failed',?)",
                     (candidate_id,template_name,recipient,subject,body,str(e)))
        conn.commit(); conn.close()
        return {"error":str(e)}

def send_custom_email(to_addr, subject, body, candidate_id=None, template_name="custom"):
    use_google = bool(session.get("google_token"))
    try:
        if use_google:
            sent_result = send_google_oauth_email(to_addr, subject, body)
            if not sent_result.get("ok"):
                return sent_result
        else:
            if not GMAIL_USER or not GMAIL_APP_PASS:
                return {"error":"Please log in with Google again before sending email."}
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"] = GMAIL_USER
            msg["To"] = to_addr
            msg.attach(MIMEText(body, "plain"))
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
                s.login(GMAIL_USER, GMAIL_APP_PASS)
                s.sendmail(GMAIL_USER, to_addr, msg.as_string())
        conn = get_db()
        conn.execute(
            "INSERT INTO email_log (candidate_id,template_name,recipient,subject,body,status) VALUES (?,?,?,?,?,'Sent')",
            (candidate_id, template_name, to_addr, subject, body)
        )
        conn.commit(); conn.close()
        return {"ok": True, "sent_to": to_addr, "sender": session.get("email") or GMAIL_USER}
    except Exception as e:
        conn = get_db()
        conn.execute(
            "INSERT INTO email_log (candidate_id,template_name,recipient,subject,body,status,error_msg) VALUES (?,?,?,?,?,'Failed',?)",
            (candidate_id, template_name, to_addr, subject, body, str(e))
        )
        conn.commit(); conn.close()
        return {"error": str(e)}

@app.route("/api/send_email", methods=["POST"])
@login_required
def api_send_email():
    d = request.json

    try:
        # The UI sends the final composed subject/body, so send that directly.
        # This avoids relying on an expired Google OAuth access token.
        if d.get("to") and d.get("subject") and d.get("body"):
            return jsonify(send_custom_email(
                d["to"],
                d["subject"],
                d["body"],
                d.get("candidate_id"),
                d.get("template_name") or "custom"
            ))

        # Template-only email flow
        if d.get("candidate_id") and d.get("template_name"):
            extra = d.get("variables", {})
            return jsonify(send_candidate_email(
                d["candidate_id"],
                d["template_name"],
                extra
            ))

        return jsonify({"error": "Invalid payload"}), 400

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/email_templates")
@login_required
def get_email_templates():
    conn = get_db()
    rows = conn.execute(
        "SELECT id,name,subject,body FROM email_templates ORDER BY name"
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/email_log",methods=["GET"])
@login_required
def get_email_log():
    conn=get_db()
    rows=conn.execute("""SELECT el.*, c.candidate_name 
                         FROM email_log el 
                         LEFT JOIN candidates c ON el.candidate_id=c.id 
                         ORDER BY sent_at DESC LIMIT 100""").fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

if __name__=="__main__":
    init_db()
    conn=get_db()
    # Create default admin user if not exists
    admin = conn.execute("SELECT id FROM app_users WHERE username='admin'").fetchone()
    if not admin and ADMIN_PASSWORD:
        conn.execute("INSERT INTO app_users (username,password,is_admin) VALUES (?,?,1)",
                     ("admin", hash_password(ADMIN_PASSWORD)))
        conn.commit()
        print("Created default admin user: admin")
    elif not admin:
        print("No default admin created. Set ADMIN_PASSWORD to create one.")
    conn.close()
    port = int(os.getenv("PORT",5001))
    print(f"✅  HR Guru ATS → http://localhost:{port}")
    app.run(debug=False,host="0.0.0.0",port=port)
