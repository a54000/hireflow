import os, re, hashlib, csv, io, sqlite3, json, smtplib, secrets, zipfile, traceback, difflib, tempfile, warnings, html, mimetypes, time, atexit, threading
import queue
import xml.etree.ElementTree as ET
import requests
from flask import send_from_directory
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from dotenv import load_dotenv
from flask import (Flask, render_template, jsonify, request, g,
                   send_file, session, redirect, url_for)
from datetime import datetime, date, timedelta
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.exceptions import HTTPException, RequestEntityTooLarge
from werkzeug.middleware.proxy_fix import ProxyFix
from functools import wraps
from authlib.integrations.flask_client import OAuth
from cryptography.fernet import Fernet, InvalidToken
from email.mime.text import MIMEText
from google.oauth2.credentials import Credentials
from google.oauth2 import service_account
from google.auth.transport.requests import Request as GoogleAuthRequest
from googleapiclient.discovery import build as google_api_build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseUpload

import base64
from urllib.parse import urlparse
load_dotenv()
from ats_schema import ensure_ats_pipeline_schema
from ats_pipeline import MATCH_PIPELINE_VERSION, parse_jd as parse_jd_structured, parse_resume as parse_resume_structured, run_hybrid_match, versioned_text_hash
from embedding_engine import deserialize_embedding, serialize_embedding
from services.match_analysis import build_match_dashboard
from services.match_pdf import build_match_pdf
from services.ai_screening_pdf import build_ai_screening_pdf
from services.screening_pdf import build_screening_questions_pdf
from skill_aliases import canonical_skill, skill_aliases_for

JD_CV_MATCHING_DISABLED = False
JD_CV_MATCHING_DISABLED_MESSAGE = "JD/CV matching is currently disabled because the matching logic is under review."


import sys
sys.path.insert(0, '/Users/surindersingh/Library/Python/3.9/lib/python3.9/site-packages')
print(f"Python: {sys.executable}", flush=True)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_OK = True
    print("openpyxl loaded!", flush=True)
except ImportError:
    XLSX_OK = False
    print("openpyxl NOT loaded", flush=True)

try:
    import cloudinary, cloudinary.uploader
    CLOUD_OK = True
except ImportError:
    CLOUD_OK = False

app = Flask(__name__)
#app = Flask(__name__, template_folder="/Users/surindersingh/Downloads/ats6-dev/templates")
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)
oauth = OAuth(app)
SLOW_REQUEST_LOG_SECONDS = float(os.getenv("ATS_SLOW_REQUEST_LOG_SECONDS", "0.75"))
PERF_RECORD_MIN_SECONDS = float(os.getenv("ATS_PERF_RECORD_MIN_SECONDS", "0"))
PERF_LOG_PATH_PREFIXES = (
    "/api/dashboard_summary",
    "/api/candidates",
    "/api/requirements",
    "/api/clients",
    "/api/me",
)
WRITE_PERF_DEBUG = os.getenv("ATS_WRITE_PERF_DEBUG", "0").strip().lower() in {"1", "true", "yes", "on"}

def perf_log(label, started_at, **fields):
    if not WRITE_PERF_DEBUG:
        return
    elapsed_ms = (time.perf_counter() - started_at) * 1000
    details = " ".join(f"{key}={value}" for key, value in fields.items() if value is not None)
    print(f"PERF {label} elapsed_ms={elapsed_ms:.1f} {details}".rstrip(), flush=True)


@app.before_request
def start_request_timer():
    g.request_started_at = time.perf_counter()


def record_performance_log(path, elapsed_ms, status_code, method=None):
    try:
        conn = get_db(timeout=2)
        conn.execute(
            """CREATE TABLE IF NOT EXISTS performance_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT,
                team_member_id INTEGER,
                recruiter_name TEXT,
                recruiter_email TEXT,
                method TEXT,
                path TEXT,
                endpoint TEXT,
                status_code INTEGER,
                elapsed_ms REAL,
                is_admin INTEGER DEFAULT 0,
                ip_address TEXT,
                user_agent TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )"""
        )
        conn.execute(
            """INSERT INTO performance_logs
               (username, team_member_id, recruiter_name, recruiter_email, method, path, endpoint,
                status_code, elapsed_ms, is_admin, ip_address, user_agent)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                session.get("username") or "",
                session.get("team_member_id"),
                session.get("recruiter_name") or "",
                session.get("recruiter_email") or session.get("email") or "",
                method or request.method,
                path,
                request.endpoint or "",
                int(status_code or 0),
                round(float(elapsed_ms or 0), 1),
                1 if session.get("is_admin") else 0,
                request.headers.get("CF-Connecting-IP") or request.remote_addr or "",
                (request.headers.get("User-Agent") or "")[:500],
            )
        )
        conn.commit()
        conn.close()
    except Exception as exc:
        print(f"PERF LOG DEBUG: unable to record performance log: {type(exc).__name__}: {exc}", flush=True)


@app.after_request
def log_slow_request(response):
    started = getattr(g, "request_started_at", None)
    if started is None:
        return response
    elapsed = time.perf_counter() - started
    path = request.path or ""
    if path.startswith(PERF_LOG_PATH_PREFIXES) and elapsed >= PERF_RECORD_MIN_SECONDS:
        record_performance_log(path, elapsed * 1000, response.status_code, request.method)
    if elapsed >= SLOW_REQUEST_LOG_SECONDS and path.startswith(PERF_LOG_PATH_PREFIXES):
        print(
            f"PERF {request.method} {path} status={response.status_code} "
            f"elapsed_ms={elapsed * 1000:.1f} user={session.get('username') or '-'}",
            flush=True,
        )
    return response
GOOGLE_OAUTH_SCOPE = "openid email profile https://www.googleapis.com/auth/gmail.send https://www.googleapis.com/auth/drive.readonly https://www.googleapis.com/auth/spreadsheets.readonly"

google = oauth.register(
    name='google',
    client_id=os.getenv("GOOGLE_CLIENT_ID"),
    client_secret=os.getenv("GOOGLE_CLIENT_SECRET"),
    authorize_url='https://accounts.google.com/o/oauth2/v2/auth',
    access_token_url='https://oauth2.googleapis.com/token',
    api_base_url='https://www.googleapis.com/oauth2/v3/',
    jwks_uri='https://www.googleapis.com/oauth2/v3/certs',
    client_kwargs={
        'scope': GOOGLE_OAUTH_SCOPE
    }
)

app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
app.config["PREFERRED_URL_SCHEME"] = os.getenv("PREFERRED_URL_SCHEME", "https")
SECRET_KEY = os.getenv("SECRET_KEY")
if not SECRET_KEY:
    raise RuntimeError("SECRET_KEY is required. Set it in your environment or .env file.")
app.secret_key = SECRET_KEY

@app.errorhandler(RequestEntityTooLarge)
def handle_file_too_large(error):
    if request.path.startswith("/api/"):
        return jsonify({
            "error": "Uploaded files are too large. Please keep the total upload under 50 MB and try again."
        }), 413
    return "Uploaded files are too large. Please keep the total upload under 50 MB and try again.", 413

@app.errorhandler(500)
def handle_internal_error(error):
    print("Unhandled server error:", error, flush=True)
    traceback.print_exc()
    if request.path.startswith("/api/"):
        return jsonify({
            "error": "Server error while processing this request. Please check the ATS terminal/logs for details."
        }), 500
    return "Server error while processing this request. Please check the ATS terminal/logs for details.", 500

@app.errorhandler(Exception)
def handle_unexpected_error(error):
    if isinstance(error, HTTPException):
        if request.path.startswith("/api/"):
            return jsonify({"error": error.description or error.name}), error.code
        return error
    print("Unhandled exception:", error, flush=True)
    traceback.print_exc()
    if request.path.startswith("/api/"):
        return jsonify({"error": str(error) or "Server error while processing this request."}), 500
    return "Server error while processing this request. Please check the ATS terminal/logs for details.", 500

BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
DB_PATH        = os.path.join(BASE_DIR, "ats.db")
APP_INSTANCE_LOCK_HANDLE = None
DB_WRITE_LOCK = threading.RLock()
ANALYTICS_SEMAPHORE = threading.BoundedSemaphore(3)
ANALYTICS_CACHE_LOCK = threading.Lock()
ANALYTICS_CACHE = {}
ANALYTICS_CACHE_TTL_SECONDS = 60
MATCH_DB_TASK_QUEUE = queue.Queue(maxsize=500)
MATCH_DB_WORKER_LOCK = threading.Lock()
MATCH_DB_WORKER_STARTED = False
AI_SCREENING_TASK_QUEUE = queue.Queue(maxsize=300)
AI_SCREENING_WORKER_LOCK = threading.Lock()
AI_SCREENING_WORKER_STARTED = False

def _match_db_worker():
    while True:
        task_name, fn, args, kwargs = MATCH_DB_TASK_QUEUE.get()
        try:
            print(f"MATCH DEBUG: background DB task started: {task_name}", flush=True)
            fn(*args, **kwargs)
            print(f"MATCH DEBUG: background DB task finished: {task_name}", flush=True)
        except Exception as exc:
            print(f"MATCH DEBUG: background DB task failed ({task_name}): {type(exc).__name__}: {exc}", flush=True)
            traceback.print_exc()
        finally:
            try:
                MATCH_DB_TASK_QUEUE.task_done()
            except Exception:
                pass

def ensure_match_db_worker():
    global MATCH_DB_WORKER_STARTED
    if MATCH_DB_WORKER_STARTED:
        return
    with MATCH_DB_WORKER_LOCK:
        if MATCH_DB_WORKER_STARTED:
            return
        worker = threading.Thread(target=_match_db_worker, name="match-db-worker", daemon=True)
        worker.start()
        MATCH_DB_WORKER_STARTED = True
        print("MATCH DEBUG: background DB worker started.", flush=True)

def queue_match_db_task(task_name, fn, *args, **kwargs):
    ensure_match_db_worker()
    try:
        MATCH_DB_TASK_QUEUE.put_nowait((task_name, fn, args, kwargs))
        print(f"MATCH DEBUG: queued background DB task: {task_name}", flush=True)
        return True
    except queue.Full:
        print(f"MATCH DEBUG: background DB queue full; running task inline: {task_name}", flush=True)
        fn(*args, **kwargs)
        return False

def _ai_screening_worker():
    while True:
        item = AI_SCREENING_TASK_QUEUE.get()
        task_name = "ai_screening"
        fn = None
        args = ()
        kwargs = {}
        if isinstance(item, tuple):
            if len(item) == 4:
                task_name, fn, args, kwargs = item
            elif len(item) == 5:
                task_name, fn, arg1, arg2, arg3 = item
                args = (arg1, arg2, arg3)
            elif len(item) >= 2:
                task_name = item[0]
                fn = item[1]
                if len(item) >= 3:
                    if isinstance(item[2], tuple):
                        args = item[2]
                    elif len(item) > 3:
                        args = tuple(item[2:])
                    else:
                        args = (item[2],)
                if len(item) >= 4 and isinstance(item[3], dict):
                    kwargs = item[3]
        if not fn:
            print(f"AI SCREENING DEBUG: invalid queued task shape: {type(item).__name__} len={len(item) if hasattr(item, '__len__') else '-'}", flush=True)
            try:
                AI_SCREENING_TASK_QUEUE.task_done()
            except Exception:
                pass
            continue
        try:
            print(f"AI SCREENING DEBUG: background task started: {task_name}", flush=True)
            fn(*args, **kwargs)
            print(f"AI SCREENING DEBUG: background task finished: {task_name}", flush=True)
        except Exception as exc:
            print(f"AI SCREENING DEBUG: background task failed ({task_name}): {type(exc).__name__}: {exc}", flush=True)
            traceback.print_exc()
        finally:
            try:
                AI_SCREENING_TASK_QUEUE.task_done()
            except Exception:
                pass

def ensure_ai_screening_worker():
    global AI_SCREENING_WORKER_STARTED
    if AI_SCREENING_WORKER_STARTED:
        return
    with AI_SCREENING_WORKER_LOCK:
        if AI_SCREENING_WORKER_STARTED:
            return
        worker = threading.Thread(target=_ai_screening_worker, name="ai-screening-worker", daemon=True)
        worker.start()
        AI_SCREENING_WORKER_STARTED = True
        print("AI SCREENING DEBUG: background screening worker started.", flush=True)

def queue_ai_screening_task(task_name, fn, *args, **kwargs):
    ensure_ai_screening_worker()
    try:
        AI_SCREENING_TASK_QUEUE.put_nowait((task_name, fn, args, kwargs))
        print(f"AI SCREENING DEBUG: queued background screening task: {task_name}", flush=True)
        return True
    except queue.Full:
        print(f"AI SCREENING DEBUG: screening queue full; running task inline: {task_name}", flush=True)
        fn(*args, **kwargs)
        return False

def schedule_candidate_ai_screening_retry(candidate_id, trigger, run_id, gemini_api_key, retry_count, delay_seconds):
    max_retries = max(0, int(float(os.getenv("GEMINI_SCREENING_RETRY_MAX_ATTEMPTS", "2") or 2)))
    if retry_count >= max_retries:
        return False
    delay_seconds = max(5, int(float(delay_seconds or 0) or 0))
    def _enqueue_retry():
        try:
            queue_ai_screening_task(
                "candidate_ai_screening_retry",
                run_candidate_ai_screening,
                candidate_id,
                trigger,
                run_id,
                gemini_api_key,
                retry_count + 1,
            )
        except Exception as exc:
            print(f"AI SCREENING DEBUG: failed to enqueue retry for candidate_id={candidate_id}: {exc}", flush=True)
    timer = threading.Timer(delay_seconds, _enqueue_retry)
    timer.daemon = True
    timer.start()
    print(
        f"AI SCREENING DEBUG: scheduled retry for candidate_id={candidate_id} run_id={run_id} in {delay_seconds}s (attempt {retry_count + 1}/{max_retries})",
        flush=True,
    )
    return True

def acquire_single_instance_lock():
    global APP_INSTANCE_LOCK_HANDLE
    lock_path = os.path.join(BASE_DIR, "ats.lock")
    handle = open(lock_path, "a+b")
    try:
        if os.name == "nt":
            import msvcrt
        else:
            import fcntl
        handle.seek(0)
        handle.write(str(os.getpid()).encode("ascii", errors="ignore"))
        handle.truncate()
        handle.flush()
        handle.seek(0)
        if os.name == "nt":
            msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
        else:
            fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
    except (OSError, ModuleNotFoundError) as e:
        try:
            handle.close()
        except Exception:
            pass
        raise RuntimeError("Another ATS instance is already running. Please stop the other process before starting a new one.") from e
    APP_INSTANCE_LOCK_HANDLE = handle
    return handle

@atexit.register
def release_single_instance_lock():
    global APP_INSTANCE_LOCK_HANDLE
    handle = APP_INSTANCE_LOCK_HANDLE
    if not handle:
        return
    try:
        try:
            handle.seek(0)
            if os.name == "nt":
                import msvcrt
                msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
            else:
                import fcntl
                fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
        except Exception:
            pass
        handle.close()
    finally:
        APP_INSTANCE_LOCK_HANDLE = None

CLOUDINARY_URL = os.getenv("CLOUDINARY_URL", "")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "")

# Gmail config for weekly summary
GMAIL_USER     = os.getenv("GMAIL_USER", "")       # your gmail address
GMAIL_APP_PASS = os.getenv("GMAIL_APP_PASS", "")   # gmail app password
ADMIN_EMAIL    = os.getenv("ADMIN_EMAIL", "")       # where to send the summary

# CV Parsing config
CV_PARSE_MODE = os.getenv("CV_PARSE_MODE", "free")  # "free" or "ai"
CV_AI_PROVIDER = os.getenv("CV_AI_PROVIDER", "")   # openai, gemini, claude
CV_AI_API_KEY = os.getenv("CV_AI_API_KEY", "")

# Candidate statuses
STATUSES = ["New", "Shortlisted", "Screening Pending", "Screen Rejected", 
            "Interviewed", "HM Rejected", "Offered", "Joined", "Dropped", "Duplicate", "OnHold"]

# Submission statuses
SUBMISSION_STATUSES = ["Pending", "Reviewing", "Interview Scheduled", "Selected", "Rejected", "Withdrawn"]

# â”€â”€ DB â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def get_db(timeout=60):
    conn = sqlite3.connect(DB_PATH, timeout=timeout)
    conn.row_factory = sqlite3.Row
    conn.execute(f"PRAGMA busy_timeout={int(max(1, float(timeout)) * 1000)}")
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("PRAGMA wal_autocheckpoint=1000")
    except sqlite3.OperationalError:
        pass
    return conn

PASSWORD_HASH_METHOD = os.getenv("PASSWORD_HASH_METHOD", "pbkdf2:sha256:260000")

def hash_password(password):
    return generate_password_hash(password, method=PASSWORD_HASH_METHOD, salt_length=16)

def password_hash_needs_upgrade(stored_hash):
    stored_method = str(stored_hash or "").split("$", 1)[0]
    return bool(stored_method and stored_method != PASSWORD_HASH_METHOD)

GEMINI_KEY_FERNET = None

def _gemini_key_cipher():
    global GEMINI_KEY_FERNET
    if GEMINI_KEY_FERNET is not None:
        return GEMINI_KEY_FERNET
    secret_source = (os.getenv("GEMINI_KEY_ENCRYPTION_SECRET") or SECRET_KEY or "").encode("utf-8")
    if not secret_source:
        raise RuntimeError("Gemini key encryption secret is not configured.")
    fernet_key = base64.urlsafe_b64encode(hashlib.sha256(secret_source).digest())
    GEMINI_KEY_FERNET = Fernet(fernet_key)
    return GEMINI_KEY_FERNET

def encrypt_gemini_api_key(api_key):
    api_key = str(api_key or "").strip()
    if not api_key:
        return ""
    try:
        return _gemini_key_cipher().encrypt(api_key.encode("utf-8")).decode("utf-8")
    except Exception as exc:
        print(f"Gemini key encryption failed: {type(exc).__name__}: {exc}", flush=True)
        return ""

def decrypt_gemini_api_key(token):
    token = str(token or "").strip()
    if not token:
        return ""
    try:
        return _gemini_key_cipher().decrypt(token.encode("utf-8")).decode("utf-8")
    except InvalidToken:
        return ""
    except Exception as exc:
        print(f"Gemini key decryption failed: {type(exc).__name__}: {exc}", flush=True)
        return ""

def mask_api_key(value):
    value = str(value or "").strip()
    if not value:
        return ""
    if len(value) <= 8:
        return "*" * len(value)
    return f"{value[:4]}…{value[-4:]}"

def _best_effort_write(conn, sql, params=(), label="write", retries=4, delay=0.2):
    last_exc = None
    for attempt in range(1, retries + 1):
        try:
            conn.execute(sql, params)
            return True
        except sqlite3.OperationalError as exc:
            last_exc = exc
            if "database is locked" not in str(exc).lower() or attempt >= retries:
                break
            time.sleep(delay * attempt)
    print(f"Database best-effort {label} skipped: {last_exc}", flush=True)
    return False

def _best_effort_fetchone(conn, sql, params=(), label="query", retries=4, delay=0.2):
    last_exc = None
    for attempt in range(1, retries + 1):
        try:
            return conn.execute(sql, params).fetchone()
        except sqlite3.OperationalError as exc:
            last_exc = exc
            if "database is locked" not in str(exc).lower() or attempt >= retries:
                break
            time.sleep(delay * attempt)
    print(f"Database best-effort {label} skipped: {last_exc}", flush=True)
    return None

def set_login_session_from_app_user(app_user, db_timeout=60):
    team_member = None
    conn = None
    if app_user["team_member_id"]:
        try:
            conn = get_db(timeout=db_timeout)
            team_member = conn.execute("SELECT * FROM team_members WHERE id=?", (app_user["team_member_id"],)).fetchone()
        finally:
            try:
                if conn is not None:
                    conn.close()
            except Exception:
                pass
    display_name = (team_member["name"] if team_member else app_user["username"]) or app_user["username"]
    email = (team_member["email"] if team_member else "") or ""
    role = (team_member["role"] if team_member else "") or ""
    role_name = role.strip().lower()
    is_admin = bool(app_user["is_admin"] or role_name == "admin")
    can_bulk_upload = bool(
        is_admin or
        app_user["is_bulk_admin"] or
        (team_member and (team_member["can_bulk_upload"] or role_name == "bulk admin"))
    )
    session["logged_in"] = True
    session["user_id"] = app_user["id"]
    session["app_user_id"] = app_user["id"]
    session["team_member_id"] = app_user["team_member_id"]
    session["username"] = display_name
    session["recruiter_name"] = display_name
    session["email"] = email
    session["recruiter_email"] = email
    session["role"] = role
    session["is_admin"] = 1 if is_admin else 0
    session["can_bulk_upload"] = 1 if can_bulk_upload else 0

def set_session_from_team_member(team_member, app_user=None, impersonation_active=False):
    role = (team_member["role"] or "").strip()
    role_name = role.lower()
    is_admin = bool((app_user and app_user["is_admin"]) or role_name == "admin")
    can_bulk_upload = bool(
        is_admin or
        (app_user and app_user["is_bulk_admin"]) or
        team_member["can_bulk_upload"] or
        role_name == "bulk admin"
    )
    session["logged_in"] = True
    session["user_id"] = app_user["id"] if app_user else session.get("user_id")
    session["app_user_id"] = app_user["id"] if app_user else session.get("app_user_id")
    session["team_member_id"] = team_member["id"]
    session["username"] = team_member["name"]
    session["recruiter_name"] = team_member["name"]
    session["email"] = team_member["email"] or ""
    session["recruiter_email"] = team_member["email"] or ""
    session["role"] = role
    session["is_admin"] = 1 if is_admin else 0
    session["can_bulk_upload"] = 1 if can_bulk_upload else 0
    session["impersonation_active"] = 1 if impersonation_active else 0

def is_client_viewer_session(current_session=None):
    current_session = current_session or session
    role = str(current_session.get("role") or "").strip().lower()
    return role in {"client viewer", "client_viewer", "external client", "external_client", "client user", "client_user"}

def is_team_leader_session(current_session=None):
    current_session = current_session or session
    role = str(current_session.get("role") or "").strip().lower()
    return role in {"team leader", "team_leader", "teamlead", "team lead", "leader"}

def post_login_redirect_endpoint():
    if is_client_viewer_session():
        return "client_candidates_page"
    if session.get("is_admin"):
        return "admin_landing_page"
    return "index"

def client_viewer_write_forbidden():
    if is_client_viewer_session():
        return jsonify({"error": "Client viewer access is read-only"}), 403
    return None

def client_viewer_candidate_clause(conn, current_session, candidate_alias="c"):
    team_member_id = current_session.get("team_member_id")
    if not team_member_id:
        return " AND 1=0", []
    rows = conn.execute("""
        SELECT lower(trim(c.client_name)) AS client_name
        FROM team_client_mappings m
        JOIN clients c ON c.id = m.client_id
        WHERE m.team_member_id=?
    """, (team_member_id,)).fetchall()
    client_names = [r["client_name"] for r in rows if r["client_name"]]
    if not client_names:
        return " AND 1=0", []
    alias = f"{candidate_alias}." if candidate_alias else ""
    placeholders = ",".join("?" * len(client_names))
    return (
        f""" AND {alias}requirement_id IN (
            SELECT id FROM requirements
            WHERE lower(trim(COALESCE(client_name,''))) IN ({placeholders})
        )""",
        client_names,
    )

def team_leader_candidate_clause(conn, current_session, candidate_alias="c"):
    leader_id = current_session.get("team_member_id")
    if not leader_id:
        return " AND 1=0", []
    leader = conn.execute(
        "SELECT id, lower(trim(COALESCE(email,''))) AS email FROM team_members WHERE id=?",
        (leader_id,),
    ).fetchone()
    rows = conn.execute("""
        SELECT tm.id, lower(trim(COALESCE(tm.email,''))) AS email
        FROM team_leader_mappings m
        JOIN team_members tm ON tm.id = m.member_team_member_id
        WHERE m.leader_team_member_id=?
    """, (leader_id,)).fetchall()
    member_ids = {int(r["id"]) for r in rows if r["id"]}
    member_emails = {r["email"] for r in rows if r["email"]}
    if leader and leader["id"]:
        member_ids.add(int(leader["id"]))
    leader_email = (current_session.get("recruiter_email") or current_session.get("email") or "").strip().lower()
    if leader and leader["email"]:
        leader_email = leader["email"]
    if leader_email:
        member_emails.add(leader_email)
    clauses = []
    params = []
    prefix = f"{candidate_alias}." if candidate_alias else ""
    if member_ids:
        ids = sorted(member_ids)
        clauses.append(f"{prefix}sourcer_id IN ({','.join('?' * len(ids))})")
        params.extend(ids)
    if member_emails:
        emails = sorted(member_emails)
        clauses.append(f"lower({prefix}recruiter_email) IN ({','.join('?' * len(emails))})")
        params.extend(emails)
    if not clauses:
        return " AND 1=0", []
    return " AND (" + " OR ".join(clauses) + ")", params

def non_admin_candidate_owner_clause(current_session, alias="c"):
    if not current_session or current_session.get("is_admin"):
        return "", []
    if is_client_viewer_session(current_session):
        conn = get_db(timeout=5)
        try:
            return client_viewer_candidate_clause(conn, current_session, alias)
        finally:
            conn.close()
    if is_team_leader_session(current_session):
        conn = get_db(timeout=5)
        try:
            return team_leader_candidate_clause(conn, current_session, alias)
        finally:
            conn.close()
    prefix = f"{alias}." if alias else ""
    team_member_id = current_session.get("team_member_id")
    recruiter_email = (current_session.get("recruiter_email") or "").strip().lower()
    if team_member_id and recruiter_email:
        return f" AND ({prefix}sourcer_id=? OR lower({prefix}recruiter_email)=?)", [team_member_id, recruiter_email]
    if team_member_id:
        return f" AND {prefix}sourcer_id=?", [team_member_id]
    if recruiter_email:
        return f" AND lower({prefix}recruiter_email)=?", [recruiter_email]
    return " AND 1=0", []

FOLLOWUP_TERMINAL_STATUSES = {"joined", "dropped", "duplicate", "hm rejected", "screen rejected", "rejected", "l1 reject"}
FOLLOWUP_STALE_RULES = {
    "new": 3,
    "shortlisted": 2,
    "screen shortlisted": 2,
    "screening pending": 2,
    "cv shared": 2,
    "interviewed": 2,
    "feedback pending": 2,
    "interview feedback pending": 2,
    "l1 scheduled": 1,
    "offered": 3,
    "onhold": 7,
    "on hold": 7,
}
CLIENT_SLA_THRESHOLDS = {
    "feedback_pending_days": 2,
    "no_recent_submission_days": 3,
    "requirement_aging_days": 7,
    "critical_aging_days": 14,
}

def normalize_status_key(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()

def parse_local_datetime(value):
    raw = str(value or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw[:len(fmt)], fmt)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(raw.replace("Z", "").replace("T", " "))
    except ValueError:
        return None

def get_candidate_status_master(conn=None):
    statuses = list(STATUSES)
    own_conn = None
    try:
        if conn is None:
            own_conn = get_db(timeout=5)
            conn = own_conn
        rows = conn.execute("""
            SELECT DISTINCT trim(status) AS status
            FROM candidates
            WHERE trim(COALESCE(status,'')) <> ''
            ORDER BY lower(trim(status))
        """).fetchall()
        seen = {normalize_status_key(s) for s in statuses}
        for row in rows:
            status = row["status"] if isinstance(row, sqlite3.Row) else row[0]
            key = normalize_status_key(status)
            if key and key not in seen:
                statuses.append(status)
                seen.add(key)
    except sqlite3.Error:
        pass
    finally:
        if own_conn:
            own_conn.close()
    return statuses

def get_candidate_statuses_for_role(conn, role_name=""):
    statuses = []
    seen = set()
    def add_many(values):
        for value in values or []:
            status = str(value or "").strip()
            key = normalize_status_key(status)
            if status and key not in seen:
                statuses.append(status)
                seen.add(key)
    try:
        pipe = conn.execute("SELECT status_list FROM pipelines WHERE role_name=? LIMIT 1", (role_name or "",)).fetchone()
        if not pipe:
            pipe = conn.execute("SELECT status_list FROM pipelines WHERE is_default=1 LIMIT 1").fetchone()
        if pipe:
            add_many(json.loads(pipe[0]))
    except (sqlite3.Error, TypeError, json.JSONDecodeError):
        pass
    add_many(get_candidate_status_master(conn))
    return statuses

def candidate_age_days(row):
    dt = parse_local_datetime(row.get("updated_at") or row.get("created_at"))
    if not dt:
        return 0
    return max(0, (datetime.now() - dt).days)

def has_candidate_cv_reference(row):
    for key in ("cv_url", "cv_filename", "cv_public_id"):
        if str(row.get(key) or "").strip():
            return True
    return False

def candidate_followup_items(current_session=None, admin=False, limit=150):
    current_session = current_session or session
    conn = get_db(timeout=10)
    valid_statuses = {normalize_status_key(s) for s in get_candidate_status_master(conn)}
    owner_sql, owner_params = ("", [])
    if not admin:
        owner_sql, owner_params = non_admin_candidate_owner_clause(current_session, "c")
    rows = conn.execute(f"""
        SELECT c.id, c.candidate_name, c.email_addr, c.phone, c.status, c.created_at, c.updated_at,
               c.cv_url, c.cv_filename, c.cv_public_id, c.requirement_id, c.recruiter_name, c.recruiter_email, c.sourcer_id,
               r.title AS requirement_title, r.client_name AS client_name
        FROM candidates c
        LEFT JOIN requirements r ON r.id = c.requirement_id
        WHERE COALESCE(c.is_duplicate, 0)=0 {owner_sql}
        ORDER BY datetime(COALESCE(NULLIF(c.updated_at,''), c.created_at)) ASC, c.id DESC
        LIMIT 600
    """, owner_params).fetchall()
    conn.close()

    items = []
    for row in rows:
        d = dict(row)
        status_key = normalize_status_key(d.get("status"))
        if status_key in FOLLOWUP_TERMINAL_STATUSES:
            continue
        age_days = candidate_age_days(d)
        reasons = []
        if status_key and status_key not in valid_statuses:
            reasons.append("Status not in master list")
        stale_after = FOLLOWUP_STALE_RULES.get(status_key)
        if stale_after is not None and age_days >= stale_after:
            reasons.append(f"No update for {age_days} day{'s' if age_days != 1 else ''}")
        if not reasons:
            continue
        items.append({
            "id": d.get("id"),
            "candidate_name": d.get("candidate_name") or "Unnamed candidate",
            "email_addr": d.get("email_addr") or "",
            "phone": d.get("phone") or "",
            "status": d.get("status") or "",
            "age_days": age_days,
            "reasons": reasons,
            "requirement_title": d.get("requirement_title") or "",
            "client_name": d.get("client_name") or "",
            "recruiter_name": d.get("recruiter_name") or "",
            "recruiter_email": d.get("recruiter_email") or "",
            "updated_at": d.get("updated_at") or d.get("created_at") or "",
        })
        if len(items) >= limit:
            break
    return items

def ensure_followup_alert_schema(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS followup_daily_alerts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            team_member_id INTEGER,
            recruiter_email TEXT,
            alert_date TEXT NOT NULL,
            followup_count INTEGER DEFAULT 0,
            shown_at TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(team_member_id, recruiter_email, alert_date)
        )
    """)

def client_sla_items(thresholds=None):
    thresholds = {**CLIENT_SLA_THRESHOLDS, **(thresholds or {})}
    conn = get_db(timeout=10)
    today_dt = datetime.now()
    active_req_statuses = {"new", "open", "in progress"}
    feedback_statuses = {"feedback pending", "interview feedback pending", "cv shared", "shortlisted", "screen shortlisted"}
    selection_statuses = {"selected", "offered", "joined", "hired"}
    terminal_candidate_statuses = {"joined", "dropped", "duplicate", "hm rejected", "screen rejected", "rejected", "l1 reject"}

    req_rows = [dict(r) for r in conn.execute("""
        SELECT id, title, client_name, status, created_at, updated_at
        FROM requirements
        WHERE trim(COALESCE(client_name,'')) <> ''
    """).fetchall()]
    cand_rows = [dict(r) for r in conn.execute("""
        SELECT c.id, c.requirement_id, c.status, c.created_at, c.updated_at,
               r.client_name, r.title AS requirement_title
        FROM candidates c
        LEFT JOIN requirements r ON r.id = c.requirement_id
        WHERE trim(COALESCE(r.client_name,'')) <> ''
          AND COALESCE(c.is_duplicate,0)=0
    """).fetchall()]
    conn.close()

    clients = {}
    def bucket(client_name):
        key = str(client_name or "").strip() or "Unmapped"
        return clients.setdefault(key, {
            "client": key,
            "active_requirements": 0,
            "pending_feedback": 0,
            "oldest_pending_feedback_age": 0,
            "no_submission_requirements": 0,
            "current_month_selections": 0,
            "active_candidates": 0,
            "aged_requirements": 0,
            "critical_aged_requirements": 0,
            "risk_level": "Healthy",
            "risk_reasons": [],
        })

    latest_submission_by_req = {}
    active_req_ids = set()
    for row in cand_rows:
        rid = row.get("requirement_id")
        if rid:
            dt = parse_local_datetime(row.get("created_at"))
            if dt and (rid not in latest_submission_by_req or dt > latest_submission_by_req[rid]):
                latest_submission_by_req[rid] = dt

    for row in req_rows:
        status_key = normalize_status_key(row.get("status") or "New")
        if status_key not in active_req_statuses:
            continue
        client = bucket(row.get("client_name"))
        client["active_requirements"] += 1
        rid = row.get("id")
        active_req_ids.add(rid)
        age = max(0, (today_dt - (parse_local_datetime(row.get("created_at")) or today_dt)).days)
        if age >= thresholds["requirement_aging_days"]:
            client["aged_requirements"] += 1
        if age >= thresholds["critical_aging_days"]:
            client["critical_aged_requirements"] += 1
        latest_submission = latest_submission_by_req.get(rid)
        stale = latest_submission is None or (today_dt - latest_submission).days >= thresholds["no_recent_submission_days"]
        if stale:
            client["no_submission_requirements"] += 1

    month_start = date.today().replace(day=1)
    for row in cand_rows:
        client = bucket(row.get("client_name"))
        status_key = normalize_status_key(row.get("status"))
        updated = parse_local_datetime(row.get("updated_at") or row.get("created_at")) or today_dt
        if row.get("requirement_id") in active_req_ids and status_key not in terminal_candidate_statuses:
            client["active_candidates"] += 1
        if status_key in feedback_statuses:
            age = max(0, (today_dt - updated).days)
            if age >= thresholds["feedback_pending_days"]:
                client["pending_feedback"] += 1
                client["oldest_pending_feedback_age"] = max(client["oldest_pending_feedback_age"], age)
        if status_key in selection_statuses and updated.date() >= month_start:
            client["current_month_selections"] += 1

    for client in clients.values():
        reasons = []
        risk = "Healthy"
        if client["pending_feedback"] > 5:
            risk = "Critical"
            reasons.append("More than 5 pending feedback candidates")
        if client["critical_aged_requirements"] > 0:
            risk = "Critical"
            reasons.append("Requirement open more than 14 days without movement")
        if risk != "Critical":
            if client["pending_feedback"] > 2:
                risk = "Watch"
                reasons.append("More than 2 pending feedback candidates")
            if client["no_submission_requirements"] > 0:
                risk = "Watch"
                reasons.append("Active requirement has no recent submission")
        if not reasons:
            reasons.append("Within SLA")
        client["risk_level"] = risk
        client["risk_reasons"] = reasons

    order = {"Critical": 0, "Watch": 1, "Healthy": 2}
    return sorted(
        clients.values(),
        key=lambda x: (order.get(x["risk_level"], 9), -x["pending_feedback"], -x["active_requirements"], x["client"].lower()),
    )

def data_quality_console_items():
    conn = get_db(timeout=10)
    valid_statuses = {normalize_status_key(s) for s in get_candidate_status_master(conn)}
    issues = []

    def add_issue(category, severity, title, count, detail="", sample=None):
        issues.append({
            "category": category,
            "severity": severity,
            "title": title,
            "count": int(count or 0),
            "detail": detail,
            "sample": sample or [],
        })

    invalid_status_rows = [dict(r) for r in conn.execute("""
        SELECT COALESCE(NULLIF(trim(status),''),'Blank') AS status, COUNT(*) AS count
        FROM candidates
        WHERE COALESCE(is_duplicate,0)=0
        GROUP BY COALESCE(NULLIF(trim(status),''),'Blank')
        ORDER BY count DESC
    """).fetchall()]
    invalid_statuses = [
        {"value": r["status"], "count": r["count"]}
        for r in invalid_status_rows
        if r["status"] == "Blank" or normalize_status_key(r["status"]) not in valid_statuses
    ]
    if invalid_statuses:
        add_issue("Statuses", "Critical", "Candidate statuses not in master list", sum(r["count"] for r in invalid_statuses), "These values may break filters, reporting, and status updates.", invalid_statuses[:10])

    missing_cv = conn.execute("""
        SELECT COUNT(*) FROM candidates
        WHERE COALESCE(is_duplicate,0)=0
          AND trim(COALESCE(phone,'')) <> ''
          AND trim(COALESCE(cv_url,''))=''
          AND trim(COALESCE(cv_filename,''))=''
          AND trim(COALESCE(cv_public_id,''))=''
          AND lower(COALESCE(status,'')) NOT IN ('duplicate','dropped','rejected','screen rejected','hm rejected')
    """).fetchone()[0]
    if missing_cv:
        sample = [dict(r) for r in conn.execute("""
            SELECT id, candidate_name, status FROM candidates
            WHERE COALESCE(is_duplicate,0)=0
              AND trim(COALESCE(cv_url,''))=''
              AND trim(COALESCE(cv_filename,''))=''
              AND trim(COALESCE(cv_public_id,''))=''
            ORDER BY datetime(created_at) DESC LIMIT 8
        """).fetchall()]
        add_issue("Candidates", "Watch", "Candidates missing CV", missing_cv, "Daily reports, follow-ups, and recruiter handoffs depend on CV availability.", sample)

    missing_req = conn.execute("""
        SELECT COUNT(*) FROM candidates c
        LEFT JOIN requirements r ON r.id=c.requirement_id
        WHERE COALESCE(c.is_duplicate,0)=0
          AND (c.requirement_id IS NULL OR r.id IS NULL)
          AND lower(COALESCE(c.status,'')) NOT IN ('duplicate','dropped','rejected','screen rejected','hm rejected')
    """).fetchone()[0]
    if missing_req:
        sample = [dict(r) for r in conn.execute("""
            SELECT c.id, c.candidate_name, c.status, c.requirement_id
            FROM candidates c
            LEFT JOIN requirements r ON r.id=c.requirement_id
            WHERE COALESCE(c.is_duplicate,0)=0
              AND (c.requirement_id IS NULL OR r.id IS NULL)
            ORDER BY datetime(c.created_at) DESC LIMIT 8
        """).fetchall()]
        add_issue("Candidates", "Critical", "Candidates missing valid requirement mapping", missing_req, "Client SLA, reports, and follow-up ownership can be wrong without requirement mapping.", sample)

    missing_contact = conn.execute("""
        SELECT COUNT(*) FROM candidates
        WHERE COALESCE(is_duplicate,0)=0
          AND trim(COALESCE(email_addr,''))=''
          AND trim(COALESCE(phone,''))=''
    """).fetchone()[0]
    if missing_contact:
        sample = [dict(r) for r in conn.execute("""
            SELECT id, candidate_name, status FROM candidates
            WHERE COALESCE(is_duplicate,0)=0
              AND trim(COALESCE(email_addr,''))=''
              AND trim(COALESCE(phone,''))=''
            ORDER BY datetime(created_at) DESC LIMIT 8
        """).fetchall()]
        add_issue("Candidates", "Watch", "Candidates missing email and phone", missing_contact, "Duplicate detection and recruiter follow-up are weaker without contact details.", sample)

    duplicate_identity = [dict(r) for r in conn.execute("""
        SELECT lower(trim(COALESCE(email_addr,''))) AS email, trim(COALESCE(phone,'')) AS phone, COUNT(*) AS count
        FROM candidates
        WHERE COALESCE(is_duplicate,0)=0
          AND (trim(COALESCE(email_addr,''))<>'' OR trim(COALESCE(phone,''))<>'')
        GROUP BY lower(trim(COALESCE(email_addr,''))), trim(COALESCE(phone,''))
        HAVING COUNT(*) > 1
        ORDER BY count DESC
        LIMIT 10
    """).fetchall()]
    if duplicate_identity:
        add_issue("Candidates", "Watch", "Possible duplicate candidate identities", sum(r["count"] for r in duplicate_identity), "Same email/phone appears on multiple active candidate records.", duplicate_identity)

    active_requirement_where = "lower(COALESCE(NULLIF(trim(status),''),'new')) IN ('new','open','in progress')"
    missing_client = conn.execute(f"""
        SELECT COUNT(*) FROM requirements
        WHERE {active_requirement_where}
          AND trim(COALESCE(client_name,''))=''
    """).fetchone()[0]
    if missing_client:
        add_issue("Requirements", "Critical", "Active requirements missing client", missing_client, "Client dashboards and recruiter mappings require client names.")

    missing_taggd = conn.execute(f"""
        SELECT COUNT(*) FROM requirements
        WHERE {active_requirement_where}
          AND (taggd_recruiter_id IS NULL OR taggd_recruiter_id=0 OR trim(COALESCE(taggd_recruiter_name,''))='')
    """).fetchone()[0]
    if missing_taggd:
        sample = [dict(r) for r in conn.execute(f"""
            SELECT id, title, client_name, status
            FROM requirements
            WHERE {active_requirement_where}
              AND (taggd_recruiter_id IS NULL OR taggd_recruiter_id=0 OR trim(COALESCE(taggd_recruiter_name,''))='')
            ORDER BY datetime(created_at) DESC LIMIT 8
        """).fetchall()]
        add_issue("Requirements", "Watch", "Active requirements missing Taggd recruiter", missing_taggd, "Requirement reports and client ownership views need Taggd recruiter mapping.", sample)

    stale_active_req = conn.execute(f"""
        SELECT COUNT(*) FROM requirements
        WHERE {active_requirement_where}
          AND julianday('now','localtime') - julianday(COALESCE(NULLIF(updated_at,''), created_at)) >= 14
    """).fetchone()[0]
    if stale_active_req:
        add_issue("Requirements", "Watch", "Active requirements not updated for 14+ days", stale_active_req, "Review whether old requirements should be closed or refreshed.")

    conn.close()
    severity_order = {"Critical": 0, "Watch": 1, "Info": 2}
    issues.sort(key=lambda x: (severity_order.get(x["severity"], 9), -x["count"], x["category"]))
    return issues

def mapped_client_names_for_current_user(conn):
    if session.get("is_admin") or not session.get("team_member_id"):
        return None
    rows = conn.execute("""
        SELECT c.client_name
        FROM team_client_mappings m
        JOIN clients c ON c.id = m.client_id
        WHERE m.team_member_id=?
    """, (session.get("team_member_id"),)).fetchall()
    mapped = {str(r["client_name"] or "").strip().lower() for r in rows}
    if mapped or is_client_viewer_session():
        return mapped
    return None

def current_user_can_use_client(conn, client_name):
    allowed = mapped_client_names_for_current_user(conn)
    if allowed is None:
        return True
    return str(client_name or "").strip().lower() in allowed

def resolve_taggd_recruiter_for_client(conn, taggd_recruiter_id, client_name):
    ensure_taggd_recruiter_schema(conn)
    try:
        taggd_recruiter_id = int(taggd_recruiter_id or 0)
    except (TypeError, ValueError):
        taggd_recruiter_id = 0
    if not taggd_recruiter_id:
        return None
    return conn.execute("""
        SELECT tr.id, tr.name, tr.email, c.client_name
        FROM taggd_recruiters tr
        JOIN clients c ON c.id = tr.client_id
        WHERE tr.id=?
          AND tr.is_active=1
          AND lower(trim(c.client_name))=lower(trim(?))
        LIMIT 1
    """, (taggd_recruiter_id, client_name or "")).fetchone()

def ensure_taggd_recruiter_schema(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS taggd_recruiters (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT,
            client_id INTEGER NOT NULL,
            is_active INTEGER DEFAULT 1,
            created_by TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(client_id, name),
            FOREIGN KEY(client_id) REFERENCES clients(id)
        )
    """)
    requirement_columns = {
        row["name"] for row in conn.execute("PRAGMA table_info(requirements)").fetchall()
    }
    if "taggd_recruiter_id" not in requirement_columns:
        conn.execute("ALTER TABLE requirements ADD COLUMN taggd_recruiter_id INTEGER")
    if "taggd_recruiter_name" not in requirement_columns:
        conn.execute("ALTER TABLE requirements ADD COLUMN taggd_recruiter_name TEXT")

def resolve_team_member_gemini_key(team_member_id=None, conn=None):
    team_member_id = team_member_id or session.get("team_member_id")
    personal_key = ""
    should_close = False
    if team_member_id:
        if conn is None:
            conn = get_db(timeout=5)
            should_close = True
        try:
            row = conn.execute(
                "SELECT gemini_api_key_enc FROM team_members WHERE id=?",
                (team_member_id,)
            ).fetchone()
            if row:
                personal_key = decrypt_gemini_api_key(row["gemini_api_key_enc"])
        except Exception as exc:
            print(f"Gemini key lookup failed for team_member_id={team_member_id}: {type(exc).__name__}: {exc}", flush=True)
        finally:
            if should_close and conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass
    org_key = (os.getenv("GEMINI_API_KEY") or "").strip()
    if personal_key:
        return personal_key, "personal"
    if org_key:
        return org_key, "org_default"
    return "", "missing"

def get_current_user_gemini_key():
    return resolve_team_member_gemini_key(session.get("team_member_id"))

def get_team_member_gemini_key_status(team_member_id=None, conn=None):
    team_member_id = team_member_id or session.get("team_member_id")
    status = {
        "has_personal_key": False,
        "key_mask": "",
        "updated_at": "",
        "source": "org_default" if (os.getenv("GEMINI_API_KEY") or "").strip() else "missing",
        "fallback_available": bool((os.getenv("GEMINI_API_KEY") or "").strip()),
    }
    if not team_member_id:
        return status
    should_close = False
    if conn is None:
        conn = get_db(timeout=5)
        should_close = True
    try:
        row = conn.execute(
            "SELECT gemini_api_key_enc, gemini_api_key_updated_at FROM team_members WHERE id=?",
            (team_member_id,)
        ).fetchone()
        if row and row["gemini_api_key_enc"]:
            key_value = decrypt_gemini_api_key(row["gemini_api_key_enc"])
            if key_value:
                status.update({
                    "has_personal_key": True,
                    "key_mask": mask_api_key(key_value),
                    "updated_at": row["gemini_api_key_updated_at"] or "",
                    "source": "personal",
                })
    finally:
        if should_close and conn is not None:
            try:
                conn.close()
            except Exception:
                pass
    return status

def normalize_client_name_value(value):
    return re.sub(r"\s+", " ", str(value or "").strip())

def find_client_by_normalized_name(conn, client_name):
    normalized = normalize_client_name_value(client_name).lower()
    if not normalized:
        return None
    return conn.execute(
        "SELECT * FROM clients WHERE lower(trim(client_name)) = ? LIMIT 1",
        (normalized,)
    ).fetchone()

def ensure_client_exists(conn, client_name):
    clean_name = normalize_client_name_value(client_name)
    if not clean_name:
        return None
    existing = find_client_by_normalized_name(conn, clean_name)
    if existing:
        return existing
    try:
        conn.execute("INSERT INTO clients (client_name) VALUES (?)", (clean_name,))
    except sqlite3.IntegrityError:
        existing = find_client_by_normalized_name(conn, clean_name)
        if existing:
            return existing
        raise
    return conn.execute(
        "SELECT * FROM clients WHERE lower(trim(client_name)) = ? LIMIT 1",
        (clean_name.lower(),)
    ).fetchone()

def record_login_audit(conn, *, app_user_id=None, team_member_id=None, username="", email="", display_name="", role="", method="google", status="success", message="", ip_address="", user_agent=""):
    conn.execute("""
        INSERT INTO user_login_audit
        (app_user_id,team_member_id,username,email,display_name,role,method,status,ip_address,user_agent,message)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (
        app_user_id,
        team_member_id,
        username or "",
        (email or "").strip().lower(),
        display_name or "",
        role or "",
        method or "",
        status or "",
        ip_address or "",
        (user_agent or "")[:300],
        message or "",
    ))

def row_get(row, key, default=None):
    if not row:
        return default
    try:
        return row[key]
    except (KeyError, IndexError):
        return default

def set_login_session_from_joined_user(app_user):
    display_name = row_get(app_user, "team_name") or app_user["username"]
    email = row_get(app_user, "team_email") or row_get(app_user, "email") or ""
    role = row_get(app_user, "team_role") or ""
    role_name = role.strip().lower()
    is_admin = bool(app_user["is_admin"] or role_name == "admin")
    can_bulk_upload = bool(
        is_admin or
        app_user["is_bulk_admin"] or
        row_get(app_user, "team_can_bulk_upload") or
        role_name == "bulk admin"
    )
    session["logged_in"] = True
    session["user_id"] = app_user["id"]
    session["app_user_id"] = app_user["id"]
    session["team_member_id"] = app_user["team_member_id"]
    session["username"] = display_name
    session["recruiter_name"] = display_name
    session["email"] = email
    session["recruiter_email"] = email
    session["role"] = role
    session["is_admin"] = 1 if is_admin else 0
    session["can_bulk_upload"] = 1 if can_bulk_upload else 0

def normalize_user_role_label(role, is_admin=False):
    clean = str(role or "").strip()
    key = clean.lower().replace("-", " ").replace("_", " ")
    if is_admin or key == "admin":
        return "Admin"
    if key in {"client", "client viewer", "client user", "external client", "external"}:
        return "Client Viewer"
    if key in {"team leader", "team lead", "teamlead", "leader"}:
        return "Team Leader"
    if key in {"bulk admin", "bulk upload"}:
        return "Bulk Admin"
    return clean or "Recruiter"

def should_update_login_timestamp(conn, table, row_id, threshold_minutes=10):
    if not row_id:
        return False
    try:
        row = conn.execute(
            f"""
            SELECT 1 AS due
            FROM {table}
            WHERE id=?
              AND (
                last_login_at IS NULL
                OR last_login_at=''
                OR datetime(last_login_at) <= datetime('now','localtime',?)
              )
            LIMIT 1
            """,
            (row_id, f"-{int(threshold_minutes)} minutes"),
        ).fetchone()
        return bool(row)
    except Exception:
        return True

def persist_password_login_artifacts(app_user_id, team_member_id, username="", email="", display_name="", role="", status="success", message="", ip_address="", user_agent="", password_hash_upgrade=None):
    conn = None
    try:
        conn = get_db(timeout=1)
        if status == "success":
            if app_user_id and password_hash_upgrade:
                _best_effort_write(
                    conn,
                    "UPDATE app_users SET password=? WHERE id=?",
                    (password_hash_upgrade, app_user_id),
                    label="password hash upgrade",
                    retries=1,
                    delay=0.05,
                )
            if should_update_login_timestamp(conn, "app_users", app_user_id):
                _best_effort_write(
                    conn,
                    "UPDATE app_users SET last_login_at=datetime('now','localtime') WHERE id=?",
                    (app_user_id,),
                    label="password app user last_login_at",
                    retries=1,
                    delay=0.05,
                )
            if team_member_id and should_update_login_timestamp(conn, "team_members", team_member_id):
                _best_effort_write(
                    conn,
                    "UPDATE team_members SET last_login_at=datetime('now','localtime') WHERE id=?",
                    (team_member_id,),
                    label="password team member last_login_at",
                    retries=1,
                    delay=0.05,
                )
        record_login_audit(
            conn,
            app_user_id=app_user_id,
            team_member_id=team_member_id,
            username=username,
            email=email,
            display_name=display_name,
            role=role,
            method="password",
            status=status,
            message=message,
            ip_address=ip_address,
            user_agent=user_agent,
        )
        conn.commit()
    except Exception as exc:
        print(f"Password login artifact write skipped: {type(exc).__name__}: {exc}", flush=True)
        try:
            if conn is not None:
                conn.rollback()
        except Exception:
            pass
    finally:
        try:
            if conn is not None:
                conn.close()
        except Exception:
            pass

def queue_password_login_artifacts(app_user_id, team_member_id, username="", email="", display_name="", role="", status="success", message="", ip_address="", user_agent="", password_hash_upgrade=None):
    def _runner():
        persist_password_login_artifacts(
            app_user_id,
            team_member_id,
            username=username,
            email=email,
            display_name=display_name,
            role=role,
            status=status,
            message=message,
            ip_address=ip_address,
            user_agent=user_agent,
            password_hash_upgrade=password_hash_upgrade,
        )
    try:
        threading.Thread(target=_runner, name="password-login-artifacts", daemon=True).start()
    except Exception as exc:
        print(f"Password login background thread skipped: {type(exc).__name__}: {exc}", flush=True)

def persist_google_login_artifacts(app_user_id, team_member_id, username="", email="", display_name="", role="", method="google", status="success", message="", ip_address="", user_agent=""):
    conn = None
    try:
        conn = get_db(timeout=1)
        ensure_ats_pipeline_schema(conn)
        if status == "success":
            if team_member_id and should_update_login_timestamp(conn, "team_members", team_member_id):
                _best_effort_write(
                    conn,
                    "UPDATE team_members SET last_login_at=datetime('now','localtime') WHERE id=?",
                    (team_member_id,),
                    label="team member last_login_at",
                    retries=1,
                    delay=0.05,
                )
            if app_user_id and should_update_login_timestamp(conn, "app_users", app_user_id):
                _best_effort_write(
                    conn,
                    "UPDATE app_users SET last_login_at=datetime('now','localtime') WHERE id=?",
                    (app_user_id,),
                    label="app user last_login_at",
                    retries=1,
                    delay=0.05,
                )
        record_login_audit(
            conn,
            app_user_id=app_user_id,
            team_member_id=team_member_id,
            username=username,
            email=email,
            display_name=display_name,
            role=role,
            method=method,
            status=status,
            message=message,
            ip_address=ip_address,
            user_agent=user_agent,
        )
        try:
            conn.commit()
        except sqlite3.OperationalError as exc:
            if "database is locked" not in str(exc).lower():
                raise
            print(f"Database best-effort google login commit skipped: {exc}", flush=True)
    except Exception as exc:
        print(f"Google login background write skipped: {type(exc).__name__}: {exc}", flush=True)
    finally:
        try:
            if conn is not None:
                conn.close()
        except Exception:
            pass

def queue_google_login_artifacts(app_user_id, team_member_id, username="", email="", display_name="", role="", method="google", status="success", message="", ip_address="", user_agent=""):
    def _runner():
        persist_google_login_artifacts(
            app_user_id,
            team_member_id,
            username=username,
            email=email,
            display_name=display_name,
            role=role,
            method=method,
            status=status,
            message=message,
            ip_address=ip_address,
            user_agent=user_agent,
        )
    try:
        threading.Thread(target=_runner, name="google-login-artifacts", daemon=True).start()
    except Exception as exc:
        print(f"Google login background thread skipped: {type(exc).__name__}: {exc}", flush=True)

def _insert_match_audit(conn, *, event_type, object_type, object_hash="", jd_hash="", resume_hash="", pipeline_version=MATCH_PIPELINE_VERSION, status="", source="", parser_confidence="", manual_review_required=False, score=None, message="", details=None, ip_address="", user_agent=""):
    conn.execute("""
        INSERT INTO match_audit_log
        (event_type, object_type, object_hash, jd_hash, resume_hash, pipeline_version, status, source, parser_confidence, manual_review_required, score, message, details_json, ip_address, user_agent)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        event_type or "",
        object_type or "",
        object_hash or "",
        jd_hash or "",
        resume_hash or "",
        pipeline_version or "",
        status or "",
        source or "",
        parser_confidence or "",
        1 if manual_review_required else 0,
        int(score) if score is not None and str(score).strip() != "" else None,
        message or "",
        json.dumps(details or {}, ensure_ascii=False),
        ip_address or "",
        user_agent or "",
    ))

def record_match_audit(conn, *, event_type, object_type, object_hash="", jd_hash="", resume_hash="", pipeline_version=MATCH_PIPELINE_VERSION, status="", source="", parser_confidence="", manual_review_required=False, score=None, message="", details=None):
    _insert_match_audit(
        conn,
        event_type=event_type,
        object_type=object_type,
        object_hash=object_hash,
        jd_hash=jd_hash,
        resume_hash=resume_hash,
        pipeline_version=pipeline_version,
        status=status,
        source=source,
        parser_confidence=parser_confidence,
        manual_review_required=manual_review_required,
        score=score,
        message=message,
        details=details,
        ip_address=request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip(),
        user_agent=(request.headers.get("User-Agent") or "")[:300],
    )

def resolve_upload_reference_path(file_url):
    file_url = str(file_url or "").strip()
    if not file_url:
        return ""
    if file_url.startswith("/uploads/"):
        return os.path.join(app.root_path, file_url.lstrip("/").replace("/", os.sep))
    if os.path.isabs(file_url) and os.path.exists(file_url):
        return file_url
    return ""

def extract_text_from_upload_reference(file_url, filename_hint=""):
    file_url = str(file_url or "").strip()
    if not file_url:
        return ""
    local_path = resolve_upload_reference_path(file_url)
    if local_path and os.path.exists(local_path):
        return extract_cv_text(local_path)
    if file_url.startswith("/api/drive_file/"):
        try:
            drive_id = file_url.rsplit("/", 1)[-1]
            content = fetch_drive_file_bytes(drive_id)
            if not content:
                return ""
            suffix = os.path.splitext(filename_hint or "")[1] or ".pdf"
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
            try:
                tmp.write(content)
                tmp.close()
                return extract_cv_text(tmp.name)
            finally:
                try:
                    os.unlink(tmp.name)
                except Exception:
                    pass
        except Exception:
            return ""
    if file_url.lower().startswith(("http://", "https://")):
        try:
            response = requests.get(file_url, timeout=20)
            response.raise_for_status()
            suffix = os.path.splitext(filename_hint or "")[1] or ".pdf"
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
            try:
                tmp.write(response.content)
                tmp.close()
                return extract_cv_text(tmp.name)
            finally:
                try:
                    os.unlink(tmp.name)
                except Exception:
                    pass
        except Exception:
            return ""
    return ""

def requirement_jd_text(requirement_row):
    requirement_row = requirement_row or {}
    parts = []
    description = str(row_value(requirement_row, "description") or row_value(requirement_row, "requirement_description") or "").strip()
    if description:
        parts.append(description)
    jd_url = row_value(requirement_row, "jd_url") or row_value(requirement_row, "requirement_jd_url") or ""
    jd_filename = row_value(requirement_row, "jd_filename") or row_value(requirement_row, "requirement_jd_filename") or ""
    jd_text = extract_text_from_upload_reference(jd_url, jd_filename)
    if jd_text.strip():
        parts.append(jd_text.strip())
    return "\n\n".join(part for part in parts if part).strip()

def candidate_cv_text(candidate_row):
    candidate_row = candidate_row or {}
    cv_text = extract_text_from_upload_reference(row_value(candidate_row, "cv_url") or "", row_value(candidate_row, "cv_filename") or "")
    if cv_text.strip():
        return cv_text.strip()
    return ""

def persist_ai_screening_log(run_id, candidate_id=None, requirement_id=None, candidate_name="", requirement_title="", stage="", status="", score=None, message="", details=None):
    details = details or {}
    try:
        conn = get_db(timeout=5)
        ensure_ats_pipeline_schema(conn)
        conn.execute(
            """INSERT INTO ai_screening_logs
               (run_id, candidate_id, requirement_id, candidate_name, requirement_title, stage, status, score, message, details_json)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (
                str(run_id or ""),
                candidate_id,
                requirement_id,
                candidate_name or "",
                requirement_title or "",
                stage or "",
                status or "",
                int(score) if score is not None and str(score).strip() != "" else None,
                message or "",
                json.dumps(details or {}, ensure_ascii=False),
            )
        )
        conn.commit()
        conn.close()
    except Exception as exc:
        print(f"AI SCREENING DEBUG: log write skipped ({stage}): {type(exc).__name__}: {exc}", flush=True)

def record_ai_screening_event(run_id, candidate_id=None, requirement_id=None, candidate_name="", requirement_title="", stage="", status="", score=None, message="", details=None):
    details = details or {}
    safe_name = candidate_name or f"candidate_{candidate_id or 'unknown'}"
    safe_req = requirement_title or f"requirement_{requirement_id or 'none'}"
    msg = message or ""
    print(
        f"AI SCREENING DEBUG: run={run_id} candidate={safe_name} requirement={safe_req} stage={stage} status={status} score={score if score is not None else '-'} {msg}",
        flush=True,
    )
    persist_ai_screening_log(
        run_id,
        candidate_id=candidate_id,
        requirement_id=requirement_id,
        candidate_name=candidate_name,
        requirement_title=requirement_title,
        stage=stage,
        status=status,
        score=score,
        message=message,
        details=details,
    )

def persist_candidate_ai_screening_result(candidate_id, *, run_id="", status="", score=None, error="", report_json=None, report_url=""):
    report_url = report_url or (f"/api/candidate/{candidate_id}/ai_screening_report" if status == "scored" and report_json else "")
    with DB_WRITE_LOCK:
        conn = get_db(timeout=5)
        ensure_ats_pipeline_schema(conn)
        conn.execute(
            """UPDATE candidates
               SET ai_screening_status=?,
                   ai_screening_score=?,
                   ai_screening_error=?,
                   ai_screening_report_json=?,
                   ai_screening_report_url=?,
                   ai_screening_run_id=?,
                   ai_screening_updated_at=datetime('now','localtime')
               WHERE id=?""",
            (
                status or "",
                int(score) if score is not None and str(score).strip() != "" else None,
                error or "",
                json.dumps(report_json, ensure_ascii=False) if report_json is not None else None,
                report_url,
                run_id or "",
                candidate_id,
            )
        )
        conn.commit()
        conn.close()

def save_candidate_ai_screening_pdf(candidate, requirement, report):
    candidate = candidate or {}
    requirement = requirement or {}
    report = report or {}
    candidate_id = candidate.get("id") or candidate.get("candidate_id") or "candidate"
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", str(candidate.get("candidate_name") or report.get("candidate_name") or "candidate")).strip("_") or "candidate"
    owner_key = _safe_path_part(
        candidate.get("recruiter_email")
        or candidate.get("sourcer_email")
        or candidate.get("recruiter_name")
        or candidate.get("sourcer_name")
        or candidate.get("email_addr")
        or candidate.get("candidate_name")
        or "screening"
    )
    upload_folder = os.path.join(
        app.root_path,
        "uploads",
        "recruiters",
        owner_key,
        "analysis",
        "screening",
        f"candidate_{candidate_id}",
    )
    os.makedirs(upload_folder, exist_ok=True)
    filename = f"ai_screening_{safe_name}_{date.today().isoformat()}.pdf"
    file_path = os.path.join(upload_folder, filename)
    buf = build_ai_screening_pdf(candidate=candidate, requirement=requirement, report=report)
    with open(file_path, "wb") as f:
        f.write(buf.read())
    rel_path = os.path.relpath(file_path, os.path.join(app.root_path, "uploads")).replace(os.sep, "/")
    return f"/uploads/{rel_path}"

def queue_candidate_ai_screening(candidate_id, trigger="candidate_saved", gemini_api_key=None, requester_team_member_id=None):
    if trigger != "manual_trigger" and not AUTO_AI_SCREENING_ON_CV_UPLOAD:
        print(f"AI SCREENING DEBUG: auto screening skipped for candidate_id={candidate_id} trigger={trigger}", flush=True)
        return False
    conn = None
    run_id = f"ai_{candidate_id}_{int(datetime.now().timestamp() * 1000)}"
    resolved_api_key = (gemini_api_key or "").strip()
    try:
        conn = get_db(timeout=5)
        ensure_ats_pipeline_schema(conn)
        row = conn.execute(
            "SELECT id, requirement_id FROM candidates WHERE id=?",
            (candidate_id,)
        ).fetchone()
        if not row or not row["requirement_id"]:
            return False
        if not resolved_api_key:
            resolved_api_key, _api_source = resolve_team_member_gemini_key(requester_team_member_id, conn=conn)
        conn.execute(
            """UPDATE candidates
               SET ai_screening_status='pending',
                   ai_screening_error='',
                   ai_screening_run_id=?,
                   ai_screening_updated_at=datetime('now','localtime')
               WHERE id=?""",
            (run_id, candidate_id)
        )
        conn.commit()
    except Exception as exc:
        print(f"AI SCREENING DEBUG: pending state skipped for candidate_id={candidate_id}: {type(exc).__name__}: {exc}", flush=True)
    finally:
        try:
            if conn is not None:
                conn.close()
        except Exception:
            pass
    ensure_ai_screening_worker()
    try:
        queue_ai_screening_task("candidate_ai_screening", run_candidate_ai_screening, candidate_id, trigger, run_id, resolved_api_key)
        print(f"AI SCREENING DEBUG: queued screening for candidate_id={candidate_id} trigger={trigger}", flush=True)
        return True
    except queue.Full:
        print(f"AI SCREENING DEBUG: screening queue full; running inline for candidate_id={candidate_id}", flush=True)
        run_candidate_ai_screening(candidate_id, trigger, run_id, resolved_api_key)
        return False

def run_candidate_ai_screening(candidate_id, trigger="candidate_saved", run_id=None, gemini_api_key=None, retry_count=0):
    run_id = run_id or f"ai_{candidate_id}_{int(datetime.now().timestamp() * 1000)}"
    record_ai_screening_event(run_id, candidate_id=candidate_id, stage="queued", status="pending", message=f"Trigger={trigger}")
    conn = None
    try:
        conn = get_db(timeout=10)
        ensure_ats_pipeline_schema(conn)
        candidate = conn.execute(
            """SELECT c.*, r.title AS requirement_title, r.description AS requirement_description, r.jd_url AS requirement_jd_url, r.jd_filename AS requirement_jd_filename, r.client_name AS requirement_client_name
               FROM candidates c
               LEFT JOIN requirements r ON r.id = c.requirement_id
               WHERE c.id=?""",
            (candidate_id,)
        ).fetchone()
        if not candidate:
            record_ai_screening_event(run_id, candidate_id=candidate_id, stage="load_candidate", status="error", message="Candidate not found")
            persist_candidate_ai_screening_result(candidate_id, run_id=run_id, status="error", score=None, error="Candidate not found", report_json=None)
            return
        candidate = dict(candidate)
        requirement_id = candidate.get("requirement_id")
        candidate_name = candidate.get("candidate_name") or ""
        requirement_title = candidate.get("requirement_title") or candidate.get("role_name") or ""
        if not requirement_id:
            persist_candidate_ai_screening_result(candidate_id, run_id=run_id, status="no_jd", score=None, error="", report_json=None)
            record_ai_screening_event(run_id, candidate_id=candidate_id, requirement_id=None, candidate_name=candidate_name, requirement_title="", stage="skip", status="no_jd", message="No JD attached")
            return
        jd_text = requirement_jd_text(candidate)
        if not jd_text.strip():
            persist_candidate_ai_screening_result(candidate_id, run_id=run_id, status="no_jd", score=None, error="", report_json=None)
            record_ai_screening_event(run_id, candidate_id=candidate_id, requirement_id=requirement_id, candidate_name=candidate_name, requirement_title=requirement_title, stage="jd_missing", status="no_jd", message="No JD attached")
            return
        cv_text = candidate_cv_text(candidate)
        if not cv_text.strip():
            persist_candidate_ai_screening_result(candidate_id, run_id=run_id, status="error", score=None, error="Could not read candidate CV.", report_json=None)
            record_ai_screening_event(run_id, candidate_id=candidate_id, requirement_id=requirement_id, candidate_name=candidate_name, requirement_title=requirement_title, stage="cv_missing", status="error", message="Could not read candidate CV.")
            return
        parsed_jd = {
            "role_title": requirement_title or "",
            "target_role": requirement_title or "",
            "description": candidate.get("requirement_description") or "",
            "client_name": candidate.get("requirement_client_name") or "",
        }
        parsed_candidate = {
            "candidate_name": candidate_name,
            "current_role": candidate.get("current_role") or "",
            "current_company": candidate.get("current_company") or "",
            "experience_years": candidate.get("experience_years") or "",
            "key_skills": candidate.get("key_skills") or "",
            "education": candidate.get("education") or "",
            "current_location": candidate.get("current_location") or "",
        }
        record_ai_screening_event(run_id, candidate_id=candidate_id, requirement_id=requirement_id, candidate_name=candidate_name, requirement_title=requirement_title, stage="gemini_request", status="running", message="Gemini screening report is being generated.")
        screening = maybe_generate_gemini_screening_report(
            jd_text,
            cv_text,
            candidate_name=candidate_name,
            target_job_title=requirement_title or candidate.get("role_name") or "",
            parsed_jd=parsed_jd,
            parsed_candidate=parsed_candidate,
            api_key=gemini_api_key,
        )
        if screening.get("ok"):
            report = screening.get("report") or {}
            score = report.get("final_score", "")
            report_url = ""
            try:
                report_url = save_candidate_ai_screening_pdf(candidate, requirement, report)
            except Exception as pdf_exc:
                print(f"AI SCREENING DEBUG: local PDF save failed for candidate_id={candidate_id}: {pdf_exc}", flush=True)
            persist_candidate_ai_screening_result(
                candidate_id,
                run_id=run_id,
                status="scored",
                score=score,
                error="",
                report_json=report,
                report_url=report_url,
            )
            record_ai_screening_event(run_id, candidate_id=candidate_id, requirement_id=requirement_id, candidate_name=candidate_name, requirement_title=requirement_title, stage="complete", status="ok", score=score, message=report.get("summary") or "AI screening completed.")
        else:
            error_message = screening.get("error", "Gemini screening failed.")
            error_type = str(screening.get("error_type") or "").strip()
            retryable = bool(screening.get("retryable"))
            if retryable:
                retry_after_seconds = screening.get("retry_after_seconds")
                scheduled = schedule_candidate_ai_screening_retry(
                    candidate_id,
                    trigger,
                    run_id,
                    gemini_api_key,
                    retry_count,
                    retry_after_seconds or os.getenv("GEMINI_SCREENING_RETRY_DELAY_SECONDS", "60"),
                )
                record_ai_screening_event(
                    run_id,
                    candidate_id=candidate_id,
                    requirement_id=requirement_id,
                    candidate_name=candidate_name,
                    requirement_title=requirement_title,
                    stage="retry_scheduled" if scheduled else "retry_skipped",
                    status="pending",
                    score=None,
                    message=f"{error_message} Retry {'scheduled' if scheduled else 'not scheduled'}.",
                    details={"error_type": error_type, "retry_count": retry_count, "retryable": True},
                )
                return
            persist_candidate_ai_screening_result(candidate_id, run_id=run_id, status="error", score=None, error=error_message, report_json=None)
            record_ai_screening_event(run_id, candidate_id=candidate_id, requirement_id=requirement_id, candidate_name=candidate_name, requirement_title=requirement_title, stage="error", status="error", message=error_message, details={"error_type": error_type, "retryable": retryable})
    except Exception as exc:
        error_message = f"{type(exc).__name__}: {exc}"
        print(f"AI SCREENING DEBUG: screening failed for candidate_id={candidate_id}: {error_message}", flush=True)
        try:
            persist_candidate_ai_screening_result(candidate_id, run_id=run_id, status="error", score=None, error=error_message, report_json=None)
        except Exception:
            pass
        try:
            record_ai_screening_event(run_id, candidate_id=candidate_id, stage="error", status="error", message=error_message)
        except Exception:
            pass
    finally:
        try:
            if conn is not None:
                conn.close()
        except Exception:
            pass

def clean_value(value, limit=120):
    return re.sub(r"\s+", " ", str(value or "")).strip(" -:|,\t\r\n")[:limit]

def row_value(row, key, default=""):
    if row is None:
        return default
    if isinstance(row, dict):
        return row.get(key, default)
    try:
        return row[key]
    except Exception:
        return getattr(row, key, default)

def split_multi_value_terms(value):
    parts = re.split(r"[,;\n\r\t]+", str(value or ""))
    return [re.sub(r"\s+", " ", part).strip() for part in parts if re.sub(r"\s+", " ", part).strip()]

def build_candidate_identity_filter_clause(names="", phones="", emails=""):
    clauses = []
    params = []

    name_terms = []
    for term in split_multi_value_terms(names):
        normalized = term.lower()
        if normalized:
            name_terms.append(normalized)
    if name_terms:
        name_clauses = []
        for term in name_terms:
            name_clauses.append("LOWER(TRIM(COALESCE(c.candidate_name, ''))) LIKE ?")
            params.append(f"%{term}%")
        clauses.append("(" + " OR ".join(name_clauses) + ")")

    phone_terms = []
    for term in split_multi_value_terms(phones):
        normalized = norm_phone(term)
        if normalized:
            phone_terms.append(normalized)
    if phone_terms:
        phone_clauses = []
        for term in phone_terms:
            phone_clauses.append("replace(replace(replace(replace(replace(lower(COALESCE(c.phone,'')),'+',''),'-',''),' ',''),'(',''),')','') LIKE ?")
            params.append(f"%{term}%")
        clauses.append("(" + " OR ".join(phone_clauses) + ")")

    email_terms = []
    for term in split_multi_value_terms(emails):
        normalized = term.strip().lower()
        if normalized:
            email_terms.append(normalized)
    if email_terms:
        email_clauses = []
        for term in email_terms:
            email_clauses.append("LOWER(TRIM(COALESCE(c.email_addr, ''))) LIKE ?")
            params.append(f"%{term}%")
        clauses.append("(" + " OR ".join(email_clauses) + ")")

    if not clauses:
        return "", []
    return " AND (" + " OR ".join(clauses) + ")", params

def _with_db_write_retry(fn, attempts=4, base_delay=0.2):
    last_exc = None
    for attempt in range(1, attempts + 1):
        try:
            with DB_WRITE_LOCK:
                return fn()
        except sqlite3.OperationalError as exc:
            last_exc = exc
            if "database is locked" not in str(exc).lower() or attempt >= attempts:
                raise
            time.sleep(min(1.5, base_delay * attempt))
    if last_exc:
        raise last_exc

def system_email(to_addr, subject, body):
    if not to_addr:
        return {"error": "No recipient configured"}
    if not GMAIL_USER or not GMAIL_APP_PASS:
        return {"error": "Email not configured"}
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = GMAIL_USER
    msg["To"] = to_addr
    msg.attach(MIMEText(body, "plain"))
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
        s.login(GMAIL_USER, GMAIL_APP_PASS)
        s.sendmail(GMAIL_USER, to_addr, msg.as_string())
    return {"ok": True}

def admin_approval_emails():
    conn = get_db()
    rows = conn.execute("""
        SELECT DISTINCT COALESCE(t.email, '') AS email
        FROM app_users u
        LEFT JOIN team_members t ON t.id = u.team_member_id
        WHERE u.is_admin=1 AND u.is_active=1
    """).fetchall()
    team_admins = conn.execute("""
        SELECT DISTINCT email
        FROM team_members
        WHERE email IS NOT NULL
          AND email != ''
          AND LOWER(COALESCE(role,'')) IN ('admin', 'bulk admin')
    """).fetchall()
    conn.close()
    emails = [r["email"] for r in rows if r["email"]]
    emails.extend([r["email"] for r in team_admins if r["email"]])
    if ADMIN_EMAIL:
        emails.append(ADMIN_EMAIL)
    return sorted(set(emails))

def normalize_login_name(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())

def resolve_user_for_password_reset(conn, identifier):
    identifier = (identifier or "").strip().lower()
    user = conn.execute("""
        SELECT u.*, t.email AS team_email
        FROM app_users u
        LEFT JOIN team_members t ON t.id = u.team_member_id
        WHERE LOWER(u.username)=LOWER(?) OR LOWER(t.email)=LOWER(?)
    """, (identifier, identifier)).fetchone()
    if user:
        return user
    team = conn.execute("""
        SELECT *
        FROM team_members
        WHERE LOWER(email)=LOWER(?) OR LOWER(name)=LOWER(?)
    """, (identifier, identifier)).fetchone()
    if not team:
        return None
    normalized_name = normalize_login_name(team["name"])
    return conn.execute("""
        SELECT u.*, ? AS team_email
        FROM app_users u
        WHERE LOWER(u.username)=LOWER(?)
           OR LOWER(u.username)=LOWER(?)
    """, (team["email"], team["name"], normalized_name)).fetchone()

def reset_recipients_for_user(user):
    recipients = []
    if user and user["team_email"]:
        recipients.append(user["team_email"])
    recipients.extend(admin_approval_emails())
    return sorted(set([email for email in recipients if email]))

def ensure_auth_workflow_schema(conn):
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS password_reset_tokens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        token TEXT UNIQUE NOT NULL,
        expires_at TEXT NOT NULL,
        used INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        FOREIGN KEY(user_id) REFERENCES app_users(id)
    );
    CREATE TABLE IF NOT EXISTS user_registration_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        email TEXT NOT NULL,
        username TEXT NOT NULL,
        password_hash TEXT NOT NULL,
        token TEXT UNIQUE NOT NULL,
        status TEXT DEFAULT 'pending',
        approved_by INTEGER,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        approved_at TEXT
    );
    """)

def init_db():
    conn = get_db()
    try:
        ensure_ats_pipeline_schema(conn)
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS job_details (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            candidate_id INTEGER NOT NULL,
            title TEXT,
            location TEXT,
            salary TEXT,
            department TEXT,
            job_id TEXT,
            notes TEXT,
            tags TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(candidate_id) REFERENCES candidates(id)
        );
        CREATE TABLE IF NOT EXISTS standard_roles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            domain TEXT NOT NULL,
            role_name TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(domain, role_name)
        );
        CREATE TABLE IF NOT EXISTS standard_skills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            skill_name TEXT UNIQUE,
            category TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_name TEXT UNIQUE,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS team_client_mappings (
            team_member_id INTEGER NOT NULL,
            client_id INTEGER NOT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            PRIMARY KEY(team_member_id, client_id),
            FOREIGN KEY(team_member_id) REFERENCES team_members(id),
            FOREIGN KEY(client_id) REFERENCES clients(id)
        );
        CREATE TABLE IF NOT EXISTS taggd_recruiters (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT,
            client_id INTEGER NOT NULL,
            is_active INTEGER DEFAULT 1,
            created_by TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(client_id, name),
            FOREIGN KEY(client_id) REFERENCES clients(id)
        );
        CREATE TABLE IF NOT EXISTS pipelines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            role_name TEXT UNIQUE,
            status_list TEXT,
            is_default INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id TEXT UNIQUE,
            title TEXT NOT NULL,
            description TEXT,
            client_name TEXT,
            status TEXT DEFAULT 'Open',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS requirements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            description TEXT,
            client_name TEXT,
            jd_filename TEXT,
            jd_url TEXT,
            jd_public_id TEXT,
            taggd_recruiter_id INTEGER,
            taggd_recruiter_name TEXT,
            assigned_sourcer_id INTEGER,
            assigned_recruiter_id INTEGER,
            daily_target INTEGER DEFAULT 3,
            status TEXT DEFAULT 'Open',
            created_by TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS requirement_checks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            requirement_id INTEGER NOT NULL,
            check_name TEXT NOT NULL,
            check_description TEXT,
            check_type TEXT DEFAULT 'boolean',
            pass_criteria TEXT,
            is_mandatory INTEGER DEFAULT 1,
            sort_order INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS candidates (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            upload_batch       TEXT,
            recruiter_name     TEXT,
            recruiter_email    TEXT,
            role_name          TEXT,
            education       TEXT,
            candidate_name     TEXT,
            email_addr         TEXT,
            phone              TEXT,
            current_company    TEXT,
            current_role       TEXT,
            experience_years   TEXT,
            key_skills         TEXT,
            notice_period      TEXT,
            current_salary     TEXT,
            expected_salary    TEXT,
            current_location   TEXT,
            preferred_location TEXT,
            remarks            TEXT,
            cv_filename        TEXT,
            cv_url             TEXT,
            cv_public_id       TEXT,
            cv_summary         TEXT,
            candidate_feedback TEXT,
            status             TEXT DEFAULT 'New',
            industry_domain    TEXT,
            tags               TEXT DEFAULT '',
            is_duplicate       INTEGER DEFAULT 0,
            duplicate_of       INTEGER,
            missing_info       TEXT,
            job_id            TEXT,
            sourcer_id         INTEGER,
            requirement_id     INTEGER,
            created_at         TEXT DEFAULT (datetime('now','localtime')),
            updated_at         TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS requirement_submissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            candidate_id INTEGER NOT NULL,
            requirement_id INTEGER NOT NULL,
            sourcer_id INTEGER,
            submitted_by TEXT,
            status TEXT DEFAULT 'Submitted',
            notes TEXT,
            recruiter_feedback TEXT,
            feedback_by INTEGER,
            submitted_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS submission_checks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            submission_id INTEGER NOT NULL,
            check_id INTEGER NOT NULL,
            check_name TEXT,
            passed INTEGER DEFAULT 0,
            notes TEXT
        );
        CREATE TABLE IF NOT EXISTS team_members (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            name      TEXT,
            email     TEXT UNIQUE,
            phone     TEXT,
            role      TEXT,
            can_bulk_upload INTEGER DEFAULT 0,
            gemini_api_key_enc TEXT,
            gemini_api_key_updated_at TEXT,
            is_fixed  INTEGER DEFAULT 0,
            added_at  TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS app_users (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            username     TEXT UNIQUE NOT NULL,
            password     TEXT NOT NULL,
            email        TEXT,
            team_member_id INTEGER,
            is_admin     INTEGER DEFAULT 0,
            is_bulk_admin INTEGER DEFAULT 0,
            is_active    INTEGER DEFAULT 1,
            created_at   TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS user_login_audit (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            app_user_id INTEGER,
            team_member_id INTEGER,
            username TEXT,
            email TEXT,
            display_name TEXT,
            role TEXT,
            method TEXT,
            status TEXT,
            ip_address TEXT,
            user_agent TEXT,
            message TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS team_leader_mappings (
            leader_team_member_id INTEGER NOT NULL,
            member_team_member_id INTEGER NOT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            PRIMARY KEY(leader_team_member_id, member_team_member_id)
        );
        CREATE TABLE IF NOT EXISTS performance_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT,
            team_member_id INTEGER,
            recruiter_name TEXT,
            recruiter_email TEXT,
            method TEXT,
            path TEXT,
            endpoint TEXT,
            status_code INTEGER,
            elapsed_ms REAL,
            is_admin INTEGER DEFAULT 0,
            ip_address TEXT,
            user_agent TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS password_reset_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            token TEXT UNIQUE NOT NULL,
            expires_at TEXT NOT NULL,
            used INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY(user_id) REFERENCES app_users(id)
        );
        CREATE TABLE IF NOT EXISTS user_registration_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            email TEXT NOT NULL,
            username TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            token TEXT UNIQUE NOT NULL,
            status TEXT DEFAULT 'pending',
            approved_by INTEGER,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            approved_at TEXT
        );
        CREATE TABLE IF NOT EXISTS alerts (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            alert_type      TEXT,
            message         TEXT,
            candidate_id    INTEGER,
            recruiter_email TEXT,
            is_read         INTEGER DEFAULT 0,
            created_at      TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS upload_log (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id         TEXT,
            recruiter_name   TEXT,
            recruiter_email  TEXT,
            filename         TEXT,
            candidates_added INTEGER DEFAULT 0,
            duplicates_found INTEGER DEFAULT 0,
            missing_count    INTEGER DEFAULT 0,
            uploaded_at      TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS google_sheet_import_state (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_id         TEXT NOT NULL,
            sheet_name       TEXT NOT NULL,
            last_row_number  INTEGER DEFAULT 0,
            updated_by       TEXT,
            updated_at       TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(sheet_id, sheet_name)
        );
        CREATE TABLE IF NOT EXISTS saved_searches (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT NOT NULL,
            filters    TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS submissions (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            candidate_id       INTEGER NOT NULL,
            recruiter_name     TEXT,
            recruiter_email    TEXT,
            hm_name            TEXT NOT NULL,
            hm_email           TEXT,
            hm_company         TEXT,
            role_submitted     TEXT,
            status             TEXT DEFAULT 'Pending',
            submitted_at       TEXT DEFAULT (datetime('now','localtime')),
            updated_at         TEXT DEFAULT (datetime('now','localtime')),
            notes              TEXT,
            FOREIGN KEY (candidate_id) REFERENCES candidates(id)
        );
        CREATE TABLE IF NOT EXISTS hiring_managers (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT NOT NULL,
            email       TEXT,
            company     TEXT,
            created_at  TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS pipelines (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            role_name   TEXT NOT NULL,
            status_list TEXT NOT NULL,
            is_default  INTEGER DEFAULT 0,
            created_at  TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS automation_rules (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            rule_name    TEXT NOT NULL,
            trigger_type TEXT NOT NULL,
            trigger_value TEXT,
            action_type  TEXT NOT NULL,
            action_config TEXT,
            is_active    INTEGER DEFAULT 1,
            created_at   TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS interviews (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            candidate_id   INTEGER NOT NULL,
            interviewer_name TEXT,
            interviewer_email TEXT,
            scheduled_at  TEXT,
            duration_mins INTEGER DEFAULT 60,
            location      TEXT,
            meeting_link  TEXT,
            status        TEXT DEFAULT 'Scheduled',
            notes         TEXT,
            created_at    TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (candidate_id) REFERENCES candidates(id)
        );
        CREATE TABLE IF NOT EXISTS approvals (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            candidate_id    INTEGER NOT NULL,
            requestor_name  TEXT,
            requestor_email TEXT,
            approval_type   TEXT,
            status          TEXT DEFAULT 'Pending',
            comments       TEXT,
            reviewed_by    TEXT,
            reviewed_at    TEXT,
            created_at     TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (candidate_id) REFERENCES candidates(id)
        );
        CREATE TABLE IF NOT EXISTS email_templates (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT NOT NULL UNIQUE,
            subject     TEXT NOT NULL,
            body        TEXT NOT NULL,
            trigger_event TEXT,
            is_default  INTEGER DEFAULT 0,
            created_at  TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS email_log (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            candidate_id   INTEGER,
            template_name  TEXT,
            recipient     TEXT NOT NULL,
            subject       TEXT,
            body          TEXT,
            status        TEXT DEFAULT 'Sent',
            error_msg     TEXT,
            sent_at       TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (candidate_id) REFERENCES candidates(id)
        );
        CREATE TABLE IF NOT EXISTS communication_campaigns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            campaign_type TEXT NOT NULL,
            name TEXT NOT NULL,
            linkedin_page_url TEXT,
            target_segment TEXT,
            send_schedule TEXT,
            tracking_link TEXT,
            follow_up_step TEXT,
            exclusion_rules_json TEXT,
            templates_json TEXT,
            analytics_json TEXT,
            status TEXT DEFAULT 'Active',
            created_by TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS communication_campaign_recipients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            campaign_id INTEGER NOT NULL,
            candidate_id INTEGER NOT NULL,
            email_addr TEXT NOT NULL,
            first_name TEXT,
            current_step INTEGER DEFAULT 1,
            max_steps INTEGER DEFAULT 3,
            status TEXT DEFAULT 'Pending',
            sent_count INTEGER DEFAULT 0,
            opened_count INTEGER DEFAULT 0,
            clicked_count INTEGER DEFAULT 0,
            tracking_token TEXT UNIQUE,
            next_send_at TEXT,
            last_sent_at TEXT,
            error_msg TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (campaign_id) REFERENCES communication_campaigns(id),
            FOREIGN KEY (candidate_id) REFERENCES candidates(id),
            UNIQUE(campaign_id, candidate_id)
        );
        CREATE TABLE IF NOT EXISTS screening_question_bank (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            skill TEXT NOT NULL,
            question TEXT NOT NULL,
            expected_signal TEXT,
            follow_up TEXT,
            source_model TEXT,
            usage_count INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            last_used_at TEXT,
            UNIQUE(skill, question)
        );
        CREATE TABLE IF NOT EXISTS skills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            raw_value TEXT,
            canonical_value TEXT NOT NULL UNIQUE,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS skill_aliases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            skill_id INTEGER NOT NULL,
            alias TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(skill_id, alias),
            FOREIGN KEY(skill_id) REFERENCES skills(id)
        );
        CREATE TABLE IF NOT EXISTS candidate_skills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            candidate_id INTEGER,
            resume_hash TEXT,
            raw_value TEXT,
            canonical_value TEXT NOT NULL,
            confidence REAL DEFAULT 0,
            skill_type TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS candidate_roles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            candidate_id INTEGER,
            resume_hash TEXT,
            raw_value TEXT,
            canonical_value TEXT NOT NULL,
            confidence REAL DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS candidate_domains (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            candidate_id INTEGER,
            resume_hash TEXT,
            raw_value TEXT,
            canonical_value TEXT NOT NULL,
            confidence REAL DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS jd_requirements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            jd_hash TEXT UNIQUE,
            role_title TEXT,
            parsed_json TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS match_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            jd_hash TEXT NOT NULL,
            resume_hash TEXT NOT NULL,
            final_score INTEGER,
            structured_score INTEGER,
            semantic_score INTEGER,
            hard_filter_score INTEGER,
            result_json TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(jd_hash, resume_hash)
        );
        CREATE TABLE IF NOT EXISTS parsed_resume_cache (
            resume_hash TEXT PRIMARY KEY,
            parsed_json TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS normalized_skill_cache (
            raw_value TEXT PRIMARY KEY,
            canonical_value TEXT NOT NULL,
            aliases_json TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS embedding_cache (
            text_hash TEXT NOT NULL,
            model TEXT NOT NULL,
            embedding_json TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            PRIMARY KEY(text_hash, model)
        );
        CREATE TABLE IF NOT EXISTS match_audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_type TEXT NOT NULL,
            object_type TEXT NOT NULL,
            object_hash TEXT,
            jd_hash TEXT,
            resume_hash TEXT,
            pipeline_version TEXT,
            status TEXT,
            source TEXT,
            parser_confidence TEXT,
            manual_review_required INTEGER DEFAULT 0,
            score INTEGER,
            message TEXT,
            details_json TEXT,
            ip_address TEXT,
            user_agent TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS ai_screening_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id TEXT,
            candidate_id INTEGER,
            requirement_id INTEGER,
            candidate_name TEXT,
            requirement_title TEXT,
            stage TEXT NOT NULL,
            status TEXT,
            score INTEGER,
            message TEXT,
            details_json TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        """)
        candidate_columns = {
            row["name"] for row in conn.execute("PRAGMA table_info(candidates)").fetchall()
        }
        if "sourcer_id" not in candidate_columns:
            conn.execute("ALTER TABLE candidates ADD COLUMN sourcer_id INTEGER")
        if "requirement_id" not in candidate_columns:
            conn.execute("ALTER TABLE candidates ADD COLUMN requirement_id INTEGER")
        if "candidate_feedback" not in candidate_columns:
            conn.execute("ALTER TABLE candidates ADD COLUMN candidate_feedback TEXT")
        if "industry_domain" not in candidate_columns:
            conn.execute("ALTER TABLE candidates ADD COLUMN industry_domain TEXT")
        for col in ["ai_screening_status", "ai_screening_score", "ai_screening_error", "ai_screening_report_json", "ai_screening_run_id", "ai_screening_updated_at"]:
            if col not in candidate_columns:
                conn.execute(f"ALTER TABLE candidates ADD COLUMN {col} TEXT")
        if "ai_screening_report_url" not in candidate_columns:
            conn.execute("ALTER TABLE candidates ADD COLUMN ai_screening_report_url TEXT")
        requirement_columns = {
            row["name"] for row in conn.execute("PRAGMA table_info(requirements)").fetchall()
        }
        for col in ["jd_filename", "jd_url", "jd_public_id"]:
            if col not in requirement_columns:
                conn.execute(f"ALTER TABLE requirements ADD COLUMN {col} TEXT")
        if "location" not in requirement_columns:
            conn.execute("ALTER TABLE requirements ADD COLUMN location TEXT")
        if "taggd_recruiter_id" not in requirement_columns:
            conn.execute("ALTER TABLE requirements ADD COLUMN taggd_recruiter_id INTEGER")
        if "taggd_recruiter_name" not in requirement_columns:
            conn.execute("ALTER TABLE requirements ADD COLUMN taggd_recruiter_name TEXT")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS taggd_recruiters (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                email TEXT,
                client_id INTEGER NOT NULL,
                is_active INTEGER DEFAULT 1,
                created_by TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                UNIQUE(client_id, name),
                FOREIGN KEY(client_id) REFERENCES clients(id)
            )
        """)
        team_columns = {
            row["name"] for row in conn.execute("PRAGMA table_info(team_members)").fetchall()
        }
        if "can_bulk_upload" not in team_columns:
            conn.execute("ALTER TABLE team_members ADD COLUMN can_bulk_upload INTEGER DEFAULT 0")
        if "gemini_api_key_enc" not in team_columns:
            conn.execute("ALTER TABLE team_members ADD COLUMN gemini_api_key_enc TEXT")
        if "gemini_api_key_updated_at" not in team_columns:
            conn.execute("ALTER TABLE team_members ADD COLUMN gemini_api_key_updated_at TEXT")
        if "last_login_at" not in team_columns:
            conn.execute("ALTER TABLE team_members ADD COLUMN last_login_at TEXT")
        user_columns = {
            row["name"] for row in conn.execute("PRAGMA table_info(app_users)").fetchall()
        }
        if "is_bulk_admin" not in user_columns:
            conn.execute("ALTER TABLE app_users ADD COLUMN is_bulk_admin INTEGER DEFAULT 0")
        if "email" not in user_columns:
            conn.execute("ALTER TABLE app_users ADD COLUMN email TEXT")
        if "last_login_at" not in user_columns:
            conn.execute("ALTER TABLE app_users ADD COLUMN last_login_at TEXT")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_app_users_username_lower ON app_users(lower(trim(username)))")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_app_users_email_lower ON app_users(lower(trim(COALESCE(email,''))))")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_app_users_team_member ON app_users(team_member_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_app_users_active ON app_users(is_active)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_team_members_email_lower ON team_members(lower(trim(COALESCE(email,''))))")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS user_login_audit (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                app_user_id INTEGER,
                team_member_id INTEGER,
                username TEXT,
                email TEXT,
                display_name TEXT,
                role TEXT,
                method TEXT,
                status TEXT,
                ip_address TEXT,
                user_agent TEXT,
                message TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        """)
        conn.execute("""CREATE TABLE IF NOT EXISTS team_leader_mappings (
            leader_team_member_id INTEGER NOT NULL,
            member_team_member_id INTEGER NOT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            PRIMARY KEY(leader_team_member_id, member_team_member_id)
        )""")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_user_login_audit_created ON user_login_audit(created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_user_login_audit_email ON user_login_audit(lower(email))")
        conn.execute("""CREATE TABLE IF NOT EXISTS performance_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT,
            team_member_id INTEGER,
            recruiter_name TEXT,
            recruiter_email TEXT,
            method TEXT,
            path TEXT,
            endpoint TEXT,
            status_code INTEGER,
            elapsed_ms REAL,
            is_admin INTEGER DEFAULT 0,
            ip_address TEXT,
            user_agent TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )""")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_performance_logs_created ON performance_logs(created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_performance_logs_user ON performance_logs(username, team_member_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_performance_logs_path ON performance_logs(path, created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_match_audit_log_object ON match_audit_log(object_type, object_hash)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_match_audit_log_hashes ON match_audit_log(jd_hash, resume_hash)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_ai_screening_logs_created ON ai_screening_logs(created_at)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_ai_screening_logs_candidate ON ai_screening_logs(candidate_id)")
        app_settings_columns = {
            row["name"] for row in conn.execute("PRAGMA table_info(app_settings)").fetchall()
        }
        if "updated_at" not in app_settings_columns:
            conn.execute("ALTER TABLE app_settings ADD COLUMN updated_at TEXT")
        ensure_ats_pipeline_schema(conn)

        # Insert default pipeline if not exists
        conn.execute("INSERT OR IGNORE INTO pipelines (role_name,status_list,is_default) VALUES ('Default',?,1)",
                     (json.dumps(["New","Shortlisted","Feedback Pending","Offered","On Hold","Joined","Rejected","Duplicate"]),))
        # Insert default email templates
        defaults = [
            ('welcome', 'Application received for {role_name}', 
             'Hi {candidate_name},\n\nThank you for applying for {role_name}. We have received your profile and will review it shortly.\n\nBest regards,\n{recruiter_name}',
             'new_candidate', 1),
            ('status_update', 'Update on your {role_name} application',
             'Hi {candidate_name},\n\nYour application for {role_name} is now marked as {status}.\n\nBest regards,\n{recruiter_name}',
             'status_change', 1),
            ('interview_invite', 'Interview invitation for {role_name}',
             'Hi {candidate_name},\n\nCongratulations! You have been shortlisted for an interview for {role_name}.\n\nInterview Details:\nDate: {scheduled_at}\nInterviewer: {interviewer_name}\nLocation: {location}\n\nBest regards,\n{recruiter_name}',
             'interview_scheduled', 1)
        ]
        for name, subject, body, trigger, is_def in defaults:
            conn.execute("INSERT OR IGNORE INTO email_templates (name,subject,body,trigger_event,is_default) VALUES (?,?,?,?,?)",
                         (name, subject, body, trigger, is_def))
            conn.execute("UPDATE email_templates SET subject=?, body=?, trigger_event=?, is_default=? WHERE name=? AND is_default=1",
                         (subject, body, trigger, is_def, name))
        # Insert standard roles if not exists
        standard_roles = [
            ('IT', 'Program Manager'), ('IT', 'Software Engineer'), ('IT', 'Senior Software Engineer'),
            ('IT', 'Tech Lead'), ('IT', 'Data Analyst'), ('IT', 'DevOps Engineer'),
            ('IT', 'QA Engineer'), ('IT', 'UI/UX Designer'), ('IT', 'System Administrator'),
            ('Manufacturing', 'Plant Head'), ('Manufacturing', 'Production Manager'),
            ('Manufacturing', 'Quality Manager'), ('Manufacturing', 'Supply Chain Manager'),
            ('Automobile', 'Design Engineer'), ('Automobile', 'Production Engineer'),
            ('Automobile', 'Quality Engineer'), ('Automobile', 'R&D Engineer'),
            ('Finance', 'Accountant'), ('Finance', 'Financial Analyst'),
            ('Finance', 'Tax Specialist'), ('Finance', 'Auditor'),
            ('HR', 'Recruiter'), ('HR', 'HR Manager'), ('HR', 'Training Coordinator'),
            ('Sales', 'Sales Manager'), ('Sales', 'Business Developer'),
            ('Sales', 'Account Manager'), ('Marketing', 'Marketing Manager'),
            ('Marketing', 'Content Writer'), ('Marketing', 'SEO Specialist'),
            ('Operations', 'Operations Manager'), ('Operations', 'Logistics Manager'),
            ('Legal', 'Legal Counsel'), ('Legal', 'Compliance Officer'),
        ]
        for domain, role in standard_roles:
            conn.execute("INSERT OR IGNORE INTO standard_roles (domain, role_name) VALUES (?,?)", (domain, role))
        # Insert standard skills if not exists
        standard_skills = [
            ('Python', 'Programming'), ('JavaScript', 'Programming'), ('Java', 'Programming'), ('C++', 'Programming'),
            ('React', 'Frontend'), ('Angular', 'Frontend'), ('Vue.js', 'Frontend'), ('HTML', 'Frontend'), ('CSS', 'Frontend'),
            ('Node.js', 'Backend'), ('Django', 'Backend'), ('Spring', 'Backend'), ('Flask', 'Backend'),
            ('SQL', 'Database'), ('MongoDB', 'Database'), ('PostgreSQL', 'Database'), ('MySQL', 'Database'),
            ('AWS', 'Cloud'), ('Azure', 'Cloud'), ('GCP', 'Cloud'), ('Docker', 'DevOps'), ('Kubernetes', 'DevOps'),
            ('Git', 'Tools'), ('Jira', 'Tools'), (' Jenkins', 'CI/CD'),
            ('Machine Learning', 'AI/ML'), ('Data Analysis', 'AI/ML'), ('NLP', 'AI/ML'),
            ('Excel', 'Office'), ('PowerPoint', 'Office'), ('Word', 'Office'),
            ('Communication', 'Soft Skills'), ('Teamwork', 'Soft Skills'), ('Leadership', 'Soft Skills'),
        ]
        for skill, category in standard_skills:
            conn.execute("INSERT OR IGNORE INTO standard_skills (skill_name, category) VALUES (?,?)", (skill, category))
        # Insert standard clients if not exists
        standard_clients = [
            'Diamanti', 'EY', 'KPMG', 'Taggd-Birla Opus', 'Taggd-HPE', 'Taggd-Hyundai',
            'Taggd-M&M', 'Taggd-Neosoft', 'Taggd-Pidilite', 'Taggd-Siemens', 'Taggd-TCPL', 'e-Zest'
        ]
        for client_name in standard_clients:
            conn.execute("INSERT OR IGNORE INTO clients (client_name) VALUES (?)", (client_name,))
        # Keep requirement creation user-driven; do not seed demo requirements.
        standard_reqs = []
        default_checks = [
            "Technical Skills", "Years of Relevant Experience", "Within Given Budget",
            "Notice Period", "Location", "Non-Poachable Employee", "Updated CV"
        ]
        for title, client, status in standard_reqs:
            # Check if requirements already exists
            existing = conn.execute("SELECT id FROM requirements WHERE title=? AND client_name=?", (title, client)).fetchone()
            if not existing:
                cursor = conn.execute("INSERT INTO requirements (title, client_name, status) VALUES (?,?,?)", (title, client, status))
                rid = cursor.lastrowid
                # Add default checks for new requirement
                for i, check_name in enumerate(default_checks):
                    conn.execute("""INSERT INTO requirement_checks 
                        (requirement_id,check_name,check_description,check_type,pass_criteria,sort_order)
                        VALUES (?,?,?,?,?,?)""",
                        (rid, check_name, "", "boolean", "Yes", i))
        for index_sql in [
            "CREATE INDEX IF NOT EXISTS idx_candidates_created_at ON candidates(created_at)",
            "CREATE INDEX IF NOT EXISTS idx_candidates_status ON candidates(status)",
            "CREATE INDEX IF NOT EXISTS idx_candidates_requirement_id ON candidates(requirement_id)",
            "CREATE INDEX IF NOT EXISTS idx_candidates_recruiter_email ON candidates(recruiter_email)",
            "CREATE INDEX IF NOT EXISTS idx_candidates_sourcer_id ON candidates(sourcer_id)",
            "CREATE INDEX IF NOT EXISTS idx_candidates_duplicate ON candidates(is_duplicate)",
            "CREATE INDEX IF NOT EXISTS idx_candidates_owner_created ON candidates(sourcer_id, created_at)",
            "CREATE INDEX IF NOT EXISTS idx_candidates_owner_status_created ON candidates(sourcer_id, status, created_at)",
            "CREATE INDEX IF NOT EXISTS idx_candidates_requirement_created ON candidates(requirement_id, created_at)",
            "CREATE INDEX IF NOT EXISTS idx_candidates_status_created ON candidates(status, created_at)",
            "CREATE INDEX IF NOT EXISTS idx_candidates_email_lower ON candidates(lower(recruiter_email))",
            "CREATE INDEX IF NOT EXISTS idx_candidates_candidate_name ON candidates(candidate_name)",
            "CREATE INDEX IF NOT EXISTS idx_candidates_email_addr ON candidates(email_addr)",
            "CREATE INDEX IF NOT EXISTS idx_candidates_phone ON candidates(phone)",
            "CREATE INDEX IF NOT EXISTS idx_candidates_phone_normalized ON candidates(replace(replace(replace(phone,'+',''),'-',''),' ',''))",
            "CREATE INDEX IF NOT EXISTS idx_candidates_role_company ON candidates(role_name, current_company)",
            "CREATE INDEX IF NOT EXISTS idx_candidates_locations ON candidates(current_location, preferred_location)",
            "CREATE INDEX IF NOT EXISTS idx_candidates_skills ON candidates(key_skills)",
            "CREATE INDEX IF NOT EXISTS idx_requirements_created_at ON requirements(created_at DESC)",
            "CREATE INDEX IF NOT EXISTS idx_requirements_status_created ON requirements(status, created_at DESC)",
            "CREATE INDEX IF NOT EXISTS idx_requirements_title_lower ON requirements(lower(COALESCE(title,'')))",
            "CREATE INDEX IF NOT EXISTS idx_requirements_client_name ON requirements(client_name)",
            "CREATE INDEX IF NOT EXISTS idx_requirements_client_id ON requirements(client_name, id)",
            "CREATE INDEX IF NOT EXISTS idx_requirements_client_lower ON requirements(lower(trim(COALESCE(client_name,''))))",
        ]:
            conn.execute(index_sql)
        date_cleanup_version = "2026-05-23-candidate-date-cleanup-v2"
        cleanup_done = conn.execute(
            "SELECT value FROM app_settings WHERE key='candidate_date_cleanup_version'"
        ).fetchone()
        if not cleanup_done or cleanup_done["value"] != date_cleanup_version:
            normalize_existing_candidate_created_at(conn)
            correct_future_swapped_candidate_dates(conn)
            conn.execute("""
                INSERT INTO app_settings (key,value,updated_at)
                VALUES ('candidate_date_cleanup_version',?,datetime('now','localtime'))
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=datetime('now','localtime')
            """, (date_cleanup_version,))
        conn.commit()

    # â”€â”€ Auth â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    finally:
        conn.close()

def initialize_app_database():
    attempts = 5
    delay_seconds = 2
    for attempt in range(1, attempts + 1):
        try:
            init_db()
            conn = get_db()
            admin = conn.execute("SELECT id FROM app_users WHERE username='admin'").fetchone()
            if not admin and ADMIN_PASSWORD:
                conn.execute(
                    "INSERT INTO app_users (username,password,is_admin) VALUES (?,?,1)",
                    ("admin", hash_password(ADMIN_PASSWORD))
                )
                conn.commit()
                print("Created default admin user: admin", flush=True)
            elif not admin:
                print("No default admin created. Set ADMIN_PASSWORD to create one.", flush=True)
            conn.close()
            return
        except sqlite3.OperationalError as e:
            if "locked" in str(e).lower() or "busy" in str(e).lower():
                if attempt < attempts:
                    print(
                        f"Database initialization busy/locked (attempt {attempt}/{attempts}); retrying in {delay_seconds}s",
                        flush=True,
                    )
                    time.sleep(delay_seconds)
                    continue
            raise


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            # API routes return 401, page routes redirect
            if request.path.startswith("/api/"):
                return jsonify({"error": "Unauthorised"}), 401
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated


def register_invoice_app():
    invoice_app_dir = os.getenv("INVOICE_APP_DIR", r"D:\HRGuruInvoiceApp")
    if invoice_app_dir not in sys.path:
        sys.path.insert(0, invoice_app_dir)

    try:
        from invoice_web_app import invoice_bp
    except ModuleNotFoundError as exc:
        if exc.name == "invoice_web_app":
            print(f"Invoice app integration skipped: invoice_web_app not found at {invoice_app_dir}", flush=True)
            return
        raise

    @invoice_bp.before_request
    def require_invoice_admin():
        if not session.get("logged_in"):
            return redirect(url_for("login_page"))
        if not session.get("is_admin"):
            return "Admin access required.", 403

    app.register_blueprint(invoice_bp, url_prefix="/invoices")


register_invoice_app()
def has_bulk_upload_access():
    if not session.get("logged_in"):
        return False
    session["can_bulk_upload"] = 1
    return True

def app_setting_bool(key, default=False):
    conn = get_db()
    try:
        row = conn.execute("SELECT value FROM app_settings WHERE key=?", (key,)).fetchone()
        if not row:
            return bool(default)
        return str(row["value"] or "").strip().lower() in ("1", "true", "yes", "on")
    finally:
        conn.close()

def set_app_setting(key, value):
    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO app_settings (key,value,updated_at)
            VALUES (?,?,datetime('now','localtime'))
            ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=datetime('now','localtime')
        """, (key, str(value)))
        conn.commit()
    finally:
        conn.close()

@app.route("/login", methods=["GET"])
def login_page():
    if session.get("logged_in"):
        #return "THIS IS NEW LOGIN PAGE"
        return redirect(url_for(post_login_redirect_endpoint()))
    return render_template("login.html")

@app.route("/login", methods=["POST"])
def login_password():
    endpoint_started = time.perf_counter()
    username = (request.form.get("username") or "").strip().lower()
    password = request.form.get("password") or ""
    login_ip = (request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip())
    login_user_agent = (request.headers.get("User-Agent") or "")[:300]
    if not username or not password:
        return render_template("login.html", error="Enter username and password."), 400
    lookup_started = time.perf_counter()
    conn = get_db(timeout=3)
    try:
        user = conn.execute(
            """SELECT u.*, t.email AS team_email, t.name AS team_name, t.role AS team_role, t.can_bulk_upload AS team_can_bulk_upload
               FROM app_users u
               LEFT JOIN team_members t ON t.id = u.team_member_id
               WHERE lower(trim(u.username))=lower(trim(?))
                  OR lower(trim(COALESCE(u.email,'')))=lower(trim(?))
                  OR lower(trim(COALESCE(t.email,'')))=lower(trim(?))
               LIMIT 1""",
            (username, username, username)
        ).fetchone()
    finally:
        conn.close()
    perf_log("login.lookup", lookup_started, user=username, found=1 if user else 0)
    password_started = time.perf_counter()
    password_ok = bool(user and check_password_hash(user["password"], password))
    perf_log("login.password_check", password_started, user=username, ok=1 if password_ok else 0)
    if not password_ok:
        artifact_started = time.perf_counter()
        queue_password_login_artifacts(
            app_user_id=user["id"] if user else None,
            team_member_id=user["team_member_id"] if user else None,
            username=username,
            email=(user["team_email"] if user else "") or "",
            display_name=(user["team_name"] if user else "") or username,
            role=(user["team_role"] if user else "") or "",
            status="failed",
            message="Invalid username or password.",
            ip_address=login_ip,
            user_agent=login_user_agent,
        )
        perf_log("login.artifacts", artifact_started, user=username, status="failed")
        perf_log("login.total", endpoint_started, user=username, status=401)
        return render_template("login.html", error="Invalid username or password."), 401
    if not user["is_active"]:
        artifact_started = time.perf_counter()
        queue_password_login_artifacts(
            app_user_id=user["id"],
            team_member_id=user["team_member_id"],
            username=user["username"],
            email=user["team_email"] or "",
            display_name=user["team_name"] or user["username"],
            role=user["team_role"] or "",
            status="failed",
            message="Login user is inactive.",
            ip_address=login_ip,
            user_agent=login_user_agent,
        )
        perf_log("login.artifacts", artifact_started, user=username, status="inactive")
        perf_log("login.total", endpoint_started, user=username, status=403)
        return render_template("login.html", error="This login is inactive. Please contact admin."), 403
    hash_started = time.perf_counter()
    upgraded_password_hash = hash_password(password) if password_hash_needs_upgrade(user["password"]) else None
    perf_log("login.hash_upgrade", hash_started, user=username, upgraded=1 if upgraded_password_hash else 0)
    set_login_session_from_joined_user(user)
    artifact_started = time.perf_counter()
    queue_password_login_artifacts(
        app_user_id=user["id"],
        team_member_id=user["team_member_id"],
        username=user["username"],
        email=user["team_email"] or "",
        display_name=user["team_name"] or user["username"],
        role=user["team_role"] or "",
        status="success",
        message="Password login successful.",
        ip_address=login_ip,
        user_agent=login_user_agent,
        password_hash_upgrade=upgraded_password_hash,
    )
    perf_log("login.artifacts", artifact_started, user=username, status="success")
    perf_log("login.total", endpoint_started, user=username, status=302)
    return redirect(url_for(post_login_redirect_endpoint()))

@app.route("/forgot_password", methods=["POST"])
def forgot_password():
    username = (request.form.get("username") or "").strip().lower()
    conn = get_db()
    ensure_auth_workflow_schema(conn)
    user = resolve_user_for_password_reset(conn, username)
    email_results = []
    if user and user["is_active"]:
        token = secrets.token_urlsafe(32)
        conn.execute(
            "INSERT INTO password_reset_tokens (user_id, token, expires_at) VALUES (?,?,datetime('now','+2 hours'))",
            (user["id"], token)
        )
        conn.commit()
        reset_url = url_for("reset_password", token=token, _external=True)
        recipients = reset_recipients_for_user(user)
        for recipient in recipients:
            try:
                result = system_email(
                    recipient,
                    "Reset your HR Guru ATS password",
                    f"Use this link to reset your password. It expires in 2 hours:\n\n{reset_url}"
                )
                email_results.append({"recipient": recipient, **result})
            except Exception as e:
                print("Password reset email error:", e)
                email_results.append({"recipient": recipient, "error": str(e)})
    conn.close()
    if user and user["is_active"] and not email_results:
        return render_template("login.html", message="Reset link was created, but no admin/user email is configured. Please set ADMIN_EMAIL or link the user to a team email.")
    if any(item.get("ok") for item in email_results):
        return render_template("login.html", message="Password reset email has been sent.")
    if email_results:
        return render_template("login.html", message="Reset link was created, but email sending failed. Please check GMAIL_USER and GMAIL_APP_PASS.")
    return render_template("login.html", message="If the account exists, a password reset email has been sent.")

@app.route("/reset_password/<token>", methods=["GET", "POST"])
def reset_password(token):
    if session.get("logged_in"):
        session.clear()
    conn = get_db()
    ensure_auth_workflow_schema(conn)
    row = conn.execute("""
        SELECT * FROM password_reset_tokens
        WHERE token=? AND used=0 AND datetime(expires_at) >= datetime('now')
    """, (token,)).fetchone()
    if not row:
        conn.close()
        return render_template("login.html", error="This reset link is invalid or expired."), 400
    if request.method == "POST":
        password = request.form.get("password") or ""
        if len(password) < 8:
            conn.close()
            return render_template("login.html", reset_token=token, error="Password must be at least 8 characters."), 400
        new_hash = hash_password(password)
        conn.execute("UPDATE app_users SET password=? WHERE id=?", (new_hash, row["user_id"]))
        conn.execute("UPDATE password_reset_tokens SET used=1 WHERE user_id=?", (row["user_id"],))
        conn.commit()
        updated = conn.execute("SELECT password FROM app_users WHERE id=?", (row["user_id"],)).fetchone()
        if not updated or not check_password_hash(updated["password"], password):
            conn.close()
            return render_template("login.html", reset_token=token, error="Password update could not be verified. Please try again."), 500
        conn.close()
        return render_template("login.html", message="Password reset successfully. You can sign in now.")
    conn.close()
    return render_template("login.html", reset_token=token)

@app.route("/register", methods=["POST"])
def register_user_request():
    name = (request.form.get("name") or "").strip()
    email = (request.form.get("email") or "").strip().lower()
    username = (request.form.get("username") or "").strip().lower()
    password = request.form.get("password") or ""
    if not email or not username or len(password) < 8:
        return render_template("login.html", register_error="Email, username, and an 8+ character password are required."), 400
    token = secrets.token_urlsafe(32)
    conn = get_db()
    ensure_auth_workflow_schema(conn)
    existing = conn.execute("SELECT id FROM app_users WHERE LOWER(username)=LOWER(?)", (username,)).fetchone()
    if existing:
        conn.close()
        return render_template("login.html", register_error="That username is already taken."), 400
    conn.execute(
        """INSERT INTO user_registration_requests
           (name, email, username, password_hash, token)
           VALUES (?,?,?,?,?)""",
        (name, email, username, hash_password(password), token)
    )
    conn.commit()
    conn.close()
    approve_url = url_for("approve_registration", token=token, _external=True)
    body = (
        f"New HR Guru ATS registration request:\n\n"
        f"Name: {name or '-'}\nEmail: {email}\nUsername: {username}\n\n"
        f"Admin approval link:\n{approve_url}\n\n"
        "You must be logged in as an admin to approve this request."
    )
    for admin_email in admin_approval_emails():
        try:
            system_email(admin_email, "Approve new HR Guru ATS user", body)
        except Exception as e:
            print("Registration approval email error:", e)
    return render_template("login.html", message="Registration request submitted. An admin must approve it before you can sign in.")

@app.route("/register/approve/<token>")
@login_required
def approve_registration(token):
    if not session.get("is_admin"):
        return "Admin approval required.", 403
    conn = get_db()
    ensure_auth_workflow_schema(conn)
    req = conn.execute("SELECT * FROM user_registration_requests WHERE token=? AND status='pending'", (token,)).fetchone()
    if not req:
        conn.close()
        return "Registration request is invalid or already processed.", 400
    member = conn.execute("SELECT id FROM team_members WHERE LOWER(email)=LOWER(?)", (req["email"],)).fetchone()
    team_member_id = member["id"] if member else None
    if not team_member_id:
        team_member_id = conn.execute(
            "INSERT INTO team_members (name,email,role) VALUES (?,?,?)",
            (req["name"] or req["username"], req["email"], "User")
        ).lastrowid
    conn.execute(
        """INSERT INTO app_users (username,password,team_member_id,is_admin,is_bulk_admin,is_active)
           VALUES (?,?,?,?,?,1)""",
        (req["username"], req["password_hash"], team_member_id, 0, 0)
    )
    conn.execute(
        "UPDATE user_registration_requests SET status='approved', approved_by=?, approved_at=datetime('now','localtime') WHERE id=?",
        (session.get("user_id"), req["id"])
    )
    conn.commit()
    conn.close()
    try:
        system_email(req["email"], "Your HR Guru ATS account is approved", "Your account has been approved. You can now sign in with the username and password you registered.")
    except Exception as e:
        print("Registration approved email error:", e)
    return redirect(url_for("index"))

@app.route("/login/google")
def login_google():
    public_base_url = (os.getenv("PUBLIC_BASE_URL") or "").strip().rstrip("/")
    request_host = (request.host or "").lower()
    is_local_request = (
        request_host.startswith("localhost")
        or request_host.startswith("127.0.0.1")
        or request_host.startswith("0.0.0.0")
    )
    if is_local_request:
        public_base_url = ""
    if not public_base_url:
        for header_name in ("X-Forwarded-Host", "Host"):
            host = (request.headers.get(header_name) or "").split(",")[0].strip()
            if host.lower().startswith("hireflow.hrgp.in"):
                public_base_url = "https://hireflow.hrgp.in"
                break
    if not public_base_url:
        for header_name in ("Referer", "Origin"):
            parsed = urlparse(request.headers.get(header_name) or "")
            if parsed.netloc.lower().startswith("hireflow.hrgp.in"):
                public_base_url = "https://hireflow.hrgp.in"
                break
    redirect_uri = (os.getenv("GOOGLE_REDIRECT_URI") or "").strip()
    if is_local_request:
        redirect_uri = ""
    if public_base_url and (
        not redirect_uri
        or "localhost" in redirect_uri
        or "127.0.0.1" in redirect_uri
        or redirect_uri.startswith("http://")
    ):
        redirect_uri = f"{public_base_url}/login/callback"
    if not redirect_uri:
        if public_base_url:
            redirect_uri = f"{public_base_url}/login/callback"
        elif request.host.lower().startswith("hireflow.hrgp.in"):
            redirect_uri = "https://hireflow.hrgp.in/login/callback"
        else:
            redirect_uri = url_for("google_callback", _external=True).replace("/auth/google/callback", "/login/callback")
    if redirect_uri.startswith("http://hireflow.hrgp.in"):
        redirect_uri = redirect_uri.replace("http://hireflow.hrgp.in", "https://hireflow.hrgp.in", 1)
    print("Google OAuth redirect_uri:", redirect_uri, flush=True)
    try:
        return google.authorize_redirect(
            redirect_uri,
            access_type="offline",
            prompt="consent",
            include_granted_scopes="true"
        )
    except Exception as e:
        print("Google login start error:", e, flush=True)
        traceback.print_exc()
        return render_template(
            "login.html",
            message="Google login could not start. Please check internet access from the ATS server and try again."
        ), 503

@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"ok": True})

@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory('uploads', filename, as_attachment=True)

@app.route("/login/callback")
@app.route("/auth/google/callback")
def google_callback():
    try:
        token = google.authorize_access_token()
        session["google_token"] = token

        #resp = google.get("userinfo")
        resp = google.get("https://www.googleapis.com/oauth2/v3/userinfo")
        user_info = resp.json()
    except Exception as e:
        print("Google login callback error:", e, flush=True)
        traceback.print_exc()
        return render_template(
            "login.html",
            message="Google login could not be completed. Please reconnect and try again."
        ), 503

    email = user_info["email"].lower()
    username = email.split("@")[0]
    full_name = user_info.get("name", username.title())
    login_ip = (request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip())
    login_user_agent = (request.headers.get("User-Agent") or "")[:300]

    print("Google userinfo:", email)
    

    conn = get_db(timeout=3)
    try:
        team_member = _best_effort_fetchone(
            conn,
            "SELECT * FROM team_members WHERE LOWER(TRIM(email)) = LOWER(TRIM(?))",
            (email,),
            label="team member lookup"
        )
        app_user = _best_effort_fetchone(
            conn,
            """
            SELECT u.*
            FROM app_users u
            LEFT JOIN team_members t ON t.id = u.team_member_id
            WHERE LOWER(TRIM(u.username)) = LOWER(TRIM(?))
               OR LOWER(TRIM(COALESCE(u.email, ''))) = LOWER(TRIM(?))
               OR LOWER(TRIM(COALESCE(t.email, ''))) = LOWER(TRIM(?))
            ORDER BY u.is_active DESC, u.id DESC
            LIMIT 1
            """,
            (username, email, email),
            label="app user lookup"
        )

        if not team_member and app_user and app_user["team_member_id"]:
            team_member = _best_effort_fetchone(
                conn,
                "SELECT * FROM team_members WHERE id=?",
                (app_user["team_member_id"],),
                label="team member by id"
            )

        if not team_member and app_user:
            team_member = _best_effort_fetchone(
                conn,
                "SELECT * FROM team_members WHERE LOWER(TRIM(email)) = LOWER(TRIM(?)) OR LOWER(TRIM(name)) = LOWER(TRIM(?))",
                (email, full_name),
                label="team member by email/name"
            )

        if not team_member:
            queue_google_login_artifacts(
                app_user_id=None,
                team_member_id=None,
                username=username,
                email=email,
                display_name=full_name,
                role="",
                method="google",
                status="denied",
                message="Google account is not provisioned in ATS team members.",
                ip_address=login_ip,
                user_agent=login_user_agent,
            )
            return "Access not provisioned. Contact Admin.", 403

        if not app_user:
            app_user = _best_effort_fetchone(
                conn,
                "SELECT * FROM app_users WHERE LOWER(TRIM(username)) = LOWER(TRIM(?)) OR LOWER(TRIM(COALESCE(email, ''))) = LOWER(TRIM(?)) LIMIT 1",
                (username, email),
                label="app user fallback lookup"
            )

        if not app_user:
            temp_password_hash = hash_password(secrets.token_urlsafe(16))
            _best_effort_write(
                conn,
                """INSERT INTO app_users (username, password, team_member_id, is_admin, is_bulk_admin, is_active, email)
                   VALUES (?,?,?,?,?,?,?)""",
                (
                    username,
                    temp_password_hash,
                    team_member["id"],
                    1 if (team_member["role"] or "").strip().lower() == "admin" else 0,
                    1 if (team_member["can_bulk_upload"] or 0) else 0,
                    1,
                    email
                ),
                label="app user create"
            )
            app_user = _best_effort_fetchone(conn, "SELECT * FROM app_users WHERE LOWER(TRIM(username)) = LOWER(TRIM(?)) LIMIT 1", (username,), label="app user refresh")

        if not app_user:
            raise RuntimeError("Unable to resolve Google user to an ATS app user.")

        if not app_user["is_active"]:
            queue_google_login_artifacts(
                app_user_id=app_user["id"],
                team_member_id=team_member["id"],
                username=username,
                email=email,
                display_name=team_member["name"],
                role=team_member["role"] or "",
                method="google",
                status="denied",
                message="Linked ATS user is inactive.",
                ip_address=login_ip,
                user_agent=login_user_agent,
            )
            return "Access blocked. Contact Admin.", 403

        if not app_user["team_member_id"]:
            _best_effort_write(conn, "UPDATE app_users SET team_member_id=? WHERE id=?", (team_member["id"], app_user["id"]), label="link app user to team member")
            app_user = _best_effort_fetchone(conn, "SELECT * FROM app_users WHERE id=?", (app_user["id"],), label="app user post-link refresh") or app_user

        set_login_session_from_app_user(app_user, db_timeout=3)
        session["username"] = team_member["name"]
        session["recruiter_name"] = team_member["name"]
        session["email"] = team_member["email"]
        session["recruiter_email"] = team_member["email"]
        role_name = (team_member["role"] or "").strip().lower()
        session["is_admin"] = 1 if role_name == "admin" or app_user["is_admin"] else 0
        session["can_bulk_upload"] = 1 if session["is_admin"] or team_member["can_bulk_upload"] or role_name == "bulk admin" or app_user["is_bulk_admin"] else 0
        queue_google_login_artifacts(
            app_user["id"],
            team_member["id"],
            username=username,
            email=team_member["email"],
            display_name=team_member["name"],
            role=team_member["role"] or "",
            method="google",
            status="success",
            message="Login successful.",
            ip_address=login_ip,
            user_agent=login_user_agent,
        )
        return redirect(url_for(post_login_redirect_endpoint()))
    except sqlite3.OperationalError as exc:
        if "database is locked" in str(exc).lower():
            print(f"Google callback retry skipped due to DB lock: {exc}", flush=True)
            return render_template(
                "login.html",
                message="ATS is busy saving data right now. Please retry the Google login in a moment."
            ), 503
        raise
    finally:
        try:
            conn.close()
        except Exception:
            pass

# â”€â”€ Column map â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
COL_MAP = {
    "candidate name":"candidate_name","name":"candidate_name","candidate":"candidate_name",
    "candidate full name":"candidate_name","full name":"candidate_name","applicant name":"candidate_name",
    "email":"email_addr","email id":"email_addr","email address":"email_addr","email-id":"email_addr",
    "candidate email":"email_addr","mail id":"email_addr","e-mail":"email_addr",
    "phone":"phone","phone number":"phone","telephone":"phone","mobile":"phone","mobile number":"phone",
    "contact":"phone","contact no":"phone","contact no.":"phone","contact number":"phone",
    "candidate phone":"phone","candidate mobile":"phone","phone number":"phone",
    "current company":"current_company","current compant":"current_company","current org.":"current_company",
    "current org":"current_company","company":"current_company","employer":"current_company",
    "organisation":"current_company","organization":"current_company","current org":"current_company",
    "client":"client_name","client name":"client_name","customer":"client_name","account":"client_name",
    "current role":"current_role","designation":"current_role","title":"current_role",
    "position":"current_role","current designation":"current_role","profile":"current_role",
    "experience":"experience_years","exp":"experience_years","exp.":"experience_years","total exp":"experience_years",
    "years of exp":"experience_years","experience (yrs)":"experience_years","experience (years)":"experience_years",
    "yrs":"experience_years","total experience":"experience_years","total experience (in yrs)":"experience_years",
    "skills":"key_skills","key skills":"key_skills","skill set":"key_skills","primary skills":"key_skills",
    "tech skills":"key_skills","technical skills":"key_skills",
    "notice":"notice_period","notice period":"notice_period","np":"notice_period",
    "current salary":"current_salary","current ctc":"current_salary","ctc":"current_salary","ctc (lacs)":"current_salary","salary":"current_salary",
    "expected salary":"expected_salary","expected ctc":"expected_salary","ectc":"expected_salary","exp salary":"expected_salary",
    "location":"current_location","current location":"current_location","city":"current_location",
    "preferred location":"preferred_location","pref location":"preferred_location","preferred city":"preferred_location",
    "job location":"preferred_location",
    "remarks":"remarks","notes":"remarks","comments":"remarks","feedback":"remarks","recruiter comments":"remarks",
    "role":"role_name","job role":"role_name","applied for":"role_name",
    "position applied":"role_name","job title":"role_name","position name":"role_name",
    "position title":"requirement_title",
    "requirement":"requirement_title","requirement title":"requirement_title",
    "requirement id":"requirement_id","req id":"requirement_id",
    "education":"education",
    "status":"status","candidate status":"status","cv status":"status","ats status":"status","stage":"status",
    "date":"created_at","added date":"created_at","added on":"created_at","created date":"created_at","created at":"created_at",
    "date added":"created_at","submission date":"created_at","profile date":"created_at",
    "industry":"industry_domain","domain":"industry_domain","industry/domain":"industry_domain",
    "industry domain":"industry_domain","vertical":"industry_domain",
    "sourcer":"recruiter_name","sourcer name":"recruiter_name","recruiter":"recruiter_name",
    "recruiter name":"recruiter_name","sourcer/recruiter name":"recruiter_name",
}
ALL_FIELDS = ["candidate_name","email_addr","phone","current_company","current_role",
              "experience_years","key_skills","notice_period","current_salary",
              "expected_salary","current_location","preferred_location","remarks",
              "role_name","requirement_title","requirement_id","client_name","education",
              "status","created_at","industry_domain","recruiter_name"]
ATS_FIELD_LABELS = {
    "": "Do not import",
    "candidate_name": "Candidate Name",
    "email_addr": "Email",
    "phone": "Phone",
    "current_company": "Current Company",
    "current_role": "Current Role",
    "experience_years": "Experience",
    "key_skills": "Key Skills",
    "notice_period": "Notice Period",
    "current_salary": "Current Salary",
    "expected_salary": "Expected Salary",
    "current_location": "Current Location",
    "preferred_location": "Preferred Location",
    "remarks": "Remarks",
    "role_name": "Role Name",
    "requirement_title": "Requirement / Position Title",
    "requirement_id": "Requirement ID",
    "client_name": "Client Name",
    "education": "Education",
    "status": "Candidate Status",
    "created_at": "Added Date",
    "industry_domain": "Industry / Domain",
    "recruiter_name": "Sourcer / Recruiter Name",
}

def norm_key(k): return COL_MAP.get(str(k).strip().lower(), None)
def norm_mapping_value(value):
    value = str(value or "").strip()
    return value if value in ALL_FIELDS else ""
def mapped_header_key(header, column_mapping=None):
    header_text = str(header or "").strip()
    column_mapping = column_mapping or {}
    if header_text in column_mapping:
        return norm_mapping_value(column_mapping.get(header_text))
    lower_map = {str(k).strip().lower(): v for k, v in column_mapping.items()}
    if header_text.lower() in lower_map:
        return norm_mapping_value(lower_map.get(header_text.lower()))
    return norm_key(header_text)
def mapping_from_raw_headers(raw_headers, column_mapping=None):
    seen = set()
    mappings = []
    for header in raw_headers or []:
        header_text = str(header or "").strip()
        if not header_text or header_text.lower() in seen:
            continue
        seen.add(header_text.lower())
        field = mapped_header_key(header_text, column_mapping)
        mappings.append({
            "sheet_header": header_text,
            "ats_field": field or "",
            "ats_label": ATS_FIELD_LABELS.get(field or "", "Do not import"),
        })
    return mappings
def merge_column_mappings(existing, new_items):
    existing = existing or []
    new_items = new_items or []
    by_key = {str(item.get("sheet_header", "")).strip().lower(): item for item in existing if item.get("sheet_header")}
    for item in new_items:
        key = str(item.get("sheet_header", "")).strip().lower()
        if key and key not in by_key:
            by_key[key] = item
    return list(by_key.values())

def header_row_index(values):
    for i, row in enumerate(values or []):
        if len([c for c in row if c is not None and str(c).strip()]) >= 3:
            return i
    return 0

def ats_field_options():
    return [{"value": key, "label": label} for key, label in ATS_FIELD_LABELS.items()]
def empty_row(): return {f: "" for f in ALL_FIELDS}
def truthy_flag(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "on", "y"}

AUTO_AI_SCREENING_ON_CV_UPLOAD = False
def normalize_candidate_created_at(value):
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, (int, float)) and XLSX_OK:
        try:
            return openpyxl.utils.datetime.from_excel(value).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass
    text = re.sub(r"\s+", " ", str(value or "").strip())
    if not text:
        return ""
    text = re.sub(r"(\d+)(st|nd|rd|th)", r"\1", text, flags=re.I)
    formats = [
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d",
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
        "%d/%m/%y %H:%M:%S", "%d/%m/%y %H:%M", "%d/%m/%y",
        "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d-%m-%Y",
        "%d-%m-%y %H:%M:%S", "%d-%m-%y %H:%M", "%d-%m-%y",
        "%d.%m.%y", "%d.%m.%Y",
        "%d %b %Y", "%d-%b-%Y", "%d/%b/%Y", "%d %B %Y", "%d-%B-%Y", "%d/%B/%Y",
        "%d %b %y", "%d-%b-%y", "%d/%b/%y", "%d %B %y", "%d-%B-%y", "%d/%B/%y",
        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y",
        "%m/%d/%y %H:%M:%S", "%m/%d/%y %H:%M", "%m/%d/%y",
    ]
    for fmt in formats:
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass
    return text

def normalize_sheet_cell_value(key, cell):
    if key == "created_at":
        return normalize_candidate_created_at(cell)
    return str(cell).strip() if cell is not None else ""
def normalize_existing_candidate_created_at(conn):
    try:
        rows = conn.execute("SELECT id,created_at FROM candidates WHERE created_at IS NOT NULL AND created_at!=''").fetchall()
    except Exception:
        return 0
    updated = 0
    for row in rows:
        current = row["created_at"]
        normalized = normalize_candidate_created_at(current)
        if normalized and normalized != current:
            conn.execute("UPDATE candidates SET created_at=? WHERE id=?", (normalized, row["id"]))
            updated += 1
    return updated
def correct_future_swapped_candidate_dates(conn):
    today = date.today()
    try:
        rows = conn.execute("SELECT id,created_at FROM candidates WHERE created_at IS NOT NULL AND created_at!=''").fetchall()
    except Exception:
        return 0
    updated = 0
    for row in rows:
        current = row["created_at"]
        try:
            parsed = datetime.strptime(str(current)[:19], "%Y-%m-%d %H:%M:%S")
        except Exception:
            continue
        if parsed.date() <= today:
            continue
        try:
            swapped = datetime(parsed.year, parsed.day, parsed.month, parsed.hour, parsed.minute, parsed.second)
        except Exception:
            continue
        if swapped.date() <= today:
            conn.execute("UPDATE candidates SET created_at=? WHERE id=?", (swapped.strftime("%Y-%m-%d %H:%M:%S"), row["id"]))
            updated += 1
    return updated
def row_sheet_position_key(row):
    return f"{row.get('_sheet_name') or ''} #{row.get('_row_number') or ''}".strip()
def google_sheet_state_rows(sheet_id):
    if not sheet_id:
        return {}
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT sheet_name,last_row_number FROM google_sheet_import_state WHERE sheet_id=?",
            (sheet_id,)
        ).fetchall()
        return {str(row["sheet_name"] or "").strip().lower(): int(row["last_row_number"] or 0) for row in rows}
    finally:
        conn.close()

def update_google_sheet_import_state(conn, sheet_id, row_items, updated_by=""):
    if not sheet_id:
        return
    max_by_sheet = {}
    for row in row_items or []:
        sheet_name = str(row.get("_sheet_name") or "").strip()
        try:
            row_num = int(row.get("_row_number") or 0)
        except Exception:
            row_num = 0
        if sheet_name and row_num > max_by_sheet.get(sheet_name, 0):
            max_by_sheet[sheet_name] = row_num
    for sheet_name, row_num in max_by_sheet.items():
        conn.execute("""
            INSERT INTO google_sheet_import_state (sheet_id,sheet_name,last_row_number,updated_by,updated_at)
            VALUES (?,?,?,?,datetime('now','localtime'))
            ON CONFLICT(sheet_id,sheet_name) DO UPDATE SET
                last_row_number=MAX(last_row_number, excluded.last_row_number),
                updated_by=excluded.updated_by,
                updated_at=datetime('now','localtime')
        """, (sheet_id, sheet_name, row_num, updated_by))
def is_template_hint_row(row):
    name = re.sub(r"\s+", " ", str(row.get("candidate_name", "") or "").strip().lower())
    email = re.sub(r"\s+", " ", str(row.get("email_addr", "") or "").strip().lower())
    return name in {"full name", "candidate name"} or email in {"email address", "email id"}

# â”€â”€ Parsers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def parse_xlsx(file_bytes, role_override=""):
    if not XLSX_OK: return [], "openpyxl not installed"
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    rows = []
    for ws in wb.worksheets:
        data = list(ws.iter_rows(values_only=True))
        if not data: continue
        hdr_idx = 0
        for i, row in enumerate(data):
            if len([c for c in row if c is not None and str(c).strip()]) >= 3:
                hdr_idx = i; break
        headers = [norm_key(str(c).strip() if c else "") for c in data[hdr_idx]]
        if "candidate_name" not in headers:
            continue
        for dr in data[hdr_idx+1:]:
            if not any(c is not None and str(c).strip() for c in dr): continue
            row = empty_row()
            for i, cell in enumerate(dr):
                if i >= len(headers): break
                key = headers[i]
                val = normalize_sheet_cell_value(key, cell)
                if key and key in ALL_FIELDS: row[key] = val
            if is_template_hint_row(row):
                continue
            if role_override and not row["role_name"]: row["role_name"] = role_override
            rows.append(row)
    return rows, None

def extract_google_sheet_id(url):
    value = (url or "").strip()
    if re.fullmatch(r"[A-Za-z0-9_-]{20,}", value):
        return value
    match = re.search(r"/spreadsheets/d/([A-Za-z0-9_-]+)", value)
    return match.group(1) if match else ""

def download_google_sheet_xlsx(sheet_url):
    sheet_id = extract_google_sheet_id(sheet_url)
    if not sheet_id:
        return None, "Please enter a valid Google Sheet URL."
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    def _is_xlsx_response(resp):
        content_type = (resp.headers.get("content-type") or "").lower()
        return resp.status_code == 200 and "text/html" not in content_type and resp.content[:2] == b"PK"
    try:
        resp = requests.get(url, timeout=90)
    except Exception as e:
        return None, f"Unable to download Google Sheet: {e}"
    if _is_xlsx_response(resp):
        return resp.content, None

    token = session.get("google_token") or {}
    access_token = token.get("access_token")
    auth_error = ""
    if access_token:
        try:
            authed = requests.get(url, headers={"Authorization": f"Bearer {access_token}"}, timeout=90)
            if _is_xlsx_response(authed):
                return authed.content, None
        except Exception as e:
            auth_error = str(e)

        try:
            creds, error = google_oauth_credentials([
                "https://www.googleapis.com/auth/drive.readonly",
                "https://www.googleapis.com/auth/spreadsheets.readonly",
            ])
            if creds:
                service = google_api_build("drive", "v3", credentials=creds, cache_discovery=False)
                exported = service.files().export_media(
                    fileId=sheet_id,
                    mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ).execute()
                if exported:
                    return exported, None
            elif error:
                auth_error = error
        except Exception as e:
            auth_error = str(e)

    service_error = ""
    try:
        service = central_drive_service()
        exported = service.files().export_media(
            fileId=sheet_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ).execute()
        if exported:
            return exported, None
    except Exception as e:
        service_error = str(e)

    detail = ""
    if auth_error:
        detail += f" Google user access failed: {auth_error[:180]}."
    if service_error:
        detail += f" Service account access failed: {service_error[:180]}."
    return None, (
        "Google Sheet could not be downloaded. If you can open it in your browser, sign out and sign in with Google again "
        "so ATS gets Sheet/Drive read permission. Otherwise share the sheet with link access or with the ATS service account, then retry."
        + detail
    )

def quote_sheet_range_name(title):
    return "'" + str(title or "").replace("'", "''") + "'"

def rows_from_sheet_values(values, sheet_name, role_override="", column_mapping=None, first_data_row_num=None):
    if not values:
        return []
    hdr_idx = header_row_index(values)
    headers = [mapped_header_key(str(c).strip() if c else "", column_mapping) for c in values[hdr_idx]]
    if not any(headers):
        return []
    rows = []
    start_row_num = first_data_row_num if first_data_row_num else hdr_idx + 2
    for row_num, dr in enumerate(values[hdr_idx + 1:], start=start_row_num):
        if not any(c is not None and str(c).strip() for c in dr):
            continue
        row = empty_row()
        for i, cell in enumerate(dr):
            if i >= len(headers):
                break
            key = headers[i]
            val = normalize_sheet_cell_value(key, cell)
            if key and key in ALL_FIELDS:
                row[key] = val
        if is_template_hint_row(row):
            continue
        if role_override and not row["role_name"]:
            row["role_name"] = role_override
        row["_sheet_name"] = sheet_name
        row["_row_number"] = row_num
        rows.append(row)
    return rows

def read_google_sheet_rows_api(sheet_url, selected_tabs=None, role_override="", column_mapping=None, start_after_rows=None):
    sheet_id = extract_google_sheet_id(sheet_url)
    if not sheet_id:
        return None, [], [], "Please enter a valid Google Sheet URL."
    creds, error = google_oauth_credentials([
        "https://www.googleapis.com/auth/drive.readonly",
        "https://www.googleapis.com/auth/spreadsheets.readonly",
    ])
    if error or not creds:
        return None, [], [], error or "Please sign in with Google again before importing this sheet."
    try:
        service = google_api_build("sheets", "v4", credentials=creds, cache_discovery=False)
        meta = service.spreadsheets().get(
            spreadsheetId=sheet_id,
            fields="sheets.properties(title,gridProperties.rowCount)"
        ).execute()
        sheet_names = [s["properties"]["title"] for s in meta.get("sheets", [])]
        row_counts = {
            s["properties"]["title"]: int((s["properties"].get("gridProperties") or {}).get("rowCount") or 0)
            for s in meta.get("sheets", [])
        }
        selected = {str(tab).strip().lower() for tab in selected_tabs or [] if str(tab).strip()}
        target_names = [name for name in sheet_names if not selected or name.strip().lower() in selected]
        rows = []
        mappings = []
        range_requests = []
        start_after_rows = start_after_rows or {}
        for name in target_names:
            last_row = int(start_after_rows.get(name.strip().lower(), 0) or 0)
            if last_row > 0:
                start_row = last_row + 1
                end_row = max(start_row, row_counts.get(name) or start_row)
                range_requests.append((name, "header", 1, f"{quote_sheet_range_name(name)}!1:10"))
                range_requests.append((name, "data", start_row, f"{quote_sheet_range_name(name)}!{start_row}:{end_row}"))
            else:
                range_requests.append((name, "full", 1, quote_sheet_range_name(name)))
        values_by_sheet = {name: {"full": None, "header": [], "data": [], "start_row": 1} for name in target_names}
        ranges = [item[3] for item in range_requests]
        if ranges:
            result = service.spreadsheets().values().batchGet(
                spreadsheetId=sheet_id,
                ranges=ranges,
                valueRenderOption="FORMATTED_VALUE"
            ).execute()
            for request_item, value_item in zip(range_requests, result.get("valueRanges", [])):
                sheet_name, part, start_row, _ = request_item
                values = value_item.get("values", [])
                if part == "full":
                    values_by_sheet[sheet_name]["full"] = values
                elif part == "header":
                    values_by_sheet[sheet_name]["header"] = values
                else:
                    values_by_sheet[sheet_name]["data"] = values
                    values_by_sheet[sheet_name]["start_row"] = start_row
        for name in target_names:
            sheet_values = values_by_sheet.get(name, {})
            values = sheet_values.get("full")
            first_data_row_num = None
            if values is None:
                header_values = sheet_values.get("header") or []
                data_values = sheet_values.get("data") or []
                if header_values:
                    hdr_idx = header_row_index(header_values)
                    values = [header_values[hdr_idx]] + data_values
                    first_data_row_num = sheet_values.get("start_row") or 1
                else:
                    values = data_values
            if values:
                hdr_idx = header_row_index(values)
                mappings = merge_column_mappings(mappings, mapping_from_raw_headers(values[hdr_idx], column_mapping))
            rows.extend(rows_from_sheet_values(values, name, role_override, column_mapping, first_data_row_num))
        return rows, sheet_names, mappings, None
    except Exception as e:
        return None, [], [], f"Google Sheets API read failed: {e}"

def parse_xlsx_sheets(file_bytes, selected_tabs=None, role_override="", column_mapping=None):
    if not XLSX_OK:
        return [], [], [], "openpyxl not installed"
    selected = {str(tab).strip().lower() for tab in selected_tabs or [] if str(tab).strip()}
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    rows = []
    mappings = []
    sheet_names = wb.sheetnames
    for ws in wb.worksheets:
        if selected and ws.title.strip().lower() not in selected:
            continue
        data = list(ws.iter_rows(values_only=True))
        if not data:
            continue
        hdr_idx = header_row_index(data)
        mappings = merge_column_mappings(mappings, mapping_from_raw_headers(data[hdr_idx], column_mapping))
        headers = [mapped_header_key(str(c).strip() if c else "", column_mapping) for c in data[hdr_idx]]
        if not any(headers):
            continue
        for row_num, dr in enumerate(data[hdr_idx + 1:], start=hdr_idx + 2):
            if not any(c is not None and str(c).strip() for c in dr):
                continue
            row = empty_row()
            for i, cell in enumerate(dr):
                if i >= len(headers):
                    break
                key = headers[i]
                val = normalize_sheet_cell_value(key, cell)
                if key and key in ALL_FIELDS:
                    row[key] = val
            if is_template_hint_row(row):
                continue
            if role_override and not row["role_name"]:
                row["role_name"] = role_override
            row["_sheet_name"] = ws.title
            row["_row_number"] = row_num
            rows.append(row)
    return rows, sheet_names, mappings, None

def parse_csv(file_bytes, role_override=""):
    rows   = []
    text   = file_bytes.decode("utf-8-sig", errors="replace")
    # Check for tabs first
    if '\t' in text:
        reader = csv.DictReader(io.StringIO(text), dialect=csv.excel_tab)
    else:
        reader = csv.DictReader(io.StringIO(text))
    for dr in reader:
        row = empty_row()
        for k, v in dr.items():
            key = norm_key(k)
            if key and key in ALL_FIELDS: row[key] = normalize_sheet_cell_value(key, v)
        if is_template_hint_row(row):
            continue
        if role_override and not row["role_name"]: row["role_name"] = role_override
        rows.append(row)
    return rows, None

# â”€â”€ CV upload â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def upload_cv_cloud(file_storage, batch_id):
    fname = secure_filename(file_storage.filename)
    ext   = os.path.splitext(fname)[1].lower()
    if ext not in (".pdf", ".doc", ".docx"): return fname, None, None
    if not CLOUD_OK or not CLOUDINARY_URL:   return fname, None, None
    cloudinary.config(cloudinary_url=CLOUDINARY_URL)
    
    result = cloudinary.uploader.upload(
        file_storage, folder="hrguru_cvs",
        public_id=f"{batch_id}_{fname}",
        resource_type="raw", use_filename=True, unique_filename=False)
    return fname, result.get("secure_url"), result.get("public_id")

def hrguru_cv_filename(original_name, upload_folder=None):
    safe_name = secure_filename(original_name or "candidate_cv")
    base, ext = os.path.splitext(safe_name)
    base = base or "candidate_cv"
    candidate = f"{base}{ext}"
    if not upload_folder:
        return candidate
    counter = 2
    while os.path.exists(os.path.join(upload_folder, candidate)):
        candidate = f"{base}_{counter}{ext}"
        counter += 1
    return candidate

def upload_cv(file_storage, batch_id):
    fname = secure_filename(file_storage.filename)
    ext = os.path.splitext(fname)[1].lower()

    if ext not in (".pdf", ".doc", ".docx"):
        return fname, None, None, None

    upload_folder = os.path.join(app.root_path, "uploads", "recruiters", current_upload_owner_key(), "candidates", "bulk", _safe_path_part(batch_id))
    os.makedirs(upload_folder, exist_ok=True)

    saved_name = f"{batch_id}_{fname}"
    file_path = os.path.join(upload_folder, saved_name)

    file_storage.save(file_path)

    rel_path = os.path.relpath(file_path, os.path.join(app.root_path, "uploads")).replace(os.sep, "/")
    file_url = f"/uploads/{rel_path}"

    return fname, file_url, saved_name, file_path

def legacy_match_cv(name, cv_files):
    """
    Match a CV file to a candidate by name.

    Strategy:
    1. Strip the filename down to lowercase alpha-only characters (removes
       underscores, digits, dots, prefixes like 'Naukri', 'IR_' etc.)
       e.g. 'IR_RahulSPawar.pdf'   â†’ 'irrahulspawar'
            'Naukri_DineshM4y.pdf' â†’ 'naukridineshmy'
    2. Check if the candidate's FIRST NAME appears as a substring in that
       cleaned string â€” this is the primary signal.
    3. Bonus points if the last name also appears.
    4. Minimum requirement: first name must match (len >= 3).
    """
    if not cv_files: return None, None, None, False
    name_parts = (name or "").strip().split()
    if not name_parts: return None, None, None, False

    first_name = name_parts[0].lower()
    last_name  = name_parts[-1].lower() if len(name_parts) > 1 else ""

    # First name must be meaningful (skip very short names)
    if len(first_name) < 3: return None, None, None, False

    best_score, best = 0, (None, None, None)
    for orig, url, pub_id in cv_files:
        # Clean filename: keep only lowercase letters for substring search
        fname_alpha = re.sub(r'[^a-z]', '', orig.lower())
        fname_lower = orig.lower()   # also keep original for exact-word check

        score = 0

        # Primary: first name as substring in alpha-stripped filename
        if first_name in fname_alpha:
            score += 10
            # Bonus: last name also present
            if last_name and len(last_name) >= 3 and last_name in fname_alpha:
                score += 5

        # Fallback: check original filename with separators
        # e.g. 'rahul_pawar.pdf' â€” first name as a word token
        elif re.search(r'(?<![a-z])' + re.escape(first_name) + r'(?![a-z])', fname_lower):
            score += 8

        if score > best_score:
            best_score, best = score, (orig, url, pub_id)

    if best_score > 0: return *best, True
    return None, None, None, False

def unmatched_cvs(cv_files, matched_ids):
    return [(o, u, p) for o, u, p in cv_files if p not in matched_ids]

def cv_filename_match_score(name, original_filename):
    legacy = [(original_filename or "", "", "")]
    _, _, _, ok = legacy_match_cv(name, legacy)
    if not ok:
        return 0
    name_parts = (name or "").strip().split()
    last_name = name_parts[-1].lower() if len(name_parts) > 1 else ""
    fname_alpha = re.sub(r'[^a-z]', '', (original_filename or "").lower())
    return 15 if last_name and len(last_name) >= 3 and last_name in fname_alpha else 10

def match_cv(row, cv_files):
    if not cv_files:
        return None, None, None, False
    row_email = (row.get("email_addr", "") or "").strip().lower()
    row_phone = norm_phone(row.get("phone", ""))
    best_score, best = 0, (None, None, None)
    for cv in cv_files:
        orig, url, pub_id = cv.get("orig"), cv.get("url"), cv.get("pub_id")
        parsed = cv.get("parsed") or {}
        cv_email = (parsed.get("email_addr", "") or "").strip().lower()
        cv_phone = norm_phone(parsed.get("phone", ""))
        score = 0
        if row_email and cv_email and row_email == cv_email:
            score += 100
        if row_phone and cv_phone and row_phone == cv_phone:
            score += 90
        if score == 0:
            score = cv_filename_match_score(row.get("candidate_name", ""), orig)
        if score > best_score:
            best_score, best = score, (orig, url, pub_id)
    return (*best, True) if best_score > 0 else (None, None, None, False)

def unmatched_cvs(cv_files, matched_ids):
    return [(cv.get("orig"), cv.get("url"), cv.get("pub_id")) for cv in cv_files if cv.get("pub_id") not in matched_ids]

def parse_huggingface(text, api_key=""):
    """Use HuggingFace Inference API for NER-based CV parsing"""
    HF_API_URL = "https://api-inference.huggingface.co/models/dslim/bert-base-NER"
    
    headers = {"Authorization": f"Bearer {api_key}"} if api_key else {}
    
    result = {"candidate_name":"","current_company":"","current_role":"",
              "experience_years":"","email_addr":"","phone":"",
              "key_skills":"","current_location":""}
    
    try:
        response = requests.post(HF_API_URL, headers=headers, json={"inputs": text[:1500]}, timeout=30)
        if response.status_code == 200:
            entities = response.json()
            current_entity = None
            current_text = []
            
            for entity in entities:
                if isinstance(entity, dict):
                    word = entity.get("word", "")
                    entity_type = entity.get("entity_group", "")
                    
                    if entity_type == current_entity:
                        current_text.append(word)
                    else:
                        if current_entity == "PER" and current_text:
                            result["candidate_name"] = " ".join(current_text)
                        elif current_entity == "ORG" and current_text:
                            result["current_company"] = " ".join(current_text)
                        current_entity = entity_type
                        current_text = [word]
            
            if current_entity == "PER" and current_text:
                result["candidate_name"] = " ".join(current_text)
            elif current_entity == "ORG" and current_text:
                result["current_company"] = " ".join(current_text)
        
        if not api_key:
            result["_hf_limited"] = True
    except Exception as e:
        result["_hf_error"] = str(e)
        result["_hf_limited"] = True
    
    return result

def extract_cv_text(cv_path):
    ext = os.path.splitext(cv_path)[1].lower()
    if ext == ".txt":
        try:
            with open(cv_path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        except Exception:
            return ""
    if ext == ".docx":
        try:
            with zipfile.ZipFile(cv_path) as z:
                xml = z.read("word/document.xml")
            root = ET.fromstring(xml)
            ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
            lines = []
            for para in root.findall(".//w:p", ns):
                parts = [t.text for t in para.findall(".//w:t", ns) if t.text]
                line = re.sub(r"\s+", " ", "".join(parts)).strip()
                if line:
                    lines.append(line)
            return "\n".join(lines)
        except Exception:
            return ""
    if ext == ".doc":
        try:
            data = open(cv_path, "rb").read()
            chunks = []
            current = []
            for idx in range(0, len(data) - 1, 2):
                code = data[idx] + (data[idx + 1] << 8)
                if code in (9, 10, 13) or 32 <= code <= 126:
                    current.append(chr(code))
                else:
                    if len(current) >= 4:
                        chunks.append("".join(current))
                    current = []
            if len(current) >= 4:
                chunks.append("".join(current))
            lines = []
            seen = set()
            for chunk in chunks:
                for part in re.split(r"[\r\n]+", chunk):
                    line = re.sub(r"\s+", " ", part).strip()
                    if len(line) < 3:
                        continue
                    if not re.search(r"[A-Za-z]", line):
                        continue
                    key = line.lower()
                    if key in seen:
                        continue
                    seen.add(key)
                    lines.append(line)
            for idx, line in enumerate(lines):
                if line.strip().lower() in {"resume", "curriculum vitae", "cv"}:
                    lines = lines[idx:]
                    break
            return "\n".join(lines)
        except Exception:
            return ""
    try:
        import pymupdf
        doc = pymupdf.open(cv_path)
        text = ""
        for page in doc:
            blocks = page.get_text("blocks") or []
            if blocks:
                sorted_blocks = sorted(blocks, key=lambda b: (round(float(b[1]) / 4) * 4, float(b[0])))
                page_text = "\n".join(re.sub(r"\s*\n\s*", "\n", str(b[4] or "").strip()) for b in sorted_blocks if str(b[4] or "").strip())
                text += page_text + "\n"
            else:
                text += page.get_text() + "\n"
        doc.close()
        return text
    except Exception:
        return ""

def keyword_set(text):
    stop = {
        "and","or","the","a","an","to","for","of","in","on","with","by","as","is","are",
        "be","this","that","from","at","it","we","you","your","our","job","role","candidate",
        "experience","years","year","skills","skill","responsibilities","requirements"
    }
    words = re.findall(r"[a-zA-Z][a-zA-Z0-9+#.-]{2,}", (text or "").lower())
    return {w.strip(".-") for w in words if w not in stop and len(w.strip(".-")) > 2}

SKILL_ALIASES = {
    "ml": "Machine Learning",
    "machine learning": "Machine Learning",
    "ai": "Artificial Intelligence",
    "artificial intelligence": "Artificial Intelligence",
    "pyspark": "Apache Spark",
    "spark": "Apache Spark",
    "apache spark": "Apache Spark",
    "js": "JavaScript",
    "javascript": "JavaScript",
    "node": "Node.js",
    "node.js": "Node.js",
    "react": "React",
    "react.js": "React",
    "angular": "Angular",
    "vue": "Vue.js",
    "python": "Python",
    "java": "Java",
    "spring boot": "Spring Boot",
    "spring": "Spring",
    "sql": "SQL",
    "mysql": "MySQL",
    "postgresql": "PostgreSQL",
    "mongodb": "MongoDB",
    "aws": "AWS",
    "amazon web services": "AWS",
    "azure": "Microsoft Azure",
    "gcp": "Google Cloud Platform",
    "google cloud": "Google Cloud Platform",
    "docker": "Docker",
    "kubernetes": "Kubernetes",
    "k8s": "Kubernetes",
    "terraform": "Terraform",
    "jenkins": "Jenkins",
    "git": "Git",
    "jira": "Jira",
    "selenium": "Selenium",
    "cypress": "Cypress",
    "playwright": "Playwright",
    "rest api": "REST API",
    "restful": "REST API",
    "graphql": "GraphQL",
    "microservices": "Microservices",
    "linux": "Linux",
    "unix": "Unix",
    "tableau": "Tableau",
    "power bi": "Power BI",
    "excel": "Microsoft Excel",
    "tensorflow": "TensorFlow",
    "pytorch": "PyTorch",
    "scikit-learn": "Scikit-learn",
    "pandas": "Pandas",
    "numpy": "NumPy",
    "snowflake": "Snowflake",
    "databricks": "Databricks",
    "airflow": "Apache Airflow",
    "apache airflow": "Apache Airflow",
    "etl": "ETL",
    "ci/cd": "CI/CD",
    "devops": "DevOps",
    "html": "HTML",
    "css": "CSS",
}

ROLE_CATEGORIES = {
    "Data Engineering": ["data engineer", "etl", "big data", "spark", "snowflake", "databricks"],
    "Data Science": ["data scientist", "machine learning", "ml engineer", "ai engineer"],
    "Software Engineering": ["software engineer", "developer", "backend", "frontend", "full stack", "java", "python"],
    "Quality Assurance": ["qa", "quality", "test engineer", "automation testing", "selenium"],
    "DevOps": ["devops", "sre", "site reliability", "cloud engineer", "kubernetes"],
    "Product Management": ["product manager", "product owner"],
    "Business Analysis": ["business analyst", "system analyst"],
    "IT Support": ["it support", "desktop support", "linux admin", "system administrator"],
    "Cybersecurity": ["security", "cyber", "penetration", "soc", "iam"],
}

DOMAIN_HINTS = {
    "Banking / Financial Services": ["bank", "finance", "fintech", "payment", "trading", "loan", "insurance"],
    "Healthcare": ["healthcare", "clinical", "hospital", "patient", "medical"],
    "Retail / E-commerce": ["retail", "ecommerce", "e-commerce", "marketplace"],
    "Telecom": ["telecom", "network operator", "5g", "lte"],
    "SaaS": ["saas", "subscription", "multi-tenant"],
    "Manufacturing": ["manufacturing", "plant", "supply chain", "erp"],
}

def normalize_skill(skill):
    cleaned = re.sub(r"\s+", " ", str(skill or "").strip(" ,.;:|/\\")).lower()
    if not cleaned:
        return ""
    return SKILL_ALIASES.get(cleaned, cleaned.title())

def unique_list(items, limit=None):
    seen, out = set(), []
    for item in items or []:
        val = str(item or "").strip()
        if not val:
            continue
        key = val.lower()
        if key not in seen:
            seen.add(key)
            out.append(val)
    return out[:limit] if limit else out

def extract_known_skills(text):
    text_l = (text or "").lower()
    found = []
    for raw, normalized in sorted(SKILL_ALIASES.items(), key=lambda x: len(x[0]), reverse=True):
        pattern = r"(?<![a-z0-9])" + re.escape(raw.lower()) + r"(?![a-z0-9])"
        if re.search(pattern, text_l):
            found.append(normalized)
    return unique_list(found)

def infer_role_category(text):
    text_l = (text or "").lower()
    best = ("", 0)
    for category, hints in ROLE_CATEGORIES.items():
        score = sum(1 for hint in hints if hint in text_l)
        if score > best[1]:
            best = (category, score)
    return best[0]

def infer_domain(text):
    text_l = (text or "").lower()
    for domain, hints in DOMAIN_HINTS.items():
        if any(h in text_l for h in hints):
            return domain
    return ""

def extract_year_range(text):
    text_l = (text or "").lower()
    ranges = re.findall(r"(\d{1,2})\s*(?:-|to)\s*(\d{1,2})\s*\+?\s*(?:years?|yrs?)", text_l)
    if ranges:
        a, b = ranges[0]
        return {"min_years": int(a), "max_years": int(b)}
    singles = re.findall(r"(\d{1,2})\s*\+?\s*(?:years?|yrs?)", text_l)
    if singles:
        val = int(singles[0])
        return {"min_years": val, "max_years": 0}
    return {"min_years": 0, "max_years": 0}

def extract_role_title(text):
    lines = [re.sub(r"\s+", " ", l).strip() for l in (text or "").splitlines() if l.strip()]
    label_patterns = [
        r"(?:job\s*title|role|position|designation)\s*[:\-]\s*(.+)",
        r"(?:hiring\s+for|opening\s+for)\s*[:\-]?\s*(.+)",
    ]
    for line in lines[:30]:
        for pattern in label_patterns:
            m = re.search(pattern, line, re.I)
            if m:
                return clean_cv_value(m.group(1), 80).title()
    for line in lines[:12]:
        if len(line) <= 90 and infer_role_category(line):
            return clean_cv_value(line, 80).title()
    return ""

def extract_location(text):
    m = re.search(r"(?:location|job\s+location|work\s+location)\s*[:\-]\s*([A-Za-z0-9, /.-]{2,80})", text or "", re.I)
    if m:
        return clean_cv_value(m.group(1), 80)
    cities = ['Bangalore', 'Bengaluru', 'Mumbai', 'Delhi', 'Hyderabad', 'Chennai', 'Pune', 'Noida', 'Gurgaon', 'Gurugram', 'Kolkata', 'Ahmedabad', 'Remote']
    text_l = (text or "").lower()
    found = [c for c in cities if c.lower() in text_l]
    return ", ".join(unique_list(found))

def extract_certifications(text):
    certs = []
    patterns = [
        r"\bAWS Certified [A-Za-z ]+",
        r"\bAzure [A-Za-z ]+",
        r"\bPMP\b",
        r"\bCSM\b",
        r"\bCISSP\b",
        r"\bCEH\b",
        r"\bISTQB\b",
        r"\bCCNA\b",
        r"\bITIL\b",
    ]
    for pattern in patterns:
        certs.extend(re.findall(pattern, text or "", re.I))
    return unique_list([c.upper() if len(c) <= 5 else c.strip() for c in certs], 10)

def extract_responsibilities(text):
    lines = [re.sub(r"\s+", " ", l).strip(" -â€¢*\t") for l in (text or "").splitlines()]
    signals = ["responsible", "develop", "design", "manage", "build", "implement", "lead", "support", "analyze", "maintain", "create", "work with"]
    picked = []
    for line in lines:
        if 20 <= len(line) <= 180 and any(sig in line.lower() for sig in signals):
            picked.append(line)
    return unique_list(picked, 8)

def split_must_nice_skills(text):
    all_skills = extract_known_skills(text)
    must, nice = [], []
    must_markers = ["must", "required", "mandatory", "should have", "need", "hands-on", "strong experience"]
    nice_markers = ["preferred", "nice to have", "good to have", "plus", "advantage", "optional"]
    sentences = re.split(r"[\n.;]", text or "")
    for sentence in sentences:
        skills = extract_known_skills(sentence)
        sent_l = sentence.lower()
        if any(m in sent_l for m in nice_markers):
            nice.extend(skills)
        elif any(m in sent_l for m in must_markers):
            must.extend(skills)
    if not must:
        must = all_skills[:8]
    for skill in all_skills:
        if skill not in must:
            nice.append(skill)
    return unique_list(must, 12), unique_list(nice, 12)

def parse_jd_requirements(jd_text):
    role_title = extract_role_title(jd_text)
    category = infer_role_category((role_title or "") + "\n" + (jd_text or ""))
    must, nice = split_must_nice_skills(jd_text)
    text_l = (jd_text or "").lower()
    employment = ""
    if "contract" in text_l:
        employment = "Contract"
    elif "part time" in text_l or "part-time" in text_l:
        employment = "Part-time"
    elif "full time" in text_l or "full-time" in text_l or "permanent" in text_l:
        employment = "Full-time"
    seniority = ""
    if re.search(r"\b(lead|principal|staff|architect|manager)\b", text_l):
        seniority = "Lead/Senior"
    elif re.search(r"\b(senior|sr\.?)\b", text_l):
        seniority = "Senior"
    elif re.search(r"\b(junior|fresher|entry)\b", text_l):
        seniority = "Junior"
    return {
        "role_title": role_title,
        "role_category": category,
        "employment_type": employment,
        "experience_required": extract_year_range(jd_text),
        "must_have_skills": must,
        "nice_to_have_skills": nice,
        "tools_technologies": unique_list(extract_known_skills(jd_text), 18),
        "domain": infer_domain(jd_text),
        "sub_domain": "",
        "responsibilities": extract_responsibilities(jd_text),
        "seniority_level": seniority,
        "certifications_required": extract_certifications(jd_text),
        "location": extract_location(jd_text),
        "keywords_expanded": unique_list(must + nice + ([category] if category else []), 25)
    }

def parse_experience_years(value):
    if isinstance(value, (int, float)):
        return float(value)
    m = re.search(r"(\d+(?:\.\d+)?)", str(value or ""))
    return float(m.group(1)) if m else 0

def extract_education_items(text):
    items = re.findall(r"\b(?:B\.?Tech|B\.?E\.?|B\.?Sc|M\.?Tech|M\.?Sc|MBA|MCA|BCA|Ph\.?D|Bachelor(?:'s)?|Master(?:'s)?)\b", text or "", re.I)
    return unique_list([i.upper().replace("BACHELOR'S", "Bachelor").replace("MASTER'S", "Master") for i in items], 8)

def parse_candidate_profile(cv_text, parsed_cv=None):
    parsed = parsed_cv or {}
    skills = extract_known_skills(cv_text)
    parsed_skills = [normalize_skill(s) for s in re.split(r"[,;/|]", parsed.get("key_skills", "")) if s.strip()]
    primary = unique_list(parsed_skills + skills, 15)
    current_role = parsed.get("current_role", "")
    role_history = []
    if current_role or parsed.get("current_company"):
        role_history.append({
            "title": current_role,
            "company": parsed.get("current_company", ""),
            "duration_years": parse_experience_years(parsed.get("experience_years", "")),
            "responsibilities": extract_responsibilities(cv_text)[:5],
            "skills_used": primary[:10]
        })
    total_exp = parse_experience_years(parsed.get("experience_years", ""))
    stability = 60 if role_history else 0
    if total_exp >= 3 and role_history:
        stability = 75
    red_flags = []
    if not current_role:
        red_flags.append("Current role not clearly stated")
    if not total_exp:
        red_flags.append("Total experience not clearly stated")
    return {
        "candidate_name": parsed.get("candidate_name", ""),
        "total_experience_years": total_exp,
        "current_role": current_role,
        "role_history": role_history,
        "primary_skills": primary[:10],
        "secondary_skills": primary[10:20],
        "tools_technologies": primary[:18],
        "domain_experience": unique_list([infer_domain(cv_text)] if infer_domain(cv_text) else []),
        "certifications": extract_certifications(cv_text),
        "education": unique_list([parsed.get("education", "")] + extract_education_items(cv_text), 8),
        "location": parsed.get("current_location", "") or extract_location(cv_text),
        "career_stability_score": stability,
        "red_flags": red_flags,
        "project_complexity_indicators": extract_project_complexity(cv_text),
        "ownership_signals": extract_ownership_signals(cv_text)
    }

def extract_project_complexity(text):
    indicators = []
    hints = ["microservices", "distributed", "large scale", "high availability", "migration", "architecture", "pipeline", "automation", "performance", "cloud"]
    for hint in hints:
        if hint in (text or "").lower():
            indicators.append(hint.title())
    return unique_list(indicators, 8)

def extract_ownership_signals(text):
    signals = []
    for phrase in ["led", "owned", "managed", "designed", "architected", "delivered", "mentored", "implemented"]:
        if re.search(r"\b" + re.escape(phrase) + r"\b", text or "", re.I):
            signals.append(phrase.title())
    return unique_list(signals, 8)

def skill_match_score(required, available):
    available_norm = {normalize_skill(s).lower() for s in available or []}
    matched, missing = [], []
    for skill in required or []:
        normalized = normalize_skill(skill)
        key = normalized.lower()
        if key in available_norm:
            matched.append(normalized)
        else:
            missing.append(normalized)
    return unique_list(matched), unique_list(missing)

def score_candidate_fit(jd, cv):
    required = jd.get("must_have_skills", [])
    available = (cv.get("primary_skills", []) or []) + (cv.get("secondary_skills", []) or []) + (cv.get("tools_technologies", []) or [])
    matched, missing = skill_match_score(required, available)
    must_score = round((len(matched) / len(required)) * 100) if required else 0
    jd_category = (jd.get("role_category") or "").lower()
    cv_role_text = (cv.get("current_role") or "") + " " + " ".join(cv.get("primary_skills", []) or [])
    cv_category = infer_role_category(cv_role_text).lower()
    role_score = 85 if jd_category and jd_category == cv_category else 35 if jd_category and cv_category else 50
    req_exp = jd.get("experience_required") or {}
    min_exp, max_exp = req_exp.get("min_years", 0) or 0, req_exp.get("max_years", 0) or 0
    cand_exp = cv.get("total_experience_years", 0) or 0
    if min_exp and cand_exp < min_exp:
        exp_score = max(0, round((cand_exp / min_exp) * 70))
    elif max_exp and cand_exp > max_exp + 3:
        exp_score = 70
    elif min_exp or max_exp:
        exp_score = 90
    else:
        exp_score = 50
    jd_domain = (jd.get("domain") or "").lower()
    cv_domains = [str(d).lower() for d in cv.get("domain_experience", []) or []]
    domain_score = 85 if jd_domain and jd_domain in cv_domains else 50 if not jd_domain else 25
    nice = jd.get("nice_to_have_skills", []) or []
    nice_matched, _ = skill_match_score(nice, available)
    secondary_score = round((len(nice_matched) / len(nice)) * 100) if nice else 50
    stability_score = int(cv.get("career_stability_score", 0) or 0)
    context_score = 70 if cv.get("project_complexity_indicators") or cv.get("ownership_signals") else 45
    penalties = []
    if missing:
        penalties.append({"reason": "Missing must-have skills: " + ", ".join(missing[:6]), "impact": min(45, len(missing) * 12)})
    if jd_category and cv_category and jd_category != cv_category:
        penalties.append({"reason": "Role category mismatch", "impact": 30})
    if min_exp and cand_exp < min_exp:
        penalties.append({"reason": f"Experience below required minimum of {min_exp} years", "impact": 20})
    if cv.get("red_flags"):
        penalties.append({"reason": "CV has missing or unclear core details", "impact": 10})
    weighted = (
        must_score * 0.34 + role_score * 0.20 + exp_score * 0.15 + domain_score * 0.10 +
        secondary_score * 0.08 + stability_score * 0.06 + context_score * 0.07
    )
    final_score = max(0, min(100, round(weighted - sum(p["impact"] for p in penalties) * 0.35)))
    if final_score >= 80:
        verdict = "Strong Match"
    elif final_score >= 65:
        verdict = "Moderate Match"
    elif final_score >= 45:
        verdict = "Weak Match"
    else:
        verdict = "Reject / Not Recommended"
    strengths = []
    if matched:
        strengths.append("Matches must-have skills: " + ", ".join(matched[:6]))
    if cv.get("ownership_signals"):
        strengths.append("Shows ownership signals: " + ", ".join(cv.get("ownership_signals", [])[:4]))
    concerns = []
    if missing:
        concerns.append("Missing must-have skills: " + ", ".join(missing[:6]))
    if role_score < 50:
        concerns.append("Current role does not align well with the JD role category")
    if exp_score < 70:
        concerns.append("Experience is below the JD requirement")
    return {
        "final_score": final_score,
        "verdict": verdict,
        "score_breakdown": {
            "must_have_skills": must_score,
            "role_relevance": role_score,
            "experience_fit": exp_score,
            "domain_fit": domain_score,
            "secondary_skills": secondary_score,
            "stability": stability_score,
            "contextual_intelligence": context_score
        },
        "matched_must_have_skills": matched,
        "missing_must_have_skills": missing,
        "strengths": strengths[:8],
        "concerns": concerns[:8],
        "red_flags": cv.get("red_flags", []),
        "penalties_applied": penalties,
        "explanation_summary": f"Candidate scored {final_score}/100. {verdict}. " + ("; ".join(concerns[:2]) if concerns else "Core hiring signals are reasonably aligned.")
    }

def build_jd_match_response(jd_json, cv_json, score_json, mode):
    score = int(score_json.get("final_score", 0) or 0)
    strengths = score_json.get("strengths", []) or []
    gaps = (score_json.get("concerns", []) or []) + [p.get("reason", "") for p in score_json.get("penalties_applied", []) if p.get("reason")]
    return {
        "score": score,
        "verdict": score_json.get("verdict", ""),
        "summary": score_json.get("explanation_summary", ""),
        "strengths": strengths[:8],
        "gaps": unique_list(gaps, 8),
        "mode": mode,
        "jd_json": jd_json,
        "cv_json": cv_json,
        "score_json": score_json
    }

def heuristic_jd_match(jd_text, cv_text, parsed_cv=None):
    jd_json = parse_jd_requirements(jd_text)
    cv_json = parse_candidate_profile(cv_text, parsed_cv)
    score_json = score_candidate_fit(jd_json, cv_json)
    return build_jd_match_response(jd_json, cv_json, score_json, "structured_keyword")

def build_candidate_summary_from_cv(cv_path, cv_text=""):
    parsed = {}
    try:
        parsed = parse_cv(cv_path) or {}
    except Exception as e:
        print("CV summary parse failed:", e)
    skills = parsed.get("key_skills", "")
    if isinstance(skills, str):
        skills_list = [s.strip() for s in skills.split(",") if s.strip()]
    else:
        skills_list = skills or []
    relevant = parsed.get("cv_summary", "")
    if not relevant and cv_text:
        lines = [re.sub(r"\s+", " ", line).strip() for line in cv_text.splitlines() if line.strip()]
        relevant = " ".join(lines[:4])[:700]
    return {
        "current_designation": parsed.get("current_role", ""),
        "key_skills": skills_list[:18],
        "total_experience": parsed.get("experience_years", ""),
        "education": parsed.get("education", ""),
        "current_company": parsed.get("current_company", ""),
        "relevant_details": relevant
    }

def extract_resume_sections(cv_text):
    section_aliases = {
        "summary": ["summary", "profile", "objective", "professional summary", "career objective"],
        "contact_details": ["contact", "personal details"],
        "technical_skills": ["skills", "technical skills", "key skills", "technologies", "tools"],
        "experience": ["experience", "work experience", "professional experience", "employment history", "career history"],
        "projects": ["projects", "project experience"],
        "education": ["education", "academic", "qualification", "qualifications"],
        "certifications": ["certifications", "certificates"],
    }
    sections = {}
    current = "header"
    sections[current] = []
    heading_lookup = {}
    for key, aliases in section_aliases.items():
        for alias in aliases:
            heading_lookup[alias] = key
    for raw_line in (cv_text or "").splitlines():
        line = clean_value(raw_line, 220)
        if not line:
            continue
        normalized = re.sub(r"[^a-z ]+", "", line.lower()).strip()
        heading_key = heading_lookup.get(normalized)
        if not heading_key:
            for alias, key in heading_lookup.items():
                if normalized == alias or normalized.startswith(alias + " "):
                    heading_key = key
                    break
        if heading_key and len(line.split()) <= 5:
            current = heading_key
            sections.setdefault(current, [])
            continue
        sections.setdefault(current, []).append(line)
    return {key: "\n".join(value[:80]) for key, value in sections.items() if value}

DEFAULT_RESUME_ANALYSIS_PROMPT = (
    "Analyze this resume for a recruiter. Return compact JSON only, no markdown. "
    "Extract: candidate_name, total_experience, current_role, current_company, contact_details "
    "{email, phone, linkedin}, organization_experience array with company, role, duration, summary, "
    "technical_skills array with skill, experience in years/months when evidence exists, and evidence, "
    "education_details array, and current_location. "
    "If skill-wise experience is not explicitly clear, say 'Not clear from resume'. "
    "Use only resume evidence and deterministic parse context."
)

def call_resume_analysis_llm(cv_text, deterministic_candidate, sections, custom_prompt=""):
    analysis_instruction = re.sub(r"\s+", " ", str(custom_prompt or DEFAULT_RESUME_ANALYSIS_PROMPT)).strip()[:4000]
    prompt = (
        f"{analysis_instruction}\n\n"
        "Return the same JSON schema even when the recruiter instruction changes.\n\n"
        f"Deterministic parse JSON: {json.dumps(deterministic_candidate, ensure_ascii=False)[:5000]}\n\n"
        f"Labelled sections JSON: {json.dumps(sections, ensure_ascii=False)[:6000]}\n\n"
        f"Resume text:\n{cv_text[:12000]}"
    )
    provider = os.getenv("LLM_PROVIDER", "openrouter").lower()
    errors = []
    attempts = []
    if provider == "openrouter":
        attempts.append(("openrouter", os.getenv("OPENROUTER_API_KEY", ""), os.getenv("OPENROUTER_API_BASE", "https://openrouter.ai/api/v1").rstrip("/"), os.getenv("OPENROUTER_MODEL", os.getenv("LLM_MODEL", "inclusionai/ring-2.6-1t:free"))))
        if os.getenv("OPENAI_API_KEY", ""):
            attempts.append(("openai", os.getenv("OPENAI_API_KEY", ""), os.getenv("OPENAI_API_BASE", "https://api.openai.com/v1").rstrip("/"), os.getenv("OPENAI_FALLBACK_MODEL", "gpt-5-mini")))
    elif provider == "openai":
        attempts.append(("openai", os.getenv("OPENAI_API_KEY", ""), os.getenv("OPENAI_API_BASE", "https://api.openai.com/v1").rstrip("/"), os.getenv("LLM_MODEL", "gpt-5-mini")))
    for provider_name, api_key, api_base, model in attempts:
        if not api_key:
            errors.append(f"{provider_name} API key is not configured")
            continue
        try:
            headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
            if provider_name == "openrouter":
                headers["HTTP-Referer"] = os.getenv("PUBLIC_BASE_URL", "https://hireflow.hrgp.in")
                headers["X-Title"] = "HRGuru ATS"
            body = {
                "model": model,
                "messages": [
                    {"role": "system", "content": "Return valid JSON only. No markdown."},
                    {"role": "user", "content": prompt},
                ],
                "response_format": {"type": "json_object"},
            }
            if provider_name == "openai":
                body["max_completion_tokens"] = 2200
            else:
                body["max_tokens"] = 2200
            response = requests.post(f"{api_base}/chat/completions", headers=headers, json=body, timeout=35)
            response.raise_for_status()
            raw = ((response.json().get("choices") or [{}])[0].get("message") or {}).get("content") or "{}"
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                parsed["_source_model"] = model
                return parsed, ""
        except Exception as e:
            errors.append(f"{provider_name}: {e}")
    return {}, "; ".join(errors)

def maybe_generate_gemini_screening_report(
    jd_text,
    cv_text,
    candidate_name="",
    target_job_title="",
    parsed_jd=None,
    parsed_candidate=None,
    batch_delay_seconds=0,
    api_key=None,
):
    """
    Optional Gemini-powered junior-recruiter screening report.
    Keeps the deterministic matcher as the source of truth for scores/ranking.
    """
    if not (os.getenv("GEMINI_API_KEY") or "").strip():
        return {"ok": False, "source": "gemini", "error": "GEMINI_API_KEY is not configured."}
    try:
        from services.gemini_screening import build_screening_report
    except Exception as e:
        return {"ok": False, "source": "gemini", "error": f"Gemini screening helper is unavailable: {e}"}
    try:
        report = build_screening_report(
            jd_text,
            cv_text,
            candidate_name=candidate_name,
            target_job_title=target_job_title,
            parsed_jd=parsed_jd or {},
            parsed_candidate=parsed_candidate or {},
            api_key=api_key,
        )
        if batch_delay_seconds:
            time.sleep(float(batch_delay_seconds))
        return report
    except Exception as e:
        return {"ok": False, "source": "gemini", "error": f"Gemini screening report failed: {e}"}

def apply_gemini_match_report(result, screening_report):
    report = screening_report.get("report") or {}
    if not report:
        return result
    usage_metadata = screening_report.get("usage_metadata") or {}
    deterministic_snapshot = {
        "final_score": result.get("final_score", result.get("score", 0)),
        "verdict": result.get("verdict", ""),
        "recommendation": result.get("recommendation", ""),
        "score_json": result.get("score_json") or {},
        "score_breakdown": result.get("score_breakdown") or {},
        "strengths": result.get("strengths") or [],
        "concerns": result.get("concerns") or [],
        "matched_must_have_skills": result.get("matched_must_have_skills") or [],
        "missing_must_have_skills": result.get("missing_must_have_skills") or [],
        "summary": result.get("summary") or "",
        "overall_recruiter_summary": result.get("overall_recruiter_summary") or "",
    }
    result["deterministic_match_snapshot"] = deterministic_snapshot
    result["deterministic_final_score"] = deterministic_snapshot["final_score"]
    result["deterministic_verdict"] = deterministic_snapshot["verdict"]
    result["deterministic_recommendation"] = deterministic_snapshot["recommendation"]
    result["deterministic_score_json"] = deterministic_snapshot["score_json"]
    result["deterministic_score_breakdown"] = deterministic_snapshot["score_breakdown"]
    result["gemini_match_report"] = report
    result["gemini_screening_report"] = report
    result["gemini_screening_source"] = screening_report.get("source", "gemini")
    result["gemini_screening_model"] = screening_report.get("model", "gemini-2.5-flash")
    result["gemini_screening_usage"] = usage_metadata
    result["scoring_source"] = screening_report.get("source", "gemini")
    result["score_source"] = screening_report.get("source", "gemini")
    result["final_score"] = int(report.get("final_score", result.get("final_score", result.get("score", 0))) or 0)
    result["score"] = result["final_score"]
    result["verdict"] = report.get("ats_verdict") or result.get("verdict")
    result["call_or_reject"] = report.get("call_or_reject") or result.get("call_or_reject", "")
    result["recommendation"] = report.get("recommendation") or result.get("recommendation")
    result["matched_must_have_skills"] = report.get("matched_must_have_skills") or result.get("matched_must_have_skills") or []
    result["missing_must_have_skills"] = report.get("missing_must_have_skills") or result.get("missing_must_have_skills") or []
    result["strengths"] = report.get("strengths") or report.get("green_flags") or result.get("strengths") or []
    result["concerns"] = report.get("concerns") or report.get("red_flags") or result.get("concerns") or []
    result["score_breakdown"] = report.get("score_breakdown") or result.get("score_breakdown") or []
    score_json = dict(result.get("score_json") or {})
    score_json.update({
        "final_score": result["final_score"],
        "verdict": result["verdict"],
        "recommendation": result["recommendation"],
        "scoring_source": result.get("scoring_source", "gemini"),
        "score_breakdown": report.get("score_breakdown") or score_json.get("score_breakdown") or [],
        "matched_must_have_skills": result["matched_must_have_skills"],
        "missing_must_have_skills": result["missing_must_have_skills"],
        "strengths": result["strengths"],
        "concerns": result["concerns"],
        "gemini_usage_metadata": usage_metadata,
        "explanation_summary": report.get("summary") or score_json.get("explanation_summary") or "",
    })
    result["score_json"] = score_json
    result["summary"] = report.get("summary") or result.get("summary") or ""
    result["overall_recruiter_summary"] = report.get("summary") or result.get("overall_recruiter_summary") or ""
    return result

def check_skills_in_resume(cv_text, deterministic_candidate, skills_text):
    requested = [canonical_skill(item.strip()) for item in re.split(r"[,;\n]", skills_text or "") if item.strip()]
    requested = list(dict.fromkeys([skill for skill in requested if skill]))
    if not requested:
        return {"checked_skills": [], "found": [], "missing": []}
    text_l = (cv_text or "").lower()
    parsed_skills = {canonical_skill(skill).lower() for skill in (deterministic_candidate.get("normalized_skills") or []) if canonical_skill(skill)}
    found, missing = [], []
    for skill in requested:
        skill_l = skill.lower()
        tokens = [token for token in re.split(r"[^a-zA-Z0-9+#.]+", skill_l) if token]
        text_match = skill_l in text_l or any(len(token) > 2 and re.search(r"\b" + re.escape(token) + r"\b", text_l) for token in tokens)
        if skill_l in parsed_skills or text_match:
            found.append(skill)
        else:
            missing.append(skill)
    return {"checked_skills": requested, "found": found, "missing": missing}

def ai_jd_match(jd_text, cv_text, candidate_name=""):
    # Deprecated: scoring is now owned by deterministic Python modules.
    # LLM extraction is handled by ats_pipeline.llm_extract and never returns scores.
    return None

def _safe_path_part(value):
    value = secure_filename(str(value or "").strip())
    return value or "default"

def current_upload_owner_key():
    return _safe_path_part(
        session.get("recruiter_email")
        or session.get("email")
        or session.get("username")
        or "system"
    )

def save_uploaded_file_to_bucket(file_storage, prefix, *bucket_parts):
    if not file_storage or not file_storage.filename:
        raise ValueError("Missing file")
    ext = os.path.splitext(file_storage.filename)[1].lower()
    if ext not in [".pdf", ".doc", ".docx", ".txt"]:
        raise ValueError("Upload PDF, DOC, DOCX, or TXT files only")
    upload_folder = os.path.join(app.root_path, "uploads", "recruiters", current_upload_owner_key())
    rel_parts = [_safe_path_part(part) for part in bucket_parts if str(part or "").strip()]
    if rel_parts:
        upload_folder = os.path.join(upload_folder, *rel_parts)
    os.makedirs(upload_folder, exist_ok=True)
    safe_name = secure_filename(file_storage.filename)
    saved_name = f"{prefix}_{int(datetime.now().timestamp())}_{safe_name}"
    path = os.path.join(upload_folder, saved_name)
    file_storage.save(path)
    return path, safe_name

def save_uploaded_analysis_file(file_storage, prefix, *bucket_parts):
    return save_uploaded_file_to_bucket(file_storage, prefix, "analysis", *bucket_parts)

def get_cached_embedding(text_hash_value, model):
    conn = get_db()
    ensure_ats_pipeline_schema(conn)
    row = conn.execute(
        "SELECT embedding_json FROM embedding_cache WHERE text_hash=? AND model=?",
        (text_hash_value, model)
    ).fetchone()
    conn.close()
    return deserialize_embedding(row["embedding_json"]) if row else None

def set_cached_embedding(text_hash_value, model, embedding):
    with DB_WRITE_LOCK:
        conn = get_db()
        ensure_ats_pipeline_schema(conn)
        conn.execute(
            "INSERT OR REPLACE INTO embedding_cache (text_hash, model, embedding_json) VALUES (?,?,?)",
            (text_hash_value, model, serialize_embedding(embedding))
        )
        conn.commit()
        conn.close()

def get_cached_match_result(jd_hash, resume_hash, pipeline_version=MATCH_PIPELINE_VERSION):
    conn = get_db()
    ensure_ats_pipeline_schema(conn)
    row = conn.execute(
        "SELECT result_json, pipeline_version FROM match_results WHERE jd_hash=? AND resume_hash=? AND COALESCE(pipeline_version,'')=?",
        (jd_hash, resume_hash, pipeline_version)
    ).fetchone()
    conn.close()
    if not row:
        return None
    try:
        return json.loads(row["result_json"])
    except Exception:
        return None

def get_cached_parsed_resume(resume_hash, pipeline_version=MATCH_PIPELINE_VERSION):
    conn = get_db()
    ensure_ats_pipeline_schema(conn)
    row = conn.execute(
        "SELECT parsed_json, pipeline_version FROM parsed_resume_cache WHERE resume_hash=? AND COALESCE(pipeline_version,'')=?",
        (resume_hash, pipeline_version)
    ).fetchone()
    conn.close()
    if not row:
        return None
    try:
        parsed = json.loads(row["parsed_json"] or "{}")
        return parsed if isinstance(parsed, dict) else None
    except Exception:
        return None

def set_cached_parsed_resume(resume_hash, parsed_candidate):
    if not parsed_candidate:
        return
    with DB_WRITE_LOCK:
        conn = get_db()
        ensure_ats_pipeline_schema(conn)
        conn.execute(
            """INSERT OR REPLACE INTO parsed_resume_cache
               (resume_hash, parsed_json, pipeline_version, updated_at)
               VALUES (?,?,?,datetime('now','localtime'))""",
            (resume_hash, json.dumps(parsed_candidate), MATCH_PIPELINE_VERSION)
        )
        conn.commit()
        conn.close()

def get_cached_parsed_jd(jd_hash, pipeline_version=MATCH_PIPELINE_VERSION):
    conn = get_db()
    ensure_ats_pipeline_schema(conn)
    row = conn.execute(
        "SELECT parsed_json, pipeline_version FROM jd_requirements WHERE jd_hash=? AND COALESCE(pipeline_version,'')=?",
        (jd_hash, pipeline_version)
    ).fetchone()
    conn.close()
    if not row:
        return None
    try:
        parsed = json.loads(row["parsed_json"] or "{}")
        return parsed if isinstance(parsed, dict) else None
    except Exception:
        return None

def set_cached_parsed_jd(jd_hash, parsed_jd):
    if not parsed_jd:
        return
    with DB_WRITE_LOCK:
        conn = get_db()
        ensure_ats_pipeline_schema(conn)
        conn.execute(
            """INSERT OR REPLACE INTO jd_requirements
               (jd_hash, role_title, parsed_json, pipeline_version, updated_at)
               VALUES (?,?,?,?,datetime('now','localtime'))""",
            (jd_hash, parsed_jd.get("role_title", ""), json.dumps(parsed_jd), MATCH_PIPELINE_VERSION)
        )
        conn.commit()
        conn.close()

def reset_matching_caches(pipeline_version=MATCH_PIPELINE_VERSION):
    with DB_WRITE_LOCK:
        conn = get_db()
        ensure_ats_pipeline_schema(conn)
        deleted = {
            "match_results": conn.execute(
                "DELETE FROM match_results WHERE COALESCE(pipeline_version,'')=?",
                (pipeline_version,)
            ).rowcount,
            "parsed_resume_cache": conn.execute(
                "DELETE FROM parsed_resume_cache WHERE COALESCE(pipeline_version,'')=?",
                (pipeline_version,)
            ).rowcount,
            "jd_requirements": conn.execute(
                "DELETE FROM jd_requirements WHERE COALESCE(pipeline_version,'')=?",
                (pipeline_version,)
            ).rowcount,
        }
        conn.commit()
        conn.close()
        return deleted

def persist_skill_aliases(conn, canonical_value):
    canonical_value = canonical_skill(canonical_value)
    if not canonical_value:
        return
    cursor = conn.execute(
        "INSERT OR IGNORE INTO skills (raw_value, canonical_value) VALUES (?,?)",
        (canonical_value, canonical_value)
    )
    row = conn.execute("SELECT id FROM skills WHERE canonical_value=?", (canonical_value,)).fetchone()
    if not row:
        return
    skill_id = row["id"]
    aliases = skill_aliases_for(canonical_value)
    conn.execute(
        "INSERT OR REPLACE INTO normalized_skill_cache (raw_value, canonical_value, aliases_json) VALUES (?,?,?)",
        (canonical_value.lower(), canonical_value, json.dumps(aliases))
    )
    for alias in aliases:
        conn.execute(
            "INSERT OR IGNORE INTO skill_aliases (skill_id, alias) VALUES (?,?)",
            (skill_id, alias)
        )

def persist_match_artifacts(jd_hash, resume_hash, result):
    with DB_WRITE_LOCK:
        conn = get_db()
        ensure_ats_pipeline_schema(conn)
        parsed_jd = result.get("parsed_jd") or result.get("jd_json") or {}
        parsed_candidate = result.get("parsed_candidate") or result.get("cv_json") or {}
        pipeline_version = result.get("pipeline_version") or MATCH_PIPELINE_VERSION
        conn.execute(
            """INSERT OR REPLACE INTO jd_requirements
               (jd_hash, role_title, parsed_json, pipeline_version, updated_at)
               VALUES (?,?,?,?,datetime('now','localtime'))""",
            (jd_hash, parsed_jd.get("role_title", ""), json.dumps(parsed_jd), pipeline_version)
        )
        conn.execute(
            """INSERT OR REPLACE INTO parsed_resume_cache
               (resume_hash, parsed_json, pipeline_version, updated_at)
               VALUES (?,?,?,datetime('now','localtime'))""",
            (resume_hash, json.dumps(parsed_candidate), pipeline_version)
        )
        conn.execute("DELETE FROM match_results WHERE jd_hash=? AND resume_hash=?", (jd_hash, resume_hash))
        conn.execute(
            """INSERT INTO match_results
               (jd_hash, resume_hash, pipeline_version, final_score, structured_score, semantic_score, hard_filter_score, result_json)
           VALUES (?,?,?,?,?,?,?,?)""",
        (
            jd_hash,
            resume_hash,
            pipeline_version,
            int(result.get("final_score", 0) or 0),
            int(result.get("structured_score", 0) or 0),
            int(result.get("semantic_score", 0) or 0),
            int(result.get("hard_filter_score", 0) or 0),
            json.dumps(result)
        )
    )
    try:
        record_match_audit(
            conn,
            event_type="match_run",
            object_type="match",
            object_hash=f"{jd_hash}:{resume_hash}",
            jd_hash=jd_hash,
            resume_hash=resume_hash,
            pipeline_version=pipeline_version,
            status=str(result.get("verdict") or ""),
            source="persist",
            parser_confidence=str((parsed_jd.get("parser_confidence") or "")),
            manual_review_required=bool((result.get("dashboard") or {}).get("manual_review", {}).get("required")),
            score=result.get("final_score", 0),
            message=str((result.get("dashboard") or {}).get("manual_review", {}).get("summary") or result.get("recruiter_summary") or ""),
            details={
                "cache_hit": bool(result.get("cache_hit")),
                "structured_score": result.get("structured_score"),
                "semantic_score": result.get("semantic_score"),
                "hard_filter_score": result.get("hard_filter_score"),
                "role_family": (result.get("dashboard") or {}).get("role_family_comparison", {}),
                "validation_gaps": (result.get("dashboard") or {}).get("validation_gaps", [])[:8],
            }
        )
    except Exception:
        pass
    conn.execute("DELETE FROM candidate_skills WHERE resume_hash=?", (resume_hash,))
    for item in parsed_candidate.get("skill_confidence_scores", []) or []:
        canonical = canonical_skill(item.get("skill"))
        persist_skill_aliases(conn, canonical)
        skill_type = "production" if canonical in (parsed_candidate.get("production_skills") or []) else "exposure"
        conn.execute(
            """INSERT INTO candidate_skills
               (resume_hash, raw_value, canonical_value, confidence, skill_type)
               VALUES (?,?,?,?,?)""",
            (resume_hash, item.get("skill", ""), canonical, float(item.get("confidence", 0) or 0), skill_type)
        )
    conn.execute("DELETE FROM candidate_roles WHERE resume_hash=?", (resume_hash,))
    for role in parsed_candidate.get("normalized_roles", []) or []:
        conn.execute(
            """INSERT INTO candidate_roles (resume_hash, raw_value, canonical_value, confidence)
               VALUES (?,?,?,?)""",
            (resume_hash, role, role, 0.8)
        )
    conn.execute("DELETE FROM candidate_domains WHERE resume_hash=?", (resume_hash,))
    for item in parsed_candidate.get("domain_confidence_scores", []) or []:
        conn.execute(
            """INSERT INTO candidate_domains (resume_hash, raw_value, canonical_value, confidence)
               VALUES (?,?,?,?)""",
            (resume_hash, item.get("domain", ""), item.get("domain", ""), float(item.get("confidence", 0) or 0))
        )
    conn.commit()
    conn.close()

def persist_match_artifacts_async(jd_hash, resume_hash, result):
    snapshot = json.loads(json.dumps(result))
    queue_match_db_task("persist_match_artifacts", persist_match_artifacts, jd_hash, resume_hash, snapshot)

def persist_match_audit_entry(task_name, *, event_type, object_type, object_hash="", jd_hash="", resume_hash="", pipeline_version=MATCH_PIPELINE_VERSION, status="", source="", parser_confidence="", manual_review_required=False, score=None, message="", details=None, ip_address="", user_agent=""):
    conn = None
    try:
        conn = get_db(timeout=3)
        ensure_ats_pipeline_schema(conn)
        _insert_match_audit(
            conn,
            event_type=event_type,
            object_type=object_type,
            object_hash=object_hash,
            jd_hash=jd_hash,
            resume_hash=resume_hash,
            pipeline_version=pipeline_version,
            status=status,
            source=source,
            parser_confidence=parser_confidence,
            manual_review_required=manual_review_required,
            score=score,
            message=message,
            details=details,
            ip_address=ip_address,
            user_agent=user_agent,
        )
        conn.commit()
    finally:
        try:
            if conn is not None:
                conn.close()
        except Exception:
            pass

def queue_match_audit(task_name, **payload):
    payload.setdefault("ip_address", request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip())
    payload.setdefault("user_agent", (request.headers.get("User-Agent") or "")[:300])
    queue_match_db_task(task_name, persist_match_audit_entry, task_name, **payload)

def clean_cv_value(value, limit=80):
    value = re.sub(r"\s+", " ", value or "").strip(" -:|,\t\r\n")
    value = re.sub(r"\b(email|phone|mobile|contact|location|address)\b.*$", "", value, flags=re.I).strip(" -:|,")
    return value[:limit]

EMAIL_DOMAIN_WORDS = {
    "gmail", "yahoo", "hotmail", "outlook", "icloud", "proton", "zoho", "rediffmail",
    "live", "msn", "aol", "mail", "com", "co", "in", "net", "org", "edu"
}

def sanitize_candidate_name(value):
    raw = str(value or "").strip()
    if not raw:
        return ""
    email_match = re.search(r"[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}", raw)
    if email_match and raw.strip() == email_match.group(0):
        raw = email_match.group(0).split("@", 1)[0]
    raw = re.sub(r"^\s*(?:name|candidate\s+name|full\s+name)\s*[:\-]\s*", "", raw, flags=re.I)
    raw = re.sub(r"\b(?:gmail|yahoo|hotmail|outlook|icloud|proton|zoho|rediffmail|live|msn|aol|mail)\s*(?:\.|\s+)?(?:com|co\.in|in|net|org|edu)\b.*$", "", raw, flags=re.I)
    raw = re.sub(r"[@_]+", " ", raw)
    raw = re.sub(r"\.+", " ", raw)
    raw = re.sub(r"\b(?:email|e-mail|mail|phone|mobile|contact)\b.*$", "", raw, flags=re.I)
    raw = re.sub(r"[^A-Za-z'\-\s]", " ", raw)
    words = [w for w in re.findall(r"[A-Za-z][A-Za-z'-]*", raw) if w.lower() not in EMAIL_DOMAIN_WORDS]
    if len(words) < 2 or len(words) > 4:
        return ""
    return " ".join(words).title()

def looks_like_person_name(line):
    if not line or any(ch in line for ch in "@:/\\|()[]{}"):
        return False
    lower = line.lower()
    blocked = [
        "resume", "curriculum", "vitae", "profile", "summary", "experience", "education",
        "skill", "skills", "core skill", "developer", "engineer", "manager", "analyst", "consultant", "specialist",
        "email", "phone", "mobile", "contact", "address", "linkedin", "github",
        "gmail", "yahoo", "hotmail", "outlook", "icloud", "proton", "zoho", "rediffmail"
    ]
    if any(word in lower for word in blocked):
        return False
    words = re.findall(r"[A-Za-z][A-Za-z'-]*", line.replace(".", " "))
    if len(words) < 2 or len(words) > 4:
        return False
    if any(w.lower() in EMAIL_DOMAIN_WORDS for w in words):
        return False
    return sum(1 for w in words if w[:1].isupper()) >= min(2, len(words))

def normalize_person_name(value):
    return sanitize_candidate_name(value)

def infer_candidate_name_from_filename(filename):
    base = os.path.splitext(os.path.basename(filename or ""))[0]
    base = re.sub(r"(?i)(_?HRGURU(?:_\d+)?)$", "", base)
    base = re.sub(r"[_\-.]+", " ", base)
    base = re.sub(r"(?i)\b(resume|cv|profile|naukri|linkedin|indeed|monster|updated|latest|final|copy)\b", " ", base)
    base = re.sub(r"\b\d+\b", " ", base)
    return normalize_person_name(base)

def infer_candidate_name_from_email(email):
    local = (email or "").split("@", 1)[0]
    local = re.sub(r"\d+", " ", local)
    local = re.sub(r"[._\-]+", " ", local)
    return normalize_person_name(local)

def clean_company_name(value):
    value = clean_cv_value(value, 70)
    value = re.sub(r"\s+", " ", value).strip(" -:|,\t\r\n")
    if not value:
        return ""
    lower = value.lower()
    blocked = {
        "that", "would", "utilise", "organization", "organisation", "professional",
        "technology", "driven", "career", "objective", "summary"
    }
    words = re.findall(r"[A-Za-z0-9&.'-]+", value)
    if not words or words[0].lower() in blocked:
        return ""
    if len(words) == 1 and words[0].lower() in blocked:
        return ""
    return " ".join(words[:8]).title()

def clean_company_display(value):
    value = re.sub(r"\s*\([^)]*\).*$", "", str(value or ""))
    value = re.sub(r"(?i)eicher\s*ve", "Eicher VE", value)
    company = clean_company_name(value)
    company = re.sub(r"\bVe\b", "VE", company)
    company = re.sub(r"\bPvt\b", "Pvt", company)
    company = re.sub(r"\bLtd\b", "Ltd", company)
    return company

def split_company_role_line(line):
    value = clean_cv_value(line, 120)
    value = re.sub(r"^\s*\d+[\).]?\s*", "", value).strip()
    match = re.match(r"(.+?)\s+[–—-]\s+(.+)$", value)
    if not match:
        return "", ""
    left, right = match.group(1), match.group(2)
    date_words = r"\b(current|present|20\d{2}|19\d{2}|jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec|january|february|march|april|june|july|august|september|october|november|december)\b"
    if re.search(date_words, left, re.I) or re.search(r"^\s*(current|present|till\s+date|to\s+date)\s*$", right, re.I):
        return "", ""
    company = clean_company_display(left)
    role = clean_role_title(right)
    return company, role

def split_company_date_line(line):
    value = clean_cv_value(line, 120)
    match = re.match(
        r"(.+?)\s*[–—-]\s*(?:\(?[A-Za-z ]+\)?\s*)?(?:(?:jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*\s+)?(?:19|20)\d{2}\b.*(?:current|present|till\s+working|till\s+date|to\s+date|working)?",
        value,
        re.I
    )
    if not match:
        return ""
    return clean_company_display(match.group(1))

def infer_presently_working_company_role(lines):
    text = "\n".join(lines[:100])
    flat = re.sub(r"\s+", " ", text)
    match = re.search(
        r"presently\s+working.*?(customer\s+service\s+manager\s*&?\s*senior\s+eos\s+executive).*?\bin\s+(.+?)(?:\s+for\s+|\s+from\s+|\s+my\s+\d|\n|$)",
        flat,
        re.I
    )
    if match:
        return clean_company_display(match.group(2)), clean_role_title(match.group(1))
    match = re.search(r"presently\s+working.*?\bin\s+([A-Za-z0-9&()., \-]+?\b(?:ltd|limited|pvt|private|motors|vehicles)\b[^.,\n]*)", flat, re.I)
    if match:
        return clean_company_display(match.group(1)), ""
    return "", ""

def looks_like_role_title_line(line):
    role = clean_role_title(line)
    if not role:
        return False
    return bool(re.search(r"\b(engineer|developer|manager|analyst|consultant|specialist|associate|administrator|officer|executive|lead|architect|designer|director|head|service|support|quality|technical)\b", role, re.I))

def infer_current_company_role_from_experience(lines):
    if not lines:
        return "", ""
    company, role = infer_presently_working_company_role(lines)
    if company:
        return company, role
    heading_idx = -1
    for idx, line in enumerate(lines[:100]):
        lower = line.lower()
        if any(key in lower for key in ["work experience", "professional experience", "employment history", "career history", "employment"]):
            heading_idx = idx
            break
    if heading_idx < 0:
        for idx in range(0, min(len(lines) - 1, 80)):
            role_line = lines[idx]
            company = split_company_date_line(lines[idx + 1])
            nearby = " ".join(lines[idx + 1:idx + 4]).lower()
            if company and looks_like_role_title_line(role_line) and re.search(r"\b(current|present|till\s+working|till\s+date|to\s+date|working)\b", nearby):
                return company, clean_role_title(role_line)
        return "", ""
    stop_words = {"education", "academic", "skills", "certification", "certifications", "projects", "personal details"}
    for idx in range(heading_idx + 1, min(len(lines), heading_idx + 50)):
        lower = lines[idx].strip().lower()
        if lower in stop_words:
            break
        company, role = split_company_role_line(lines[idx])
        if not company:
            continue
        nearby = " ".join(lines[idx:idx + 4]).lower()
        if re.search(r"\b(current|present|till\s+date|to\s+date)\b", nearby):
            return company, role
    return "", ""

def infer_current_company_from_experience(lines):
    if not lines:
        return ""
    split_company, _ = infer_current_company_role_from_experience(lines)
    if split_company:
        return split_company
    heading_idx = -1
    for idx, line in enumerate(lines[:80]):
        lower = line.lower()
        if any(key in lower for key in ["work experience", "professional experience", "employment history", "career history", "employment"]):
            heading_idx = idx
            break
    if heading_idx < 0:
        return ""

    stop_words = {
        "career objective", "professional synopsis", "education", "academic", "skills",
        "certification", "certifications", "projects", "personal details"
    }
    role_words = {
        "associate", "consultant", "analyst", "engineer", "developer", "manager",
        "accountant", "auditor", "executive", "lead", "architect", "specialist"
    }
    for idx in range(heading_idx + 1, min(len(lines), heading_idx + 45)):
        line = clean_cv_value(lines[idx], 90)
        lower = line.lower()
        if not line:
            continue
        if lower in stop_words:
            break
        if re.search(r"\b(current|present|till\s+date|to\s+date)\b", lower):
            for back in range(idx - 1, heading_idx, -1):
                candidate = clean_cv_value(lines[back], 90)
                cand_lower = candidate.lower()
                if not candidate or re.search(r"\b(current|present|20\d{2}|19\d{2}|jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b", cand_lower):
                    continue
                if any(word in cand_lower for word in role_words):
                    continue
                if len(candidate.split()) <= 6:
                    cleaned = clean_company_name(candidate)
                    if cleaned:
                        return cleaned
    return ""

def clean_role_title(value):
    value = clean_cv_value(value, 90)
    value = re.sub(r"\s*&\s*", " & ", value)
    value = re.sub(r"\s+", " ", value).strip(" -:|,\t\r\n")
    if not value:
        return ""
    lower = value.lower()
    blocked_fragments = [
        "professional and technology", "career objective", "would utilise", "growth of the organisation",
        "challenging position", "looking for", "seeking", "aspiring"
    ]
    if any(fragment in lower for fragment in blocked_fragments):
        return ""
    words = re.findall(r"[A-Za-z0-9&/.'+-]+", value)
    if len(words) < 1 or len(words) > 9:
        return ""
    role = " ".join(words).title()
    role = re.sub(r"\bEos\b", "EOS", role)
    return role

def infer_current_role_from_experience(lines):
    if not lines:
        return ""
    _, split_role = infer_current_company_role_from_experience(lines)
    if split_role:
        return split_role
    heading_idx = -1
    for idx, line in enumerate(lines[:80]):
        lower = line.lower()
        if any(key in lower for key in ["work experience", "professional experience", "employment history", "career history", "employment"]):
            heading_idx = idx
            break
    if heading_idx < 0:
        return ""

    role_words = {
        "associate", "consultant", "analyst", "engineer", "developer", "manager",
        "accountant", "auditor", "executive", "lead", "architect", "specialist",
        "administrator", "director", "officer"
    }
    for idx in range(heading_idx + 1, min(len(lines), heading_idx + 45)):
        line = clean_cv_value(lines[idx], 90)
        lower = line.lower()
        if re.search(r"\b(current|present|till\s+date|to\s+date)\b", lower):
            for forward in range(idx + 1, min(len(lines), idx + 5)):
                candidate = clean_cv_value(lines[forward], 90).strip(",")
                cand_lower = candidate.lower()
                if not candidate:
                    continue
                if re.search(r"\b(20\d{2}|19\d{2}|jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec|current|present)\b", cand_lower):
                    continue
                if any(word in cand_lower for word in role_words):
                    cleaned = clean_role_title(candidate)
                    if cleaned:
                        return cleaned
            for back in range(idx - 1, heading_idx, -1):
                candidate = clean_cv_value(lines[back], 90).strip(",")
                cand_lower = candidate.lower()
                if any(word in cand_lower for word in role_words):
                    cleaned = clean_role_title(candidate)
                    if cleaned:
                        return cleaned
    return ""

def clean_header_role(value):
    role = clean_role_title(value)
    if not role:
        return ""
    if not re.search(r"\b(engineer|developer|manager|analyst|consultant|specialist|associate|administrator|officer|executive|lead|architect|designer|director|head|service)\b", role, re.I):
        return ""
    return role

def extract_pdf_header_fields(cv_path):
    fields = {}
    if os.path.splitext(cv_path or "")[1].lower() != ".pdf":
        return fields
    try:
        import pymupdf
        doc = pymupdf.open(cv_path)
        if not doc:
            return fields
        page = doc[0]
        blocks = []
        for block in page.get_text("blocks") or []:
            x0, y0, x1, y1, text = block[:5]
            if y0 <= min(page.rect.height * 0.28, 180) and str(text or "").strip():
                for line in str(text or "").splitlines():
                    cleaned = re.sub(r"\s+", " ", line).strip()
                    if cleaned:
                        blocks.append((float(y0), float(x0), cleaned))
        doc.close()
    except Exception:
        return fields
    lines = [item[2] for item in sorted(blocks, key=lambda b: (round(b[0] / 4) * 4, b[1]))]
    header_text = "\n".join(lines)
    email_match = re.search(r"(?<![\w.-])[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}(?![\w.-])", header_text)
    if email_match:
        fields["email_addr"] = email_match.group(0).strip().lower()
    phone_match = re.search(r"(?:\+?91[\s\-]?)?[6-9]\d{4}[\s\-]?\d{5}|\+?\d[\d\s\-]{9,}\d", header_text)
    if phone_match:
        phone = re.sub(r"[^\d+]", "", phone_match.group(0))
        fields["phone"] = phone[-10:] if phone.startswith("91") and len(phone) > 10 else phone
    exp_match = re.search(r"(\d+(?:\.\d+)?)\s*\+?\s*(?:years?|yrs?)\s+(?:of\s+)?(?:experience|exp)", header_text, re.I)
    if exp_match:
        years = float(exp_match.group(1))
        if years <= 40:
            fields["experience_years"] = f"{years:g} years"
    for line in lines[:8]:
        candidate = normalize_person_name(line)
        if candidate and looks_like_person_name(candidate):
            fields["candidate_name"] = candidate
            break
    if fields.get("candidate_name"):
        name_idx = next((i for i, line in enumerate(lines[:8]) if normalize_person_name(line) == fields["candidate_name"]), -1)
        for line in lines[name_idx + 1:name_idx + 6]:
            if "@" in line or re.search(r"\b(contact|phone|mobile|email|experience)\b", line, re.I):
                continue
            role = clean_header_role(line)
            if role:
                fields["current_role"] = role
                break
    cities = ['Bangalore', 'Bengaluru', 'Mumbai', 'Delhi', 'Hyderabad', 'Chennai', 'Pune', 'Noida', 'Gurgaon', 'Gurugram', 'Kolkata', 'Ahmedabad']
    text_l = header_text.lower()
    for city in cities:
        if city.lower() in text_l:
            fields["current_location"] = city
            break
    return fields

def clean_skill_phrase(value):
    value = re.sub(r"^[\s\-*•??\d.)]+", "", str(value or ""))
    value = re.sub(r"\s+", " ", value).strip(" .;:-|")
    value = re.sub(r"^(?:core\s+)?(?:skills?|competenc(?:y|ies)|expertise|areas?\s+of\s+expertise)\s*[:\-]\s*", "", value, flags=re.I)
    if not value or len(value) < 3:
        return ""
    phrase = value.title() if value.isupper() else value
    sentence_patterns = [
        (r"ability to handle all kinds of automobile customers.*", "Customer Complaint Handling"),
        (r".*\bcomplaints?\b.*\bsolution\b.*customer satisfaction.*", "Customer Complaint Handling"),
        (r".*\bautomobile(?:s)?\b.*\bparts\b.*\brepair\b.*", "Automobile Parts & Repair Knowledge"),
        (r".*\bcommunication\b.*\borganizational skills\b.*", "Communication & Organization Skills"),
        (r".*\btaking charge of work\b.*", "Accountability & Responsibility Handling"),
        (r".*\bverbal\b.*\bcommunication\b.*\breporting\b.*", "Communication & Reporting"),
        (r".*\bhandling projects\b.*", "Project Handling"),
        (r".*\bms word\b.*\bms excel\b.*\bms outlook\b.*", "MS Word, MS Excel, MS Outlook"),
        (r".*\bfi[- ]mm integration\b.*", "FI-MM Integration"),
        (r".*\bmonth end\b.*", "Month-End Closing"),
        (r".*\byear[- ]end\b.*", "Year-End Closing"),
        (r".*\bopening\b.*\bclosing posting\b.*", "Opening & Closing Posting Periods"),
        (r".*\bconfiguration\b.*\bfico\b.*", "SAP FICO Configuration"),
        (r".*\binvoice processing\b.*", "Invoice Processing"),
        (r".*\bpayment runs?\b.*", "Payment Runs"),
    ]
    for pattern, replacement in sentence_patterns:
        if re.match(pattern, phrase, re.I):
            return replacement
    if len(value) > 80:
        return ""
    lower = value.lower()
    blocked = [
        "page ", "resume", "curriculum vitae", "professional experience", "work experience",
        "education", "personal details", "contact number", "email", "references",
        "willing to work on challenging assignments"
    ]
    if any(fragment in lower for fragment in blocked):
        return ""
    if re.search(r"\b(19|20)\d{2}\b", value) or re.search(r"@|www\.|https?://", value, re.I):
        return ""
    words = re.findall(r"[A-Za-z][A-Za-z0-9+#.-]*", value)
    allowed_single = {"SAP", "FICO", "Excel", "GST", "TDS", "Python", "Java", "React", "Angular", "Docker", "Kubernetes", "SQL"}
    if len(words) == 1 and words[0].upper() not in {x.upper() for x in allowed_single}:
        return ""
    return phrase

def extract_skill_section_phrases(lines, limit=12):
    headings = [
        "core skill", "core skills", "key skill", "key skills", "technical skills",
        "functional skills", "sap functional skills", "skills", "skill set",
        "areas of expertise", "area of expertise", "competencies", "core competencies",
        "professional skills", "domain skills"
    ]
    stops = [
        "professional experience", "work experience", "employment history", "career history",
        "education", "academic", "certification", "certifications", "projects",
        "personal details", "declaration", "languages known", "hobbies", "responsibilities"
    ]
    found = []
    for idx, line in enumerate(lines[:160]):
        normalized = re.sub(r"[^a-z0-9 ]+", " ", line.lower()).strip()
        if normalized not in headings:
            continue
        items = []
        current = ""
        for item in lines[idx + 1:min(len(lines), idx + 35)]:
            item_norm = re.sub(r"[^a-z0-9 ]+", " ", item.lower()).strip()
            if item_norm in stops or item_norm in headings or looks_like_role_title_line(item) or split_company_date_line(item):
                break
            if re.fullmatch(r"[\s\-*•??]+", item):
                if current:
                    items.append(current)
                current = ""
                continue
            if current and re.match(r"^[a-z,;&(). ]+$", item.strip()):
                current += " " + item.strip()
            else:
                if current:
                    items.append(current)
                current = item.strip()
        if current:
            items.append(current)
        for item in items:
            whole_skill = clean_skill_phrase(item)
            if whole_skill and whole_skill != item.strip() and whole_skill.lower() not in {x.lower() for x in found}:
                found.append(whole_skill)
                if len(found) >= limit:
                    return found
                continue
            parts = re.split(r"[,;|]+", item)
            for part in parts:
                skill = clean_skill_phrase(part)
                if skill and skill.lower() not in {x.lower() for x in found}:
                    found.append(skill)
                    if len(found) >= limit:
                        return found
    return found

def extract_domain_skill_phrases(text, limit=12):
    phrases = [
        "Field Breakdown Analysis", "Warranty Handling", "Service Operations", "Dealer Management",
        "Customer Complaint Handling", "Root Cause Analysis", "Diesel Engines", "Engine Diagnostics",
        "Preventive Maintenance", "Technical Support", "After Sales Service", "Team Handling",
        "Vendor Management", "Client Relationship Management", "Business Development", "Channel Sales",
        "Key Account Management", "Lead Generation", "Negotiation", "Recruitment", "Talent Acquisition",
        "Sourcing", "Screening", "Interview Coordination", "Payroll", "Employee Relations",
        "Accounts Payable", "Accounts Receivable", "General Ledger", "GST", "TDS", "Bank Reconciliation",
        "Financial Reporting", "SAP FICO", "Month End Closing", "Invoice Processing", "Payment Runs",
        "Customer Satisfaction Index", "Post Service Feedback", "Retrofitments", "AMC", "Technical Issues",
        "E-Learning Monitoring", "Value Added Services", "KMPL Checkup", "Onsite Services",
        "AutoCAD", "Siebel", "SAP", "CRM DMS", "Vehicle Management", "Service Advisor",
        "Job Card Opening", "Vehicle Delivery", "Breakdown Service"
    ]
    text_l = (text or "").lower()
    found = []
    for phrase in phrases:
        pattern = r"(?<![a-z0-9])" + re.escape(phrase.lower()) + r"(?![a-z0-9])"
        if re.search(pattern, text_l):
            found.append(phrase)
            if len(found) >= limit:
                break
    return found

def extract_resume_skills(text, lines, skills_db, limit=12):
    section_skills = extract_skill_section_phrases(lines, limit)
    domain_skills = extract_domain_skill_phrases(text, limit)
    combined = unique_list(section_skills + domain_skills, limit)
    if combined:
        return ", ".join(combined)

    found_skills = set()
    text_lower = (text or "").lower()
    for category, skills in skills_db.items():
        for skill in skills:
            pattern = r"(?<![a-z0-9])" + re.escape(skill.lower()) + r"(?![a-z0-9])"
            if re.search(pattern, text_lower):
                display = skill.replace('r programming', 'R').replace('c programming', 'C')
                found_skills.add(display.title())
    if found_skills:
        sorted_skills = sorted(found_skills, key=lambda x: len(x), reverse=True)
        return ", ".join(sorted_skills[:limit])
    return ""

def parse_cv(cv_path, use_ai=False, ai_provider="", ai_api_key=""):
    """
    CV parsing with HuggingFace NER (when available) + enhanced regex/heuristics fallback.
    Extracts: name, email, phone, company, role, experience, skills, location, education, notice period, salary, summary.
    """
    result = {"candidate_name":"","email_addr":"","phone":"",
              "current_company":"","current_role":"","experience_years":"",
              "key_skills":"","current_location":"","cv_summary":"",
              "education":"","notice_period":"","current_salary":"","expected_salary":""}
    try:
        text = extract_cv_text(cv_path)
        if not text.strip():
            if os.path.splitext(cv_path or "")[1].lower() == ".pdf":
                result["_parse_warning"] = "No selectable text was found in this PDF. It looks like a scanned/image resume, so ATS could not auto-read it. Please enter details manually or use OCR before upload."
            else:
                result["_parse_warning"] = "ATS could not read text from this CV file. Please check the file or enter details manually."
            return result
        
        import re
        text_lower = text.lower()
        
        # Step 1: Try HuggingFace NER (free tier - no API key needed)
        hf_result = parse_huggingface(text, ai_api_key or "")
        if not hf_result.get("_hf_limited") and not hf_result.get("_hf_error"):
            # HF worked! Use its results and fill remaining with heuristics
            if hf_result.get("candidate_name"):
                result["candidate_name"] = normalize_person_name(hf_result["candidate_name"])
            if hf_result.get("current_company"):
                result["current_company"] = hf_result["current_company"]
        else:
            # HF not available (rate limit or no API key) - use full heuristics
            pass
        
        # Step 2: Always run enhanced heuristics to fill gaps + extract other fields
        lines_list = [l.strip() for l in text.split('\n') if l.strip()]
        header_fields = extract_pdf_header_fields(cv_path)
        for field in ("candidate_name", "email_addr", "phone", "current_role", "experience_years", "current_location"):
            if header_fields.get(field):
                result[field] = header_fields[field]
        
        # ===== EMAIL =====
        if not result.get("email_addr"):
            email_match = re.search(r"(?<![\w.-])[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}(?![\w.-])", text)
            if email_match:
                result["email_addr"] = email_match.group(0).strip().lower()
        
        # ===== PHONE =====
        if not result.get("phone"):
            phone_patterns = [
                r"(?:\+?91[\s\-]?)?[6-9]\d{4}[\s\-]?\d{5}",  # Indian mobile
                r"\+?1[\s\-]?\(?\d{3}\)[\s\-]?\d{3}[\s\-]?\d{4}",  # US
                r"\+?\d[\d\s\-]{9,}\d",  # Generic international
                r"\(?\d{2,4}\)?[\s\-]?\d{3,4}[\s\-]?\d{3,4}",  # Landline
            ]
            for pattern in phone_patterns:
                match = re.search(pattern, text)
                if match:
                    phone = re.sub(r"[^\d+]", "", match.group(0))
                    result["phone"] = phone[-10:] if phone.startswith("91") and len(phone) > 10 else phone
                    break
        
        # ===== NAME =====
        if not result.get("candidate_name"):
            name_patterns = [
                r"(?:^|\n)\s*(?:name|candidate\s+name|full\s+name)\s*[:\-]\s*([A-Za-z][A-Za-z'.\- \t]{3,70})",
                r"(?:^|\n)\s*([A-Za-z][A-Za-z'.-]+[ \t]+[A-Za-z][A-Za-z'.-]+(?:[ \t]+[A-Za-z][A-Za-z'.-]+){0,2})\s*(?:\n|$)",
            ]
            for pattern in name_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    candidate = clean_cv_value(match.group(1), 60)
                    normalized = normalize_person_name(candidate)
                    if normalized and looks_like_person_name(normalized):
                        result["candidate_name"] = normalized
                        break

        if not result.get("candidate_name"):
            for line in lines_list[:18]:
                candidate = clean_cv_value(line, 60)
                normalized = normalize_person_name(candidate)
                if normalized and looks_like_person_name(normalized):
                    result["candidate_name"] = normalized
                    break

        if not result.get("candidate_name") and result.get("email_addr"):
            email_name = infer_candidate_name_from_email(result.get("email_addr"))
            if email_name:
                result["candidate_name"] = email_name
        
        # ===== COMPANY =====
        email_domains = ['gmail', 'yahoo', 'hotmail', 'outlook', 'aol', 'icloud', 'proton', 'zoho', 'mail', 'live', 'msn']
        def is_not_email(s):
            return not any(d in s.lower() for d in email_domains)

        experience_company = infer_current_company_from_experience(lines_list)
        if experience_company:
            result["current_company"] = experience_company
        
        company_suffixes = ['inc', 'llc', 'corp', 'pvt', 'ltd', 'llp', 'co', 'company', 'technologies', 'tech',
                          'solutions', 'systems', 'services', 'labs', 'studios', 'digital', 'software',
                          'consulting', 'group', 'international', 'global', 'innovations', 'partners',
                          'ventures', 'info', 'institute', 'bank', 'finance', 'enterprises', 'holdings',
                          'analytics', 'data', 'ai', 'learning']
        
        # Pattern 1: Explicit labels
        if not result.get("current_company"):
            label_patterns = [
                r"(?:current\s+company|current\s+employer|company|employer)[:\s-]*([A-Za-z0-9&.,' -]{2,60})",
                r"(?:works?\s+(?:at|with)|employed\s+(?:at|by)|associated\s+with)[:\s-]*([A-Za-z0-9&.,' -]{2,60})",
                r"(?:currently\s+(?:at|with|working\s+at))[:\s-]*([A-Za-z0-9&.,' -]{2,60})",
            ]
            for pattern in label_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match and is_not_email(match.group(1)):
                    result["current_company"] = clean_company_name(match.group(1))
                    if not result["current_company"]:
                        continue
                    break
        
        # Pattern 2: Look for company in experience entries
        if not result.get("current_company"):
            exp_section = False
            for i, line in enumerate(lines_list[:50]):
                lower = line.lower()
                if any(kw in lower for kw in ['experience', 'employment', 'work history', 'professional experience']):
                    exp_section = True
                    continue
                if exp_section and line.strip():
                    line_clean = line.strip()
                    if len(line_clean) > 3:
                        words_in_line = line_clean.split()
                        if len(words_in_line) >= 2:
                            for suffix in company_suffixes:
                                if suffix in lower:
                                    match = re.match(r"([A-Za-z][\w\s&.,]+)", line_clean)
                                    if match and is_not_email(line_clean):
                                        result["current_company"] = match.group(1).strip().title()[:45]
                                        break
                        if result.get("current_company"):
                            break
                        if len(line.split()) > 6:
                            break
        
        # Pattern 3: Company name with suffix
        if not result.get("current_company"):
            for line in lines_list[:30]:
                for suffix in company_suffixes:
                    if suffix in line.lower() and len(line) > 5 and len(line) < 60:
                        if is_not_email(line):
                            match = re.match(r"([A-Za-z][\w\s&.,]+)", line.strip())
                            if match:
                                result["current_company"] = match.group(1).strip().title()[:45]
                                break
                if result.get("current_company"):
                    break
        
        # ===== ROLE =====
        role_suffixes = ['engineer', 'developer', 'manager', 'analyst', 'designer', 'architect', 
                        'consultant', 'lead', 'specialist', 'administrator', 'director', 'head',
                        'associate', 'coordinator', 'officer', 'executive', 'president', 'vp']
        
        experience_role = infer_current_role_from_experience(lines_list)
        if experience_role:
            result["current_role"] = experience_role

        # Pattern 1: Explicit labels
        if not result.get("current_role"):
            role_patterns = [
                r"(?:current\s+role|current\s+designation|designation|job\s+title|role|title)[:\s-]*([A-Za-z][^\n,]{2,60})",
                r"(?:working\s+as)[:\s-]*([A-Za-z][^\n,]{2,60})",
            ]
            for pattern in role_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    result["current_role"] = clean_role_title(match.group(1))
                    if not result["current_role"]:
                        continue
                    break
        
        # Pattern 2: Look for role titles in experience section
        if not result.get("current_role"):
            exp_section = False
            for i, line in enumerate(lines_list[:60]):
                lower = line.lower()
                if any(kw in lower for kw in ['experience', 'employment']):
                    exp_section = True
                    continue
                if exp_section:
                    for suffix in role_suffixes:
                        if suffix in lower:
                            match = re.search(r"([A-Za-z][\w\s/\-]+(?:\s+(?:engineer|developer|manager|analyst|designer|architect|consultant|lead|specialist))?)", line, re.IGNORECASE)
                            if match:
                                result["current_role"] = clean_role_title(match.group(1))
                                break
                    if result.get("current_role"):
                        break
        
        # Pattern 3: Common title formats anywhere
        if not result.get("current_role"):
            title_patterns = [
                r"((?:Senior|Junior|Lead|Principal|Staff|Associate)\s+(?:Software|Frontend|Backend|Full\s*Stack|Data|ML|DevOps)\s+Engineer)",
                r"((?:Software|Full\s*Stack|Frontend|Backend)\s+Developer)",
                r"((?:Product|Project|Program|Technical|Engineering)\s+Manager)",
                r"((?:Business|System|Data|Business\s+Intelligence)\s+Analyst)",
                r"((?:DevOps|Cloud|Site\s*Reliability)\s+Engineer)",
                r"((?:UI|UX|Graphic)\s+Designer)",
                r"((?:Machine\s*Learning|ML)\s+Engineer)",
                r"((?:Data|Business)\s+Scientist)",
            ]
            for pattern in title_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    result["current_role"] = match.group(1).strip().title()[:55]
                    break
        
        # ===== SKILLS =====
        skills_db = {
            'languages': ['python', 'java', 'javascript', 'c++', 'c#', 'c programming', 'php', 'ruby', 'go', 'golang', 'rust', 'swift', 'kotlin', 'scala', 'r programming', 'matlab', 'perl', 'shell', 'bash'],
            'frameworks': ['react', 'angular', 'vue', 'node', 'node.js', 'express', 'django', 'flask', 'spring', 'spring boot', 'rails', 'laravel', 'next.js', 'gatsby', 'nuxt'],
            'databases': ['sql', 'mysql', 'postgresql', 'mongodb', 'oracle', 'redis', 'elasticsearch', 'cassandra', 'dynamodb', 'sqlite', 'sql server', 'plsql'],
            'cloud': ['aws', 'amazon web services', 'azure', 'google cloud', 'gcp', 'heroku', 'digitalocean', 'docker', 'kubernetes', 'k8s', 'terraform'],
            'tools': ['git', 'github', 'gitlab', 'jenkins', 'ci/cd', 'jira', 'agile', 'scrum', 'jira', 'confluence', 'maven', 'gradle', 'npm', 'yarn', 'webpack'],
            'data': ['machine learning', 'ml', 'ai', 'artificial intelligence', 'deep learning', 'tensorflow', 'pytorch', 'pandas', 'numpy', 'scikit-learn', 'data science', 'tableau', 'power bi', 'excel', 'statistics'],
            'web': ['html', 'css', 'rest api', 'graphql', 'json', 'xml', 'web services', 'sass', 'less', 'bootstrap', 'tailwind'],
            'other': ['linux', 'unix', 'windows server', 'networking', 'security', 'devops', 'microservices', 'api', 'backend', 'frontend', 'full stack'],
        }
        
        result["key_skills"] = extract_resume_skills(text, lines_list, skills_db, 12)
        
        # ===== EXPERIENCE =====
        exp_patterns = [
            r"(?:total|overall)\s+(\d+(?:\.\d+)?)\s*-\s*years?",
            r"(?:total|overall)\s+(?:experience\s+)?(?:of\s+)?(\d+(?:\.\d+)?)\s*\+?\s*years?",
            r"(?:experience|total\s+experience)[:\s]*(\d+(?:\.\d+)?)\s*\+?\s*years?",
            r"(\d+(?:\.\d+)?)\s*\+?\s*years?\s+(?:of\s+)?(?:experience|work)",
            r"(\d+(?:\.\d+)?)\s*years?\s+(?:in|of|working)",
            r"(\d+(?:\.\d+)?)\+?\s*yrs?\s+(?:exp|experience)",
        ]
        for pattern in exp_patterns:
            match = re.search(pattern, text_lower)
            if match:
                years = float(match.group(1))
                if years <= 40:
                    result["experience_years"] = f"{years:g} years"
                    break
        
        # Fallback: count months
        if not result.get("experience_years"):
            months = len(re.findall(r"\d+\s*(?:months?|yrs?|years?)", text_lower))
            if months > 0:
                total_years = max(1, months // 12)
                result["experience_years"] = f"{total_years} years"
        
        # ===== LOCATION =====
        indian_cities = ['bangalore', 'bengaluru', 'mumbai', 'delhi', 'hyderabad', 'chennai', 'pune', 
                         'kolkata', 'ahmedabad', 'noida', 'gurgaon', 'gurugram', 'chandigarh', 'kolkata',
                         'jaipur', 'surat', 'lucknow', 'indore', 'coimbatore', 'vizag', 'kochi']
        for city in indian_cities:
            if city in text_lower:
                result["current_location"] = city.title()
                break
        
        if not result.get("current_location"):
            loc_match = re.search(r"(?:location|based\s+in|residing\s+in|city)[:\s]*([A-Za-z][a-z]+)", text_lower)
            if loc_match:
                result["current_location"] = loc_match.group(1).title()
        
        # ===== EDUCATION =====
        edu_patterns = [
            r"(?:b\.?tech|b\.?e\.?|b\.?sc\.?|m\.?tech|m\.?sc\.?|mba|m\.?ca\.?|ph\.?d\.?|b\.?a\.?|m\.?com\.?)",
        ]
        for pattern in edu_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                result["education"] = match.group(0).upper()
                break
        
        # ===== NOTICE PERIOD =====
        notice_patterns = [
            r"(?:notice\s+period|np)[:\s]*(\d+)\s*(?:days?|months?)",
            r"(?:serving\s+notice|serving\s+np)[:\s]*(\d+)",
            r"can\s+join\s+(?:within|in)\s+(\d+)\s*(?:days?|months?)",
            r"(\d+)\s*(?:days?|months?)\s+notice",
        ]
        for pattern in notice_patterns:
            match = re.search(pattern, text_lower)
            if match:
                result["notice_period"] = match.group(0).strip()
                break
        
        # ===== SALARY =====
        ctc_patterns = [
            r"(?:current\s+ctc|present\s+ctc|existing\s+ctc|salary)[:\s]*â‚¹?\s*(\d+(?:\.\d+)?)\s*(?:lpa|lakh|lacs?|inr)?",
            r"(?:ctc)[:\s]*â‚¹?\s*(\d+(?:\.\d+)?)\s*(?:lpa|lakh|lacs?)",
        ]
        for pattern in ctc_patterns:
            match = re.search(pattern, text_lower)
            if match:
                try:
                    val = float(match.group(1))
                    if val < 1000:
                        val *= 100000
                    result["current_salary"] = f"â‚¹{val/100000:.1f} LPA"
                except:
                    pass
                break
        
        exp_sal_patterns = [
            r"(?:expected\s+ctc|expected\s+salary|notice\s+salary)[:\s]*â‚¹?\s*(\d+(?:\.\d+)?)\s*(?:lpa|lakh|lacs?)?",
        ]
        for pattern in exp_sal_patterns:
            match = re.search(pattern, text_lower)
            if match:
                try:
                    val = float(match.group(1))
                    if val < 1000:
                        val *= 100000
                    result["expected_salary"] = f"â‚¹{val/100000:.1f} LPA"
                except:
                    pass
                break
        
        # ===== SUMMARY =====
        summary_parts = []
        
        line1_parts = []
        if result.get("current_role"):
            line1_parts.append(result["current_role"])
        if result.get("experience_years"):
            line1_parts.append(f"{result['experience_years']}")
        if result.get("current_company"):
            line1_parts.append(f"@ {result['current_company']}")
        if line1_parts:
            summary_parts.append(" â€¢ ".join(line1_parts))
        
        line2_parts = []
        if result.get("key_skills"):
            skills = result["key_skills"]
            if len(skills) > 55:
                skill_list = skills.split(", ")
                skills = ", ".join(skill_list[:6])
                if len(skill_list) > 6:
                    skills += f" +{len(skill_list)-6}"
            line2_parts.append(skills)
        if result.get("current_location"):
            line2_parts.append(f"ðŸ“ {result['current_location']}")
        if line2_parts:
            summary_parts.append(" â€¢ ".join(line2_parts))
        
        result["candidate_name"] = normalize_person_name(result.get("candidate_name", ""))
        result["cv_summary"] = " | ".join(summary_parts) if summary_parts else ""
        
        return result
    except Exception as e:
        return {"error": f"CV parsing failed: {str(e)}"}
def norm_phone(p): return re.sub(r'[^\d]', '', p or "")[-10:]

def check_dup(conn, row):
    phone = norm_phone(row.get("phone", ""))
    email = (row.get("email_addr", "") or "").strip().lower()
    req_id = row.get("requirement_id")
    req_clause = ""
    params_prefix = []
    if req_id:
        req_clause = " AND requirement_id=?"
        params_prefix.append(req_id)
    if email:
        r = conn.execute(
            """SELECT id, candidate_name, email_addr, phone, current_company, current_role
               FROM candidates
               WHERE email_addr=? AND is_duplicate=0""" + req_clause + " LIMIT 1",
            [email] + params_prefix).fetchone()
        if r: return True, dict(r), "email"
    if phone and len(phone) >= 8:
        r = conn.execute(
            """SELECT id, candidate_name, email_addr, phone, current_company, current_role
               FROM candidates
               WHERE phone=? AND is_duplicate=0""" + req_clause + " LIMIT 1",
            [phone] + params_prefix).fetchone()
        if r: return True, dict(r), "phone"
    if phone and len(phone) >= 8:
        r = conn.execute(
            """SELECT id, candidate_name, email_addr, phone, current_company, current_role
               FROM candidates
               WHERE replace(replace(replace(phone,'+',''),'-',''),' ','')=? AND is_duplicate=0""" + req_clause + " LIMIT 1",
            [phone] + params_prefix).fetchone()
        if r: return True, dict(r), "phone"
    return False, None, None

def build_duplicate_candidate_message(row, duplicate_row, dup_why):
    current_name = str(row.get("candidate_name") or "candidate").strip() or "candidate"
    dup_name = str((duplicate_row or {}).get("candidate_name") or "existing candidate").strip() or "existing candidate"
    dup_id = (duplicate_row or {}).get("id")
    dup_company = str((duplicate_row or {}).get("current_company") or "").strip()
    dup_role = str((duplicate_row or {}).get("current_role") or "").strip()
    dup_email = str((duplicate_row or {}).get("email_addr") or "").strip()
    dup_phone = str((duplicate_row or {}).get("phone") or "").strip()
    context_bits = []
    if dup_company:
        context_bits.append(f"company: {dup_company}")
    if dup_role:
        context_bits.append(f"role: {dup_role}")
    if dup_email:
        context_bits.append(f"email: {dup_email}")
    if dup_phone:
        context_bits.append(f"phone: {dup_phone}")
    context = f" ({'; '.join(context_bits)})" if context_bits else ""
    action = "Please skip this row, or update the existing candidate if this is a correction."
    reason = f"matched by {dup_why}" if dup_why else "matched on the same candidate details"
    return f"{current_name} is a duplicate of {dup_name} (ID #{dup_id}){context}. {reason}. {action}"

def normalize_requirement_key(value):
    return re.sub(r"\s+", " ", str(value or "").strip().lower())

def normalize_requirement_file_key(value):
    value = os.path.splitext(os.path.basename(str(value or "")))[0]
    value = re.sub(r"(?i)\b(jd|job\s*description|requirement|req|profile)\b", " ", value)
    value = re.sub(r"[_\-.]+", " ", value)
    value = re.sub(r"[^a-zA-Z0-9+# ]+", " ", value)
    return normalize_requirement_key(value)

def load_requirement_lookup(conn):
    rows = conn.execute("SELECT id,title,client_name,status FROM requirements ORDER BY title").fetchall()
    lookup = {}
    for r in rows:
        rid = str(r["id"])
        title = normalize_requirement_key(r["title"])
        client = normalize_requirement_key(r["client_name"])
        keys = {rid, title}
        if client:
            keys.add(f"{title} - {client}")
            keys.add(f"{title} / {client}")
        for key in keys:
            if key:
                lookup[key] = r
    return rows, lookup

def resolve_requirement_id(row, requirement_lookup):
    raw_id = str(row.get("requirement_id") or "").strip()
    raw_title = str(row.get("requirement_title") or row.get("role_name") or "").strip()
    raw_client = str(row.get("client_name") or "").strip()
    match = None
    if raw_id:
        match = requirement_lookup.get(normalize_requirement_key(raw_id))
    if not match and raw_title:
        match = requirement_lookup.get(normalize_requirement_key(raw_title))
    if not match and raw_title and raw_client:
        combined = normalize_requirement_key(f"{raw_title} - {raw_client}")
        alt_combined = normalize_requirement_key(f"{raw_title} / {raw_client}")
        match = requirement_lookup.get(combined) or requirement_lookup.get(alt_combined)
    if not match:
        return None
    row["requirement_id"] = match["id"]
    if not row.get("role_name"):
        row["role_name"] = match["title"]
    return match["id"]

def suggest_requirement_matches(raw_title, raw_client, requirement_lookup, limit=3):
    target_title = normalize_requirement_title_text(raw_title).lower().strip()
    target_client = normalize_requirement_client_key(raw_client)
    seen_ids = set()
    suggestions = []
    requirements = list({row["id"]: row for row in requirement_lookup.values()}.values())
    for req in requirements:
        rid = row_value(req, "id")
        if rid in seen_ids:
            continue
        seen_ids.add(rid)
        req_title = normalize_requirement_title_text(row_value(req, "title")).lower().strip()
        req_client = normalize_requirement_client_key(row_value(req, "client_name"))
        title_ratio = difflib.SequenceMatcher(None, target_title, req_title).ratio() if target_title else 0.0
        combined_ratio = difflib.SequenceMatcher(None, f"{target_title} {target_client}".strip(), f"{req_title} {req_client}".strip()).ratio() if (target_title or target_client) else title_ratio
        score = max(title_ratio, combined_ratio)
        if target_client and req_client and target_client == req_client:
            score = min(1.0, score + 0.12)
        suggestions.append((score, req))
    suggestions.sort(key=lambda item: (item[0], str(row_value(item[1], "title") or "")), reverse=True)
    out = []
    for score, req in suggestions[:limit]:
        label = str(row_value(req, "title") or "").strip()
        client = str(row_value(req, "client_name") or "").strip()
        if client:
            label = f"{label} - {client}"
        out.append({"id": row_value(req, "id"), "title": row_value(req, "title"), "client_name": row_value(req, "client_name"), "match_score": round(score, 3), "label": label})
    return out

def save_bulk_jd_files(conn, jd_file_list, requirement_lookup, batch_id):
    result = {"attached": 0, "warnings": []}
    upload_folder = os.path.join(app.root_path, "uploads", "recruiters", current_upload_owner_key(), "requirements", "bulk", _safe_path_part(batch_id))
    os.makedirs(upload_folder, exist_ok=True)
    requirements = list({row["id"]: row for row in requirement_lookup.values()}.values())
    for jd in jd_file_list or []:
        if not jd or not jd.filename:
            continue
        safe_name = secure_filename(jd.filename)
        ext = os.path.splitext(safe_name)[1].lower()
        if ext not in (".pdf", ".doc", ".docx", ".txt"):
            result["warnings"].append(f"{safe_name} (invalid file type)")
            continue
        file_key = normalize_requirement_file_key(safe_name)
        match = None
        if file_key:
            match = requirement_lookup.get(file_key)
        if not match and file_key:
            for req in requirements:
                title_key = normalize_requirement_key(req["title"])
                client_key = normalize_requirement_key(req["client_name"])
                combined = normalize_requirement_key(f"{req['title']} {req['client_name'] or ''}")
                if title_key and (title_key in file_key or file_key in title_key or combined in file_key):
                    match = req
                    break
                if client_key and title_key and title_key in file_key and client_key in file_key:
                    match = req
                    break
        if not match:
            result["warnings"].append(safe_name)
            continue
        saved_name = f"bulk_jd_{batch_id}_{match['id']}_{int(datetime.now().timestamp())}_{safe_name}"
        file_path = os.path.join(upload_folder, saved_name)
        jd.save(file_path)
        conn.execute(
            """UPDATE requirements
               SET jd_filename=?, jd_url=?, jd_public_id=?, updated_at=datetime('now','localtime')
               WHERE id=?""",
            (safe_name, f"/uploads/{os.path.relpath(file_path, os.path.join(app.root_path, 'uploads')).replace(os.sep, '/')}", saved_name, match["id"])
        )
        result["attached"] += 1
    return result

def default_requirement_checks():
    return [
        "Technical Skills", "Years of Relevant Experience", "Within Given Budget",
        "Notice Period", "Location", "Non-Poachable Employee", "Updated CV"
    ]

def find_requirement_by_title_client(conn, title, client_name):
    title_key = normalize_requirement_key(title)
    client_key = normalize_requirement_client_key(client_name)
    for row in conn.execute("SELECT id,title,client_name FROM requirements").fetchall():
        if normalize_requirement_key(row["title"]) == title_key and normalize_requirement_client_key(row["client_name"]) == client_key:
            return row
    return None

def create_requirement_from_sheet(conn, title, client_name, created_by="Google Sheet Import"):
    title = str(title or "").strip()
    client_name = str(client_name or "").strip() or "Google Sheet Import"
    if not title:
        return None
    existing = find_requirement_by_title_client(conn, title, client_name)
    if existing:
        return existing
    conn.execute("INSERT OR IGNORE INTO clients (client_name) VALUES (?)", (client_name,))
    rid = conn.execute("""
        INSERT INTO requirements (title, description, client_name, status, created_by)
        VALUES (?,?,?,?,?)
    """, (title, "", client_name, "Open", created_by)).lastrowid
    for i, check_name in enumerate(default_requirement_checks()):
        conn.execute("""
            INSERT INTO requirement_checks
                (requirement_id,check_name,check_description,check_type,pass_criteria,sort_order)
            VALUES (?,?,?,?,?,?)
        """, (rid, check_name, "", "boolean", "Yes", i))
    return conn.execute("SELECT id,title,client_name FROM requirements WHERE id=?", (rid,)).fetchone()

def check_missing(row):
    m = []
    if not row.get("candidate_name"): m.append("name")
    if not row.get("phone") and not row.get("email_addr"): m.append("phone/email")
    elif not row.get("phone"):       m.append("phone")
    elif not row.get("email_addr"):  m.append("email")
    return m

EXCEL_REQUIRED_FIELDS = [
    ("candidate_name", "Candidate Name"),
    ("email_addr", "Email"),
    ("phone", "Phone"),
    ("current_company", "Current Company"),
    ("current_role", "Current Role"),
    ("experience_years", "Experience (Years)"),
    ("key_skills", "Key Skills"),
    ("notice_period", "Notice Period"),
    ("current_salary", "Current Salary"),
    ("expected_salary", "Expected Salary"),
    ("current_location", "Current Location"),
    ("preferred_location", "Preferred Location"),
]

def missing_excel_fields(row):
    return [label for field, label in EXCEL_REQUIRED_FIELDS if not str(row.get(field) or "").strip()]

# â”€â”€ Upload handler â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def process_upload(recruiter_name, recruiter_email, role_override, sourcer_id, excel_file, cv_file_list, jd_file_list=None):
    batch_id = hashlib.md5(f"{recruiter_email}{datetime.now().isoformat()}".encode()).hexdigest()[:10]
    result   = {"added":0,"duplicates":0,"missing":0,"errors":[],"cv_warnings":[],"jd_attached":0,"jd_warnings":[]}

    fname = excel_file.filename.lower()
    data  = excel_file.read()
    if   fname.endswith(".csv"):           rows, err = parse_csv(data, role_override)
    elif fname.endswith((".xlsx",".xls")): rows, err = parse_xlsx(data, role_override)
    else: return {"error": "Please upload a .xlsx or .csv file"}
    if err:   return {"error": err}
    if not rows: return {"error": "No rows found. Check the file has a header row with data below it."}

    conn = get_db()
    requirements, requirement_lookup = load_requirement_lookup(conn)
    if not requirements:
        conn.close()
        return {"error": "Create at least one Requirement before bulk upload. The bulk template now includes a Requirement column and every uploaded row must map to an existing Requirement."}

    saved_cvs = []
    for cv in cv_file_list:
        if cv and cv.filename:
            orig, url, pub_id, file_path = upload_cv(cv, batch_id)
            parsed_cv = {}
            if file_path:
                try:
                    parsed_cv = parse_cv(file_path) or {}
                except Exception as e:
                    print("Bulk CV parse error:", e)
            if orig:
                saved_cvs.append({"orig": orig, "url": url, "pub_id": pub_id, "parsed": parsed_cv})

    validation_errors = []
    seen = {}
    prepared_rows = []
    matched_ids = set()
    for idx, row in enumerate(rows, start=2):
        if not any(str(v).strip() for v in row.values()):
            continue
        cname = row.get("candidate_name", "").strip()
        raw_req_title = str(row.get("requirement_title") or row.get("role_name") or row.get("requirement") or "").strip()
        raw_req_client = str(row.get("client_name") or "").strip()
        req_id = resolve_requirement_id(row, requirement_lookup)
        if not req_id:
            suggestions = suggest_requirement_matches(raw_req_title, raw_req_client, requirement_lookup, limit=3)
            if suggestions:
                suggestion_text = "; ".join(s["label"] for s in suggestions if s.get("label"))
                validation_errors.append(
                    f"Row {idx}: Requirement '{raw_req_title or row.get('requirement_id') or row.get('role_name') or 'blank'}' did not match an existing Requirement. "
                    f"Closest matches: {suggestion_text}."
                )
            else:
                validation_errors.append(
                    f"Row {idx}: Requirement '{raw_req_title or row.get('requirement_id') or row.get('role_name') or 'blank'}' could not be found in ATS. "
                    "Check the title spelling, spacing, or client name."
                )
        missing_fields = missing_excel_fields(row)
        if missing_fields:
            validation_errors.append(f"Row {idx}: Missing required columns: {', '.join(missing_fields)}.")
        cv_orig, cv_url, cv_pub, cv_ok = match_cv(row, saved_cvs)
        if not cv_ok:
            validation_errors.append(f"Row {idx}: CV is required and could not be matched for {cname or 'candidate'}. Email or phone in CV should match the spreadsheet; filename matching is used only as fallback.")
        phone = norm_phone(row.get("phone", ""))
        email = (row.get("email_addr", "") or "").strip().lower()
        duplicate_keys = []
        if phone and len(phone) >= 8:
            duplicate_keys.append(("phone", req_id, phone))
        if email:
            duplicate_keys.append(("email", req_id, email))
        for key in duplicate_keys:
            if key in seen:
                validation_errors.append(f"Row {idx}: Duplicate {key[0]} within this upload for the same Requirement (also row {seen[key]}).")
            else:
                seen[key] = idx
        if req_id:
            is_dup, dup_row, dup_why = check_dup(conn, row)
            if is_dup:
                validation_errors.append(f"Row {idx}: {build_duplicate_candidate_message(row, dup_row, dup_why)}")
        row["_cv"] = (cv_orig, cv_url, cv_pub)
        prepared_rows.append(row)
        if cv_pub:
            matched_ids.add(cv_pub)

    if validation_errors:
        conn.close()
        return {
            "error": "Bulk upload blocked. Fix these issues and upload again:\n" + "\n".join(validation_errors[:20]),
            "errors": validation_errors,
        }

    def _bulk_write():
        conn = get_db()
        try:
            jd_result = save_bulk_jd_files(conn, jd_file_list or [], requirement_lookup, batch_id)
            result["jd_attached"] = jd_result["attached"]
            result["jd_warnings"] = jd_result["warnings"]

            conn.execute("INSERT OR IGNORE INTO team_members (name,email,is_fixed) VALUES (?,?,0)",
                         (recruiter_name, recruiter_email.lower()))

            for row in prepared_rows:
                if not any(str(v).strip() for v in row.values()):
                    continue
                cname = row.get("candidate_name", "").strip()
                cv_orig, cv_url, cv_pub = row.get("_cv", ("", "", ""))
                is_dup, dup_row, dup_why = check_dup(conn, row)
                dup_id = (dup_row or {}).get("id")
                missing = check_missing(row)

                cid = conn.execute("""INSERT INTO candidates
                    (upload_batch,recruiter_name,recruiter_email,sourcer_id,role_name,candidate_name,
                     email_addr,phone,current_company,current_role,experience_years,key_skills,
                     notice_period,current_salary,expected_salary,current_location,
                     preferred_location,remarks,cv_filename,cv_url,cv_public_id,cv_summary,
                     status,tags,is_duplicate,duplicate_of,missing_info,requirement_id)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (batch_id,recruiter_name,recruiter_email.lower(),sourcer_id,
                     row.get("role_name",""),cname,row.get("email_addr",""),row.get("phone",""),
                     row.get("current_company",""),row.get("current_role",""),
                     row.get("experience_years",""),row.get("key_skills",""),
                     row.get("notice_period",""),row.get("current_salary",""),
                     row.get("expected_salary",""),row.get("current_location",""),
                     row.get("preferred_location",""),row.get("remarks",""),
                     cv_orig,cv_url,cv_pub,row.get("cv_summary",""),"New","",
                     1 if is_dup else 0,dup_id,
                     ",".join(missing) if missing else None,
                     row.get("requirement_id"))).lastrowid
                conn.commit()

                if is_dup:
                    conn.execute("INSERT INTO alerts (alert_type,message,candidate_id,recruiter_email) VALUES (?,?,?,?)",
                        ("duplicate",f"Duplicate: {cname or '?'} matched via {dup_why} (original ID #{dup_id})",cid,recruiter_email))
                    result["duplicates"] += 1
                if missing:
                    conn.execute("INSERT INTO alerts (alert_type,message,candidate_id,recruiter_email) VALUES (?,?,?,?)",
                        ("missing_info",f"Missing {', '.join(missing)} for {cname or 'unnamed'}",cid,recruiter_email))
                    result["missing"] += 1
                result["added"] += 1
                if row.get("requirement_id") and (cv_url or cv_pub or cv_orig):
                    queue_candidate_ai_screening(cid, trigger="bulk_upload")

            for orig, url, pid in unmatched_cvs(saved_cvs, matched_ids):
                conn.execute("INSERT INTO alerts (alert_type,message,recruiter_email) VALUES (?,?,?)",
                    ("cv_mismatch",
                     f"CV '{orig}' by {recruiter_name} doesn't match any candidate in the Excel. Please re-upload with correct filename.",
                     recruiter_email))
                result["cv_warnings"].append(orig)

            conn.execute("""INSERT INTO upload_log
                (batch_id,recruiter_name,recruiter_email,filename,candidates_added,duplicates_found,missing_count)
                VALUES (?,?,?,?,?,?,?)""",
                (batch_id,recruiter_name,recruiter_email.lower(),excel_file.filename,
                 result["added"],result["duplicates"],result["missing"]))

            today = date.today().isoformat()
            submitted = {r[0] for r in conn.execute(
                "SELECT DISTINCT recruiter_email FROM candidates WHERE date(created_at)=?",(today,)).fetchall()}
            for member in conn.execute("SELECT name,email FROM team_members").fetchall():
                if member["email"] not in submitted:
                    if not conn.execute(
                        "SELECT 1 FROM alerts WHERE alert_type='no_submission' AND recruiter_email=? AND date(created_at)=?",
                        (member["email"],today)).fetchone():
                        conn.execute("INSERT INTO alerts (alert_type,message,recruiter_email) VALUES (?,?,?)",
                            ("no_submission",f"{member['name']} has not submitted any profiles today",member["email"]))
            conn.commit()
            return result
        finally:
            conn.close()

    return _with_db_write_retry(_bulk_write)

# â”€â”€ Weekly summary email â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def google_sheet_rows_from_request(data):
    sheet_url = data.get("sheet_url") or data.get("url") or ""
    selected_tabs = data.get("tabs") or []
    column_mapping = data.get("column_mapping") or {}
    if not isinstance(column_mapping, dict):
        column_mapping = {}
    if isinstance(selected_tabs, str):
        selected_tabs = [tab.strip() for tab in re.split(r"[,;\n]+", selected_tabs) if tab.strip()]
    sheet_id = extract_google_sheet_id(sheet_url)
    smart_import = truthy_flag(data.get("smart_import"), True)
    snapshot_import = truthy_flag(data.get("snapshot_import"), False)
    start_after_rows = google_sheet_state_rows(sheet_id) if smart_import else {}
    if not snapshot_import:
        rows, sheet_names, mappings, api_err = read_google_sheet_rows_api(
            sheet_url, selected_tabs, data.get("role_override") or "", column_mapping, start_after_rows
        )
        if not api_err:
            return rows, sheet_names, mappings, None
        api_err_text = str(api_err)
        if ("403" in api_err_text or "PERMISSION_DENIED" in api_err_text) and "SERVICE_DISABLED" not in api_err_text and "sheets.googleapis.com" not in api_err_text:
            return None, sheet_names or [], mappings or [], (
                "Google denied access to this Sheet. Reconnect Google Permissions using the same Google account that can open the Sheet, "
                "or share the Sheet with that ATS Google account. Team members viewing or editing the Sheet at the same time does not cause this permission error. "
                + api_err_text
            )
    else:
        api_err_text = ""
    file_bytes, err = download_google_sheet_xlsx(sheet_url)
    if err:
        prefix = (api_err_text + " ") if api_err_text else ""
        return None, [], [], prefix + err
    rows, sheet_names, mappings, err = parse_xlsx_sheets(
        file_bytes, selected_tabs, data.get("role_override") or "", column_mapping
    )
    if err:
        return None, sheet_names, mappings, err
    if smart_import and start_after_rows:
        rows = [
            row for row in rows
            if int(row.get("_row_number") or 0) > int(start_after_rows.get(str(row.get("_sheet_name") or "").strip().lower(), 0) or 0)
        ]
    return rows, sheet_names, mappings, None

def local_candidate_sheet_rows_from_request(req):
    upload = req.files.get("local_excel_file") or req.files.get("file")
    if not upload or not upload.filename:
        return None, [], [], "Please select a local Excel candidate sheet."
    filename = upload.filename.lower()
    if not filename.endswith((".xlsx", ".xls")):
        return None, [], [], "Please upload a .xlsx or .xls file."
    selected_tabs = req.form.get("tabs") or ""
    selected_tabs = [tab.strip() for tab in re.split(r"[,;\n]+", selected_tabs) if tab.strip()]
    try:
        column_mapping = json.loads(req.form.get("column_mapping") or "{}")
    except Exception:
        column_mapping = {}
    if not isinstance(column_mapping, dict):
        column_mapping = {}
    rows, sheet_names, mappings, err = parse_xlsx_sheets(
        upload.read(), selected_tabs, req.form.get("role_override") or "", column_mapping
    )
    return rows, sheet_names, mappings, err

def candidate_sheet_preview_response(data, rows, sheet_names, mappings):
    conn = get_db()
    prepared, summary = prepare_google_sheet_candidates(conn, rows or [], data.get("default_client_name") or data.get("client_name") or "")
    conn.close()
    if summary.get("error"):
        return jsonify({"error": summary["error"], "sheet_names": sheet_names, "column_mappings": mappings, "field_options": ats_field_options()}), 400
    sample = [{
        "sheet": row.get("_sheet_name") or "",
        "row": row.get("_row_number") or "",
        "candidate_name": row.get("candidate_name") or "",
        "email_addr": row.get("email_addr") or "",
        "phone": row.get("phone") or "",
        "role": row.get("role_name") or row.get("current_role") or "",
        "requirement_id": row.get("requirement_id") or "",
        "status": row.get("_skip_reason") or "Ready",
    } for row in prepared[:20]]
    return jsonify({
        "ok": True,
        "sheet_names": sheet_names,
        "summary": summary,
        "sample": sample,
        "column_mappings": mappings,
        "field_options": ats_field_options(),
    })

def prepare_google_sheet_candidates(conn, rows, default_client_name="", create_requirements=False, created_by="Google Sheet Import"):
    _, requirement_lookup = load_requirement_lookup(conn)
    prepared = []
    summary = {"total_rows": 0, "eligible": 0, "duplicates": 0, "missing_requirement": 0, "missing_contact": 0, "skipped_blank": 0, "new_requirements": 0}
    planned_requirements = set()
    for row in rows:
        if not any(str(v).strip() for k, v in row.items() if not str(k).startswith("_")):
            summary["skipped_blank"] += 1
            continue
        summary["total_rows"] += 1
        req_id = resolve_requirement_id(row, requirement_lookup)
        if not req_id:
            title = (row.get("requirement_title") or row.get("role_name") or row.get("current_role") or "").strip()
            client_name = (row.get("client_name") or default_client_name or "Google Sheet Import").strip()
            if title:
                existing = find_requirement_by_title_client(conn, title, client_name)
                if existing:
                    row["requirement_id"] = existing["id"]
                    row["role_name"] = row.get("role_name") or existing["title"]
                    req_id = existing["id"]
                elif create_requirements:
                    created = create_requirement_from_sheet(conn, title, client_name, created_by)
                    row["requirement_id"] = created["id"]
                    row["role_name"] = row.get("role_name") or created["title"]
                    req_id = created["id"]
                    summary["new_requirements"] += 1
                else:
                    row["_will_create_requirement"] = f"{title} - {client_name}"
                    planned_requirements.add((normalize_requirement_key(title), normalize_requirement_client_key(client_name)))
                    row["_skip_reason"] = f"Requirement not found: {title} - {client_name}. Create the Requirement first, then upload again."
        if not req_id:
            row["_skip_reason"] = row.get("_skip_reason") or "Missing Requirement. Create/map an existing Requirement before upload."
            summary["missing_requirement"] += 1
            prepared.append(row)
            continue
        if not row.get("candidate_name") or (not row.get("email_addr") and not row.get("phone")):
            row["_skip_reason"] = "Missing candidate name or email/phone"
            summary["missing_contact"] += 1
            prepared.append(row)
            continue
        is_dup, dup_row, dup_why = check_dup(conn, row)
        if is_dup:
            row["_skip_reason"] = build_duplicate_candidate_message(row, dup_row, dup_why)
            summary["duplicates"] += 1
            prepared.append(row)
            continue
        row["_skip_reason"] = ""
        summary["eligible"] += 1
        prepared.append(row)
    if not create_requirements:
        summary["new_requirements"] = len(planned_requirements)
    return prepared, summary

@app.route("/api/google_sheet/preview", methods=["POST"])
@login_required
def api_google_sheet_preview():
    if not has_bulk_upload_access():
        return jsonify({"error": "Bulk upload permission required"}), 403
    data = request.get_json(silent=True) or {}
    rows, sheet_names, mappings, err = google_sheet_rows_from_request(data)
    if err:
        return jsonify({"error": err}), 400
    return candidate_sheet_preview_response(data, rows, sheet_names, mappings)

@app.route("/api/local_candidate_sheet/preview", methods=["POST"])
@login_required
def api_local_candidate_sheet_preview():
    if not has_bulk_upload_access():
        return jsonify({"error": "Bulk upload permission required"}), 403
    rows, sheet_names, mappings, err = local_candidate_sheet_rows_from_request(request)
    if err:
        return jsonify({"error": err}), 400
    data = {
        "default_client_name": request.form.get("default_client_name") or request.form.get("client_name") or "",
        "client_name": request.form.get("client_name") or "",
    }
    return candidate_sheet_preview_response(data, rows, sheet_names, mappings)

@app.route("/api/google_sheet/import", methods=["POST"])
@login_required
def api_google_sheet_import():
    if not has_bulk_upload_access():
        return jsonify({"error": "Bulk upload permission required"}), 403
    data = request.get_json(silent=True) or {}
    try:
        offset = max(0, int(data.get("offset") or 0))
    except Exception:
        offset = 0
    try:
        limit = int(data.get("limit") or 100)
    except Exception:
        limit = 100
    limit = max(1, min(limit, 300))
    smart_import = truthy_flag(data.get("smart_import"), True)
    sheet_id = extract_google_sheet_id(data.get("sheet_url") or data.get("url") or "")
    rows, sheet_names, mappings, err = google_sheet_rows_from_request(data)
    if err:
        return jsonify({"error": err}), 400
    sourcer_id = session.get("team_member_id")
    recruiter_name = (data.get("recruiter_name") or session.get("recruiter_name") or session.get("username") or "System").strip()
    recruiter_email = (data.get("recruiter_email") or session.get("recruiter_email") or session.get("email") or "system@hrguru.com").strip().lower()
    conn = get_db()
    prepared, summary = prepare_google_sheet_candidates(
        conn,
        rows or [],
        data.get("default_client_name") or data.get("client_name") or "",
        create_requirements=False,
        created_by=session.get("username") or session.get("email") or "Google Sheet Import"
    )
    if summary.get("error"):
        conn.close()
        return jsonify({"error": summary["error"], "sheet_names": sheet_names}), 400
    total_ready = len([row for row in prepared if not row.get("_skip_reason")])
    batch_id = (data.get("batch_id") or "").strip()
    if not batch_id:
        sheet_id = extract_google_sheet_id(data.get("sheet_url") or data.get("url") or "")
        batch_seed = json.dumps({
            "sheet_id": sheet_id,
            "tabs": data.get("tabs") or "",
            "client": data.get("default_client_name") or data.get("client_name") or "",
            "mapping": data.get("column_mapping") or {},
            "user": recruiter_email,
        }, sort_keys=True)
        batch_id = "google-sheet-" + hashlib.md5(batch_seed.encode()).hexdigest()[:12]
    chunk_rows = prepared[offset:offset + limit]
    next_offset = offset + len(chunk_rows)
    has_more = next_offset < len(prepared)
    imported = 0
    skipped = []
    conn.execute("INSERT OR IGNORE INTO team_members (name,email,is_fixed) VALUES (?,?,0)", (recruiter_name, recruiter_email))
    for row in chunk_rows:
        if row.get("_skip_reason"):
            skipped.append({"sheet": row.get("_sheet_name"), "row": row.get("_row_number"), "candidate_name": row.get("candidate_name"), "reason": row.get("_skip_reason")})
            continue
        row_marker = row_sheet_position_key(row)
        if row_marker and conn.execute(
            "SELECT 1 FROM candidates WHERE upload_batch=? AND remarks LIKE ? LIMIT 1",
            (batch_id, f"%Google Sheet Row: {row_marker}%")
        ).fetchone():
            skipped.append({"sheet": row.get("_sheet_name"), "row": row.get("_row_number"), "candidate_name": row.get("candidate_name"), "reason": "Already imported in this Google Sheet batch"})
            continue
        missing = missing_excel_fields(row)
        if "CV" not in missing:
            missing.append("CV")
        row_recruiter_name = (row.get("recruiter_name") or recruiter_name).strip()
        row_status = (row.get("status") or "New").strip()
        row_created_at = normalize_candidate_created_at(row.get("created_at")) or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row_remarks = (row.get("remarks") or "").strip()
        if row_marker:
            row_remarks = (row_remarks + "\n" if row_remarks else "") + f"Google Sheet Row: {row_marker}"
        cid = conn.execute("""INSERT INTO candidates
            (upload_batch,recruiter_name,recruiter_email,sourcer_id,role_name,candidate_name,
             email_addr,phone,current_company,current_role,experience_years,key_skills,
             notice_period,current_salary,expected_salary,current_location,
             preferred_location,remarks,cv_filename,cv_url,cv_public_id,cv_summary,
             status,industry_domain,tags,is_duplicate,duplicate_of,missing_info,requirement_id,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (batch_id, row_recruiter_name, recruiter_email, sourcer_id,
             row.get("role_name",""), row.get("candidate_name",""), row.get("email_addr",""), row.get("phone",""),
             row.get("current_company",""), row.get("current_role",""), row.get("experience_years",""), row.get("key_skills",""),
             row.get("notice_period",""), row.get("current_salary",""), row.get("expected_salary",""), row.get("current_location",""),
             row.get("preferred_location",""), row_remarks, "", "", "", row.get("cv_summary",""), row_status,
             row.get("industry_domain",""), "google_sheet", 0, None, ",".join(missing) if missing else None,
             row.get("requirement_id"), row_created_at)).lastrowid
        if missing:
            conn.execute("INSERT INTO alerts (alert_type,message,candidate_id,recruiter_email) VALUES (?,?,?,?)",
                ("missing_info", f"Google Sheet import missing {', '.join(missing)} for {row.get('candidate_name') or 'unnamed'}", cid, recruiter_email))
        imported += 1
        if row.get("requirement_id"):
            queue_candidate_ai_screening(cid, trigger="google_sheet_import")
    if smart_import:
        update_google_sheet_import_state(conn, sheet_id, chunk_rows, recruiter_email)
    conn.execute("""INSERT INTO upload_log
        (batch_id,recruiter_name,recruiter_email,filename,candidates_added,duplicates_found,missing_count)
        VALUES (?,?,?,?,?,?,?)""",
        (batch_id, recruiter_name, recruiter_email, "Google Sheet Import", imported, summary.get("duplicates", 0), 0))
    conn.commit()
    conn.close()
    return jsonify({
        "ok": True,
        "imported": imported,
        "skipped": skipped[:50],
        "summary": summary,
        "sheet_names": sheet_names,
        "batch_id": batch_id,
        "offset": offset,
        "next_offset": 0 if smart_import else next_offset,
        "limit": limit,
        "processed": len(chunk_rows),
        "total_rows": len(prepared),
        "total_ready": total_ready,
        "has_more": has_more,
        "smart_import": smart_import,
    })

@app.route("/api/local_candidate_sheet/import", methods=["POST"])
@login_required
def api_local_candidate_sheet_import():
    if not has_bulk_upload_access():
        return jsonify({"error": "Bulk upload permission required"}), 403
    try:
        offset = max(0, int(request.form.get("offset") or 0))
    except Exception:
        offset = 0
    try:
        limit = int(request.form.get("limit") or 100)
    except Exception:
        limit = 100
    limit = max(1, min(limit, 300))
    rows, sheet_names, mappings, err = local_candidate_sheet_rows_from_request(request)
    if err:
        return jsonify({"error": err}), 400
    sourcer_id = session.get("team_member_id")
    recruiter_name = (request.form.get("recruiter_name") or session.get("recruiter_name") or session.get("username") or "System").strip()
    recruiter_email = (request.form.get("recruiter_email") or session.get("recruiter_email") or session.get("email") or "system@hrguru.com").strip().lower()
    conn = get_db()
    prepared, summary = prepare_google_sheet_candidates(
        conn,
        rows or [],
        request.form.get("default_client_name") or request.form.get("client_name") or "",
        create_requirements=False,
        created_by=session.get("username") or session.get("email") or "Local Excel Import"
    )
    if summary.get("error"):
        conn.close()
        return jsonify({"error": summary["error"], "sheet_names": sheet_names}), 400
    total_ready = len([row for row in prepared if not row.get("_skip_reason")])
    batch_id = (request.form.get("batch_id") or "").strip()
    upload = request.files.get("local_excel_file") or request.files.get("file")
    filename = upload.filename if upload and upload.filename else "Local Candidate Sheet"
    if not batch_id:
        batch_seed = json.dumps({
            "filename": filename,
            "tabs": request.form.get("tabs") or "",
            "client": request.form.get("default_client_name") or request.form.get("client_name") or "",
            "mapping": request.form.get("column_mapping") or "{}",
            "user": recruiter_email,
        }, sort_keys=True)
        batch_id = "local-sheet-" + hashlib.md5(batch_seed.encode()).hexdigest()[:12]
    chunk_rows = prepared[offset:offset + limit]
    next_offset = offset + len(chunk_rows)
    has_more = next_offset < len(prepared)
    imported = 0
    skipped = []
    conn.execute("INSERT OR IGNORE INTO team_members (name,email,is_fixed) VALUES (?,?,0)", (recruiter_name, recruiter_email))
    for row in chunk_rows:
        if row.get("_skip_reason"):
            skipped.append({"sheet": row.get("_sheet_name"), "row": row.get("_row_number"), "candidate_name": row.get("candidate_name"), "reason": row.get("_skip_reason")})
            continue
        row_marker = row_sheet_position_key(row)
        if row_marker and conn.execute(
            "SELECT 1 FROM candidates WHERE upload_batch=? AND remarks LIKE ? LIMIT 1",
            (batch_id, f"%Local Excel Row: {row_marker}%")
        ).fetchone():
            skipped.append({"sheet": row.get("_sheet_name"), "row": row.get("_row_number"), "candidate_name": row.get("candidate_name"), "reason": "Already imported in this local Excel batch"})
            continue
        missing = missing_excel_fields(row)
        if "CV" not in missing:
            missing.append("CV")
        row_recruiter_name = (row.get("recruiter_name") or recruiter_name).strip()
        row_status = (row.get("status") or "New").strip()
        row_created_at = normalize_candidate_created_at(row.get("created_at")) or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row_remarks = (row.get("remarks") or "").strip()
        if row_marker:
            row_remarks = (row_remarks + "\n" if row_remarks else "") + f"Local Excel Row: {row_marker}"
        cid = conn.execute("""INSERT INTO candidates
            (upload_batch,recruiter_name,recruiter_email,sourcer_id,role_name,candidate_name,
             email_addr,phone,current_company,current_role,experience_years,key_skills,
             notice_period,current_salary,expected_salary,current_location,
             preferred_location,remarks,cv_filename,cv_url,cv_public_id,cv_summary,
             status,industry_domain,tags,is_duplicate,duplicate_of,missing_info,requirement_id,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (batch_id, row_recruiter_name, recruiter_email, sourcer_id,
             row.get("role_name",""), row.get("candidate_name",""), row.get("email_addr",""), row.get("phone",""),
             row.get("current_company",""), row.get("current_role",""), row.get("experience_years",""), row.get("key_skills",""),
             row.get("notice_period",""), row.get("current_salary",""), row.get("expected_salary",""), row.get("current_location",""),
             row.get("preferred_location",""), row_remarks, "", "", "", row.get("cv_summary",""), row_status,
             row.get("industry_domain",""), "local_excel", 0, None, ",".join(missing) if missing else None,
             row.get("requirement_id"), row_created_at)).lastrowid
        if missing:
            conn.execute("INSERT INTO alerts (alert_type,message,candidate_id,recruiter_email) VALUES (?,?,?,?)",
                ("missing_info", f"Local Excel import missing {', '.join(missing)} for {row.get('candidate_name') or 'unnamed'}", cid, recruiter_email))
        imported += 1
        if row.get("requirement_id"):
            queue_candidate_ai_screening(cid, trigger="local_excel_import")
    conn.execute("""INSERT INTO upload_log
        (batch_id,recruiter_name,recruiter_email,filename,candidates_added,duplicates_found,missing_count)
        VALUES (?,?,?,?,?,?,?)""",
        (batch_id, recruiter_name, recruiter_email, filename, imported, summary.get("duplicates", 0), 0))
    conn.commit()
    conn.close()
    return jsonify({
        "ok": True,
        "imported": imported,
        "skipped": skipped[:50],
        "summary": summary,
        "sheet_names": sheet_names,
        "batch_id": batch_id,
        "offset": offset,
        "next_offset": next_offset,
        "limit": limit,
        "processed": len(chunk_rows),
        "total_rows": len(prepared),
        "total_ready": total_ready,
        "has_more": has_more,
        "smart_import": False,
    })

def build_weekly_summary():
    conn  = get_db()
    since = (date.today() - timedelta(days=7)).isoformat()
    today = date.today().isoformat()

    total   = conn.execute("SELECT COUNT(*) FROM candidates WHERE date(created_at)>=? AND is_duplicate=0",(since,)).fetchone()[0]
    dups    = conn.execute("SELECT COUNT(*) FROM candidates WHERE date(created_at)>=? AND is_duplicate=1",(since,)).fetchone()[0]
    missing = conn.execute("SELECT COUNT(*) FROM candidates WHERE date(created_at)>=? AND missing_info IS NOT NULL",(since,)).fetchone()[0]

    # Per recruiter
    perf = conn.execute("""
        SELECT recruiter_name, COUNT(*) as total,
               SUM(CASE WHEN is_duplicate=0 THEN 1 ELSE 0 END) as unique_c
        FROM candidates WHERE date(created_at)>=?
        GROUP BY recruiter_email ORDER BY unique_c DESC""",(since,)).fetchall()

    # Status breakdown
    statuses = conn.execute("""
        SELECT status, COUNT(*) as cnt FROM candidates
        WHERE is_duplicate=0 GROUP BY status ORDER BY cnt DESC""").fetchall()

    # Who didn't submit this week
    all_team  = conn.execute("SELECT name,email FROM team_members WHERE is_fixed=1").fetchall()
    submitted = {r[0] for r in conn.execute(
        "SELECT DISTINCT recruiter_email FROM candidates WHERE date(created_at)>=?",(since,)).fetchall()}
    no_sub = [m["name"] for m in all_team if m["email"] not in submitted]

    conn.close()

    # Build HTML email
    perf_rows = "".join(f"<tr><td>{r['recruiter_name']}</td><td>{r['total']}</td><td>{r['unique_c']}</td></tr>" for r in perf)
    status_rows = "".join(f"<tr><td>{r['status']}</td><td>{r['cnt']}</td></tr>" for r in statuses)
    no_sub_txt = ", ".join(no_sub) if no_sub else "Everyone submitted this week âœ“"

    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;background:#f9f9f9;padding:24px;border-radius:8px">
      <h2 style="color:#e8643a;margin-bottom:4px">HR Guru ATS â€” Weekly Summary</h2>
      <p style="color:#666;font-size:13px">Week ending {date.today().strftime('%d %B %Y')}</p>

      <div style="display:flex;gap:12px;margin:20px 0">
        <div style="background:#fff;border-radius:8px;padding:16px 20px;flex:1;text-align:center;border:1px solid #eee">
          <div style="font-size:28px;font-weight:800;color:#e8643a">{total}</div>
          <div style="font-size:12px;color:#888">Candidates Added</div>
        </div>
        <div style="background:#fff;border-radius:8px;padding:16px 20px;flex:1;text-align:center;border:1px solid #eee">
          <div style="font-size:28px;font-weight:800;color:#e83a3a">{dups}</div>
          <div style="font-size:12px;color:#888">Duplicates</div>
        </div>
        <div style="background:#fff;border-radius:8px;padding:16px 20px;flex:1;text-align:center;border:1px solid #eee">
          <div style="font-size:28px;font-weight:800;color:#e8c53a">{missing}</div>
          <div style="font-size:12px;color:#888">Missing Info</div>
        </div>
      </div>

      <h3 style="color:#333;font-size:14px;margin-bottom:8px">Recruiter Performance</h3>
      <table style="width:100%;border-collapse:collapse;font-size:13px;background:#fff;border-radius:8px;overflow:hidden;margin-bottom:20px">
        <thead><tr style="background:#1c2030;color:#e8643a">
          <th style="padding:8px 12px;text-align:left">Recruiter</th>
          <th style="padding:8px 12px;text-align:left">Submitted</th>
          <th style="padding:8px 12px;text-align:left">Unique</th>
        </tr></thead>
        <tbody>{perf_rows if perf_rows else '<tr><td colspan="3" style="padding:8px 12px;color:#888">No submissions this week</td></tr>'}</tbody>
      </table>

      <h3 style="color:#333;font-size:14px;margin-bottom:8px">Pipeline Status (All Time)</h3>
      <table style="width:100%;border-collapse:collapse;font-size:13px;background:#fff;border-radius:8px;overflow:hidden;margin-bottom:20px">
        <thead><tr style="background:#1c2030;color:#e8643a">
          <th style="padding:8px 12px;text-align:left">Status</th>
          <th style="padding:8px 12px;text-align:left">Count</th>
        </tr></thead>
        <tbody>{status_rows}</tbody>
      </table>

      <div style="background:#fff;border-radius:8px;padding:14px 16px;border:1px solid #eee;font-size:13px;margin-bottom:20px">
        <strong>Not submitted this week:</strong><br/>
        <span style="color:#e8c53a">{no_sub_txt}</span>
      </div>

      <p style="font-size:11px;color:#aaa;text-align:center">HR Guru ATS Â· Auto-generated weekly report</p>
    </div>"""
    return html

def send_weekly_email():
    if not GMAIL_USER or not GMAIL_APP_PASS or not ADMIN_EMAIL:
        return {"error": "Gmail not configured â€” set GMAIL_USER, GMAIL_APP_PASS, ADMIN_EMAIL in wsgi.py"}
    try:
        html = build_weekly_summary()
        msg  = MIMEMultipart("alternative")
        msg["Subject"] = f"HR Guru Weekly Summary â€” {date.today().strftime('%d %b %Y')}"
        msg["From"]    = GMAIL_USER
        msg["To"]      = ADMIN_EMAIL
        msg.attach(MIMEText(html, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(GMAIL_USER, GMAIL_APP_PASS)
            s.sendmail(GMAIL_USER, ADMIN_EMAIL, msg.as_string())
        return {"ok": True, "sent_to": ADMIN_EMAIL}
    except Exception as e:
        return {"error": str(e)}

# â”€â”€ Template â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def make_template():
    if not XLSX_OK: return None
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Candidates"
    headers = ["Requirement","Candidate Name","Email","Phone","Current Company","Current Role","Education",
               "Experience (Years)","Key Skills","Notice Period","Current Salary",
               "Expected Salary","Current Location","Preferred Location","Role","Remarks"]
    hfill = PatternFill("solid",fgColor="1C2030"); hfont = Font(bold=True,color="E8643A")
    for i,h in enumerate(headers,1):
        c = ws.cell(row=1,column=i,value=h); c.font=hfont; c.fill=hfill
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[c.column_letter].width = max(len(h)+4,18)
    hints = ["â† Full name","Email address","10-digit mobile","Current employer","Job title","Degree / qualification",
             "e.g. 5 years","Comma separated","e.g. 30 days","e.g. 12 LPA","e.g. 18 LPA",
             "Current city","Preferred city","Role being screened","Any notes"]
    ifill = PatternFill("solid",fgColor="0D1017"); ifont = Font(italic=True,color="6B7494")
    hints = ["Existing requirement title or ID","Full name","Email address","10-digit mobile",
             "Current employer","Job title","Degree / qualification","e.g. 5 years","Comma separated","e.g. 30 days",
             "e.g. 12 LPA","e.g. 18 LPA","Current city","Preferred city","Role being screened","Any notes"]
    for i,h in enumerate(hints,1):
        c = ws.cell(row=2,column=i,value=h); c.font=ifont; c.fill=ifill
    try:
        conn = get_db()
        reqs = conn.execute("SELECT id,title,client_name,status FROM requirements ORDER BY title").fetchall()
        conn.close()
        req_ws = wb.create_sheet("Requirements")
        req_headers = ["Requirement ID", "Requirement Title", "Client", "Status"]
        for i,h in enumerate(req_headers,1):
            c = req_ws.cell(row=1,column=i,value=h); c.font=hfont; c.fill=hfill
            req_ws.column_dimensions[c.column_letter].width=max(len(h)+6,18)
        for ri,r in enumerate(reqs,2):
            req_ws.cell(row=ri,column=1,value=r["id"])
            req_ws.cell(row=ri,column=2,value=r["title"])
            req_ws.cell(row=ri,column=3,value=r["client_name"])
            req_ws.cell(row=ri,column=4,value=r["status"])
    except Exception as e:
        print("Template requirements sheet error:", e)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# â”€â”€ Export â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
EXPORT_COLS = ["candidate_name","email_addr","phone","current_company","current_role",
               "experience_years","key_skills","notice_period","current_salary",
               "expected_salary","current_location","preferred_location",
               "role_name","status","tags","recruiter_name","remarks","created_at"]
EXPORT_HDR  = ["Candidate Name","Email","Phone","Current Company","Current Role",
               "Experience","Key Skills","Notice Period","Current Salary","Expected Salary",
               "Current Location","Preferred Location","Role","Status","Tags",
               "Recruiter","Remarks","Date Added"]

def build_query(args, current_session=None, select_clause=None):
    role=args.get("role",""); sender=(args.get("sender","") or "").strip().lower(); loc=args.get("location","")
    notice=args.get("notice",""); q=args.get("q",""); show_d=args.get("show_dups","1")
    status=args.get("status",""); tag=args.get("tag","")
    exp_min=args.get("exp_min",""); exp_max=args.get("exp_max","")
    sal_min=args.get("sal_min",""); sal_max=args.get("sal_max","")
    sort=args.get("sort","newest")
    req_id=args.get("requirement_id","")
    req_ids=args.get("requirement_ids","")
    client=args.get("client","")
    client_like=args.get("client_like","")
    date_preset=args.get("date_preset","")
    date_from=args.get("date_from","")
    date_to=args.get("date_to","")
    name_filter=args.get("name","")
    requirement_filter=args.get("requirement","")
    recruiter_name_filter=args.get("recruiter_name","")
    current_role_filter=args.get("current_role","")
    phone_filter=args.get("phone","")
    email_filter=args.get("email","")
    names_filter=args.get("names","")
    phones_filter=args.get("phones","")
    emails_filter=args.get("emails","")
    created_at_filter=args.get("created_at","")
    select_clause = select_clause or "c.*, r.title as requirement_title, r.client_name as client_name"
    sql=f"SELECT {select_clause} FROM candidates c LEFT JOIN requirements r ON c.requirement_id = r.id WHERE 1=1"; p=[]
    
    # Filter by sourcer_id for non-admin users
    owner_sql, owner_params = non_admin_candidate_owner_clause(current_session, "c")
    sql += owner_sql
    p.extend(owner_params)
        
    if role:   sql+=" AND role_name=?";   p.append(role)
    if sender: sql+=" AND lower(c.recruiter_email)=?"; p.append(sender)
    if loc:    sql+=" AND (current_location LIKE ? OR preferred_location LIKE ?)"; p+=[f"%{loc}%"]*2
    if notice: sql+=" AND notice_period LIKE ?"; p.append(f"%{notice}%")
    if status: sql+=" AND c.status=?"; p.append(status)
    if tag:    sql+=" AND (',' || tags || ',') LIKE ?"; p.append(f"%,{tag},%")
    if req_ids:
        id_list = [x.strip() for x in str(req_ids).split(",") if x.strip().isdigit()]
        if id_list:
            sql += " AND c.requirement_id IN (" + ",".join("?" * len(id_list)) + ")"
            p.extend(id_list)
    elif req_id:
        sql+=" AND requirement_id=?"; p.append(req_id)
    if client: sql+=" AND r.client_name=?"; p.append(client)
    if client_like: sql+=" AND r.client_name LIKE ?"; p.append(f"%{client_like}%")
    if date_preset == "today":
        sql+=" AND date(c.created_at)=date('now','localtime')"
    elif date_preset == "yesterday":
        sql+=" AND date(c.created_at)=date('now','localtime','-1 day')"
    else:
        if date_from:
            sql+=" AND date(c.created_at)>=?"
            p.append(date_from)
        if date_to:
            sql+=" AND date(c.created_at)<=?"
            p.append(date_to)
    if show_d=="0": sql+=" AND is_duplicate=0"
    # Search by name, skills, company, role, email, phone
    if q:
        q_text = str(q or "").strip()
        if any(sep in q_text for sep in [",", ";", "\n", "\r", "\t"]):
            term_clauses = []
            for term in split_multi_value_terms(q_text):
                term_sql_bits = []
                term_l = term.lower()
                term_phone = norm_phone(term)
                if term_phone and len(term_phone) >= 8 and term_phone.isdigit():
                    term_sql_bits.append("replace(replace(replace(replace(replace(lower(COALESCE(phone,'')),'+',''),'-',''),' ',''),'(',''),')','') LIKE ?")
                    p.append(f"%{term_phone}%")
                if "@" in term_l:
                    term_sql_bits.append("LOWER(COALESCE(email_addr,'')) LIKE ?")
                    p.append(f"%{term_l}%")
                if not term_phone or not term_phone.isdigit():
                    term_sql_bits.append("(candidate_name LIKE ? OR key_skills LIKE ? OR current_company LIKE ? OR current_role LIKE ? OR email_addr LIKE ? OR phone LIKE ?)")
                    p.extend([f"%{term}%"] * 6)
                if term_sql_bits:
                    term_clauses.append("(" + " OR ".join(term_sql_bits) + ")")
            if term_clauses:
                sql += " AND (" + " OR ".join(term_clauses) + ")"
        else:
            sql+=" AND (candidate_name LIKE ? OR key_skills LIKE ? OR current_company LIKE ? OR current_role LIKE ? OR email_addr LIKE ? OR phone LIKE ?)"
            p += [f"%{q}%"]*6
    if name_filter:
        sql+=" AND c.candidate_name LIKE ?"
        p.append(f"%{name_filter}%")
    if requirement_filter:
        sql+=" AND r.title LIKE ?"
        p.append(f"%{requirement_filter}%")
    if recruiter_name_filter:
        sql+=" AND c.recruiter_name LIKE ?"
        p.append(f"%{recruiter_name_filter}%")
    if current_role_filter:
        sql+=" AND c.current_role LIKE ?"
        p.append(f"%{current_role_filter}%")
    if phone_filter:
        sql+=" AND c.phone LIKE ?"
        p.append(f"%{phone_filter}%")
    if email_filter:
        sql+=" AND c.email_addr LIKE ?"
        p.append(f"%{email_filter}%")
    identity_sql, identity_params = build_candidate_identity_filter_clause(names_filter, phones_filter, emails_filter)
    if identity_sql:
        sql += identity_sql
        p.extend(identity_params)
    if created_at_filter:
        sql+=" AND c.created_at LIKE ?"
        p.append(f"%{created_at_filter}%")
    # Skills filter
    skills=args.get("skills","")
    if skills:
        skill_list = [s.strip().lower() for s in skills.split(",") if s.strip()]
        for sk in skill_list:
            sql+=" AND LOWER(key_skills) LIKE ?"
            p.append(f"%{sk}%")
    # Experience range filter (e.g. "6-10")
    exp_range=args.get("exp_range","")
    if exp_range and '-' in exp_range:
        parts = exp_range.split('-')
        exp_min = parts[0].strip() if parts[0].strip() else None
        exp_max = parts[1].strip() if len(parts) > 1 and parts[1].strip() else None
        # Extract numeric from experience_years (handle "6 Years", "6 years", "6", etc)
        if exp_min:
            sql+=" AND CAST(replace(replace(replace(replace(lower(experience_years),'years',''),'year',''),' ',''),'yrs','') AS REAL) >= ?"
            p.append(float(exp_min))
        if exp_max:
            sql+=" AND CAST(replace(replace(replace(replace(lower(experience_years),'years',''),'year',''),' ',''),'yrs','') AS REAL) <= ?"
            p.append(float(exp_max))
    # Sorting
    sort_map = {
        "newest": "c.created_at DESC",
        "oldest": "c.created_at ASC",
        "created_desc": "c.created_at DESC",
        "created_asc": "c.created_at ASC",
        "date_desc": "c.created_at DESC",
        "date_asc": "c.created_at ASC",
        "requirement_asc": "LOWER(r.title) ASC",
        "requirement_desc": "LOWER(r.title) DESC",
        "name": "LOWER(c.candidate_name) ASC",
        "name_asc": "LOWER(c.candidate_name) ASC",
        "name_desc": "LOWER(c.candidate_name) DESC",
        "role_asc": "LOWER(c.current_role) ASC",
        "role_desc": "LOWER(c.current_role) DESC",
        "client_asc": "LOWER(r.client_name) ASC",
        "client_desc": "LOWER(r.client_name) DESC",
        "company_asc": "LOWER(c.current_company) ASC",
        "company_desc": "LOWER(c.current_company) DESC",
        "email_asc": "LOWER(c.email_addr) ASC",
        "email_desc": "LOWER(c.email_addr) DESC",
        "phone_asc": "c.phone ASC",
        "phone_desc": "c.phone DESC",
        "status": "LOWER(c.status) ASC",
        "status_asc": "LOWER(c.status) ASC",
        "status_desc": "LOWER(c.status) DESC",
        "experience_asc": "CAST(replace(replace(replace(replace(lower(c.experience_years),'years',''),'year',''),' ',''),'yrs','') AS REAL) ASC",
        "experience_desc": "CAST(replace(replace(replace(replace(lower(c.experience_years),'years',''),'year',''),' ',''),'yrs','') AS REAL) DESC",
        "location_asc": "LOWER(c.current_location) ASC",
        "location_desc": "LOWER(c.current_location) DESC",
        "notice_asc": "LOWER(c.notice_period) ASC",
        "notice_desc": "LOWER(c.notice_period) DESC",
        "recruiter_asc": "LOWER(c.recruiter_name) ASC",
        "recruiter_desc": "LOWER(c.recruiter_name) DESC",
    }
    sql += " ORDER BY " + sort_map.get(sort, "c.created_at DESC")
    return sql, p

def strip_order_by(sql):
    marker = " ORDER BY "
    idx = sql.upper().rfind(marker)
    return sql[:idx] if idx >= 0 else sql

def parse_positive_int(value, default, max_value=None):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    if parsed < 1:
        return default
    return min(parsed, max_value) if max_value else parsed

GENERIC_REQUIREMENT_TITLE_WORDS = {"jd", "job", "requirement", "opening", "urgent", "hiring", "needed", "new"}

def normalize_requirement_title_text(value):
    text = re.sub(r"\s+", " ", str(value or "").strip())
    text = re.sub(r"\s*[-/|]\s*", " - ", text)
    return text

def normalize_requirement_title_key(value):
    return re.sub(r"[^a-z]+", " ", normalize_requirement_title_text(value).lower()).strip()

def normalize_requirement_client_key(value):
    text = re.sub(r"\s+", " ", str(value or "").strip().lower())
    text = re.sub(r"^taggd\s*[- ]\s*", "", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()

def validate_requirement_title(title, client_name=""):
    clean = normalize_requirement_title_text(title)
    errors = []
    if len(clean) < 5 or len(clean) > 90:
        errors.append("Requirement title should be 5-90 characters.")
    if re.search(r"\d", clean):
        errors.append("Do not include numbers in the requirement title.")
    if not re.match(r"^[A-Za-z][A-Za-z .,&()+/#-]*$", clean):
        errors.append("Use only letters, spaces, and common separators like -, /, &, +, #.")
    words = re.findall(r"[A-Za-z]+", clean.lower())
    if len([w for w in words if w not in GENERIC_REQUIREMENT_TITLE_WORDS]) < 2:
        errors.append("Use a clear role name, for example 'Java Developer' or 'Finance Analyst'.")
    generic_used = sorted({w for w in words if w in GENERIC_REQUIREMENT_TITLE_WORDS})
    if generic_used:
        errors.append("Do not include generic words in the title: " + ", ".join(generic_used) + ".")
    client_core = normalize_requirement_client_key(client_name)
    client_core_parts = [part for part in client_core.split() if len(part) >= 3]
    clean_l = clean.lower()
    if client_core_parts and any(re.search(r"(?<![a-z0-9])" + re.escape(part) + r"(?![a-z0-9])", clean_l) for part in client_core_parts):
        errors.append("Do not include the client name in the requirement title; use the Client field.")
    return clean, errors

def requirement_semantic_text(title, description):
    text = f"{title or ''}\n{description or ''}".lower()
    text = re.sub(r"(?im)^(industry|degree|experience|languages|primary skills|secondary skills):", " ", text)
    text = re.sub(r"[^a-z0-9+#.]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()

def requirement_similarity(title_a, desc_a, title_b, desc_b):
    title_ratio = difflib.SequenceMatcher(None, normalize_requirement_title_text(title_a).lower(), normalize_requirement_title_text(title_b).lower()).ratio()
    text_a = requirement_semantic_text(title_a, desc_a)
    text_b = requirement_semantic_text(title_b, desc_b)
    tokens_a = {t for t in text_a.split() if len(t) > 2}
    tokens_b = {t for t in text_b.split() if len(t) > 2}
    token_ratio = len(tokens_a & tokens_b) / max(len(tokens_a | tokens_b), 1)
    text_ratio = difflib.SequenceMatcher(None, text_a[:1200], text_b[:1200]).ratio()
    return max(title_ratio * 0.65 + token_ratio * 0.35, text_ratio * 0.55 + token_ratio * 0.45)

def find_similar_requirement(conn, title, client_name, description, exclude_id=None):
    params = []
    sql = "SELECT id, title, client_name, description FROM requirements WHERE 1=1"
    if exclude_id:
        sql += " AND id<>?"
        params.append(exclude_id)
    best = None
    best_score = 0
    target_client = normalize_requirement_client_key(client_name)
    for row in conn.execute(sql, params).fetchall():
        if normalize_requirement_client_key(row["client_name"]) != target_client:
            continue
        score = requirement_similarity(title, description, row["title"], row["description"])
        if score > best_score:
            best_score = score
            best = row
    if best and best_score >= 0.78:
        return best, best_score
    return None, best_score

# â”€â”€ Routes â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/")
@login_required
def index():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not session.get("is_admin"):
        return render_template(
            "recruiter_landing.html",
            can_bulk_upload=has_bulk_upload_access(),
            is_client_viewer=is_client_viewer_session(),
            is_team_leader=is_team_leader_session(),
        )
    return redirect(url_for("admin_landing_page"))

@app.route("/app")
@login_required
def app_page():
    return admin_landing_page()

@app.route("/admin")
@login_required
def admin_landing_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not session.get("is_admin"):
        return redirect(url_for("index"))
    return render_template("admin_landing.html")

@app.route("/admin/workspace")
@login_required
def admin_workspace_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not session.get("is_admin"):
        return redirect(url_for("index"))
    return render_template(
        "index.html",
        hide_team_topbar=False,
        initial_app_route="",
        app_route_pending=False,
    )

@app.route("/admin/users")
@login_required
def admin_users_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not session.get("is_admin"):
        return redirect(url_for("index"))
    return render_template("admin_users.html")

@app.route("/admin/team")
@login_required
def admin_team_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not session.get("is_admin"):
        return redirect(url_for("index"))
    return render_template("admin_team.html")

@app.route("/admin/reports")
@login_required
def admin_reports_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not session.get("is_admin"):
        return redirect(url_for("index"))
    return render_template("admin_reports.html")

@app.route("/admin/impersonify")
@login_required
def admin_impersonify_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not session.get("is_admin") or session.get("impersonation_active"):
        return redirect(url_for("index"))
    return render_template("admin_impersonify.html")

@app.route("/admin/team-leader-mapping")
@login_required
def admin_team_leader_mapping_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not session.get("is_admin"):
        return redirect(url_for("index"))
    return render_template("admin_team_leader_mapping.html")

@app.route("/admin/followups")
@login_required
def admin_followups_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not session.get("is_admin"):
        return redirect(url_for("index"))
    return render_template("admin_followups.html")

@app.route("/admin/client-sla")
@login_required
def admin_client_sla_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not session.get("is_admin"):
        return redirect(url_for("index"))
    return render_template("admin_client_sla.html")

@app.route("/admin/data-quality")
@login_required
def admin_data_quality_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not session.get("is_admin"):
        return redirect(url_for("index"))
    return render_template("admin_data_quality.html")

def render_app_route(initial_app_route, route_pending=True):
    return render_template(
        "index.html",
        hide_team_topbar=not session.get("is_admin"),
        initial_app_route=initial_app_route,
        app_route_pending=route_pending,
    )

@app.route("/add-candidate")
@login_required
def add_candidate_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    return render_template("add_candidate.html")

@app.route("/add-requirement")
@login_required
def add_requirement_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    return render_template("add_requirement.html")

@app.route("/daily-reports")
@login_required
def daily_reports_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    return render_template("daily_reports.html")

@app.route("/weekly-performance")
@login_required
def weekly_performance_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    return render_template("recruiter_landing.html", open_weekly_performance=True)

@app.route("/my-followups")
@login_required
def my_followups_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if session.get("is_admin"):
        return redirect(url_for("admin_followups_page"))
    return render_template("my_followups.html")

@app.route("/team-reports")
@login_required
def team_reports_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not (session.get("is_admin") or is_team_leader_session()):
        return redirect(url_for("index"))
    return render_template("team_reports.html")

@app.route("/team-selection-report")
@login_required
def team_selection_report_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not (session.get("is_admin") or is_team_leader_session()):
        return redirect(url_for("index"))
    return render_template("team_selection_report.html")

@app.route("/team-analytics")
@login_required
def team_analytics_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not (session.get("is_admin") or is_team_leader_session()):
        return redirect(url_for("index"))
    return render_template("team_analytics.html")

@app.route("/candidate-search")
@login_required
def candidate_search_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    return render_template("candidate_search.html")

@app.route("/power-search")
@login_required
def power_search_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not session.get("is_admin"):
        return redirect(url_for("candidate_search_page"))
    return render_template("power_search.html")

@app.route("/ai-screening")
@login_required
def ai_screening_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not session.get("is_admin"):
        return redirect(url_for("index"))
    return render_template("ai_screening.html")

@app.route("/candidates")
@login_required
def candidates_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    return render_template("candidate_list.html")

@app.route("/client/candidates")
@login_required
def client_candidates_page():
    if not is_client_viewer_session():
        return redirect(url_for("index"))
    return render_template("client_candidates.html")

@app.route("/profile")
@login_required
def profile_page():
    return render_template("profile.html")

@app.route("/guide")
def guide(): return render_template("guide.html")

@app.route("/requirements")
@login_required
def requirements_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    return render_template("requirement_list.html")

@app.route("/taggd-recruiters")
@login_required
def taggd_recruiters_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    return render_template(
        "taggd_recruiters.html",
        dashboard_href="/admin" if session.get("is_admin") else "/",
        page_scope="Admin" if session.get("is_admin") else ("Team Leader" if is_team_leader_session() else "Recruiter"),
    )

@app.route("/upload")
@login_required
def upload_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    if not has_bulk_upload_access():
        return redirect("/")
    return render_template("upload.html")

@app.route("/bulk-upload-guide")
@login_required
def bulk_upload_guide_page():
    if is_client_viewer_session():
        return redirect(url_for("client_candidates_page"))
    return render_template("upload_instructions.html")

@app.route("/healthz")
def health():
    started = time.perf_counter()
    payload = {
        "ok": True,
        "app": "ok",
        "database": "not_checked",
        "elapsed_ms": 0,
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    try:
        conn = get_db(timeout=2)
        conn.execute("SELECT 1").fetchone()
        conn.close()
        payload["database"] = "ok"
    except Exception as exc:
        payload["ok"] = False
        payload["database"] = "error"
        payload["error"] = f"{type(exc).__name__}: {exc}"
    payload["elapsed_ms"] = round((time.perf_counter() - started) * 1000, 1)
    status = 200 if payload["ok"] else 503
    return jsonify(payload), status

@app.route("/api/me")
@login_required
def api_me():
    phone = ""
    role = session.get("role") or ("Admin" if session.get("is_admin") else "Recruiter")
    if session.get("team_member_id"):
        conn = get_db()
        member = conn.execute("SELECT phone, role FROM team_members WHERE id=?", (session.get("team_member_id"),)).fetchone()
        conn.close()
        if member:
            phone = member["phone"] or ""
            role = member["role"] or role
    return jsonify({
        "user_id": session.get("user_id"),
        "team_member_id": session.get("team_member_id"),
        "username": session.get("username"),
        "recruiter_name": session.get("recruiter_name"),
        "email": session.get("email"),
        "recruiter_email": session.get("recruiter_email"),
        "phone": phone,
        "role": role,
        "notes": session.get("profile_notes", ""),
        "is_admin": session.get("is_admin"),
        "is_client_viewer": 1 if is_client_viewer_session() else 0,
        "impersonation_active": 1 if session.get("impersonation_active") else 0,
        "original_admin_username": session.get("original_admin_username", ""),
        "can_bulk_upload": 1 if has_bulk_upload_access() else 0
    }
)

@app.route("/api/me", methods=["PATCH"])
@login_required
def api_update_me():
    data = request.get_json(silent=True) or {}
    name = clean_value(data.get("name") or data.get("username") or "", 120)
    email = clean_value(data.get("email") or "", 180).lower()
    phone = clean_value(data.get("phone") or "", 40)
    notes = clean_value(data.get("notes") or "", 1000)
    if not name:
        return jsonify({"error": "Name is required."}), 400
    if email and not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email):
        return jsonify({"error": "Please enter a valid email address."}), 400

    conn = get_db()
    try:
        team_member_id = session.get("team_member_id")
        app_user_id = session.get("app_user_id") or session.get("user_id")
        role = "Admin" if session.get("is_admin") else "Recruiter"
        if team_member_id:
            existing = conn.execute(
                "SELECT id FROM team_members WHERE lower(email)=lower(?) AND id<>?",
                (email, team_member_id)
            ).fetchone() if email else None
            if existing:
                return jsonify({"error": "Email is already used by another user."}), 400
            conn.execute(
                "UPDATE team_members SET name=?, email=?, phone=? WHERE id=?",
                (name, email, phone, team_member_id)
            )
            member = conn.execute("SELECT role FROM team_members WHERE id=?", (team_member_id,)).fetchone()
            if member and member["role"]:
                role = member["role"]
        elif app_user_id:
            existing = conn.execute(
                "SELECT id FROM app_users WHERE lower(username)=lower(?) AND id<>?",
                (name, app_user_id)
            ).fetchone()
            if existing:
                return jsonify({"error": "Username is already used by another user."}), 400
            conn.execute("UPDATE app_users SET username=? WHERE id=?", (name, app_user_id))
        conn.commit()
    finally:
        conn.close()

    session["username"] = name
    session["recruiter_name"] = name
    if email:
        session["email"] = email
        session["recruiter_email"] = email
    session["profile_notes"] = notes
    return jsonify({
        "ok": True,
        "user": {
            "user_id": session.get("user_id"),
            "team_member_id": session.get("team_member_id"),
            "username": session.get("username"),
            "recruiter_name": session.get("recruiter_name"),
            "email": session.get("email"),
            "recruiter_email": session.get("recruiter_email"),
            "phone": phone,
            "role": role,
            "notes": notes,
            "is_admin": session.get("is_admin"),
            "can_bulk_upload": 1 if has_bulk_upload_access() else 0
        }
    })

@app.route("/api/performance/summary")
@login_required
def api_performance_summary():
    if not session.get("is_admin"):
        return jsonify({"error": "Admin only"}), 403
    try:
        days = max(1, min(30, int(request.args.get("days", 7))))
    except (TypeError, ValueError):
        days = 7
    path = (request.args.get("path") or "").strip()
    slow_ms = 750.0
    try:
        slow_ms = max(0.0, float(request.args.get("slow_ms", slow_ms)))
    except (TypeError, ValueError):
        pass
    where = ["datetime(created_at) >= datetime('now','localtime', ?)"]
    params = [f"-{days} days"]
    if path:
        where.append("path=?")
        params.append(path)
    where_sql = " AND ".join(where)
    conn = get_db(timeout=5)
    conn.execute("""CREATE TABLE IF NOT EXISTS performance_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT,
        team_member_id INTEGER,
        recruiter_name TEXT,
        recruiter_email TEXT,
        method TEXT,
        path TEXT,
        endpoint TEXT,
        status_code INTEGER,
        elapsed_ms REAL,
        is_admin INTEGER DEFAULT 0,
        ip_address TEXT,
        user_agent TEXT,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    )""")
    by_user = [
        dict(row)
        for row in conn.execute(
            f"""
            SELECT
                COALESCE(NULLIF(username,''), NULLIF(recruiter_name,''), NULLIF(recruiter_email,''), 'Unknown') AS user,
                team_member_id,
                COUNT(*) AS requests,
                ROUND(AVG(elapsed_ms), 1) AS avg_ms,
                ROUND(MAX(elapsed_ms), 1) AS max_ms,
                SUM(CASE WHEN elapsed_ms>=? THEN 1 ELSE 0 END) AS slow_requests
            FROM performance_logs
            WHERE {where_sql}
            GROUP BY COALESCE(NULLIF(username,''), NULLIF(recruiter_name,''), NULLIF(recruiter_email,''), 'Unknown'), team_member_id
            ORDER BY avg_ms DESC, requests DESC
            LIMIT 100
            """,
            [slow_ms] + params,
        ).fetchall()
    ]
    by_endpoint = [
        dict(row)
        for row in conn.execute(
            f"""
            SELECT
                method,
                path,
                COUNT(*) AS requests,
                ROUND(AVG(elapsed_ms), 1) AS avg_ms,
                ROUND(MAX(elapsed_ms), 1) AS max_ms,
                SUM(CASE WHEN elapsed_ms>=? THEN 1 ELSE 0 END) AS slow_requests
            FROM performance_logs
            WHERE {where_sql}
            GROUP BY method, path
            ORDER BY avg_ms DESC, requests DESC
            LIMIT 100
            """,
            [slow_ms] + params,
        ).fetchall()
    ]
    recent = [
        dict(row)
        for row in conn.execute(
            f"""
            SELECT created_at, username, team_member_id, method, path, status_code, ROUND(elapsed_ms, 1) AS elapsed_ms
            FROM performance_logs
            WHERE {where_sql}
            ORDER BY id DESC
            LIMIT 50
            """,
            params,
        ).fetchall()
    ]
    conn.close()
    return jsonify({
        "days": days,
        "path": path,
        "slow_ms": slow_ms,
        "by_user": by_user,
        "by_endpoint": by_endpoint,
        "recent": recent
    })

@app.route("/api/upload", methods=["POST"])
@login_required
def api_upload():
    forbidden = client_viewer_write_forbidden()
    if forbidden:
        return forbidden
    if not has_bulk_upload_access():
        return jsonify({"error":"Bulk upload permission required"}),403
    sourcer_id = session.get("team_member_id")
    name  = request.form.get("recruiter_name","").strip() or session.get("recruiter_name","System")
    email = request.form.get("recruiter_email","").strip() or session.get("recruiter_email","")
    if not name or name == "System": name = session.get("recruiter_name","System")
    if not email: email = session.get("recruiter_email","system@hrguru.com")
    email = email.lower()
    role  = request.form.get("role_override","").strip()
    if "excel_file" not in request.files: return jsonify({"error":"No file attached"}),400
    return jsonify(process_upload(name,email,role,sourcer_id,
                   request.files["excel_file"],request.files.getlist("cv_files"),request.files.getlist("jd_files")))

@app.route("/api/template")
def api_template():
    buf = make_template()
    if not buf: return "openpyxl not installed",500
    return send_file(buf,download_name="candidate_template.xlsx",as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/candidates")
@login_required
def api_candidates():
    conn = get_db()
    select_clause = None
    view = str(request.args.get("view", "")).strip().lower()
    if view == "reporting":
        select_clause = (
            "c.id, c.candidate_name, c.email_addr, c.current_role, c.role_name, "
            "c.status, c.recruiter_name, c.recruiter_email, c.created_at, "
            "r.title as requirement_title, r.client_name as client_name"
        )
    elif view == "list":
        select_clause = (
            "c.id, c.candidate_name, c.email_addr, c.phone, c.current_company, c.current_role, "
            "c.experience_years, c.key_skills, c.notice_period, c.current_salary, c.expected_salary, "
            "c.current_location, c.preferred_location, c.remarks, c.role_name, c.status, "
            "c.recruiter_name, c.recruiter_email, c.created_at, c.cv_url, c.cv_filename, "
            "r.title as requirement_title, r.client_name as client_name"
        )
    sql, p = build_query(request.args, session, select_clause=select_clause)
    if str(request.args.get("all", "")).lower() in {"1", "true", "yes"}:
        rows = [dict(r) for r in conn.execute(sql, p).fetchall()]
        conn.close()
        return jsonify(rows)

    page = parse_positive_int(request.args.get("page"), 1)
    page_size = parse_positive_int(request.args.get("page_size"), 15, 500)
    count_sql = "SELECT COUNT(*) FROM (" + strip_order_by(sql) + ") filtered_candidates"
    total = conn.execute(count_sql, p).fetchone()[0]
    paged_sql = sql + " LIMIT ? OFFSET ?"
    rows = [dict(r) for r in conn.execute(paged_sql, p + [page_size, (page - 1) * page_size]).fetchall()]
    conn.close()
    return jsonify({
        "rows": rows,
        "page": page,
        "page_size": page_size,
        "total": total,
        "total_pages": max(1, (total + page_size - 1) // page_size) if page_size else 1
    })

@app.route("/api/client/candidates")
@login_required
def api_client_candidates():
    if not is_client_viewer_session():
        return jsonify({"error": "Client access only"}), 403
    conn = get_db()
    owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
    where = ["1=1" + owner_sql]
    params = list(owner_params)
    q = clean_value(request.args.get("q") or "", 120).lower()
    status = clean_value(request.args.get("status") or "", 80)
    if q:
        where.append("""(
            lower(c.candidate_name) LIKE ?
            OR lower(COALESCE(r.title,'')) LIKE ?
            OR lower(COALESCE(c.current_company,'')) LIKE ?
            OR lower(COALESCE(c.current_role,'')) LIKE ?
        )""")
        like = f"%{q}%"
        params.extend([like, like, like, like])
    if status:
        where.append("c.status=?")
        params.append(status)
    sql = f"""
        SELECT c.id, c.candidate_name, c.current_company, c.current_role,
               c.experience_years, c.current_location, c.status,
               c.candidate_feedback, c.created_at,
               r.title AS requirement_title, r.client_name AS client_name
        FROM candidates c
        LEFT JOIN requirements r ON r.id = c.requirement_id
        WHERE {' AND '.join(where)}
        ORDER BY datetime(c.created_at) DESC, c.id DESC
    """
    page = parse_positive_int(request.args.get("page"), 1)
    page_size = parse_positive_int(request.args.get("page_size"), 15, 100)
    total = conn.execute("SELECT COUNT(*) FROM (" + strip_order_by(sql) + ") client_candidates", params).fetchone()[0]
    rows = [dict(r) for r in conn.execute(sql + " LIMIT ? OFFSET ?", params + [page_size, (page - 1) * page_size]).fetchall()]
    conn.close()
    return jsonify({
        "rows": rows,
        "page": page,
        "page_size": page_size,
        "total": total,
        "total_pages": max(1, (total + page_size - 1) // page_size) if page_size else 1
    })

@app.route("/api/candidates/export")
@login_required
def export_candidates():
    fmt  = request.args.get("format","csv")
    ids  = request.args.get("ids","")
    conn = get_db()
    if ids:
        id_list = [int(x) for x in ids.split(",") if x.isdigit()]
        if id_list:
            sql = "SELECT c.*, r.title as requirement_title FROM candidates c LEFT JOIN requirements r ON c.requirement_id = r.id WHERE c.id IN (" + ",".join("?" * len(id_list)) + ")"
            p = id_list
            owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
            sql += owner_sql
            p.extend(owner_params)
        else:
            sql, p = build_query(request.args, session)
    else:
        sql, p = build_query(request.args, session)
    rows = [dict(r) for r in conn.execute(sql,p).fetchall()]; conn.close()

    if fmt=="xlsx" and XLSX_OK:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="Candidates"
        hfill=PatternFill("solid",fgColor="1C2030"); hfont=Font(bold=True,color="E8643A")
        for i,h in enumerate(EXPORT_HDR,1):
            c=ws.cell(row=1,column=i,value=h); c.font=hfont; c.fill=hfill
            ws.column_dimensions[c.column_letter].width=max(len(h)+4,16)
        for ri,row in enumerate(rows,2):
            for ci,col in enumerate(EXPORT_COLS,1):
                v=row.get(col,""); ws.cell(row=ri,column=ci,value=str(v) if v else "")
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf,download_name=f"candidates_{date.today()}.xlsx",as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    buf=io.StringIO(); writer=csv.writer(buf); writer.writerow(EXPORT_HDR)
    for row in rows: writer.writerow([str(row.get(c,"") or "") for c in EXPORT_COLS])
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode()),
                     download_name=f"candidates_{date.today()}.csv",
                     as_attachment=True,mimetype="text/csv")

def candidate_report_rows(args, current_session=None):
    ids = str(args.get("ids", "") or "")
    conn = get_db()
    if ids:
        id_list = [int(x) for x in ids.split(",") if x.isdigit()]
        if id_list:
            sql = "SELECT c.*, r.title as requirement_title, r.client_name as client_name FROM candidates c LEFT JOIN requirements r ON c.requirement_id = r.id WHERE c.id IN (" + ",".join("?" * len(id_list)) + ")"
            p = id_list
            owner_sql, owner_params = non_admin_candidate_owner_clause(current_session, "c")
            sql += owner_sql
            p.extend(owner_params)
        else:
            sql, p = build_query(args, current_session)
    else:
        sql, p = build_query(args, current_session)
    rows = [dict(r) for r in conn.execute(sql, p).fetchall()]
    conn.close()
    return rows

def daily_submission_context(rows, report_date=""):
    report_date = report_date or feedback_request_shared_date(rows, {})
    try:
        date_label = datetime.fromisoformat(report_date).strftime("%d %b %Y")
    except Exception:
        date_label = report_date or date.today().strftime("%d %b %Y")
    clients = unique_list([row.get("client_name") or "" for row in rows if row.get("client_name")], 3)
    client_label = clients[0] if len(clients) == 1 else ("selected clients" if clients else "client")
    return client_label, date_label

EMAIL_CANDIDATE_COLUMNS = [
    ("date", "Date", lambda row: str(row.get("created_at") or "").split(" ", 1)[0] or "-"),
    ("name", "Name", lambda row: row.get("candidate_name") or "-"),
    ("designation", "Current Designation", lambda row: row.get("current_role") or row.get("role_name") or "-"),
    ("current_company", "Current Company", lambda row: row.get("current_company") or "-"),
    ("requirement", "Requirement", lambda row: row.get("requirement_title") or row.get("role_name") or "-"),
    ("client", "Client", lambda row: row.get("client_name") or "-"),
    ("email", "Email-Id", lambda row: row.get("email_addr") or "-"),
    ("phone", "Phone", lambda row: row.get("phone") or "-"),
    ("experience", "Experience", lambda row: row.get("experience_years") or "-"),
    ("key_skills", "Key Skills", lambda row: row.get("key_skills") or "-"),
    ("current_salary", "Current Salary", lambda row: row.get("current_salary") or "-"),
    ("expected_salary", "Expected Salary", lambda row: row.get("expected_salary") or "-"),
    ("notice_period", "Notice Period", lambda row: row.get("notice_period") or "-"),
    ("qualification", "Qualification", lambda row: row.get("education") or row.get("qualification") or "-"),
    ("status", "Status", lambda row: row.get("status") or "New"),
    ("recruiter", "Recruiter", lambda row: row.get("recruiter_name") or "-"),
    ("location", "Location", lambda row: row.get("current_location") or "-"),
]
DEFAULT_EMAIL_CANDIDATE_COLUMNS = ["date", "name", "current_company", "designation", "requirement", "email", "phone", "experience", "current_salary", "expected_salary", "notice_period"]

def normalize_email_candidate_columns(columns=None):
    by_key = {key: (key, label, getter) for key, label, getter in EMAIL_CANDIDATE_COLUMNS}
    requested = [str(col or "").strip() for col in (columns or [])]
    selected = []
    for col in requested:
        if col != "date" and col in by_key and col not in selected:
            selected.append(col)
    if not selected:
        selected = [col for col in DEFAULT_EMAIL_CANDIDATE_COLUMNS if col != "date"]
    selected.insert(0, "date")
    return [by_key[key] for key in selected]

def build_daily_work_report_email(rows, sender_name="", filters_applied=False, to_addr="", recipient_name="", report_date="", columns=None):
    client_label, date_label = daily_submission_context(rows, report_date)
    subject = f"Daily submissions for {client_label} on {date_label}"
    recruiter = sender_name or "Recruiter"
    recipient = feedback_request_recipient_name(to_addr, recipient_name) if to_addr or recipient_name else "Team"
    lines = [
        f"Hi {recipient},",
        "",
        f"Please find below the daily submissions shared for {client_label} on {date_label}.",
        "Candidate CVs are attached for your review.",
        "",
    ]
    lines.append("Candidate details:")
    if not rows:
        lines.append("No candidates matched the selected criteria.")
    else:
        email_columns = normalize_email_candidate_columns(columns)
        for idx, row in enumerate(rows, 1):
            parts = [
                f"{idx}. " + " | ".join(f"{label}: {getter(row)}" for _, label, getter in email_columns),
            ]
            lines.append(parts[0])
    lines.extend([
        "",
        "Regards,",
        recruiter,
        "HR Guru Placement Services",
    ])
    return subject, "\n".join(lines)

def candidate_email_detail_cells(row, columns=None):
    email_columns = columns if columns and isinstance(columns[0], tuple) else normalize_email_candidate_columns(columns)
    return [getter(row) for _, _, getter in email_columns]

def candidate_email_table_rows(rows, columns=None):
    table_rows = ""
    email_columns = columns if columns and isinstance(columns[0], tuple) else normalize_email_candidate_columns(columns)
    for idx, row in enumerate(rows, 1):
        cells = [str(idx)] + candidate_email_detail_cells(row, email_columns)
        table_rows += "<tr>" + "".join(
            f"<td style=\"border:1px solid #d8dee9;padding:8px;text-align:left;vertical-align:top\">{html_escape(cell)}</td>"
            for cell in cells
        ) + "</tr>"
    return table_rows

def candidate_email_table_header(columns=None):
    email_columns = columns if columns and isinstance(columns[0], tuple) else normalize_email_candidate_columns(columns)
    headers = ["#"] + [label for _, label, _ in email_columns]
    return "".join(
        f"<th style=\"border:1px solid #d8dee9;padding:8px;text-align:left\">{html_escape(header)}</th>"
        for header in headers
    )

def build_daily_work_report_html(rows, sender_name="", filters_applied=False, to_addr="", recipient_name="", report_date="", columns=None):
    client_label, date_label = daily_submission_context(rows, report_date)
    recruiter = sender_name or "Recruiter"
    recipient = feedback_request_recipient_name(to_addr, recipient_name) if to_addr or recipient_name else "Team"
    email_columns = normalize_email_candidate_columns(columns)
    table_rows = candidate_email_table_rows(rows, email_columns)
    if not table_rows:
        table_rows = f'<tr><td colspan="{len(email_columns) + 1}" style="padding:10px;border:1px solid #d8dee9">No candidates matched the selected criteria.</td></tr>'
    return f"""
    <div style="font-family:Arial,sans-serif;color:#1f2937;font-size:14px;line-height:1.5">
      <p>Hi {html_escape(recipient)},</p>
      <p>Please find below the daily submissions shared for <strong>{html_escape(client_label)}</strong> on <strong>{html_escape(date_label)}</strong>.</p>
      <p>Candidate CVs are attached for your review. The candidate tracker is also attached as an Excel file.</p>
      <table style="border-collapse:collapse;width:100%;font-size:12px;margin:14px 0">
        <thead>
          <tr style="background:#1c2030;color:#ffffff">
            {candidate_email_table_header(email_columns)}
          </tr>
        </thead>
        <tbody>{table_rows}</tbody>
      </table>
      <p>Regards,<br>{html_escape(recruiter)}<br>HR Guru Placement Services</p>
    </div>
    """

def feedback_request_recipient_name(email, provided_name=""):
    name = (provided_name or "").strip()
    if name:
        return name
    local = (email or "").split("@", 1)[0]
    local = re.sub(r"[._\-]+", " ", local).strip()
    return local.title() if local else "Team"

def html_escape(value):
    return html.escape(str(value or ""), quote=True)

def display_person_name(value):
    raw = str(value or "").strip()
    if not raw:
        return ""
    local = raw.split("@", 1)[0]
    if any(ch in local for ch in "._-"):
        parts = [part for part in re.split(r"[._\-\s]+", local) if part]
        if parts:
            return " ".join(part[:1].upper() + part[1:].lower() for part in parts)
    if raw.islower() or raw.isupper():
        return raw.title()
    return raw

def daily_report_sender_name():
    for value in (
        session.get("recruiter_name"),
        session.get("name"),
        session.get("email"),
        session.get("recruiter_email"),
        session.get("username"),
    ):
        name = display_person_name(value)
        if name:
            return name
    return "Recruiter"

def build_feedback_request_email(rows, to_addr, recipient_name="", sender_name="", report_date="", columns=None):
    report_date = report_date or date.today().isoformat()
    try:
        subject_date = datetime.fromisoformat(report_date).strftime("%d %b %Y")
    except Exception:
        subject_date = report_date
    clients = unique_list([row.get("client_name") or "" for row in rows if row.get("client_name")], 3)
    client_label = clients[0] if len(clients) == 1 else ("selected clients" if clients else "client")
    subject = f"Feedback request for profiles shared for {client_label} on {subject_date}"
    name = feedback_request_recipient_name(to_addr, recipient_name)
    sender = sender_name or "Recruiter"
    lines = [
        f"Hi {name},",
        "",
        f"We are writing to request your feedback on the profiles shared for {client_label} on {subject_date}.",
        "Please review the candidates below and share your comments/status update so the recruiting team can take the next action quickly.",
        "",
        "Candidates shared:",
    ]
    if not rows:
        lines.append("No candidates were selected.")
    else:
        email_columns = normalize_email_candidate_columns(columns)
        headers = ["#"] + [label for _, label, _ in email_columns]
        table_rows = []
        for idx, row in enumerate(rows, 1):
            table_rows.append([str(idx)] + candidate_email_detail_cells(row, email_columns))
        widths = [len(h) for h in headers]
        for row in table_rows:
            for i, cell in enumerate(row):
                widths[i] = min(max(widths[i], len(str(cell))), 34)
        def clip(value, width):
            value = str(value or "")
            return value if len(value) <= width else value[:width - 3] + "..."
        fmt = " | ".join("{:<" + str(width) + "}" for width in widths)
        lines.append(fmt.format(*headers))
        lines.append("-+-".join("-" * width for width in widths))
        for row in table_rows:
            lines.append(fmt.format(*[clip(cell, widths[i]) for i, cell in enumerate(row)]))
    lines.extend([
        "",
        "Kindly share feedback against each profile, including whether the candidate should move ahead, be held, or be rejected with comments.",
        "",
        "Regards,",
        sender,
        "HR Guru Placement Services",
    ])
    plain_body = "\n".join(lines)
    email_columns = normalize_email_candidate_columns(columns)
    table_rows = candidate_email_table_rows(rows, email_columns)
    if not table_rows:
        table_rows = f'<tr><td colspan="{len(email_columns) + 1}" style="padding:10px;border:1px solid #d8dee9">No candidates were selected.</td></tr>'
    html_body = f"""
    <div style="font-family:Arial,sans-serif;color:#1f2937;font-size:14px;line-height:1.5">
      <p>Hi {html_escape(name)},</p>
      <p>We are writing to request your feedback on the profiles shared for <strong>{html_escape(client_label)}</strong> on <strong>{html_escape(subject_date)}</strong>.</p>
      <p>Please review the candidates below and share your comments/status update so the recruiting team can take the next action quickly.</p>
      <table style="border-collapse:collapse;width:100%;font-size:13px;margin:14px 0">
        <thead>
          <tr style="background:#1c2030;color:#ffffff">
            {candidate_email_table_header(email_columns)}
          </tr>
        </thead>
        <tbody>{table_rows}</tbody>
      </table>
      <p>Candidate CVs are attached for your review.</p>
      <p>Kindly share feedback against each profile, including whether the candidate should move ahead, be held, or be rejected with comments.</p>
      <p>Regards,<br>{html_escape(sender)}<br>HR Guru Placement Services</p>
    </div>
    """
    return subject, plain_body, html_body

def feedback_request_shared_date(rows, filters=None):
    dates = []
    for row in rows or []:
        raw = str(row.get("created_at") or row.get("added_date") or "").strip()
        if not raw:
            continue
        date_part = raw.split(" ", 1)[0]
        try:
            dates.append(datetime.fromisoformat(date_part).date().isoformat())
        except Exception:
            continue
    if dates:
        return max(dates)
    filters = filters or {}
    return filters.get("date_to") or filters.get("date_from") or date.today().isoformat()

def build_candidate_export_xlsx(rows):
    if not XLSX_OK:
        raise RuntimeError("Excel export is not available because openpyxl is not installed.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates"
    hfill = PatternFill("solid", fgColor="1C2030")
    hfont = Font(bold=True, color="E8643A")
    for i, h in enumerate(EXPORT_HDR, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hfont
        c.fill = hfill
        ws.column_dimensions[c.column_letter].width = max(len(h) + 4, 16)
    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(EXPORT_COLS, 1):
            ws.cell(row=ri, column=ci, value=str(row.get(col, "") or ""))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def fetch_drive_file_bytes(file_id):
    if not file_id:
        return None
    try:
        service = central_drive_service()
        return service.files().get_media(fileId=file_id, supportsAllDrives=True).execute()
    except Exception as e:
        print("Drive file fetch error:", e)
        return None

def candidate_cv_attachments(rows):
    upload_folder = os.path.abspath(os.path.join(app.root_path, "uploads"))
    attachments = []
    for row in rows:
        filename = row.get("cv_public_id") or ""
        cv_url = row.get("cv_url") or ""
        relative_upload_path = ""
        if not filename and cv_url.startswith("/uploads/"):
            relative_upload_path = cv_url[len("/uploads/"):].replace("/", os.sep)
            filename = os.path.basename(relative_upload_path)
        elif cv_url.startswith("/uploads/"):
            relative_upload_path = cv_url[len("/uploads/"):].replace("/", os.sep)
        drive_id = ""
        if str(cv_url).startswith("/api/drive_file/"):
            drive_id = str(cv_url).rsplit("/", 1)[-1]
        elif filename and not os.path.splitext(os.path.basename(filename))[1] and not str(cv_url).startswith("/uploads/"):
            drive_id = filename
        candidate_label = secure_filename(row.get("candidate_name") or f"candidate_{row.get('id') or len(attachments) + 1}") or f"candidate_{len(attachments) + 1}"
        if drive_id:
            content = fetch_drive_file_bytes(drive_id)
            if content:
                display_name = row.get("cv_filename") or f"candidate_cv_{drive_id}"
                base, ext = os.path.splitext(display_name)
                ext = ext or ".pdf"
                attachments.append({
                    "filename": f"{candidate_label}_{secure_filename(base) or 'CV'}{ext}",
                    "content": content,
                    "mimetype": mimetypes.guess_type(display_name)[0] or "application/octet-stream",
                    "candidate_id": row.get("id"),
                    "candidate_name": row.get("candidate_name"),
                })
            continue
        if not filename and not str(cv_url).lower().startswith(("http://", "https://")):
            continue
        path = os.path.abspath(os.path.join(upload_folder, relative_upload_path or os.path.basename(filename)))
        display_name = row.get("cv_filename") or os.path.basename(path)
        if path.startswith(upload_folder + os.sep) and os.path.exists(path):
            with open(path, "rb") as fh:
                content = fh.read()
        elif str(cv_url).lower().startswith(("http://", "https://")):
            try:
                response = requests.get(cv_url, timeout=20)
                response.raise_for_status()
                content = response.content
            except Exception as e:
                print("CV attachment fetch error:", e)
                continue
        else:
            continue
        base, ext = os.path.splitext(display_name)
        if not ext:
            ext = os.path.splitext(path)[1] or ".pdf"
        attachments.append({
            "filename": f"{candidate_label}_{secure_filename(base) or 'CV'}{ext}",
            "content": content,
            "mimetype": mimetypes.guess_type(display_name)[0] or "application/octet-stream",
            "candidate_id": row.get("id"),
            "candidate_name": row.get("candidate_name"),
        })
    return attachments

def candidate_cv_email_attachments(rows):
    return candidate_cv_attachments(rows)

def candidate_cv_attachment_summary(rows):
    attachments = candidate_cv_attachments(rows)
    return {
        "attached_cvs": len(attachments),
        "cv_filenames": [a.get("filename") or f"candidate_cv_{i}" for i, a in enumerate(attachments, 1)],
    }

def candidate_cv_attachment_validation(rows):
    attachments = candidate_cv_attachments(rows)
    attached_ids = {str(a.get("candidate_id") or "") for a in attachments if a.get("candidate_id")}
    attached_names = {str(a.get("candidate_name") or "").strip().lower() for a in attachments if a.get("candidate_name")}
    missing = []
    for row in rows or []:
        row_id = str(row.get("id") or "")
        if row_id and row_id in attached_ids:
            continue
        row_name = str(row.get("candidate_name") or "").strip().lower()
        if not row_id and row_name and row_name in attached_names:
            continue
        missing.append(row.get("candidate_name") or f"Candidate #{row_id or len(missing) + 1}")
    return attachments, missing

@app.route("/api/candidates/daily_report_export")
@login_required
def export_daily_report_bundle():
    ids = request.args.get("ids", "")
    if not ids:
        return jsonify({"error": "Select at least one candidate to export."}), 400
    rows = candidate_report_rows({"ids": ids}, session)
    if not rows:
        return jsonify({"error": "No exportable candidates found for your selection."}), 404

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(
            f"daily_candidate_report_{date.today().isoformat()}.xlsx",
            build_candidate_export_xlsx(rows)
        )
        used_names = set()
        for index, attachment in enumerate(candidate_cv_attachments(rows), 1):
            raw_name = secure_filename(attachment.get("filename") or f"candidate_cv_{index}") or f"candidate_cv_{index}"
            base, ext = os.path.splitext(raw_name)
            cv_name = f"CVs/{raw_name}"
            suffix = 2
            while cv_name.lower() in used_names:
                cv_name = f"CVs/{base}_{suffix}{ext}"
                suffix += 1
            used_names.add(cv_name.lower())
            zf.writestr(cv_name, attachment.get("content") or b"")
    zip_buf.seek(0)
    return send_file(
        zip_buf,
        download_name=f"daily_candidate_report_{date.today().isoformat()}.zip",
        as_attachment=True,
        mimetype="application/zip"
    )

@app.route("/api/reports")
@login_required
def api_reports():
    report_type = request.args.get("type", "summary")
    from_date = request.args.get("from_date", "")
    to_date = request.args.get("to_date", "")
    conn = get_db()
    
    date_filter = ""
    params = []
    if from_date:
        date_filter += " AND created_at >= ?"
        params.append(from_date)
    if to_date:
        date_filter += " AND created_at <= ?"
        params.append(to_date + " 23:59:59")
    
    user_filter = ""
    user_params = []
    if not session.get("is_admin") and session.get("recruiter_email"):
        user_filter = " AND recruiter_email = ?"
        user_params = [session["recruiter_email"].lower()]
    
    base_where = "1=1" + date_filter + user_filter
    
    if report_type == "summary":
        total = conn.execute("SELECT COUNT(*) FROM candidates WHERE " + base_where, params + user_params).fetchone()[0]
        by_status = conn.execute("SELECT status, COUNT(*) as cnt FROM candidates WHERE " + base_where + " GROUP BY status", params + user_params).fetchall()
        by_role = conn.execute("SELECT role_name, COUNT(*) as cnt FROM candidates WHERE role_name!=''" + date_filter + user_filter + " GROUP BY role_name ORDER BY cnt DESC LIMIT 10", params + user_params).fetchall()
        return jsonify({"total": total, "by_status": [dict(r) for r in by_status], "by_role": [dict(r) for r in by_role]})
    
    elif report_type == "daily_dashboard":
        if not session.get("is_admin"):
            return jsonify({"error":"Admin only"}),403
        today = date.today().isoformat()
        rows = conn.execute("""
            SELECT
                COALESCE(NULLIF(c.recruiter_name,''), tm.name, c.recruiter_email, 'Unassigned') as recruiter,
                COALESCE(NULLIF(r.title,''), NULLIF(c.role_name,''), 'No Requirement') as requirement,
                COUNT(c.id) as submissions
            FROM candidates c
            LEFT JOIN team_members tm ON c.sourcer_id=tm.id
            LEFT JOIN requirements r ON c.requirement_id=r.id
            WHERE date(c.created_at)=?
            GROUP BY recruiter, requirement
            ORDER BY recruiter ASC, submissions DESC
        """, (today,)).fetchall()
        totals = conn.execute("""
            SELECT
                COUNT(*) as submissions,
                COUNT(DISTINCT COALESCE(NULLIF(recruiter_email,''), sourcer_id)) as recruiters,
                COUNT(DISTINCT COALESCE(requirement_id, role_name)) as requirements
            FROM candidates
            WHERE date(created_at)=?
        """, (today,)).fetchone()
        return jsonify({
            "date": today,
            "totals": dict(totals),
            "rows": [dict(r) for r in rows]
        })
    
    elif report_type == "sourcer":
        rows = conn.execute("""SELECT tm.name as sourcer, COUNT(c.id) as submissions, 
            SUM(CASE WHEN c.status IN ('Shortlisted','Hired','Joined') THEN 1 ELSE 0 END) as success
            FROM candidates c LEFT JOIN team_members tm ON c.sourcer_id = tm.id
            WHERE c.sourcer_id IS NOT NULL""" + date_filter + user_filter + """ GROUP BY c.sourcer_id
            ORDER BY submissions DESC""", params + user_params).fetchall()
        return jsonify([dict(r) for r in rows])

    elif report_type == "sourcer_today":
        today = date.today().isoformat()
        rows = conn.execute("""
            SELECT
                COALESCE(NULLIF(tm.name,''), NULLIF(c.recruiter_name,''), c.recruiter_email, 'Unassigned') as sourcer,
                COUNT(c.id) as submissions
            FROM candidates c
            LEFT JOIN team_members tm ON c.sourcer_id = tm.id
            WHERE date(c.created_at)=?
        """ + user_filter + """
            GROUP BY COALESCE(NULLIF(tm.name,''), NULLIF(c.recruiter_name,''), c.recruiter_email, 'Unassigned')
            ORDER BY submissions DESC, sourcer ASC
        """, [today] + user_params).fetchall()
        totals = conn.execute("""
            SELECT COUNT(*) as submissions,
                   COUNT(DISTINCT COALESCE(NULLIF(tm.name,''), NULLIF(c.recruiter_name,''), c.recruiter_email, 'Unassigned')) as sourcers
            FROM candidates c
            LEFT JOIN team_members tm ON c.sourcer_id = tm.id
            WHERE date(c.created_at)=?
        """ + user_filter, [today] + user_params).fetchone()
        return jsonify({
            "date": today,
            "totals": dict(totals) if totals else {"submissions": 0, "sourcers": 0},
            "rows": [dict(r) for r in rows]
        })
    
    elif report_type == "status":
        rows = conn.execute("SELECT status, COUNT(*) as cnt FROM candidates WHERE " + base_where + " GROUP BY status ORDER BY cnt DESC", params + user_params).fetchall()
        return jsonify([{"status": r[0], "count": r[1]} for r in rows])
    
    elif report_type == "skills":
        all_skills = conn.execute("SELECT key_skills FROM candidates WHERE key_skills!=''" + date_filter + user_filter, params + user_params).fetchall()
        skill_counts = {}
        for row in all_skills:
            for s in row[0].split(","):
                s = s.strip().lower()
                if s:
                    skill_counts[s] = skill_counts.get(s, 0) + 1
        sorted_skills = sorted(skill_counts.items(), key=lambda x: x[1], reverse=True)[:20]
        return jsonify([{"skill": k, "count": v} for k, v in sorted_skills])
    
    elif report_type == "submissions":
        rows = conn.execute("""SELECT DATE(created_at) as date, COUNT(*) as submissions
            FROM candidates WHERE """ + base_where + """ GROUP BY DATE(created_at)
            ORDER BY date DESC LIMIT 30""", params + user_params).fetchall()
        return jsonify([{"date": str(r[0]), "count": r[1]} for r in rows])
    
    elif report_type == "status_detail":
        status = request.args.get("status", "New")
        q = [status]
        sql = "SELECT role_name, recruiter_name, COUNT(*) as cnt FROM candidates WHERE status=?"
        if from_date:
            sql += " AND created_at >= ?"
            q.append(from_date)
        if to_date:
            sql += " AND created_at <= ?"
            q.append(to_date + " 23:59:59")
        if not session.get("is_admin") and session.get("team_member_id"):
            sql += " AND sourcer_id = ?"
            q.append(session["team_member_id"])
        sql += " GROUP BY role_name, recruiter_name ORDER BY cnt DESC"
        rows = conn.execute(sql, q).fetchall()
        return jsonify([{"role_name": r[0] or "-", "recruiter_name": r[1] or "-", "cnt": r[2]} for r in rows])
    
    return jsonify({"error": "Unknown report type"})

@app.route("/api/team-reports/today-by-recruiter")
@login_required
def api_team_report_today_by_recruiter():
    if is_client_viewer_session():
        return jsonify({"error": "Team reports are not available for this user type."}), 403
    if not (session.get("is_admin") or is_team_leader_session()):
        return jsonify({"error": "Team reports are available only for admins and team leaders."}), 403

    today = date.today().isoformat()
    client_filter = (request.args.get("client") or "").strip()
    conn = get_db()
    owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
    client_sql = ""
    query_params = list(owner_params)
    if client_filter:
        client_sql = " AND r.client_name=?"
        query_params.append(client_filter)
    rows = conn.execute(f"""
        SELECT
            COALESCE(NULLIF(tm.name,''), NULLIF(c.recruiter_name,''), c.recruiter_email, 'Unassigned') AS recruiter,
            COALESCE(NULLIF(r.client_name,''), 'Unassigned') AS client_name,
            COALESCE(NULLIF(r.title,''), NULLIF(c.role_name,''), 'No Requirement') AS requirement,
            COALESCE(NULLIF(tg.name,''), NULLIF(r.taggd_recruiter_name,''), 'Not assigned') AS taggd_recruiter_name,
            COUNT(c.id) AS submissions,
            SUM(CASE WHEN COALESCE(c.status,'') IN ('Selected','Offered','Joined','Hired') THEN 1 ELSE 0 END) AS selections
        FROM candidates c
        LEFT JOIN team_members tm ON c.sourcer_id = tm.id
        LEFT JOIN requirements r ON c.requirement_id = r.id
        LEFT JOIN taggd_recruiters tg ON tg.id = r.taggd_recruiter_id
        WHERE date(c.created_at)=date('now','localtime')
          AND COALESCE(c.is_duplicate,0)=0
          {owner_sql}
          {client_sql}
        GROUP BY
            COALESCE(NULLIF(tm.name,''), NULLIF(c.recruiter_name,''), c.recruiter_email, 'Unassigned'),
            COALESCE(NULLIF(r.client_name,''), 'Unassigned'),
            COALESCE(NULLIF(r.title,''), NULLIF(c.role_name,''), 'No Requirement'),
            COALESCE(NULLIF(tg.name,''), NULLIF(r.taggd_recruiter_name,''), 'Not assigned')
        ORDER BY recruiter ASC, client_name ASC, submissions DESC, requirement ASC
    """, query_params).fetchall()
    totals = conn.execute(f"""
        SELECT
            COUNT(c.id) AS submissions,
            COUNT(DISTINCT COALESCE(c.sourcer_id, c.recruiter_email)) AS recruiters,
            COUNT(DISTINCT c.requirement_id) AS requirements_worked,
            SUM(CASE WHEN COALESCE(c.status,'') IN ('Selected','Offered','Joined','Hired') THEN 1 ELSE 0 END) AS selections
        FROM candidates c
        LEFT JOIN requirements r ON c.requirement_id = r.id
        WHERE date(c.created_at)=date('now','localtime')
          AND COALESCE(c.is_duplicate,0)=0
          {owner_sql}
          {client_sql}
    """, query_params).fetchone()
    clients = [
        r["client_name"] for r in conn.execute(f"""
            SELECT DISTINCT r.client_name
            FROM candidates c
            LEFT JOIN requirements r ON c.requirement_id = r.id
            WHERE date(c.created_at)=date('now','localtime')
              AND COALESCE(c.is_duplicate,0)=0
              AND COALESCE(r.client_name,'')!=''
              {owner_sql}
            ORDER BY r.client_name
        """, owner_params).fetchall()
    ]
    conn.close()
    return jsonify({
        "date": today,
        "client": client_filter,
        "clients": clients,
        "totals": dict(totals) if totals else {
            "submissions": 0,
            "recruiters": 0,
            "requirements_worked": 0,
            "selections": 0,
        },
        "rows": [dict(r) for r in rows],
    })

@app.route("/api/team-reports/selection-summary")
@login_required
def api_team_report_selection_summary():
    if is_client_viewer_session():
        return jsonify({"error": "Team reports are not available for this user type."}), 403
    if not (session.get("is_admin") or is_team_leader_session()):
        return jsonify({"error": "Team reports are available only for admins and team leaders."}), 403

    selection_statuses = ["Selected", "Offered", "Joined", "Hired"]
    placeholders = ",".join("?" * len(selection_statuses))
    conn = get_db()
    owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
    status_params = selection_statuses + owner_params
    rows = conn.execute(f"""
        SELECT
            COALESCE(NULLIF(tm.name,''), NULLIF(c.recruiter_name,''), c.recruiter_email, 'Unassigned') AS recruiter,
            SUM(CASE WHEN date(c.created_at)>=date('now','localtime','start of month') THEN 1 ELSE 0 END) AS current_month,
            COUNT(c.id) AS overall
        FROM candidates c
        LEFT JOIN team_members tm ON c.sourcer_id = tm.id
        WHERE COALESCE(c.is_duplicate,0)=0
          AND COALESCE(c.status,'') IN ({placeholders})
          {owner_sql}
        GROUP BY COALESCE(NULLIF(tm.name,''), NULLIF(c.recruiter_name,''), c.recruiter_email, 'Unassigned')
        ORDER BY current_month DESC, overall DESC, recruiter ASC
    """, status_params).fetchall()
    totals = conn.execute(f"""
        SELECT
            SUM(CASE WHEN date(c.created_at)>=date('now','localtime','start of month') THEN 1 ELSE 0 END) AS current_month,
            COUNT(c.id) AS overall,
            COUNT(DISTINCT COALESCE(c.sourcer_id, c.recruiter_email)) AS recruiters
        FROM candidates c
        WHERE COALESCE(c.is_duplicate,0)=0
          AND COALESCE(c.status,'') IN ({placeholders})
          {owner_sql}
    """, status_params).fetchone()
    conn.close()
    return jsonify({
        "month": date.today().strftime("%B %Y"),
        "statuses": selection_statuses,
        "totals": dict(totals) if totals else {"current_month": 0, "overall": 0, "recruiters": 0},
        "rows": [dict(r) for r in rows],
    })

@app.route("/api/team-analytics/submissions")
@login_required
def api_team_analytics_submissions():
    if is_client_viewer_session():
        return jsonify({"error": "Team analytics is not available for this user type."}), 403
    if not (session.get("is_admin") or is_team_leader_session()):
        return jsonify({"error": "Team analytics is available only for admins and team leaders."}), 403

    if not ANALYTICS_SEMAPHORE.acquire(blocking=False):
        return jsonify({
            "error": "Team analytics is busy right now. Please try again in a minute."
        }), 429

    conn = None
    try:
        conn = get_db()
        owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
        base_where = "WHERE COALESCE(c.is_duplicate,0)=0" + owner_sql
        today = date.today()
        month_start = today.replace(day=1)

        last14_rows = conn.execute(
            f"""
            SELECT date(c.created_at) AS day, COUNT(*) AS count
            FROM candidates c
            {base_where}
              AND date(c.created_at)>=date('now','localtime','-13 days')
            GROUP BY date(c.created_at)
            ORDER BY day
            """,
            owner_params,
        ).fetchall()
        month_rows = conn.execute(
            f"""
            SELECT date(c.created_at) AS day, COUNT(*) AS count
            FROM candidates c
            {base_where}
              AND date(c.created_at)>=date('now','localtime','start of month')
            GROUP BY date(c.created_at)
            ORDER BY day
            """,
            owner_params,
        ).fetchall()
        last14_counts = {row["day"]: int(row["count"] or 0) for row in last14_rows}
        month_counts = {row["day"]: int(row["count"] or 0) for row in month_rows}
        weekly_recruiter_rows = conn.execute(
            f"""
            SELECT
                COALESCE(NULLIF(TRIM(c.recruiter_name),''), tm.name, c.recruiter_email, 'Unassigned') AS recruiter,
                COALESCE(NULLIF(TRIM(c.recruiter_email),''), tm.email, '') AS email,
                COUNT(*) AS submissions
            FROM candidates c
            LEFT JOIN team_members tm
              ON tm.id = c.sourcer_id
              OR lower(tm.email)=lower(COALESCE(c.recruiter_email,''))
            {base_where}
              AND date(c.created_at)>=date('now','localtime','-6 days')
            GROUP BY
                COALESCE(NULLIF(TRIM(c.recruiter_name),''), tm.name, c.recruiter_email, 'Unassigned'),
                COALESCE(NULLIF(TRIM(c.recruiter_email),''), tm.email, '')
            ORDER BY submissions DESC, recruiter ASC
            """,
            owner_params,
        ).fetchall()

        last14 = []
        last14_total = 0
        for offset in range(13, -1, -1):
            day = today - timedelta(days=offset)
            iso_day = day.isoformat()
            count = last14_counts.get(iso_day, 0)
            last14_total += count
            last14.append({
                "date": iso_day,
                "label": day.strftime("%d %b"),
                "short_label": day.strftime("%d"),
                "count": count,
            })

        month = []
        month_total = 0
        day = month_start
        while day <= today:
            iso_day = day.isoformat()
            count = month_counts.get(iso_day, 0)
            month_total += count
            month.append({
                "date": iso_day,
                "label": day.strftime("%d %b"),
                "short_label": day.strftime("%d"),
                "count": count,
            })
            day += timedelta(days=1)

        return jsonify({
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "last14_total": last14_total,
            "month_total": month_total,
            "month_label": today.strftime("%B %Y"),
            "last14": last14,
            "month": month,
            "weekly_recruiters": [
                {
                    "recruiter": row["recruiter"] or "Unassigned",
                    "email": row["email"] or "",
                    "submissions": int(row["submissions"] or 0),
                }
                for row in weekly_recruiter_rows
            ],
        })
    finally:
        if conn:
            conn.close()
        ANALYTICS_SEMAPHORE.release()

@app.route("/api/filters")
@login_required
def api_filters():
    conn = get_db()
    roles   = [r[0] for r in conn.execute("SELECT DISTINCT role_name FROM candidates WHERE role_name!='' ORDER BY role_name").fetchall()]
    senders = [{"recruiter_email":r[0],"recruiter_name":r[1]} for r in
               conn.execute("SELECT DISTINCT recruiter_email,recruiter_name FROM candidates WHERE recruiter_email!='' ORDER BY recruiter_name").fetchall()]
    locs    = [r[0] for r in conn.execute("SELECT DISTINCT current_location FROM candidates WHERE current_location!='' ORDER BY current_location").fetchall()]
    hms     = [dict(r) for r in conn.execute("SELECT * FROM hiring_managers ORDER BY name").fetchall()]
    # All unique tags
    all_tags_raw = conn.execute("SELECT tags FROM candidates WHERE tags!=''").fetchall()
    tag_set = set()
    for row in all_tags_raw:
        for t in (row[0] or "").split(","):
            t = t.strip()
            if t: tag_set.add(t)
    # Pipelines query before closing
    pipelines = [dict(r) for r in conn.execute("SELECT * FROM pipelines ORDER BY role_name").fetchall()]
    for p in pipelines: p["status_list"] = json.loads(p["status_list"])
    conn.close()
    # Return all data
    return jsonify({"roles":roles,"senders":senders,"locations":locs,
                    "statuses":STATUSES,"tags":sorted(tag_set),
                    "submission_statuses":SUBMISSION_STATUSES,"hiring_managers":hms,
                    "pipelines":pipelines})

@app.route("/api/candidate_search_filters")
@login_required
def api_candidate_search_filters():
    conn = get_db()
    owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
    clients = [
        r[0] for r in conn.execute(f"""
            SELECT DISTINCT r.client_name
            FROM candidates c
            LEFT JOIN requirements r ON r.id=c.requirement_id
            WHERE COALESCE(r.client_name,'')!=''{owner_sql}
            ORDER BY r.client_name
        """, owner_params).fetchall()
    ]
    sourcers = [dict(r) for r in conn.execute(f"""
        SELECT DISTINCT c.recruiter_email AS email, c.recruiter_name AS name
        FROM candidates c
        WHERE COALESCE(c.recruiter_email,'')!=''{owner_sql}
        ORDER BY c.recruiter_name
    """, owner_params).fetchall()]
    conn.close()
    return jsonify({"clients": clients, "sourcers": sourcers})

@app.route("/api/stats")
@login_required
def api_stats():
    conn = get_db()
    sql, params = build_query(request.args, session)
    filtered_sql = strip_order_by(sql)
    total = conn.execute("SELECT COUNT(*) FROM (" + filtered_sql + " AND COALESCE(is_duplicate,0)=0) filtered_candidates", params).fetchone()[0]
    dups = conn.execute("SELECT COUNT(*) FROM (" + filtered_sql + " AND COALESCE(is_duplicate,0)<>0) filtered_candidates", params).fetchone()[0]
    roles = conn.execute("SELECT COUNT(DISTINCT role_name) FROM (" + filtered_sql + " AND COALESCE(is_duplicate,0)=0 AND COALESCE(role_name,'')<>'') filtered_candidates", params).fetchone()[0]
    unread = conn.execute("SELECT COUNT(*) FROM alerts WHERE is_read=0").fetchone()[0]
    last = conn.execute("SELECT uploaded_at FROM upload_log ORDER BY id DESC LIMIT 1").fetchone()
    status_counts = {
        (row["status"] or "New"): row["count"]
        for row in conn.execute(
            "SELECT COALESCE(status,'New') AS status, COUNT(*) AS count FROM (" + filtered_sql + " AND COALESCE(is_duplicate,0)=0) filtered_candidates GROUP BY COALESCE(status,'New')",
            params
        ).fetchall()
    }
    conn.close()

    return jsonify({
        "total": total,
        "duplicates": dups,
        "roles": roles,
        "unread_alerts": unread,
        "last_upload": last[0] if last else None,
        "status_counts": status_counts
    })

@app.route("/api/dashboard_summary")
@login_required
def api_dashboard_summary():
    endpoint_started = time.perf_counter()
    light = str(request.args.get("light") or "").strip().lower() in {"1", "true", "yes"}
    recruiter_view = "today" if str(request.args.get("recruiter_view") or "").strip().lower() == "today" else "month"
    client_view = "today" if str(request.args.get("client_view") or "").strip().lower() == "today" else "month"
    conn = get_db()
    owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
    base_where = "WHERE COALESCE(c.is_duplicate,0)=0" + owner_sql
    today_sql = base_where + " AND date(c.created_at)=date('now','localtime')"
    week_sql = base_where + " AND date(c.created_at)>=date('now','localtime','-6 days')"
    month_sql = base_where + " AND date(c.created_at)>=date('now','localtime','start of month')"
    selected_statuses = ["Selected", "Offered", "Joined", "Hired"]
    placeholders = ",".join("?" * len(selected_statuses))
    daily = conn.execute(f"SELECT COUNT(*) FROM candidates c {today_sql}", owner_params).fetchone()[0]
    weekly = conn.execute(f"SELECT COUNT(*) FROM candidates c {week_sql}", owner_params).fetchone()[0]
    monthly = conn.execute(f"SELECT COUNT(*) FROM candidates c {month_sql}", owner_params).fetchone()[0]
    monthly_selections = conn.execute(
        f"""
        SELECT COUNT(DISTINCT c.id)
        FROM candidates c
        {month_sql}
          AND COALESCE(c.status,'') IN ({placeholders})
        """,
        owner_params + selected_statuses
    ).fetchone()[0]
    requirements_worked = conn.execute(
        f"""
        SELECT COUNT(DISTINCT c.requirement_id)
        FROM candidates c
        {month_sql}
          AND c.requirement_id IS NOT NULL
        """,
        owner_params
    ).fetchone()[0]
    if light:
        conn.close()
        print(
            f"PERF dashboard_summary light elapsed_ms={(time.perf_counter() - endpoint_started) * 1000:.1f} "
            f"user={session.get('username') or '-'}",
            flush=True,
        )
        return jsonify({
            "light": True,
            "daily_submissions": daily,
            "weekly_submissions": weekly,
            "monthly_submissions": monthly,
            "monthly_selections": monthly_selections,
            "monthly_requirements_worked": requirements_worked,
            "daily_trend": [],
            "selection_trend": [],
            "status_breakdown": [],
            "requirement_breakdown": [],
            "recruiter_breakdown": [],
            "client_breakdown": [],
            "no_submission_today": [],
            "recruiter_view": recruiter_view,
            "client_view": client_view
        })
    trend_rows = conn.execute(
        f"""
        SELECT date(c.created_at) AS day, COUNT(*) AS count
        FROM candidates c
        {base_where}
          AND date(c.created_at)>=date('now','localtime','-13 days')
        GROUP BY date(c.created_at)
        ORDER BY day
        """,
        owner_params
    ).fetchall()
    trend_counts = {row["day"]: row["count"] for row in trend_rows}
    today = date.today()
    daily_trend = []
    for offset in range(13, -1, -1):
        day = today - timedelta(days=offset)
        iso_day = day.isoformat()
        daily_trend.append({
            "date": iso_day,
            "label": day.strftime("%d %b"),
            "count": trend_counts.get(iso_day, 0)
        })
    status_breakdown = [
        {"label": row["status"] or "New", "count": row["count"]}
        for row in conn.execute(
            f"""
            SELECT COALESCE(NULLIF(TRIM(c.status),''),'New') AS status, COUNT(*) AS count
            FROM candidates c
            {month_sql}
            GROUP BY COALESCE(NULLIF(TRIM(c.status),''),'New')
            ORDER BY count DESC, status
            LIMIT 6
            """,
            owner_params
        ).fetchall()
    ]
    requirement_breakdown = [
        {"label": row["title"] or "Unassigned", "count": row["count"]}
        for row in conn.execute(
            f"""
            SELECT COALESCE(NULLIF(TRIM(r.title),''),'Unassigned') AS title, COUNT(*) AS count
            FROM candidates c
            LEFT JOIN requirements r ON r.id = c.requirement_id
            {month_sql}
              AND c.requirement_id IS NOT NULL
            GROUP BY c.requirement_id, COALESCE(NULLIF(TRIM(r.title),''),'Unassigned')
            ORDER BY count DESC, title
            LIMIT 6
            """,
            owner_params
        ).fetchall()
    ]
    recruiter_breakdown = [
        {"label": row["recruiter"] or "Unassigned", "count": row["count"]}
        for row in conn.execute(
            f"""
            SELECT
                COALESCE(NULLIF(TRIM(c.recruiter_name),''), tm.name, c.recruiter_email, 'Unassigned') AS recruiter,
                COUNT(*) AS count
            FROM candidates c
            LEFT JOIN team_members tm ON c.sourcer_id = tm.id
            {month_sql}
            GROUP BY COALESCE(NULLIF(TRIM(c.recruiter_name),''), tm.name, c.recruiter_email, 'Unassigned')
            ORDER BY count DESC, recruiter
            LIMIT 8
            """,
            owner_params
        ).fetchall()
    ]
    selection_rows = conn.execute(
        f"""
        SELECT date(c.created_at) AS day, COUNT(DISTINCT c.id) AS count
        FROM candidates c
        {base_where}
          AND date(c.created_at)>=date('now','localtime','-13 days')
          AND COALESCE(c.status,'') IN ({placeholders})
        GROUP BY date(c.created_at)
        ORDER BY day
        """,
        owner_params + selected_statuses
    ).fetchall()
    selection_counts = {row["day"]: row["count"] for row in selection_rows}
    selection_trend = []
    for offset in range(13, -1, -1):
        day = today - timedelta(days=offset)
        iso_day = day.isoformat()
        selection_trend.append({
            "date": iso_day,
            "label": day.strftime("%d %b"),
            "count": selection_counts.get(iso_day, 0)
        })
    client_period_sql = today_sql if client_view == "today" else month_sql
    client_breakdown = [
        {"label": row["client"] or "Unassigned", "count": row["count"]}
        for row in conn.execute(
            f"""
            SELECT COALESCE(NULLIF(TRIM(r.client_name),''), 'Unassigned') AS client, COUNT(*) AS count
            FROM candidates c
            LEFT JOIN requirements r ON r.id = c.requirement_id
            {client_period_sql}
            GROUP BY COALESCE(NULLIF(TRIM(r.client_name),''), 'Unassigned')
            ORDER BY count DESC, client
            LIMIT 8
            """,
            owner_params
        ).fetchall()
    ]
    conn.close()
    print(
        f"PERF dashboard_summary full elapsed_ms={(time.perf_counter() - endpoint_started) * 1000:.1f} "
        f"user={session.get('username') or '-'}",
        flush=True,
    )
    return jsonify({
        "light": False,
        "daily_submissions": daily,
        "weekly_submissions": weekly,
        "monthly_submissions": monthly,
        "monthly_selections": monthly_selections,
        "monthly_requirements_worked": requirements_worked,
        "daily_trend": daily_trend,
        "selection_trend": selection_trend,
        "status_breakdown": status_breakdown,
        "requirement_breakdown": requirement_breakdown,
        "recruiter_breakdown": recruiter_breakdown,
        "client_breakdown": client_breakdown,
        "no_submission_today": [],
        "recruiter_view": recruiter_view,
        "client_view": client_view
    })

@app.route("/api/my/analytics")
@login_required
def api_my_analytics():
    if is_client_viewer_session():
        return jsonify({"error": "Analytics is not available for this user type."}), 403

    cache_key = (
        session.get("is_admin"),
        session.get("team_member_id"),
        (session.get("recruiter_email") or "").strip().lower(),
        (session.get("role") or "").strip().lower(),
    )
    now_ts = time.time()
    with ANALYTICS_CACHE_LOCK:
        cached = ANALYTICS_CACHE.get(cache_key)
        if cached and now_ts - cached["ts"] < ANALYTICS_CACHE_TTL_SECONDS:
            payload = dict(cached["payload"])
            payload["cached"] = True
            return jsonify(payload)

    if not ANALYTICS_SEMAPHORE.acquire(blocking=False):
        return jsonify({
            "error": "Analytics is busy right now. Please try again in a minute."
        }), 429

    conn = None
    try:
        conn = get_db()
        owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
        base_where = "WHERE COALESCE(c.is_duplicate,0)=0" + owner_sql
        today_sql = base_where + " AND date(c.created_at)=date('now','localtime')"
        week_sql = base_where + " AND date(c.created_at)>=date('now','localtime','-6 days')"
        month_sql = base_where + " AND date(c.created_at)>=date('now','localtime','start of month')"
        selection_statuses = ["Selected", "Offered", "Joined", "Hired"]
        placeholders = ",".join("?" * len(selection_statuses))

        daily = conn.execute(f"SELECT COUNT(*) FROM candidates c {today_sql}", owner_params).fetchone()[0]
        weekly = conn.execute(f"SELECT COUNT(*) FROM candidates c {week_sql}", owner_params).fetchone()[0]
        monthly = conn.execute(f"SELECT COUNT(*) FROM candidates c {month_sql}", owner_params).fetchone()[0]
        monthly_selections = conn.execute(
            f"""
            SELECT COUNT(DISTINCT c.id)
            FROM candidates c
            {month_sql}
              AND COALESCE(c.status,'') IN ({placeholders})
            """,
            owner_params + selection_statuses
        ).fetchone()[0]
        requirements_worked = conn.execute(
            f"""
            SELECT COUNT(DISTINCT c.requirement_id)
            FROM candidates c
            {month_sql}
              AND c.requirement_id IS NOT NULL
            """,
            owner_params
        ).fetchone()[0]

        payload = {
            "daily_submissions": daily,
            "weekly_submissions": weekly,
            "monthly_submissions": monthly,
            "monthly_selections": monthly_selections,
            "monthly_requirements_worked": requirements_worked,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "cached": False,
        }
        with ANALYTICS_CACHE_LOCK:
            ANALYTICS_CACHE[cache_key] = {"ts": now_ts, "payload": payload}
        return jsonify(payload)
    finally:
        if conn:
            conn.close()
        ANALYTICS_SEMAPHORE.release()

@app.route("/api/my/weekly_performance")
@login_required
def api_my_weekly_performance():
    if is_client_viewer_session():
        return jsonify({"error": "Weekly performance is not available for this user type."}), 403

    if not ANALYTICS_SEMAPHORE.acquire(blocking=False):
        return jsonify({
            "error": "Performance snapshot is busy right now. Please try again in a minute."
        }), 429

    conn = None
    try:
        conn = get_db()
        owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
        base_where = "WHERE COALESCE(c.is_duplicate,0)=0" + owner_sql
        rows = conn.execute(
            f"""
            SELECT date(c.created_at) AS day, COUNT(*) AS count
            FROM candidates c
            {base_where}
              AND date(c.created_at)>=date('now','localtime','-6 days')
            GROUP BY date(c.created_at)
            ORDER BY day
            """,
            owner_params,
        ).fetchall()
        counts = {row["day"]: row["count"] for row in rows}
        today = date.today()
        trend = []
        total = 0
        for offset in range(6, -1, -1):
            day = today - timedelta(days=offset)
            iso_day = day.isoformat()
            count = int(counts.get(iso_day, 0) or 0)
            total += count
            trend.append({
                "date": iso_day,
                "label": day.strftime("%d %b"),
                "short_label": day.strftime("%a"),
                "count": count,
            })
        return jsonify({
            "weekly_submissions": total,
            "trend": trend,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
    finally:
        if conn:
            conn.close()
        ANALYTICS_SEMAPHORE.release()

@app.route("/api/roles")
@login_required
def get_roles():
    conn = get_db()
    # Get standard roles grouped by domain
    std_roles = conn.execute("SELECT domain, role_name FROM standard_roles ORDER BY domain, role_name").fetchall()
    # Get dynamic roles from candidates
    if not session.get("is_admin") and session.get("team_member_id"):
        dyn_rows = conn.execute("SELECT DISTINCT current_role as role_name FROM candidates WHERE current_role != '' AND sourcer_id = ? ORDER BY current_role", (session["team_member_id"],)).fetchall()
    else:
        dyn_rows = conn.execute("SELECT DISTINCT current_role as role_name FROM candidates WHERE current_role != '' ORDER BY current_role").fetchall()
    conn.close()
    # Group standard roles by domain
    grouped = {}
    for r in std_roles:
        d = dict(r)
        domain = d.get('domain', 'Other')
        role = d.get('role_name', '')
        if domain not in grouped:
            grouped[domain] = []
        grouped[domain].append(role)
    return jsonify({"grouped": grouped, "dynamic": [dict(r) for r in dyn_rows]})

@app.route("/api/roles", methods=["POST"])
@login_required
def add_role():
    data = request.json or {}
    domain = data.get("domain", "").strip()
    role = data.get("role_name", "").strip()
    if not domain or not role:
        return jsonify({"error": "domain and role_name required"}), 400
    conn = get_db()
    conn.execute("INSERT OR IGNORE INTO standard_roles (domain, role_name) VALUES (?, ?)", (domain, role))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/statuses")
@login_required
def get_statuses():
    return jsonify([{"status": s} for s in get_candidate_status_master()])

@app.route("/api/admin/candidate-followups")
@login_required
def api_admin_candidate_followups():
    if not session.get("is_admin") or is_client_viewer_session():
        return jsonify({"error": "Admin access required"}), 403
    try:
        requested_limit = int(request.args.get("limit") or 150)
    except (TypeError, ValueError):
        requested_limit = 150
    limit = min(300, max(25, requested_limit))
    items = candidate_followup_items(session, admin=True, limit=limit)
    return jsonify({
        "items": items,
        "count": len(items),
        "valid_statuses": get_candidate_status_master(),
        "rules": {
            "stale_after_days": FOLLOWUP_STALE_RULES,
            "terminal_statuses": sorted(FOLLOWUP_TERMINAL_STATUSES),
        },
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })

@app.route("/api/admin/client-sla")
@login_required
def api_admin_client_sla():
    if not session.get("is_admin") or is_client_viewer_session():
        return jsonify({"error": "Admin access required"}), 403
    items = client_sla_items()
    return jsonify({
        "items": items,
        "count": len(items),
        "thresholds": CLIENT_SLA_THRESHOLDS,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })

@app.route("/api/admin/data-quality")
@login_required
def api_admin_data_quality():
    if not session.get("is_admin") or is_client_viewer_session():
        return jsonify({"error": "Admin access required"}), 403
    issues = data_quality_console_items()
    return jsonify({
        "issues": issues,
        "count": len(issues),
        "critical": sum(1 for item in issues if item["severity"] == "Critical"),
        "watch": sum(1 for item in issues if item["severity"] == "Watch"),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })

@app.route("/api/my/followups/today")
@login_required
def api_my_followups_today():
    if session.get("is_admin") or is_client_viewer_session():
        return jsonify({"show_alert": False, "count": 0, "items": []})
    items = candidate_followup_items(session, admin=False, limit=300)
    today = date.today().isoformat()
    team_member_id = session.get("team_member_id")
    recruiter_email = (session.get("recruiter_email") or session.get("email") or "").strip().lower()
    show_alert = False
    if items and (team_member_id or recruiter_email):
        conn = get_db(timeout=5)
        ensure_followup_alert_schema(conn)
        existing = conn.execute("""
            SELECT id FROM followup_daily_alerts
            WHERE COALESCE(team_member_id,0)=COALESCE(?,0)
              AND lower(COALESCE(recruiter_email,''))=lower(COALESCE(?, ''))
              AND alert_date=?
        """, (team_member_id, recruiter_email, today)).fetchone()
        show_alert = existing is None
        if show_alert:
            conn.execute("""
                INSERT OR IGNORE INTO followup_daily_alerts
                    (team_member_id, recruiter_email, alert_date, followup_count)
                VALUES (?, ?, ?, ?)
            """, (team_member_id, recruiter_email, today, len(items)))
            conn.commit()
        conn.close()
    return jsonify({
        "show_alert": show_alert,
        "count": len(items),
        "items": items[:8],
        "valid_statuses": get_candidate_status_master(),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })

@app.route("/api/my/followups")
@login_required
def api_my_followups():
    if session.get("is_admin") or is_client_viewer_session():
        return jsonify({"items": [], "count": 0, "valid_statuses": get_candidate_status_master()})
    try:
        requested_limit = int(request.args.get("limit") or 150)
    except (TypeError, ValueError):
        requested_limit = 150
    limit = min(300, max(25, requested_limit))
    items = candidate_followup_items(session, admin=False, limit=limit)
    return jsonify({
        "items": items,
        "count": len(items),
        "valid_statuses": get_candidate_status_master(),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })

@app.route("/api/skills")
@login_required
def get_skills():
    conn = get_db()
    rows = conn.execute("SELECT id, skill_name, category FROM standard_skills ORDER BY category, skill_name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/clients")
@login_required
def get_clients():
    conn = get_db()
    show_all = str(request.args.get("all") or "").strip().lower() in {"1", "true", "yes", "all"}
    if show_all or session.get("is_admin") or not session.get("team_member_id"):
        rows = conn.execute("SELECT id, client_name FROM clients ORDER BY client_name").fetchall()
    else:
        rows = conn.execute("""
            SELECT c.id, c.client_name
            FROM clients c
            JOIN team_client_mappings m ON m.client_id = c.id
            WHERE m.team_member_id=?
            ORDER BY c.client_name
        """, (session.get("team_member_id"),)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/clients", methods=["POST"])
@login_required
def add_client():
    data = request.json or {}
    client_name = normalize_client_name_value(data.get("client_name", ""))
    if not client_name:
        return jsonify({"error": "client_name required"}), 400
    conn = get_db()
    existing = find_client_by_normalized_name(conn, client_name)
    if existing:
        conn.close()
        return jsonify({"ok": True, "client": dict(existing), "created": False})
    client = ensure_client_exists(conn, client_name)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "client": dict(client) if client else {"client_name": client_name}, "created": True})

@app.route("/api/taggd-recruiters", methods=["GET", "POST"])
@login_required
def taggd_recruiters():
    if request.method == "POST":
        forbidden = client_viewer_write_forbidden()
        if forbidden:
            return forbidden
    conn = get_db()
    ensure_taggd_recruiter_schema(conn)
    if request.method == "GET":
        client_name = (request.args.get("client_name") or "").strip()
        include_inactive = str(request.args.get("include_inactive") or "").strip().lower() in {"1", "true", "yes"}
        where_parts = ["1=1"] if include_inactive else ["tr.is_active=1"]
        params = []
        if client_name:
            where_parts.append("lower(trim(c.client_name))=lower(trim(?))")
            params.append(client_name)
        allowed_clients = mapped_client_names_for_current_user(conn)
        if allowed_clients is not None:
            if not allowed_clients:
                conn.close()
                return jsonify([])
            placeholders = ",".join("?" * len(allowed_clients))
            where_parts.append(f"lower(trim(c.client_name)) IN ({placeholders})")
            params.extend(sorted(allowed_clients))
        rows = conn.execute(f"""
            SELECT tr.id, tr.name, tr.email, tr.is_active, c.client_name
            FROM taggd_recruiters tr
            JOIN clients c ON c.id=tr.client_id
            WHERE {" AND ".join(where_parts)}
            ORDER BY c.client_name, tr.name
        """, params).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip()
    client_name = normalize_client_name_value(data.get("client_name") or "")
    if not name or not client_name:
        conn.close()
        return jsonify({"error": "Taggd recruiter name and client are required."}), 400
    if not current_user_can_use_client(conn, client_name):
        conn.close()
        return jsonify({"error": "You can create Taggd recruiters only for clients mapped to you."}), 403
    client = ensure_client_exists(conn, client_name)
    try:
        rid = conn.execute("""
            INSERT INTO taggd_recruiters (name, email, client_id, created_by)
            VALUES (?,?,?,?)
        """, (name, email, client["id"], session.get("username") or "")).lastrowid
    except sqlite3.IntegrityError:
        row = conn.execute("""
            SELECT id FROM taggd_recruiters WHERE client_id=? AND lower(trim(name))=lower(trim(?)) LIMIT 1
        """, (client["id"], name)).fetchone()
        rid = row["id"] if row else None
        if rid:
            conn.execute("UPDATE taggd_recruiters SET is_active=1, email=? WHERE id=?", (email, rid))
    conn.commit()
    row = conn.execute("""
        SELECT tr.id, tr.name, tr.email, tr.is_active, c.client_name
        FROM taggd_recruiters tr
        JOIN clients c ON c.id=tr.client_id
        WHERE tr.id=?
    """, (rid,)).fetchone()
    conn.close()
    return jsonify({"ok": True, "taggd_recruiter": dict(row) if row else {"id": rid, "name": name, "client_name": client_name}})

@app.route("/api/taggd-recruiters/<int:rid>", methods=["PATCH"])
@login_required
def update_taggd_recruiter(rid):
    forbidden = client_viewer_write_forbidden()
    if forbidden:
        return forbidden
    data = request.get_json(silent=True) or {}
    conn = get_db()
    ensure_taggd_recruiter_schema(conn)
    current = conn.execute("""
        SELECT tr.*, c.client_name
        FROM taggd_recruiters tr
        JOIN clients c ON c.id=tr.client_id
        WHERE tr.id=?
    """, (rid,)).fetchone()
    if not current:
        conn.close()
        return jsonify({"error": "Taggd recruiter not found."}), 404
    if not current_user_can_use_client(conn, current["client_name"]):
        conn.close()
        return jsonify({"error": "You can update Taggd recruiters only for clients mapped to you."}), 403
    updates = []
    params = []
    if "name" in data:
        name = (data.get("name") or "").strip()
        if not name:
            conn.close()
            return jsonify({"error": "Taggd recruiter name is required."}), 400
        duplicate = conn.execute("""
            SELECT id FROM taggd_recruiters
            WHERE id<>? AND client_id=? AND lower(trim(name))=lower(trim(?))
            LIMIT 1
        """, (rid, current["client_id"], name)).fetchone()
        if duplicate:
            conn.close()
            return jsonify({"error": "A Taggd recruiter with this name already exists for this client."}), 409
        updates.append("name=?")
        params.append(name)
    if "email" in data:
        updates.append("email=?")
        params.append((data.get("email") or "").strip())
    if "is_active" in data:
        updates.append("is_active=?")
        params.append(1 if data.get("is_active") else 0)
    if not updates:
        conn.close()
        return jsonify({"ok": True})
    params.append(rid)
    conn.execute(f"UPDATE taggd_recruiters SET {', '.join(updates)} WHERE id=?", params)
    if "name" in data:
        conn.execute("""
            UPDATE requirements
            SET taggd_recruiter_name=?
            WHERE taggd_recruiter_id=?
        """, ((data.get("name") or "").strip(), rid))
    conn.commit()
    row = conn.execute("""
        SELECT tr.id, tr.name, tr.email, tr.is_active, c.client_name
        FROM taggd_recruiters tr
        JOIN clients c ON c.id=tr.client_id
        WHERE tr.id=?
    """, (rid,)).fetchone()
    conn.close()
    return jsonify({"ok": True, "taggd_recruiter": dict(row) if row else None})

@app.route("/api/clients/<int:client_id>", methods=["PATCH"])
@login_required
def update_client(client_id):
    if not session.get("is_admin"):
        return jsonify({"error": "Admin only"}), 403
    data = request.json or {}
    client_name = (data.get("client_name") or "").strip()
    if not client_name:
        return jsonify({"error": "client_name required"}), 400
    conn = get_db()
    current = conn.execute("SELECT client_name FROM clients WHERE id=?", (client_id,)).fetchone()
    if not current:
        conn.close()
        return jsonify({"error": "Client not found"}), 404
    duplicate = conn.execute(
        "SELECT id FROM clients WHERE id<>? AND lower(trim(client_name))=lower(trim(?)) LIMIT 1",
        (client_id, client_name)
    ).fetchone()
    if duplicate:
        conn.close()
        return jsonify({"error": "Client already exists"}), 409
    old_name = current["client_name"]
    conn.execute("UPDATE clients SET client_name=? WHERE id=?", (client_name, client_id))
    conn.execute("UPDATE requirements SET client_name=? WHERE client_name=?", (client_name, old_name))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/team-client-mappings", methods=["GET"])
@login_required
def get_team_client_mappings():
    if not session.get("is_admin"):
        return jsonify({"error": "Admin only"}), 403
    conn = get_db()
    conn.execute("""CREATE TABLE IF NOT EXISTS team_client_mappings (
        team_member_id INTEGER NOT NULL,
        client_id INTEGER NOT NULL,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        PRIMARY KEY(team_member_id, client_id)
    )""")
    rows = conn.execute("""
        SELECT m.team_member_id, m.client_id, c.client_name
        FROM team_client_mappings m
        JOIN clients c ON c.id = m.client_id
        ORDER BY c.client_name
    """).fetchall()
    conn.close()
    mappings = {}
    for row in rows:
        key = str(row["team_member_id"])
        mappings.setdefault(key, []).append({"id": row["client_id"], "client_name": row["client_name"]})
    return jsonify(mappings)

@app.route("/api/my/clients", methods=["GET", "PUT"])
@login_required
def my_clients():
    team_member_id = session.get("team_member_id")
    if not team_member_id:
        return jsonify({"error": "No recruiter profile linked to this user"}), 400
    conn = get_db()
    conn.execute("""CREATE TABLE IF NOT EXISTS team_client_mappings (
        team_member_id INTEGER NOT NULL,
        client_id INTEGER NOT NULL,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        PRIMARY KEY(team_member_id, client_id)
    )""")
    if request.method == "GET":
        rows = conn.execute("""
            SELECT c.id, c.client_name
            FROM team_client_mappings m
            JOIN clients c ON c.id = m.client_id
            WHERE m.team_member_id=?
            ORDER BY c.client_name
        """, (team_member_id,)).fetchall()
        conn.close()
        return jsonify({
            "team_member_id": team_member_id,
            "clients": [dict(r) for r in rows]
        })

    data = request.json or {}
    raw_names = data.get("client_names") or []
    raw_ids = data.get("client_ids") or []
    client_ids = []
    client_names = []
    for raw_id in raw_ids:
        try:
            client_ids.append(int(raw_id))
        except (TypeError, ValueError):
            pass
    for raw_name in raw_names:
        clean_name = normalize_client_name_value(raw_name)
        if clean_name:
            client_names.append(clean_name)
    if client_names and not client_ids:
        for name in client_names:
            client = ensure_client_exists(conn, name)
            if client:
                client_ids.append(int(client["id"]))
    client_ids = sorted(set(client_ids))
    if client_ids:
        placeholders = ",".join("?" * len(client_ids))
        found = conn.execute(f"SELECT id FROM clients WHERE id IN ({placeholders})", client_ids).fetchall()
        if len(found) != len(client_ids):
            conn.close()
            return jsonify({"error": "One or more clients were not found"}), 400
    conn.execute("DELETE FROM team_client_mappings WHERE team_member_id=?", (team_member_id,))
    for client_id in client_ids:
        conn.execute(
            "INSERT OR IGNORE INTO team_client_mappings (team_member_id, client_id) VALUES (?,?)",
            (team_member_id, client_id)
        )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "count": len(client_ids)})

@app.route("/api/my/gemini_key", methods=["GET", "PUT"])
@login_required
def my_gemini_key():
    team_member_id = session.get("team_member_id")
    if not team_member_id:
        return jsonify({"error": "No recruiter profile linked to this user"}), 400
    conn = get_db(timeout=5)
    ensure_ats_pipeline_schema(conn)
    if request.method == "GET":
        payload = get_team_member_gemini_key_status(team_member_id, conn=conn)
        conn.close()
        return jsonify({
            "team_member_id": team_member_id,
            **payload,
        })

    data = request.get_json(silent=True) or {}
    clear = bool(data.get("clear"))
    api_key = str(data.get("api_key") or "").strip()
    if not clear and not api_key:
        conn.close()
        return jsonify({"error": "Paste a Gemini API key or use Clear to remove your personal key."}), 400

    encrypted_key = "" if clear else encrypt_gemini_api_key(api_key)
    if not clear and not encrypted_key:
        conn.close()
        return jsonify({"error": "Unable to encrypt Gemini API key. Please try again."}), 500

    try:
        with DB_WRITE_LOCK:
            conn.execute(
                """UPDATE team_members
                   SET gemini_api_key_enc=?,
                       gemini_api_key_updated_at=datetime('now','localtime')
                   WHERE id=?""",
                (encrypted_key or None, team_member_id)
            )
            conn.commit()
    finally:
        try:
            conn.close()
        except Exception:
            pass
    return jsonify({
        "ok": True,
        "team_member_id": team_member_id,
        **get_team_member_gemini_key_status(team_member_id)
    })

@app.route("/api/team/<int:team_member_id>/clients", methods=["GET", "PUT"])
@login_required
def team_member_clients(team_member_id):
    if not session.get("is_admin"):
        return jsonify({"error": "Admin only"}), 403
    conn = get_db()
    conn.execute("""CREATE TABLE IF NOT EXISTS team_client_mappings (
        team_member_id INTEGER NOT NULL,
        client_id INTEGER NOT NULL,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        PRIMARY KEY(team_member_id, client_id)
    )""")
    member = conn.execute("SELECT id FROM team_members WHERE id=?", (team_member_id,)).fetchone()
    if not member:
        conn.close()
        return jsonify({"error": "Team member not found"}), 404
    if request.method == "GET":
        rows = conn.execute("""
            SELECT c.id, c.client_name
            FROM team_client_mappings m
            JOIN clients c ON c.id = m.client_id
            WHERE m.team_member_id=?
            ORDER BY c.client_name
        """, (team_member_id,)).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])

    data = request.json or {}
    client_ids = []
    for raw_id in data.get("client_ids", []):
        try:
            client_ids.append(int(raw_id))
        except (TypeError, ValueError):
            pass
    client_ids = sorted(set(client_ids))
    if client_ids:
        placeholders = ",".join("?" * len(client_ids))
        found = conn.execute(f"SELECT id FROM clients WHERE id IN ({placeholders})", client_ids).fetchall()
        if len(found) != len(client_ids):
            conn.close()
            return jsonify({"error": "One or more clients were not found"}), 400
    conn.execute("DELETE FROM team_client_mappings WHERE team_member_id=?", (team_member_id,))
    for client_id in client_ids:
        conn.execute(
            "INSERT OR IGNORE INTO team_client_mappings (team_member_id, client_id) VALUES (?,?)",
            (team_member_id, client_id)
        )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "count": len(client_ids)})

@app.route("/api/team-leader-mappings", methods=["GET"])
@login_required
def get_team_leader_mappings():
    if not session.get("is_admin"):
        return jsonify({"error": "Admin only"}), 403
    conn = get_db()
    conn.execute("""CREATE TABLE IF NOT EXISTS team_leader_mappings (
        leader_team_member_id INTEGER NOT NULL,
        member_team_member_id INTEGER NOT NULL,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        PRIMARY KEY(leader_team_member_id, member_team_member_id)
    )""")
    rows = conn.execute("""
        SELECT m.leader_team_member_id, m.member_team_member_id,
               leader.name AS leader_name, leader.email AS leader_email,
               member.name AS member_name, member.email AS member_email
        FROM team_leader_mappings m
        JOIN team_members leader ON leader.id = m.leader_team_member_id
        JOIN team_members member ON member.id = m.member_team_member_id
        ORDER BY leader.name, member.name
    """).fetchall()
    conn.close()
    mappings = {}
    for row in rows:
        key = str(row["leader_team_member_id"])
        mappings.setdefault(key, {
            "leader_team_member_id": row["leader_team_member_id"],
            "leader_name": row["leader_name"],
            "leader_email": row["leader_email"],
            "members": [],
        })
        mappings[key]["members"].append({
            "id": row["member_team_member_id"],
            "name": row["member_name"],
            "email": row["member_email"],
        })
    return jsonify(list(mappings.values()))

@app.route("/api/team-leader/<int:leader_team_member_id>/members", methods=["GET", "PUT"])
@login_required
def team_leader_members(leader_team_member_id):
    if not session.get("is_admin"):
        return jsonify({"error": "Admin only"}), 403
    conn = get_db()
    conn.execute("""CREATE TABLE IF NOT EXISTS team_leader_mappings (
        leader_team_member_id INTEGER NOT NULL,
        member_team_member_id INTEGER NOT NULL,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        PRIMARY KEY(leader_team_member_id, member_team_member_id)
    )""")
    leader = conn.execute("SELECT id, role FROM team_members WHERE id=?", (leader_team_member_id,)).fetchone()
    if not leader:
        conn.close()
        return jsonify({"error": "Team leader not found"}), 404
    if not is_team_leader_session({"role": leader["role"], "team_member_id": leader["id"]}):
        conn.execute("UPDATE team_members SET role=? WHERE id=?", ("Team Leader", leader_team_member_id))

    if request.method == "GET":
        rows = conn.execute("""
            SELECT tm.id, tm.name, tm.email, tm.role
            FROM team_leader_mappings m
            JOIN team_members tm ON tm.id = m.member_team_member_id
            WHERE m.leader_team_member_id=?
            ORDER BY tm.name
        """, (leader_team_member_id,)).fetchall()
        conn.close()
        return jsonify({
            "leader_team_member_id": leader_team_member_id,
            "members": [dict(r) for r in rows]
        })

    data = request.get_json(silent=True) or {}
    raw_ids = data.get("member_team_member_ids") or data.get("member_ids") or []
    member_ids = []
    for raw_id in raw_ids:
        try:
            member_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        if member_id and member_id != leader_team_member_id:
            member_ids.append(member_id)
    member_ids = sorted(set(member_ids))
    if member_ids:
        placeholders = ",".join("?" * len(member_ids))
        found = conn.execute(f"SELECT id FROM team_members WHERE id IN ({placeholders})", member_ids).fetchall()
        if len(found) != len(member_ids):
            conn.close()
            return jsonify({"error": "One or more selected team members were not found"}), 400
    conn.execute("DELETE FROM team_leader_mappings WHERE leader_team_member_id=?", (leader_team_member_id,))
    for member_id in member_ids:
        conn.execute(
            "INSERT OR IGNORE INTO team_leader_mappings (leader_team_member_id, member_team_member_id) VALUES (?,?)",
            (leader_team_member_id, member_id)
        )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "leader_team_member_id": leader_team_member_id, "count": len(member_ids)})

@app.route("/api/candidate/<int:cid>/tags", methods=["PATCH"])
@login_required
def update_tags(cid):
    tags = request.json.get("tags","")
    # Normalise: strip whitespace from each tag
    clean = ",".join(t.strip() for t in tags.split(",") if t.strip())
    conn  = get_db()
    conn.execute("UPDATE candidates SET tags=?, updated_at=datetime('now','localtime') WHERE id=?",
                 (clean, cid))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/alerts")
@login_required
def api_alerts():
    conn  = get_db()
    atype = request.args.get("type","")
    sql   = "SELECT * FROM alerts WHERE 1=1"; p=[]
    if atype: sql+=" AND alert_type=?"; p.append(atype)
    sql  += " ORDER BY created_at DESC LIMIT 150"
    rows  = conn.execute(sql,p).fetchall(); conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/alerts/read",methods=["POST"])
@login_required
def mark_all_read():
    conn=get_db(); conn.execute("UPDATE alerts SET is_read=1"); conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/alerts/<int:aid>/read",methods=["POST"])
@login_required
def mark_one_read(aid):
    conn=get_db(); conn.execute("UPDATE alerts SET is_read=1 WHERE id=?",(aid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# â”€â”€ Saved searches â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/saved_searches", methods=["GET"])
@login_required
def get_saved_searches():
    conn = get_db()
    rows = conn.execute("SELECT * FROM saved_searches ORDER BY created_at DESC").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/saved_searches", methods=["POST"])
@login_required
def save_search():
    data = request.json
    name    = (data.get("name") or "").strip()
    filters = data.get("filters", {})
    if not name: return jsonify({"error":"Name required"}),400
    conn = get_db()
    sid = conn.execute("INSERT INTO saved_searches (name,filters) VALUES (?,?)",
                       (name, json.dumps(filters))).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":sid})

@app.route("/api/candidates/bulk_delete", methods=["POST"])
@login_required
def bulk_delete_candidates():
    forbidden = client_viewer_write_forbidden()
    if forbidden:
        return forbidden
    data = request.json or {}
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"error": "no ids provided"}), 400
    
    conn = get_db()
    
    # Ownership check for non-admins
    if not session.get("is_admin"):
        tm_id = session.get("team_member_id")
        placeholders = ",".join(["?"]*len(ids))
        # Only allow deleting candidates owned by this user
        if tm_id:
            conn.execute(f"DELETE FROM candidates WHERE id IN ({placeholders}) AND sourcer_id=?", ids + [tm_id])
        else:
            conn.execute(f"DELETE FROM candidates WHERE id IN ({placeholders}) AND sourcer_id IS NULL", ids)
    else:
        placeholders = ",".join(["?"]*len(ids))
        conn.execute(f"DELETE FROM candidates WHERE id IN ({placeholders})", ids)
    
    conn.commit(); conn.close()
    return jsonify({"ok": True, "deleted": len(ids)})

# â”€â”€ Job Details â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/job_details", methods=["GET"])
@login_required
def get_job_details():
    cid = request.args.get("candidate_id") or request.args.get("cid")
    if not cid:
        return jsonify([])
    conn = get_db()
    rows = conn.execute("SELECT * FROM job_details WHERE candidate_id=? ORDER BY created_at DESC", (cid,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/job_details", methods=["POST"])
@login_required
def add_job_details():
    data = request.json or {}
    cid = data.get("candidate_id")
    if not cid:
        return jsonify({"error":"candidate_id required"}), 400
    conn = get_db()
    jd = conn.execute(
        "INSERT INTO job_details (candidate_id,title,location,salary,department,job_id,notes,tags) VALUES (?,?,?,?,?,?,?,?)",
        (cid, data.get("title",""), data.get("location",""), data.get("salary",""),
         data.get("department",""), data.get("job_id",""), data.get("notes",""), data.get("tags",""))
    ).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok": True, "id": jd})

@app.route("/api/saved_searches/<int:sid>", methods=["DELETE"])
@login_required
def delete_saved_search(sid):
    conn = get_db(); conn.execute("DELETE FROM saved_searches WHERE id=?",(sid,))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

# â”€â”€ Reporting â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/reporting/recruiter_performance")
@login_required
def recruiter_performance():
    conn  = get_db()
    days  = int(request.args.get("days",30))
    since = (date.today()-timedelta(days=days)).isoformat()
    rows  = conn.execute("""
        SELECT recruiter_name,recruiter_email,COUNT(*) as total,
            SUM(CASE WHEN is_duplicate=0 THEN 1 ELSE 0 END) as unique_candidates,
            SUM(CASE WHEN is_duplicate=1 THEN 1 ELSE 0 END) as duplicates,
            COUNT(DISTINCT role_name) as roles_worked,
            COUNT(DISTINCT date(created_at)) as active_days
        FROM candidates WHERE created_at>=?
        GROUP BY recruiter_email ORDER BY unique_candidates DESC""",(since,)).fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/reporting/daily_trend")
@login_required
def daily_trend():
    conn  = get_db()
    days  = int(request.args.get("days",14))
    since = (date.today()-timedelta(days=days)).isoformat()
    rows  = conn.execute("""
        SELECT date(created_at) as day,COUNT(*) as total,
            SUM(CASE WHEN is_duplicate=0 THEN 1 ELSE 0 END) as unique_c,
            COUNT(DISTINCT recruiter_email) as active_recruiters
        FROM candidates WHERE created_at>=?
        GROUP BY date(created_at) ORDER BY day""",(since,)).fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/reporting/no_submission_today")
@login_required
def no_submission_today():
    conn  = get_db(); today=date.today().isoformat()
    submitted={r[0] for r in conn.execute(
        "SELECT DISTINCT recruiter_email FROM candidates WHERE date(created_at)=?",(today,)).fetchall()}
    excluded_names = {
        "surindersingh",
        "reetusaini",
        "surinder",
        "reetu",
        "prithra",
    }
    missing=[]
    for r in conn.execute("SELECT name,email FROM team_members").fetchall():
        name = str(r["name"] or "").strip()
        email = str(r["email"] or "").strip()
        normalized_name = re.sub(r"[^a-z0-9]+", "", name.lower())
        if normalized_name in excluded_names:
            continue
        if email not in submitted:
            missing.append({"name": name, "email": email})
    conn.close(); return jsonify(missing)

@app.route("/api/reporting/send_weekly", methods=["POST"])
@login_required
def api_send_weekly():
    return jsonify(send_weekly_email())

@app.route("/api/upload_cv", methods=["POST"])
@login_required
def upload_cv_file():
    forbidden = client_viewer_write_forbidden()
    if forbidden:
        return forbidden
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files["file"]

    if not f.filename:
        return jsonify({"error": "Empty filename"}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in [".pdf", ".doc", ".docx"]:
        return jsonify({"error": "Invalid file type"}), 400

    try:
        upload_folder = os.path.join(app.root_path, "uploads", "recruiters", current_upload_owner_key(), "candidates", "single", "unassigned")
        os.makedirs(upload_folder, exist_ok=True)

        saved_name = hrguru_cv_filename(f.filename, upload_folder)
        file_path = os.path.join(upload_folder, saved_name)

        f.save(file_path)

        file_url = f"/uploads/{os.path.relpath(file_path, os.path.join(app.root_path, 'uploads')).replace(os.sep, '/')}"

        parsed = None
        try:
            parsed = parse_cv(file_path)
            if parsed:
                parsed["candidate_name"] = normalize_person_name(parsed.get("candidate_name", ""))
                if not parsed.get("candidate_name"):
                    parsed["candidate_name"] = infer_candidate_name_from_filename(f.filename)
                parsed["candidate_name"] = normalize_person_name(parsed.get("candidate_name", ""))
                parsed = {
                    "candidate_name": parsed.get("candidate_name", ""),
                    "email_addr": parsed.get("email_addr", ""),
                    "phone": parsed.get("phone", ""),
                    "current_company": parsed.get("current_company", ""),
                    "current_role": parsed.get("current_role", ""),
                    "experience_years": parsed.get("experience_years", ""),
                    "key_skills": parsed.get("key_skills", ""),
                    "notice_period": parsed.get("notice_period", ""),
                    "current_salary": parsed.get("current_salary", ""),
                    "expected_salary": parsed.get("expected_salary", ""),
                    "current_location": parsed.get("current_location", ""),
                    "education": parsed.get("education", ""),
                    "cv_summary": parsed.get("cv_summary", ""),
                    "_parse_warning": parsed.get("_parse_warning", "")
                }
        except Exception as pe:
            print("CV Parse Error:", pe)

        response = jsonify({
            "ok": True,
            "original_filename": f.filename,
            "parse_token": request.form.get("parse_token", ""),
            "filename": saved_name,
            "url": file_url,
            "public_id": saved_name,
            "parsed": parsed or {}
        })
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        return response

    except Exception as e:
        print("Upload Error:", str(e))
        return jsonify({"error": str(e)}), 500

@app.route("/api/candidate/<int:cid>/upload_cv", methods=["POST"])
@login_required
def upload_candidate_cv(cid):
    forbidden = client_viewer_write_forbidden()
    if forbidden:
        return forbidden
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    if not session.get("is_admin"):
        conn = get_db()
        owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
        c = conn.execute(f"SELECT c.id FROM candidates c WHERE c.id=?{owner_sql}", [cid] + owner_params).fetchone()
        conn.close()
        if not c:
            return jsonify({"error": "Permission denied"}), 403
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "Empty filename"}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in [".pdf", ".doc", ".docx"]:
        return jsonify({"error": "Invalid file type"}), 400
    upload_folder = os.path.join(app.root_path, "uploads", "recruiters", current_upload_owner_key(), "candidates", "single", f"candidate_{cid}")
    os.makedirs(upload_folder, exist_ok=True)
    saved_name = hrguru_cv_filename(f.filename, upload_folder)
    file_path = os.path.join(upload_folder, saved_name)
    f.save(file_path)
    data = {
        "ok": True,
        "original_filename": f.filename,
        "parse_token": request.form.get("parse_token", ""),
        "filename": saved_name,
        "url": f"/uploads/{os.path.relpath(file_path, os.path.join(app.root_path, 'uploads')).replace(os.sep, '/')}",
        "public_id": saved_name,
        "parsed": {}
    }
    try:
        parsed = parse_cv(file_path)
        if parsed:
            parsed["candidate_name"] = normalize_person_name(parsed.get("candidate_name", ""))
            if not parsed.get("candidate_name"):
                parsed["candidate_name"] = infer_candidate_name_from_filename(f.filename)
            parsed["candidate_name"] = normalize_person_name(parsed.get("candidate_name", ""))
            data["parsed"] = {
                "candidate_name": parsed.get("candidate_name", ""),
                "email_addr": parsed.get("email_addr", ""),
                "phone": parsed.get("phone", ""),
                "current_company": parsed.get("current_company", ""),
                "current_role": parsed.get("current_role", ""),
                "experience_years": parsed.get("experience_years", ""),
                "key_skills": parsed.get("key_skills", ""),
                "notice_period": parsed.get("notice_period", ""),
                "current_salary": parsed.get("current_salary", ""),
                "expected_salary": parsed.get("expected_salary", ""),
                "current_location": parsed.get("current_location", ""),
                "education": parsed.get("education", ""),
                "cv_summary": parsed.get("cv_summary", ""),
                "_parse_warning": parsed.get("_parse_warning", "")
            }
    except Exception as pe:
        print("CV Parse Error:", pe)
    conn = get_db()
    conn.execute(
        """UPDATE candidates
           SET cv_filename=?, cv_url=?, cv_public_id=?, updated_at=datetime('now','localtime')
           WHERE id=?""",
        (data.get("filename", ""), data.get("url", ""), data.get("public_id", ""), cid)
    )
    conn.commit()
    conn.close()
    queue_candidate_ai_screening(cid, trigger="candidate_cv_upload")
    response = jsonify(data)
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response

@app.route("/api/jd_match", methods=["POST"])
@login_required
def api_jd_match():
    if JD_CV_MATCHING_DISABLED:
        return jsonify({"error": JD_CV_MATCHING_DISABLED_MESSAGE, "disabled": True}), 410
    jd_text_input = (request.form.get("jd_text") or "").strip()
    custom_hard_filters = (request.form.get("custom_hard_filters") or "").strip()
    override_min_years = (request.form.get("override_min_years") or "").strip()
    override_max_years = (request.form.get("override_max_years") or "").strip()
    reviewed_jd_json = (request.form.get("reviewed_jd_json") or request.form.get("parsed_jd_json") or "").strip()
    if not jd_text_input and "jd_file" not in request.files:
        return jsonify({"error": "Please upload a JD file or paste the JD text."}), 400
    cv_files = request.files.getlist("cv_files") or request.files.getlist("cv_file")
    cv_files = [file for file in cv_files if file and file.filename]
    if not cv_files:
        return jsonify({"error": "Please upload the candidate CV file."}), 400

    match_request_started = time.perf_counter()
    def log_match_stage(stage, started_at=None, extra=""):
        baseline = started_at or match_request_started
        elapsed_ms = (time.perf_counter() - baseline) * 1000
        suffix = f" {extra}" if extra else ""
        print(f"MATCH TIMING: {stage} took {elapsed_ms:.0f} ms{suffix}", flush=True)

    try:
        schema_conn = get_db()
        ensure_ats_pipeline_schema(schema_conn)
        schema_conn.commit()
        schema_conn.close()
        jd_path, jd_name = "", "Pasted JD"
        if "jd_file" in request.files and request.files["jd_file"].filename:
            jd_path, jd_name = save_uploaded_analysis_file(request.files["jd_file"], "jd", "match")
        jd_text = jd_text_input or extract_cv_text(jd_path)
        if not jd_text.strip():
            return jsonify({"error": "Could not read text from the job description."}), 400
        log_match_stage("JD text extraction", None, f"file={jd_name}")
        reviewed_jd = None
        if reviewed_jd_json:
            try:
                reviewed_jd = json.loads(reviewed_jd_json)
                if not isinstance(reviewed_jd, dict):
                    return jsonify({"error": "Reviewed JD payload is invalid."}), 400
            except Exception:
                return jsonify({"error": "Reviewed JD payload is invalid."}), 400
        override_key = f"\nOVERRIDE_EXPERIENCE:{override_min_years}:{override_max_years}"
        reviewed_marker = ("\nREVIEWED_JD:\n" + json.dumps(reviewed_jd, sort_keys=True, ensure_ascii=False)) if reviewed_jd else ""
        jd_hash = versioned_text_hash(jd_text + "\nCUSTOM_HARD_FILTERS:\n" + custom_hard_filters + reviewed_marker + override_key)
        if reviewed_jd:
            print("MATCH DEBUG: Recruiter-reviewed JD is being used.", flush=True)
            parsed_jd = json.loads(json.dumps(reviewed_jd))
            parsed_jd["parse_source"] = "recruiter_reviewed"
            parsed_jd["parser_confidence"] = parsed_jd.get("parser_confidence") or "reviewed"
            parsed_jd["review_applied"] = True
            if override_min_years or override_max_years:
                try:
                    parsed_jd["experience_required"] = {
                        "min_years": float(override_min_years) if override_min_years else 0,
                        "max_years": float(override_max_years) if override_max_years else 0
                    }
                    parsed_jd["experience_override_applied"] = True
                except ValueError:
                    return jsonify({"error": "Experience override must be a number."}), 400
            set_cached_parsed_jd(jd_hash, parsed_jd)
        else:
            jd_parse_started = time.perf_counter()
            parsed_jd = get_cached_parsed_jd(jd_hash)
            if not parsed_jd:
                print("MATCH DEBUG: No cached parsed JD found; parsing JD now.", flush=True)
                parsed_jd = parse_jd_structured(jd_text)
                if override_min_years or override_max_years:
                    try:
                        parsed_jd["experience_required"] = {
                            "min_years": float(override_min_years) if override_min_years else 0,
                            "max_years": float(override_max_years) if override_max_years else 0
                        }
                        parsed_jd["experience_override_applied"] = True
                    except ValueError:
                        return jsonify({"error": "Experience override must be a number."}), 400
                set_cached_parsed_jd(jd_hash, parsed_jd)
            else:
                print("MATCH DEBUG: Cached parsed JD is being used.", flush=True)
                if override_min_years or override_max_years:
                    parsed_jd = json.loads(json.dumps(parsed_jd))
                    try:
                        parsed_jd["experience_required"] = {
                            "min_years": float(override_min_years) if override_min_years else 0,
                            "max_years": float(override_max_years) if override_max_years else 0
                        }
                        parsed_jd["experience_override_applied"] = True
                        print("MATCH DEBUG: Cached parsed JD is being used with experience override applied.", flush=True)
                    except ValueError:
                        return jsonify({"error": "Experience override must be a number."}), 400
        log_match_stage("JD parse/cache resolution", jd_parse_started, f"source={parsed_jd.get('parse_source', '')}")
        results = []
        errors = []
        for cv_index, cv_file in enumerate(cv_files):
            candidate_started = time.perf_counter()
            cv_path, cv_name = save_uploaded_analysis_file(cv_file, "cv", "match")
            log_match_stage(f"CV[{cv_index + 1}/{len(cv_files)}] upload save", candidate_started, f"file={cv_name}")
            cv_text_started = time.perf_counter()
            cv_text = extract_cv_text(cv_path)
            if not cv_text.strip():
                errors.append({"cv_filename": cv_name, "error": "Could not read text from this CV."})
                continue
            log_match_stage(f"CV[{cv_index + 1}/{len(cv_files)}] text extraction", cv_text_started, f"file={cv_name}")
            cache_started = time.perf_counter()
            resume_hash = versioned_text_hash(cv_text)
            cached = get_cached_match_result(jd_hash, resume_hash)
            log_match_stage(f"CV[{cv_index + 1}/{len(cv_files)}] match cache lookup", cache_started, f"file={cv_name} hit={bool(cached)}")
            if cached:
                print(f"MATCH DEBUG: Cached match result is being used for {cv_name}.", flush=True)
                cached["dashboard"] = build_match_dashboard(cached)
                cached["candidate_summary"] = build_candidate_summary_from_cv(cv_path, cv_text)
                cached["jd_filename"] = jd_name
                cached["cv_filename"] = cv_name
                cached["cache_hit"] = True
                if not cached.get("gemini_screening_report"):
                    if (os.getenv("GEMINI_API_KEY") or "").strip():
                        print(f"MATCH DEBUG: Gemini screening report is being generated for cached candidate {cv_name}.", flush=True)
                    gemini_api_key, _gemini_api_source = resolve_team_member_gemini_key(session.get("team_member_id"))
                    screening = maybe_generate_gemini_screening_report(
                        jd_text,
                        cv_text,
                        candidate_name=(cached.get("dashboard") or {}).get("candidate_snapshot", {}).get("candidate_name", "") or (cached.get("parsed_candidate") or {}).get("candidate_name", ""),
                        target_job_title=(cached.get("parsed_jd") or {}).get("role_title", "") or (cached.get("parsed_jd") or {}).get("title", "") or (cached.get("dashboard") or {}).get("role_family_comparison", {}).get("jd_family", ""),
                        parsed_jd=cached.get("parsed_jd") or cached.get("jd_json") or {},
                        parsed_candidate=cached.get("parsed_candidate") or cached.get("cv_json") or {},
                        api_key=gemini_api_key,
                    )
                    if screening.get("ok"):
                        cached = apply_gemini_match_report(cached, screening)
                        cached["dashboard"] = build_match_dashboard(cached)
                        cached.setdefault("dashboard", {})["gemini_screening_report"] = cached["gemini_screening_report"]
                        cached.setdefault("dashboard", {})["gemini_screening_source"] = cached["gemini_screening_source"]
                        cached.setdefault("dashboard", {})["gemini_screening_model"] = cached["gemini_screening_model"]
                        cached.setdefault("dashboard", {})["scoring_source"] = cached.get("scoring_source", "gemini")
                        persist_match_artifacts_async(jd_hash, resume_hash, cached)
                    else:
                        print(f"MATCH DEBUG: Gemini screening report fallback is being used for cached candidate {cv_name}: {screening.get('error', 'unknown error')}", flush=True)
                        cached["gemini_screening_error"] = screening.get("error", "Gemini report unavailable.")
                queue_match_audit(
                    "match_cache_hit",
                    event_type="match_cache_hit",
                    object_type="match",
                    object_hash=f"{jd_hash}:{resume_hash}",
                    jd_hash=jd_hash,
                    resume_hash=resume_hash,
                    pipeline_version=MATCH_PIPELINE_VERSION,
                    status=str(cached.get("verdict") or ""),
                    source="cache",
                    parser_confidence=str((cached.get("parsed_jd") or {}).get("parser_confidence", "")),
                    manual_review_required=bool((cached.get("dashboard") or {}).get("manual_review", {}).get("required")),
                    score=cached.get("final_score", 0),
                    message="Cached match result reused",
                    details={
                        "validation_gaps": (cached.get("dashboard") or {}).get("validation_gaps", [])[:8],
                    }
                )
                results.append(cached)
                log_match_stage(f"CV[{cv_index + 1}/{len(cv_files)}] total cached candidate flow", candidate_started, f"file={cv_name}")
                continue
            parse_started = time.perf_counter()
            parsed_candidate = get_cached_parsed_resume(resume_hash)
            if parsed_candidate:
                print(f"MATCH DEBUG: Cached parsed resume is being used for {cv_name}.", flush=True)
                parsed_cv = {}
            elif (os.getenv("GEMINI_API_KEY") or "").strip():
                print(f"MATCH DEBUG: Gemini resume parsing will be used for {cv_name}; skipping ATS CV pre-parse.", flush=True)
                parsed_cv = {}
            else:
                print(f"MATCH DEBUG: ATS CV parser fallback is being used for {cv_name} before structured resume parsing.", flush=True)
                parsed_cv = parse_cv(cv_path) or {}
            log_match_stage(f"CV[{cv_index + 1}/{len(cv_files)}] pre-parse decision", parse_started, f"file={cv_name} parsed_cache={bool(parsed_candidate)}")
            hybrid_started = time.perf_counter()
            result = run_hybrid_match(
                jd_text,
                cv_text,
                parsed_cv,
                cache_get=get_cached_embedding,
                cache_set=set_cached_embedding,
                custom_hard_filters=custom_hard_filters,
                parsed_jd=parsed_jd,
                parsed_candidate=parsed_candidate
            )
            log_match_stage(f"CV[{cv_index + 1}/{len(cv_files)}] deterministic match", hybrid_started, f"file={cv_name} score={result.get('final_score', result.get('score', 0))}")
            if not parsed_candidate:
                set_cached_parsed_resume(resume_hash, result.get("parsed_candidate") or result.get("cv_json") or {})
            dashboard_started = time.perf_counter()
            result["dashboard"] = build_match_dashboard(result)
            log_match_stage(f"CV[{cv_index + 1}/{len(cv_files)}] dashboard build", dashboard_started, f"file={cv_name}")
            screening_delay = 0
            if len(cv_files) > 1:
                try:
                    screening_delay = float(os.getenv("GEMINI_BULK_DELAY_SECONDS", "5"))
                except Exception:
                    screening_delay = 5
            if (os.getenv("GEMINI_API_KEY") or "").strip():
                print(f"MATCH DEBUG: Gemini screening report is being generated for {cv_name}.", flush=True)
            screening_started = time.perf_counter()
            gemini_api_key, _gemini_api_source = resolve_team_member_gemini_key(session.get("team_member_id"))
            screening = maybe_generate_gemini_screening_report(
                jd_text,
                cv_text,
                candidate_name=(result.get("dashboard") or {}).get("candidate_snapshot", {}).get("candidate_name", "") or (result.get("parsed_candidate") or {}).get("candidate_name", ""),
                target_job_title=(result.get("parsed_jd") or {}).get("role_title", "") or (result.get("parsed_jd") or {}).get("title", "") or (result.get("dashboard") or {}).get("role_family_comparison", {}).get("jd_family", ""),
                parsed_jd=result.get("parsed_jd") or result.get("jd_json") or {},
                parsed_candidate=result.get("parsed_candidate") or result.get("cv_json") or {},
                batch_delay_seconds=screening_delay if len(cv_files) > 1 and cv_index < len(cv_files) - 1 else 0,
                api_key=gemini_api_key,
            )
            log_match_stage(f"CV[{cv_index + 1}/{len(cv_files)}] Gemini screening", screening_started, f"file={cv_name} ok={bool(screening.get('ok'))}")
            if screening.get("ok"):
                result = apply_gemini_match_report(result, screening)
                result["dashboard"] = build_match_dashboard(result)
                result.setdefault("dashboard", {})["gemini_screening_report"] = result["gemini_screening_report"]
                result.setdefault("dashboard", {})["gemini_screening_source"] = result["gemini_screening_source"]
                result.setdefault("dashboard", {})["gemini_screening_model"] = result["gemini_screening_model"]
                result.setdefault("dashboard", {})["scoring_source"] = result.get("scoring_source", "gemini")
            else:
                print(f"MATCH DEBUG: Gemini screening report fallback is being used for {cv_name}: {screening.get('error', 'unknown error')}", flush=True)
                result["gemini_screening_error"] = screening.get("error", "Gemini report unavailable.")
            persist_started = time.perf_counter()
            persist_match_artifacts_async(jd_hash, resume_hash, result)
            log_match_stage(f"CV[{cv_index + 1}/{len(cv_files)}] persist artifacts", persist_started, f"file={cv_name}")
            result["candidate_summary"] = build_candidate_summary_from_cv(cv_path, cv_text)
            result["jd_filename"] = jd_name
            result["jd_source"] = "text" if jd_text_input else "file"
            result["cv_filename"] = cv_name
            result["jd_hash"] = jd_hash
            result["resume_hash"] = resume_hash
            result["pipeline_version"] = MATCH_PIPELINE_VERSION
            result["custom_hard_filters"] = custom_hard_filters
            if override_min_years or override_max_years:
                result["experience_override"] = {
                    "min_years": float(override_min_years) if override_min_years else 0,
                    "max_years": float(override_max_years) if override_max_years else 0
                }
            result["cache_hit"] = False
            results.append(result)
            log_match_stage(f"CV[{cv_index + 1}/{len(cv_files)}] total candidate flow", candidate_started, f"file={cv_name} score={result.get('final_score', result.get('score', 0))}")
        results.sort(key=lambda item: int(item.get("final_score", item.get("score", 0)) or 0), reverse=True)
        log_match_stage("Match request total", match_request_started, f"candidates={len(results)} errors={len(errors)}")
        if len(cv_files) == 1 and results and not request.files.getlist("cv_files"):
            return jsonify(results[0])
        return jsonify({
            "batch": True,
            "jd_filename": jd_name,
            "jd_hash": jd_hash,
            "count": len(results),
            "errors": errors,
            "results": results,
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print("JD match error:", e)
        return jsonify({"error": str(e)}), 500

@app.route("/api/jd_match/cache/reset", methods=["POST"])
@login_required
def api_reset_jd_match_cache():
    if not session.get("is_admin"):
        return jsonify({"error": "Admin access required."}), 403
    data = request.get_json(silent=True) or {}
    scope = str(data.get("scope") or "matching").strip().lower()
    if scope not in {"matching", "all"}:
        scope = "matching"
    try:
        deleted = reset_matching_caches()
        conn = get_db()
        ensure_ats_pipeline_schema(conn)
        queue_match_audit(
            "match_cache_reset",
            event_type="match_cache_reset",
            object_type="match_cache",
            object_hash=scope,
            pipeline_version=MATCH_PIPELINE_VERSION,
            status="cleared",
            source="admin_reset",
            message="Match caches reset",
            details={
                "scope": scope,
                "deleted_rows": deleted,
            }
        )
        return jsonify({"ok": True, "scope": scope, "deleted": deleted})
    except Exception as e:
        print("Match cache reset error:", e, flush=True)
        traceback.print_exc()
        return jsonify({"error": str(e) or "Unable to reset match cache."}), 500

@app.route("/api/parse_jd", methods=["POST"])
@login_required
def api_parse_jd_structured():
    data = request.get_json(silent=True) or {}
    jd_text = (data.get("jd_text") or request.form.get("jd_text") or "").strip()
    if not jd_text and "jd_file" in request.files and request.files["jd_file"].filename:
        jd_path, _ = save_uploaded_analysis_file(request.files["jd_file"], "jd", "parse")
        jd_text = extract_cv_text(jd_path)
    if not jd_text:
        return jsonify({"error": "Please upload a JD file or paste JD text."}), 400
    parsed = parse_jd_structured(jd_text)
    queue_match_audit(
        "parse_jd",
        event_type="parse_jd",
        object_type="jd",
        object_hash=versioned_text_hash(jd_text),
        pipeline_version=MATCH_PIPELINE_VERSION,
        status="parsed",
        source=parsed.get("parse_source", ""),
        parser_confidence=parsed.get("parser_confidence", ""),
        manual_review_required=bool(parsed.get("manual_review_required")),
        message="JD parsed via structured parser",
        details={
            "warnings": parsed.get("parser_warnings", []),
            "role_title": parsed.get("role_title", ""),
            "role_family": parsed.get("role_family", ""),
            "domain_family": parsed.get("domain_family", ""),
        }
    )
    return jsonify({"ok": True, "parsed_jd": parsed})

@app.route("/api/parse_cv", methods=["POST"])
@login_required
def api_parse_cv_structured():
    data = request.get_json(silent=True) or {}
    cv_text = (data.get("cv_text") or request.form.get("cv_text") or "").strip()
    if not cv_text:
        return jsonify({"error": "Please provide CV text."}), 400
    parsed = parse_resume_structured(cv_text, {})
    queue_match_audit(
        "parse_cv",
        event_type="parse_cv",
        object_type="cv",
        object_hash=versioned_text_hash(cv_text),
        pipeline_version=MATCH_PIPELINE_VERSION,
        status="parsed",
        source=parsed.get("parse_source", ""),
        parser_confidence=parsed.get("experience_metrics", {}).get("experience_confidence", ""),
        manual_review_required=bool(parsed.get("manual_review_required")),
        message="CV parsed via structured parser",
        details={
            "warnings": parsed.get("parser_warnings", []),
            "candidate_name": parsed.get("candidate_name", ""),
            "current_role": parsed.get("current_role", ""),
            "role_family": parsed.get("role_family", ""),
            "domain_family": parsed.get("domain_family", ""),
        }
    )
    return jsonify(parsed)

@app.route("/api/candidate_resume/analyze", methods=["POST"])
@login_required
def api_candidate_resume_analyze():
    if "resume_file" not in request.files or not request.files["resume_file"].filename:
        return jsonify({"error": "Please upload a resume file."}), 400
    try:
        cv_path, cv_name = save_uploaded_analysis_file(request.files["resume_file"], "resume", "analysis")
        cv_text = extract_cv_text(cv_path)
        if not cv_text.strip():
            return jsonify({"error": "Could not read text from this resume."}), 400
        parsed_cv = parse_cv(cv_path) or {}
        deterministic_candidate = parse_resume_structured(cv_text, parsed_cv)
        sections = extract_resume_sections(cv_text)
        skill_check = check_skills_in_resume(cv_text, deterministic_candidate, request.form.get("skills_to_check", ""))
        llm_summary, llm_error = call_resume_analysis_llm(
            cv_text,
            deterministic_candidate,
            sections,
            request.form.get("custom_prompt", ""),
        )
        if isinstance(llm_summary.get("summary"), dict):
            llm_summary = llm_summary["summary"]
        fallback_summary = {
            "candidate_name": deterministic_candidate.get("candidate_name", ""),
            "total_experience": str(deterministic_candidate.get("total_experience_years", "")) + " years" if deterministic_candidate.get("total_experience_years") else "",
            "current_role": deterministic_candidate.get("current_role", ""),
            "current_company": deterministic_candidate.get("current_company", ""),
            "contact_details": deterministic_candidate.get("contact", {}),
            "organization_experience": deterministic_candidate.get("role_history", []),
            "technical_skills": [{"skill": skill, "experience": "Not clear from resume", "evidence": ""} for skill in (deterministic_candidate.get("normalized_skills") or [])[:15]],
            "education_details": deterministic_candidate.get("education", []),
            "current_location": deterministic_candidate.get("location", ""),
        }
        summary = {**fallback_summary, **{k: v for k, v in llm_summary.items() if v}}
        queue_match_audit(
            "resume_analysis",
            event_type="resume_analysis",
            object_type="cv",
            object_hash=versioned_text_hash(cv_text),
            pipeline_version=MATCH_PIPELINE_VERSION,
            status="analyzed",
            source=deterministic_candidate.get("parse_source", ""),
            parser_confidence=deterministic_candidate.get("experience_metrics", {}).get("experience_confidence", ""),
            manual_review_required=bool(deterministic_candidate.get("manual_review_required")),
            message="Candidate resume analysis completed",
            details={
                "warnings": deterministic_candidate.get("parser_warnings", []),
                "llm_error": llm_error,
                "skill_check": skill_check,
            }
        )
        return jsonify({
            "ok": True,
            "filename": cv_name,
            "sections": sections,
            "summary": summary,
            "skill_check": skill_check,
            "parsed_candidate": deterministic_candidate,
            "llm_error": llm_error,
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print("Candidate resume analysis error:", e)
        return jsonify({"error": str(e)}), 500

@app.route("/api/candidate_resume/default_prompt")
@login_required
def api_candidate_resume_default_prompt():
    return jsonify({"prompt": DEFAULT_RESUME_ANALYSIS_PROMPT})

@app.route("/api/match_jd_cv", methods=["POST"])
@login_required
def api_match_jd_cv_structured():
    if JD_CV_MATCHING_DISABLED:
        return jsonify({"error": JD_CV_MATCHING_DISABLED_MESSAGE, "disabled": True}), 410
    data = request.get_json(silent=True) or {}
    jd_text = (data.get("jd_text") or "").strip()
    cv_text = (data.get("cv_text") or "").strip()
    if not jd_text or not cv_text:
        return jsonify({"error": "Please provide both jd_text and cv_text."}), 400
    custom_hard_filters = (data.get("custom_hard_filters") or "").strip()
    result = run_hybrid_match(
        jd_text,
        cv_text,
        {},
        cache_get=get_cached_embedding,
        cache_set=set_cached_embedding,
        custom_hard_filters=custom_hard_filters
    )
    result["dashboard"] = build_match_dashboard(result)
    queue_match_audit(
        "direct_match",
        event_type="direct_match",
        object_type="match",
        object_hash=f"{versioned_text_hash(jd_text)}:{versioned_text_hash(cv_text)}",
        jd_hash=versioned_text_hash(jd_text),
        resume_hash=versioned_text_hash(cv_text),
        pipeline_version=MATCH_PIPELINE_VERSION,
        status=str(result.get("verdict") or ""),
        source="api_match_jd_cv",
        parser_confidence=str((result.get("parsed_jd") or {}).get("parser_confidence", "")),
        manual_review_required=bool((result.get("dashboard") or {}).get("manual_review", {}).get("required")),
        score=result.get("final_score", 0),
        message=str((result.get("dashboard") or {}).get("manual_review", {}).get("summary") or result.get("recruiter_summary") or ""),
        details={
            "validation_gaps": (result.get("dashboard") or {}).get("validation_gaps", [])[:8],
            "role_family": (result.get("dashboard") or {}).get("role_family_comparison", {}),
        }
    )
    return jsonify(result)

def save_screening_questions_to_bank(questions, source_model=""):
    clean_questions = [q for q in questions or [] if q.get("skill") and q.get("question")]
    if not clean_questions:
        return 0
    conn = get_db()
    saved = 0
    for q in clean_questions:
        skill = canonical_skill(q.get("skill") or "")
        question = clean_value(q.get("question") or "", 500)
        expected_signal = clean_value(q.get("expected_signal") or "", 500)
        follow_up = clean_value(q.get("follow_up") or "", 500)
        if not skill or not question:
            continue
        cursor = conn.execute(
            """INSERT OR IGNORE INTO screening_question_bank
               (skill, question, expected_signal, follow_up, source_model)
               VALUES (?,?,?,?,?)""",
            (skill, question, expected_signal, follow_up, source_model)
        )
        saved += cursor.rowcount if cursor.rowcount and cursor.rowcount > 0 else 0
    conn.commit()
    conn.close()
    return saved

def get_screening_questions_from_bank(skills, per_skill=2):
    conn = get_db()
    questions = []
    used_ids = []
    for skill in skills or []:
        canonical = canonical_skill(skill)
        rows = conn.execute(
            """SELECT id, skill, question, expected_signal, follow_up
               FROM screening_question_bank
               WHERE lower(skill)=lower(?)
               ORDER BY usage_count ASC, created_at DESC
               LIMIT ?""",
            (canonical, per_skill)
        ).fetchall()
        for row in rows:
            questions.append({
                "skill": row["skill"],
                "question": row["question"],
                "expected_signal": row["expected_signal"],
                "follow_up": row["follow_up"],
            })
            used_ids.append(row["id"])
    if used_ids:
        placeholders = ",".join("?" for _ in used_ids)
        conn.execute(
            f"""UPDATE screening_question_bank
                SET usage_count=usage_count+1, last_used_at=datetime('now','localtime')
                WHERE id IN ({placeholders})""",
            used_ids
        )
        conn.commit()
    conn.close()
    return questions

@app.route("/api/match/export_pdf", methods=["POST"])
@login_required
def export_match_pdf():
    if JD_CV_MATCHING_DISABLED:
        return jsonify({"error": JD_CV_MATCHING_DISABLED_MESSAGE, "disabled": True}), 410
    try:
        analysis = request.get_json(silent=True) or {}
        if not analysis:
            return jsonify({"error": "No match analysis data received."}), 400
        dashboard = analysis.get("dashboard") or build_match_dashboard(analysis)
        analysis["dashboard"] = dashboard
        buf = build_match_pdf(analysis)
        candidate = (dashboard.get("candidate_snapshot") or {}).get("candidate_name") or "candidate"
        safe_candidate = re.sub(r"[^A-Za-z0-9_-]+", "_", candidate).strip("_") or "candidate"
        return send_file(
            buf,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"match_analysis_{safe_candidate}_{date.today()}.pdf"
        )
    except Exception as e:
        print("Match PDF export error:", e)
        return jsonify({"error": str(e)}), 500

@app.route("/api/match/screening_questions", methods=["POST"])
@login_required
def generate_match_screening_questions():
    data = request.get_json(silent=True) or {}
    if request.form:
        data = dict(request.form)
        if data.get("analysis_json"):
            try:
                data["analysis"] = json.loads(data.get("analysis_json") or "{}")
            except Exception:
                data["analysis"] = {}
    analysis = data.get("analysis") or {}
    parsed_jd = analysis.get("parsed_jd") or analysis.get("jd_json") or {}
    jd_text = (data.get("jd_text") or "").strip()
    if not jd_text and "jd_file" in request.files and request.files["jd_file"].filename:
        jd_path, _ = save_uploaded_analysis_file(request.files["jd_file"], "jd", "screening")
        jd_text = extract_cv_text(jd_path)
    if not parsed_jd and jd_text:
        parsed_jd = parse_jd_structured(jd_text)
    must_skills = parsed_jd.get("must_have_skills") or [
        item.get("skill") for item in parsed_jd.get("must_have_skills_weighted", []) or [] if item.get("skill")
    ]
    must_skills = [skill for skill in must_skills if skill][:8]
    if not must_skills:
        return jsonify({"error": "No mandatory skills found in the JD."}), 400
    candidate = analysis.get("parsed_candidate") or analysis.get("cv_json") or {}
    jd_context = {
        "role_title": parsed_jd.get("role_title") or parsed_jd.get("title") or "",
        "experience_required": parsed_jd.get("experience_required") or {},
        "location": parsed_jd.get("location") or "",
        "employment_type": parsed_jd.get("employment_type") or "",
        "responsibilities": (parsed_jd.get("responsibilities") or [])[:8],
        "domain": parsed_jd.get("domain") or "",
        "education_required": parsed_jd.get("education_required") or [],
        "certifications_required": parsed_jd.get("certifications_required") or [],
    }
    prompt = (
        "Generate first-level recruiter phone-screen questions from the actual job description context. "
        "These questions are for a recruiter, not a technical panel, so keep them simple, conversational, and easy to ask over a phone call. "
        "Generate 2 to 3 short screening questions for each mandatory skill. "
        "Do not ask definition-style questions. Do not make the recruiter debug code or judge deep architecture. "
        "Each question should help a recruiter check whether the candidate has real hands-on exposure, recent usage, project context, and confidence level. "
        "Return JSON with a questions array. Each item must have skill, question, expected_signal, and follow_up. "
        f"JD context JSON: {json.dumps(jd_context, ensure_ascii=False)}. "
        f"Mandatory skills: {', '.join(must_skills)}. "
        f"Candidate current role: {candidate.get('current_role','')}. "
        f"Candidate skills: {', '.join((candidate.get('normalized_skills') or [])[:20])}."
    )
    llm_error_reason = ""
    try:
        provider = os.getenv("LLM_PROVIDER", "openrouter").lower()
        if provider == "openrouter":
            api_key = os.getenv("OPENROUTER_API_KEY", "")
            if not api_key:
                raise RuntimeError("OPENROUTER_API_KEY is not configured.")
            response = requests.post(
                f"{os.getenv('OPENROUTER_API_BASE', 'https://openrouter.ai/api/v1').rstrip('/')}/chat/completions",
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                    "HTTP-Referer": os.getenv("PUBLIC_BASE_URL", "https://hireflow.hrgp.in"),
                    "X-Title": "HRGuru ATS",
                },
                json={
                    "model": os.getenv("OPENROUTER_MODEL", os.getenv("LLM_MODEL", "inclusionai/ring-2.6-1t:free")),
                    "messages": [
                        {"role": "system", "content": "Return compact JSON only. No markdown."},
                        {"role": "user", "content": prompt},
                    ],
                    "response_format": {"type": "json_object"},
                    "max_tokens": 2200,
                },
                timeout=float(os.getenv("OPENROUTER_QUESTION_TIMEOUT", "30"))
            )
            response.raise_for_status()
            raw = ((response.json().get("choices") or [{}])[0].get("message") or {}).get("content") or "{}"
        elif provider == "openai":
            api_key = os.getenv("OPENAI_API_KEY", "")
            if not api_key:
                raise RuntimeError("OPENAI_API_KEY is not configured.")
            response = requests.post(
                f"{os.getenv('OPENAI_API_BASE', 'https://api.openai.com/v1').rstrip('/')}/chat/completions",
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json={
                    "model": os.getenv("LLM_MODEL", "gpt-5-mini"),
                    "messages": [
                        {"role": "system", "content": "Return compact JSON only. No markdown."},
                        {"role": "user", "content": prompt},
                    ],
                    "response_format": {"type": "json_object"},
                    "max_completion_tokens": 2200,
                },
                timeout=float(os.getenv("OPENAI_QUESTION_TIMEOUT", "25"))
            )
            response.raise_for_status()
            raw = ((response.json().get("choices") or [{}])[0].get("message") or {}).get("content") or "{}"
        else:
            response = requests.post(
                f"{os.getenv('LLM_API_BASE', 'http://localhost:11434').rstrip('/')}/api/generate",
                json={"model": os.getenv("LLM_MODEL", "qwen2.5:7b"), "prompt": prompt, "format": "json", "stream": False},
                timeout=float(os.getenv("OLLAMA_QUESTION_TIMEOUT", "12"))
            )
            response.raise_for_status()
            raw = response.json().get("response") or "{}"
        parsed = json.loads(raw)
        questions = parsed.get("questions", []) if isinstance(parsed, dict) else []
        if questions:
            save_screening_questions_to_bank(questions, os.getenv("OPENROUTER_MODEL", os.getenv("LLM_MODEL", "")) if provider == "openrouter" else os.getenv("LLM_MODEL", ""))
    except Exception as e:
        llm_error_reason = str(e)
        if os.getenv("LLM_PROVIDER", "openrouter").lower() == "openrouter" and os.getenv("OPENAI_API_KEY", ""):
            try:
                response = requests.post(
                    f"{os.getenv('OPENAI_API_BASE', 'https://api.openai.com/v1').rstrip('/')}/chat/completions",
                    headers={"Authorization": f"Bearer {os.getenv('OPENAI_API_KEY')}", "Content-Type": "application/json"},
                    json={
                        "model": os.getenv("OPENAI_FALLBACK_MODEL", "gpt-5-mini"),
                        "messages": [
                            {"role": "system", "content": "Return compact JSON only. No markdown."},
                            {"role": "user", "content": prompt},
                        ],
                        "response_format": {"type": "json_object"},
                        "max_completion_tokens": 2200,
                    },
                    timeout=float(os.getenv("OPENAI_QUESTION_TIMEOUT", "25"))
                )
                response.raise_for_status()
                raw = ((response.json().get("choices") or [{}])[0].get("message") or {}).get("content") or "{}"
                parsed = json.loads(raw)
                questions = parsed.get("questions", []) if isinstance(parsed, dict) else []
                if questions:
                    save_screening_questions_to_bank(questions, os.getenv("OPENAI_FALLBACK_MODEL", "gpt-5-mini"))
            except Exception as fallback_error:
                llm_error_reason = f"{llm_error_reason}; OpenAI fallback: {fallback_error}"
                questions = []
        else:
            questions = []
    if not questions:
        bank_questions = get_screening_questions_from_bank(must_skills, per_skill=2)
        if bank_questions:
            print("MATCH DEBUG: Screening question bank fallback is being used.", flush=True)
            return jsonify({"ok": True, "questions": bank_questions[:24], "skills": must_skills, "source": "question_bank", "note": "The configured LLM was not reachable, so saved screening questions from the question bank were used."})
        question_patterns = [
            (
                "Have you used {skill} in a real project? Briefly tell me what you used it for.",
                "Look for recent hands-on exposure and a concrete project example.",
                "Was this something you owned directly or supported as part of a team?"
            ),
            (
                "When did you last work with {skill}, and for how long?",
                "Look for recency and actual duration of usage.",
                "Was it used in production or only during training/proof of concept?"
            ),
            (
                "What was the most important task you completed using {skill}?",
                "Look for a clear business or delivery outcome, not a keyword mention.",
                "Who reviewed or used the output of your work?"
            ),
        ]
        questions = []
        for skill in must_skills[:8]:
            for question, expected_signal, follow_up in question_patterns[:2]:
                questions.append({
                    "skill": skill,
                    "question": question.format(skill=skill),
                    "expected_signal": expected_signal,
                    "follow_up": follow_up
                })
        reason = "Please verify the configured LLM provider, API key, and model name in environment settings."
        if "404" in (llm_error_reason or ""):
            reason = "The configured LLM model or endpoint was not found. Please verify OPENROUTER_MODEL/LLM_MODEL and the API base URL."
        print("MATCH DEBUG: Simple screening question fallback is being used.", flush=True)
        return jsonify({"ok": True, "questions": questions[:24], "skills": must_skills, "source": "fallback", "note": f"The configured LLM was not reachable, so simple recruiter screening questions were generated. {reason}"})
    return jsonify({"ok": True, "questions": questions[:24], "skills": must_skills, "source": "llm"})

@app.route("/api/match/screening_questions_pdf", methods=["POST"])
@login_required
def export_screening_questions_pdf():
    try:
        payload = request.get_json(silent=True) or {}
        questions = payload.get("questions") or []
        if not questions:
            return jsonify({"error": "No screening questions received."}), 400
        buf = build_screening_questions_pdf(payload)
        return send_file(
            buf,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"interview_screening_questions_{date.today()}.pdf"
        )
    except Exception as e:
        print("Screening PDF export error:", e)
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/switch-profile", methods=["POST"])
@login_required
def api_admin_switch_profile():
    data = request.get_json(silent=True) or {}
    action = str(data.get("action") or "").strip().lower()
    if action == "restore":
        original = session.get("original_admin_session") or {}
        if not session.get("impersonation_active") or not original:
            return jsonify({"error": "No switched profile session is active."}), 400
        google_token = session.get("google_token")
        session.clear()
        for key, value in original.items():
            session[key] = value
        if google_token:
            session["google_token"] = google_token
        session["impersonation_active"] = 0
        return jsonify({"ok": True, "restored": True})

    if not session.get("is_admin") or session.get("impersonation_active"):
        return jsonify({"error": "Switch profile is available only from the admin profile."}), 403

    team_member_id = parse_positive_int(data.get("team_member_id"), 0)
    if not team_member_id:
        return jsonify({"error": "Please select a profile to switch into."}), 400

    conn = get_db()
    try:
        member = conn.execute("SELECT * FROM team_members WHERE id=?", (team_member_id,)).fetchone()
        if not member:
            return jsonify({"error": "Selected profile was not found."}), 404
        app_user = conn.execute(
            """
            SELECT *
            FROM app_users
            WHERE team_member_id=?
               OR lower(trim(COALESCE(email,'')))=lower(trim(?))
            ORDER BY is_active DESC, id DESC
            LIMIT 1
            """,
            (team_member_id, member["email"] or "")
        ).fetchone()
        if app_user and not app_user["is_active"]:
            return jsonify({"error": "Selected user's login is inactive."}), 400
    finally:
        conn.close()

    keep_keys = [
        "logged_in", "user_id", "app_user_id", "team_member_id", "username",
        "recruiter_name", "email", "recruiter_email", "is_admin",
        "can_bulk_upload", "profile_notes"
    ]
    session["original_admin_session"] = {key: session.get(key) for key in keep_keys if key in session}
    session["original_admin_username"] = session.get("username", "Admin")
    set_session_from_team_member(member, app_user, impersonation_active=True)
    session["switched_profile_name"] = member["name"]
    return jsonify({
        "ok": True,
        "team_member_id": member["id"],
        "username": member["name"],
        "email": member["email"] or "",
        "is_admin": session.get("is_admin"),
        "impersonation_active": 1
    })

@app.route("/admin/restore-profile")
@login_required
def restore_admin_profile_page():
    original = session.get("original_admin_session") or {}
    if session.get("impersonation_active") and original:
        google_token = session.get("google_token")
        session.clear()
        for key, value in original.items():
            session[key] = value
        if google_token:
            session["google_token"] = google_token
        session["impersonation_active"] = 0
    return redirect(url_for("app_page" if session.get("is_admin") else "index"))
        
@app.route("/api/team",methods=["GET"])
@login_required
def get_team():
    conn=get_db()
    rows=conn.execute("SELECT * FROM team_members ORDER BY name").fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/team",methods=["POST"])
@login_required
def add_team():
    d=request.json; conn=get_db()
    conn.execute("INSERT OR IGNORE INTO team_members (name,email,phone,role,can_bulk_upload,is_fixed) VALUES (?,?,?,?,?,1)",
                 (d["name"],(d["email"] or "").lower(),d.get("phone",""),d.get("role",""),
                  1 if d.get("can_bulk_upload") else 0))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/team/<int:tid>",methods=["PATCH"])
@login_required
def update_team_member(tid):
    if not session.get("is_admin"):
        return jsonify({"error":"Admin only"}),403
    d=request.json or {}
    allowed=["name","email","phone","role","can_bulk_upload"]
    sets=", ".join(f"{k}=?" for k in d if k in allowed)
    vals=[]
    for k,v in d.items():
        if k not in allowed:
            continue
        if k == "email":
            v = (v or "").lower()
        if k == "can_bulk_upload":
            v = 1 if v else 0
        vals.append(v)
    if not sets:
        return jsonify({"error":"nothing to update"}),400
    vals.append(tid)
    conn=get_db()
    conn.execute(f"UPDATE team_members SET {sets} WHERE id=?", vals)
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/team/bulk",methods=["POST"])
@login_required
def bulk_add_team():
    if 'file' not in request.files:
        return jsonify({"error":"No file attached"}),400
    f = request.files['file']
    data = f.read()
    fname = f.filename.lower()
    rows = []
    if fname.endswith('.csv'):
        import io as io_module
        reader = csv.DictReader(io_module.StringIO(data.decode('utf-8', errors='ignore')))
        for row in reader:
            rows.append(row)
    elif fname.endswith(('.xlsx','.xls')):
        try:
            import openpyxl
            import io as io_module
            wb = openpyxl.load_workbook(io_module.BytesIO(data))
            ws = wb.active
            headers = [h.value or h for h in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell for cell in row):
                    d = dict(zip(headers, row))
                    rows.append(d)
        except: pass
    if not rows:
        return jsonify({"error":"No valid data found in file"}),400
    conn=get_db()
    added = 0; skipped = 0
    for row in rows:
        name = (row.get("Name") or row.get("name") or row.get("Name ","") or "").strip()
        email = (row.get("Email") or row.get("email") or row.get("Email Id") or row.get("Email-Id") or row.get("Email Id ") or row.get("Email-id") or "").strip().lower()
        phone = (row.get("Phone") or row.get("phone") or row.get("Phone Number") or row.get("Phone number") or row.get("Mobile") or "").strip().replace(",", "")
        role = (row.get("Role") or row.get("role") or row.get("Designation") or row.get("Title") or "").strip()
        if name and email:
            try:
                conn.execute("INSERT INTO team_members (name,email,phone,role,is_fixed) VALUES (?,?,?,?,1)",
                             (name,email,phone,role))
                added += 1
            except:
                skipped += 1
        else:
            skipped += 1
    conn.commit(); conn.close()
    return jsonify({"ok":True,"added":added,"skipped":skipped})

@app.route("/api/team/<int:tid>",methods=["DELETE"])
@login_required
def del_team(tid):
    if not session.get("is_admin"):
        return jsonify({"error":"Admin only"}),403
    conn=get_db(); conn.execute("DELETE FROM team_members WHERE id=?",(tid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# â”€â”€ User Management â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/users",methods=["GET"])
@login_required
def get_users():
    if not session.get("is_admin"):
        return jsonify({"error":"Admin only"}),403
    conn=get_db()
    rows=conn.execute("""
        SELECT u.id,u.username,u.email,u.team_member_id,u.is_admin,u.is_bulk_admin,u.is_active,u.created_at,u.last_login_at,
               t.name AS team_name,t.email AS team_email,t.role AS team_role,t.last_login_at AS team_last_login_at
        FROM app_users u
        LEFT JOIN team_members t ON t.id=u.team_member_id
        ORDER BY COALESCE(u.last_login_at,t.last_login_at,u.created_at) DESC
    """).fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/users/login_report")
@login_required
def user_login_report():
    if not session.get("is_admin"):
        return jsonify({"error":"Admin only"}),403
    limit = parse_positive_int(request.args.get("limit"), 100, 500)
    status = (request.args.get("status") or "").strip()
    email = (request.args.get("email") or "").strip().lower()
    params = []
    where = []
    if status:
        where.append("status=?")
        params.append(status)
    if email:
        where.append("lower(email) LIKE ?")
        params.append(f"%{email}%")
    sql_where = ("WHERE " + " AND ".join(where)) if where else ""
    conn=get_db()
    rows=conn.execute(f"""
        SELECT id, app_user_id, team_member_id, username, email, display_name, role,
               method, status, ip_address, user_agent, message, created_at
        FROM user_login_audit
        {sql_where}
        ORDER BY datetime(created_at) DESC, id DESC
        LIMIT ?
    """, params + [limit]).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/users",methods=["POST"])
@login_required
def add_user():
    if not session.get("is_admin"):
        return jsonify({"error":"Admin only"}),403
    d=request.json or {}
    username = (d.get("username") or "").strip().lower()
    email = (d.get("email") or "").strip().lower()
    password = (d.get("password") or "").strip()
    role = normalize_user_role_label(d.get("role"), bool(d.get("is_admin")))
    name = (d.get("name") or "").strip()
    if not username:
        return jsonify({"error":"Username is required"}),400
    if not password and not email:
        return jsonify({"error":"Password is required when email is not provided"}),400
    if password and len(password) < 8:
        return jsonify({"error":"Password must be at least 8 characters"}),400
    password_hash = hash_password(password) if password else "google-oauth-user"
    conn=get_db()
    try:
        existing_user = conn.execute(
            """SELECT id FROM app_users
               WHERE lower(trim(username))=lower(trim(?))
                  OR (COALESCE(?, '')<>'' AND lower(trim(COALESCE(email,'')))=lower(trim(?)))
               LIMIT 1""",
            (username, email, email),
        ).fetchone()
        if existing_user:
            conn.close()
            return jsonify({"error":"Username or email already exists"}),400
        team_member_id = None
        if email:
            member = conn.execute(
                "SELECT id, role FROM team_members WHERE lower(trim(email))=? LIMIT 1",
                (email,),
            ).fetchone()
            if member:
                team_member_id = member["id"]
                conn.execute(
                    "UPDATE team_members SET name=COALESCE(NULLIF(?,''),name), role=?, can_bulk_upload=?, is_ex_employee=0 WHERE id=?",
                    (name, role, 1 if d.get("is_bulk_admin") else 0, team_member_id),
                )
            else:
                member_name = (name or username or email.split("@")[0]).strip()
                team_member_id = conn.execute(
                    "INSERT INTO team_members (name,email,role,can_bulk_upload) VALUES (?,?,?,?)",
                    (member_name, email, role, 1 if d.get("is_bulk_admin") else 0),
                ).lastrowid
        elif not d.get("is_admin"):
            member_name = (name or username).strip()
            team_member_id = conn.execute(
                "INSERT INTO team_members (name,email,role,can_bulk_upload) VALUES (?,?,?,?)",
                (member_name, None, role, 1 if d.get("is_bulk_admin") else 0),
            ).lastrowid
        conn.execute("""
            INSERT INTO app_users (username,password,email,team_member_id,is_admin,is_bulk_admin,is_active)
            VALUES (?,?,?,?,?,?,1)
        """, (
            username,
            password_hash,
            email or None,
            team_member_id,
            1 if d.get("is_admin") else 0,
            1 if d.get("is_bulk_admin") else 0,
        ))
        conn.commit()
        conn.close()
        return jsonify({"ok":True, "team_member_id": team_member_id})
    except Exception as e:
        conn.close()
        message = str(e)
        if "UNIQUE" in message.upper():
            return jsonify({"error":"Username or email already exists"}),400
        return jsonify({"error":"Unable to add user"}),400

@app.route("/api/users/<int:uid>",methods=["DELETE"])
@login_required
def del_user(uid):
    if not session.get("is_admin"):
        return jsonify({"error":"Admin only"}),403
    if uid == session.get("user_id"):
        return jsonify({"error":"Cannot delete yourself"}),400
    conn=get_db(); conn.execute("DELETE FROM app_users WHERE id=? AND is_admin=0",(uid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/users/create-from-team", methods=["POST"])
@login_required
def create_users_from_team():
    if not session.get("is_admin"):
        return jsonify({"error":"Admin only"}),403
    
    conn = get_db()
    
    # Get all team members without app users
    rows = conn.execute("""
        SELECT t.id, t.name, t.email, t.role 
        FROM team_members t 
        LEFT JOIN app_users u ON t.id = u.team_member_id 
        WHERE u.id IS NULL AND t.email IS NOT NULL AND t.email != ''
        ORDER BY t.name
    """).fetchall()
    
    created = []
    skipped = []
    
    for member in rows:
        name = member["name"]
        email = member["email"]
        role = member["role"]
        
        # Generate username from name
        username = name.lower().replace(" ", "").replace(".", "")
        
        temp_password = secrets.token_urlsafe(10)
        password_hash = hash_password(temp_password)
        
        # Reetu is admin
        is_admin = 1 if name.lower() == "reetu" else 0
        
        try:
            conn.execute("""
                INSERT INTO app_users (username, password, email, team_member_id, is_admin) 
                VALUES (?, ?, ?, ?, ?)
            """, (username, password_hash, email, member["id"], is_admin))
            created.append({
                "name": name,
                "username": username,
                "email": email,
                "role": role,
                "is_admin": is_admin,
                "password": temp_password
            })
        except Exception as e:
            skipped.append({"name": name, "error": str(e)})
    
    conn.commit()
    conn.close()
    
    return jsonify({"ok": True, "created": created, "skipped": skipped, "total_created": len(created)})

@app.route("/api/users/<int:uid>",methods=["PATCH"])
@login_required
def update_user(uid):
    if not session.get("is_admin"):
        return jsonify({"error":"Admin only"}),403
    d=request.json
    conn=get_db()
    try:
        user = conn.execute("SELECT * FROM app_users WHERE id=?", (uid,)).fetchone()
        if not user:
            conn.close()
            return jsonify({"error":"User not found"}),404
        if uid == session.get("user_id") and d.get("is_active") is not None and not d.get("is_active"):
            conn.close()
            return jsonify({"error":"Cannot deactivate yourself"}),400

        updates = []
        params = []
        if "username" in d:
            username = (d.get("username") or "").strip().lower()
            if not username:
                conn.close()
                return jsonify({"error":"Username is required"}),400
            exists = conn.execute(
                "SELECT id FROM app_users WHERE lower(trim(username))=lower(trim(?)) AND id<>?",
                (username, uid),
            ).fetchone()
            if exists:
                conn.close()
                return jsonify({"error":"Username already exists"}),400
            updates.append("username=?")
            params.append(username)

        email = None
        if "email" in d:
            email = (d.get("email") or "").strip().lower()
            if email:
                exists = conn.execute(
                    """SELECT u.id
                       FROM app_users u
                       LEFT JOIN team_members t ON t.id=u.team_member_id
                       WHERE (lower(trim(COALESCE(u.email,'')))=lower(trim(?))
                              OR lower(trim(COALESCE(t.email,'')))=lower(trim(?)))
                         AND u.id<>?
                       LIMIT 1""",
                    (email, email, uid),
                ).fetchone()
                if exists:
                    conn.close()
                    return jsonify({"error":"Email already belongs to another login user"}),400
            updates.append("email=?")
            params.append(email or None)

        role = None
        if "role" in d or "is_admin" in d:
            role = normalize_user_role_label(d.get("role"), bool(d.get("is_admin")))

        if d.get("password"):
            password = str(d["password"])
            if len(password) < 8:
                conn.close()
                return jsonify({"error":"Password must be at least 8 characters"}),400
            updates.append("password=?")
            params.append(hash_password(password))
        if d.get("is_active") is not None:
            updates.append("is_active=?")
            params.append(1 if d["is_active"] else 0)
        if d.get("is_bulk_admin") is not None:
            updates.append("is_bulk_admin=?")
            params.append(1 if d["is_bulk_admin"] else 0)
        if d.get("is_admin") is not None:
            updates.append("is_admin=?")
            params.append(1 if d["is_admin"] else 0)

        if updates:
            conn.execute(f"UPDATE app_users SET {', '.join(updates)} WHERE id=?", params + [uid])

        target = conn.execute("SELECT * FROM app_users WHERE id=?", (uid,)).fetchone()
        team_member_id = target["team_member_id"] if target else None
        if team_member_id:
            member_updates = []
            member_params = []
            if email is not None:
                member_updates.append("email=?")
                member_params.append(email or None)
            if role is not None:
                member_updates.append("role=?")
                member_params.append(role)
            if d.get("is_bulk_admin") is not None:
                member_updates.append("can_bulk_upload=?")
                member_params.append(1 if d["is_bulk_admin"] else 0)
            if member_updates:
                conn.execute(f"UPDATE team_members SET {', '.join(member_updates)} WHERE id=?", member_params + [team_member_id])
        elif email:
            name = (d.get("name") or target["username"] or email.split("@")[0]).strip()
            team_member_id = conn.execute(
                "INSERT INTO team_members (name,email,role,can_bulk_upload) VALUES (?,?,?,?)",
                (name, email, role or "Recruiter", 1 if d.get("is_bulk_admin") else 0),
            ).lastrowid
            conn.execute("UPDATE app_users SET team_member_id=? WHERE id=?", (team_member_id, uid))

        conn.commit()
        conn.close()
        return jsonify({"ok":True})
    except Exception as e:
        conn.close()
        message = str(e)
        if "UNIQUE" in message.upper():
            return jsonify({"error":"Username or email already exists"}),400
        return jsonify({"error":"Unable to update user"}),400

# â”€â”€ Candidate CRUD â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/candidate/<int:cid>",methods=["PATCH"])
@login_required
def update_candidate(cid):
    forbidden = client_viewer_write_forbidden()
    if forbidden:
        return forbidden
    data=request.json
    allowed=["candidate_name","email_addr","phone","current_company","current_role",
             "experience_years","key_skills","notice_period","current_salary",
             "expected_salary","current_location","preferred_location","remarks","role_name","status","requirement_id","candidate_feedback","education"]
    
    # Ownership check for non-admins
    if not session.get("is_admin"):
        conn = get_db()
        owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
        c = conn.execute(f"SELECT c.id FROM candidates c WHERE c.id=?{owner_sql}", [cid] + owner_params).fetchone()
        conn.close()
        if not c:
            return jsonify({"error": "Permission denied"}), 403
    elif data.get("sourcer_id"):
        allowed.append("sourcer_id")
        conn = get_db()
        member = conn.execute("SELECT name,email FROM team_members WHERE id=?", (data.get("sourcer_id"),)).fetchone()
        conn.close()
        if member:
            data["recruiter_name"] = member["name"]
            data["recruiter_email"] = (member["email"] or "").lower()
            allowed.extend(["recruiter_name", "recruiter_email"])
    
    sets=", ".join(f"{k}=?" for k in data if k in allowed)
    vals=[v for k,v in data.items() if k in allowed]+[cid]
    if not sets: return jsonify({"error":"nothing to update"}),400
    conn=get_db()
    conn.execute(f"UPDATE candidates SET {sets}, updated_at=datetime('now','localtime') WHERE id=?",vals)
    conn.commit()
    conn.close()
    return jsonify({"ok":True})

@app.route("/api/candidate",methods=["POST"])
@login_required
def add_single_candidate():
    endpoint_started = time.perf_counter()
    forbidden = client_viewer_write_forbidden()
    if forbidden:
        return forbidden
    d=request.json or {}
    required_fields = {
        "requirement_id": "Requirement",
        "candidate_name": "Candidate name",
        "email_addr": "Email",
        "phone": "Phone",
        "current_location": "Current location",
        "preferred_location": "Preferred location",
        "notice_period": "Notice period",
        "experience_years": "Experience",
        "key_skills": "Skills",
        "current_salary": "Current salary",
        "expected_salary": "Expected salary",
        "cv_filename": "CV/Resume"
    }
    validation_started = time.perf_counter()
    missing_required = [
        label for field, label in required_fields.items()
        if not str(d.get(field) or "").strip()
    ]
    if not (d.get("sourcer_id") or session.get("team_member_id")):
        missing_required.append("Recruiter")
    if missing_required:
        perf_log("candidate_add.validation", validation_started, status="missing", missing=len(missing_required))
        perf_log("candidate_add.total", endpoint_started, status=400)
        return jsonify({
            "ok": False,
            "error": "Missing required fields: " + ", ".join(missing_required),
            "missing_fields": missing_required
        }), 400
    perf_log("candidate_add.validation", validation_started, status="ok")
    
    # Auto-populate recruiter info from session if not provided
    recruiter_name = d.get("recruiter_name","") or session.get("recruiter_name","")
    recruiter_email = d.get("recruiter_email","") or session.get("recruiter_email","")
    
    if not recruiter_name:
        recruiter_name = "Unknown"
    if not recruiter_email:
        recruiter_email = "unknown@system.com"
    
    def _write_candidate():
        write_started = time.perf_counter()
        conn = get_db()
        try:
            dup_started = time.perf_counter()
            is_dup, dup_row, dup_why = check_dup(conn, d)
            perf_log("candidate_add.duplicate_check", dup_started, duplicate=1 if is_dup else 0, why=dup_why or "")
            dup_id = (dup_row or {}).get("id")
            missing = check_missing(d)
            if is_dup:
                perf_log("candidate_add.write", write_started, status="duplicate")
                return {"ok": False, "error": build_duplicate_candidate_message(d, dup_row, dup_why)}
            # CV fields
            cv_filename = d.get("cv_filename","")
            cv_url      = d.get("cv_url","")
            cv_public_id = d.get("cv_public_id","")
            sourcer_id = d.get("sourcer_id") or session.get("team_member_id")
            insert_started = time.perf_counter()
            cid = conn.execute("""INSERT INTO candidates
                (upload_batch,recruiter_name,recruiter_email,role_name,candidate_name,
                 email_addr,phone,current_company,current_role,experience_years,key_skills,
                 notice_period,current_salary,expected_salary,current_location,
                 preferred_location,remarks,cv_filename,cv_url,cv_public_id,cv_summary,
                 status,tags,is_duplicate,duplicate_of,missing_info,job_id,sourcer_id,requirement_id,education)
                 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                ("single_"+datetime.now().strftime("%Y%m%d%H%M%S"),
                 recruiter_name,recruiter_email,d.get("role_name",""),d.get("candidate_name",""),
                 d.get("email_addr",""),d.get("phone",""),
                 d.get("current_company",""),d.get("current_role",""),
                 d.get("experience_years",""),d.get("key_skills",""),
                 d.get("notice_period",""),d.get("current_salary",""),
                 d.get("expected_salary",""),d.get("current_location",""),
                 d.get("preferred_location",""),d.get("remarks",""),
                 cv_filename,cv_url,cv_public_id,d.get("cv_summary",""),
                 d.get("status","New"),d.get("tags",""),
                 1 if is_dup else 0,
                 dup_id if is_dup else None,
                 ",".join(missing) if missing else None,
                 d.get("job_id",""),
                 sourcer_id,
                 d.get("requirement_id"),
                 d.get("education",""))).lastrowid
            conn.commit()
            perf_log("candidate_add.insert_commit", insert_started, candidate_id=cid)
            if is_dup:
                conn.execute("INSERT INTO alerts (alert_type,message,candidate_id,recruiter_email) VALUES (?,?,?,?)",
                    ("duplicate",f"Duplicate: {d.get('candidate_name','?')} via {dup_why} (original ID #{dup_id})",cid,recruiter_email))
            if missing:
                conn.execute("INSERT INTO alerts (alert_type,message,candidate_id,recruiter_email) VALUES (?,?,?,?)",
                    ("missing_info",f"Missing {','.join(missing)} for {d.get('candidate_name','unnamed')}",cid,recruiter_email))
            conn.commit()
            perf_log("candidate_add.write", write_started, status="ok", candidate_id=cid)
            return {
                "ok": True,
                "id": cid,
                "duplicate": is_dup,
                "missing": missing,
                "cv_filename": cv_filename,
                "cv_url": cv_url,
                "cv_public_id": cv_public_id,
            }
        finally:
            conn.close()

    retry_started = time.perf_counter()
    write_result = _with_db_write_retry(_write_candidate)
    perf_log("candidate_add.write_retry", retry_started, ok=1 if write_result.get("ok") else 0)
    if not write_result.get("ok"):
        perf_log("candidate_add.total", endpoint_started, status=400)
        return jsonify(write_result), 400
    cid = write_result["id"]
    perf_log("candidate_add.total", endpoint_started, status=200, candidate_id=cid)
    return jsonify({"ok":True,"id":cid})

@app.route("/api/candidate/<int:cid>/status", methods=["PATCH"])
@login_required
def update_status(cid):
    forbidden = client_viewer_write_forbidden()
    if forbidden:
        return forbidden
    data   = request.json
    status = data.get("status","")
    feedback = clean_value(data.get("candidate_feedback") or data.get("feedback") or "", 2000)
    if not status: return jsonify({"error":"Status required"}),400
    conn = get_db()
    owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
    c = conn.execute(
        f"SELECT c.recruiter_email, c.role_name FROM candidates c WHERE c.id=?{owner_sql}",
        [cid] + owner_params
    ).fetchone()
    if not c:
        conn.close()
        return jsonify({"error": "Not found or permission denied"}), 404
    
    valid_statuses = get_candidate_statuses_for_role(conn, c["role_name"] or "")
    if valid_statuses and normalize_status_key(status) not in {normalize_status_key(s) for s in valid_statuses}:
        conn.close()
        return jsonify({"error":f"Invalid status. Must be one of: {', '.join(valid_statuses)}"}),400
    conn.execute("UPDATE candidates SET status=?, candidate_feedback=?, updated_at=datetime('now','localtime') WHERE id=?",
                 (status, feedback, cid))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/client/candidate/<int:cid>/status", methods=["PATCH"])
@login_required
def update_client_candidate_status(cid):
    if not is_client_viewer_session():
        return jsonify({"error": "Client access only"}), 403
    data = request.json or {}
    status = clean_value(data.get("status") or "", 80)
    feedback = clean_value(data.get("candidate_feedback") or data.get("feedback") or "", 2000)
    if not status:
        return jsonify({"error": "Status required"}), 400
    conn = get_db()
    owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
    row = conn.execute(
        f"""SELECT c.id, c.role_name, c.candidate_name, c.recruiter_email, c.recruiter_name,
                   r.client_name, r.title AS requirement_title
            FROM candidates c
            LEFT JOIN requirements r ON r.id = c.requirement_id
            WHERE c.id=?{owner_sql}""",
        [cid] + owner_params
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "Not found or permission denied"}), 404
    valid_statuses = get_candidate_statuses_for_role(conn, row["role_name"] or "")
    if valid_statuses and normalize_status_key(status) not in {normalize_status_key(s) for s in valid_statuses}:
        conn.close()
        return jsonify({"error": f"Invalid status. Must be one of: {', '.join(valid_statuses)}"}), 400
    conn.execute(
        "UPDATE candidates SET status=?, candidate_feedback=?, updated_at=datetime('now','localtime') WHERE id=?",
        (status, feedback, cid)
    )
    recruiter_email = (row["recruiter_email"] or "").strip().lower()
    if recruiter_email:
        client_name = row["client_name"] or "Client"
        candidate_name = row["candidate_name"] or f"Candidate #{cid}"
        requirement_title = row["requirement_title"] or row["role_name"] or "Requirement"
        feedback_snippet = feedback[:180] + ("..." if len(feedback) > 180 else "")
        message = f"{client_name} updated {candidate_name} ({requirement_title}) to {status}."
        if feedback_snippet:
            message += f" Feedback: {feedback_snippet}"
        conn.execute(
            "INSERT INTO alerts (alert_type,message,candidate_id,recruiter_email) VALUES (?,?,?,?)",
            ("client_status_update", message, cid, recruiter_email)
        )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/candidate/<int:cid>",methods=["GET"])
@login_required
def get_candidate(cid):
    conn = get_db()
    select_clause = "c.*, r.title as requirement_title, r.client_name as client_name"
    if str(request.args.get("view", "")).strip().lower() == "edit":
        select_clause = (
            "c.id, c.candidate_name, c.email_addr, c.phone, c.current_company, c.current_role, "
            "c.experience_years, c.key_skills, c.notice_period, c.current_salary, c.expected_salary, "
            "c.current_location, c.preferred_location, c.remarks, c.role_name, c.status, "
            "c.recruiter_name, c.recruiter_email, c.created_at, c.cv_url, c.cv_filename, "
            "r.title as requirement_title, r.client_name as client_name"
        )

    if not session.get("is_admin"):
        owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
        row = conn.execute(
            f"""SELECT {select_clause}
               FROM candidates c
               LEFT JOIN requirements r ON c.requirement_id = r.id
               WHERE c.id=?{owner_sql}""",
            [cid] + owner_params
        ).fetchone()
    else:
        row = conn.execute(
            f"""SELECT {select_clause}
               FROM candidates c
               LEFT JOIN requirements r ON c.requirement_id = r.id
               WHERE c.id=?""",
            (cid,)
        ).fetchone()

    conn.close()

    if row:
        return jsonify(dict(row))

    return jsonify({"error":"Not found"}),404

@app.route("/api/ai_screening_logs", methods=["GET"])
@login_required
def get_ai_screening_logs():
    limit = parse_positive_int(request.args.get("limit"), 10, 10)
    conn = get_db()
    ensure_ats_pipeline_schema(conn)
    rows = conn.execute(
        """SELECT *
           FROM ai_screening_logs
           ORDER BY id DESC
           LIMIT ?""",
        (limit,)
    ).fetchall()
    conn.close()
    logs = []
    for row in rows:
        item = dict(row)
        try:
            item["details"] = json.loads(item.get("details_json") or "{}")
        except Exception:
            item["details"] = {}
        logs.append(item)
    return jsonify(logs)

@app.route("/api/candidate/<int:cid>/ai_screening_report", methods=["GET"])
@login_required
def get_candidate_ai_screening_report(cid):
    conn = get_db(timeout=5)
    owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
    row = conn.execute(
        f"""SELECT c.*, r.title AS requirement_title, r.client_name AS requirement_client_name, r.description AS requirement_description
            FROM candidates c
            LEFT JOIN requirements r ON r.id = c.requirement_id
            WHERE c.id=?{owner_sql}""",
        [cid] + owner_params
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "Not found or permission denied"}), 404
    report_raw = row["ai_screening_report_json"] or ""
    if not report_raw:
        return jsonify({"error": "No AI screening report available for this candidate."}), 404
    try:
        report = json.loads(report_raw)
    except Exception:
        return jsonify({"error": "Stored AI screening report is invalid."}), 500
    requirement = {
        "title": row["requirement_title"] or row["role_name"] or "",
        "client_name": row["requirement_client_name"] or "",
        "description": row["requirement_description"] or "",
    }
    candidate = dict(row)
    try:
        buf = build_ai_screening_pdf(candidate=candidate, requirement=requirement, report=report)
        safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", (row["candidate_name"] or "candidate")).strip("_") or "candidate"
        return send_file(
            buf,
            mimetype="application/pdf",
            as_attachment=False,
            download_name=f"ai_screening_{safe_name}_{date.today()}.pdf"
        )
    except Exception as e:
        print("AI screening PDF export error:", e, flush=True)
        return jsonify({"error": str(e)}), 500

@app.route("/api/candidate/<int:cid>/ai_screening", methods=["POST"])
@login_required
def trigger_candidate_ai_screening(cid):
    conn = get_db(timeout=5)
    owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
    row = conn.execute(
        f"SELECT c.id, c.requirement_id, c.ai_screening_score, c.ai_screening_status FROM candidates c WHERE c.id=?{owner_sql}",
        [cid] + owner_params
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "Not found or permission denied"}), 404
    row_data = dict(row)
    if not row_data.get("requirement_id"):
        return jsonify({"ok": True, "status": "no_jd", "message": "No JD attached"}), 200
    if not session.get("is_admin") and row_data.get("ai_screening_score") not in (None, "", "None"):
        return jsonify({"error": "AI screening already completed for this candidate. Normal users cannot re-run it."}), 403
    if not session.get("is_admin") and str(row_data.get("ai_screening_status") or "").lower() == "scored":
        return jsonify({"error": "AI screening already completed for this candidate. Normal users cannot re-run it."}), 403
    queued = queue_candidate_ai_screening(cid, trigger="manual_trigger")
    return jsonify({
        "ok": True,
        "queued": bool(queued),
        "status": "pending" if queued else "running",
        "message": "AI screening queued" if queued else "AI screening started inline"
    })

@app.route("/api/candidate/<int:cid>",methods=["DELETE"])
@login_required
def del_candidate(cid):
    forbidden = client_viewer_write_forbidden()
    if forbidden:
        return forbidden
    conn = get_db()
    owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
    c = conn.execute(
        f"SELECT c.recruiter_email FROM candidates c WHERE c.id=?{owner_sql}",
        [cid] + owner_params
    ).fetchone()
    if not c:
        conn.close()
        return jsonify({"error": "Not found or permission denied"}), 404

    conn.execute("DELETE FROM candidates WHERE id=?",(cid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# â”€â”€ Hiring Managers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/hiring_managers",methods=["GET"])
@login_required
def get_hiring_managers():
    conn=get_db()
    rows=conn.execute("SELECT * FROM hiring_managers ORDER BY name").fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/hiring_managers",methods=["POST"])
@login_required
def add_hiring_manager():
    d=request.json
    if not d.get("name"): return jsonify({"error":"Name required"}),400
    conn=get_db()
    hid = conn.execute("INSERT INTO hiring_managers (name,email,company) VALUES (?,?,?)",
                       (d["name"],d.get("email",""),d.get("company",""))).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":hid})

@app.route("/api/hiring_managers/<int:hid>",methods=["DELETE"])
@login_required
def del_hiring_manager(hid):
    conn=get_db(); conn.execute("DELETE FROM hiring_managers WHERE id=?",(hid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# â”€â”€ Pipelines â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/pipelines",methods=["GET"])
@login_required
def get_pipelines():
    conn=get_db()
    rows=conn.execute("SELECT * FROM pipelines ORDER BY role_name").fetchall()
    result=[dict(r) for r in rows]
    for r in result: r["status_list"]=json.loads(r["status_list"])
    conn.close(); return jsonify(result)

@app.route("/api/pipelines",methods=["POST"])
@login_required
def add_pipeline():
    d=request.json
    if not d.get("role_name") or not d.get("status_list"):
        return jsonify({"error":"role_name and status_list required"}),400
    conn=get_db()
    pid=conn.execute("INSERT INTO pipelines (role_name,status_list) VALUES (?,?)",
                     (d["role_name"],json.dumps(d["status_list"]))).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":pid})

@app.route("/api/pipelines/<int:pid>",methods=["PATCH"])
@login_required
def update_pipeline(pid):
    d=request.json
    if "status_list" in d:
        conn=get_db()
        conn.execute("UPDATE pipelines SET status_list=? WHERE id=?",(json.dumps(d["status_list"]),pid))
        conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/pipelines/<int:pid>",methods=["DELETE"])
@login_required
def del_pipeline(pid):
    conn=get_db(); conn.execute("DELETE FROM pipelines WHERE id=?",(pid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

@app.route("/api/pipelines/for_role")
@login_required
def get_pipeline_for_role():
    role=request.args.get("role","")
    conn=get_db()
    statuses = get_candidate_statuses_for_role(conn, role)
    conn.close()
    return jsonify(statuses)

# â”€â”€ Jobs â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/jobs",methods=["GET"])
@login_required
def get_jobs():
    conn=get_db()
    rows=conn.execute("SELECT * FROM jobs ORDER BY created_at DESC").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/jobs",methods=["POST"])
@login_required
def create_job():
    d=request.json
    if not d.get("title"):
        return jsonify({"error":"title required"}),400
    conn=get_db()
    # Generate job_id if not provided
    job_id = d.get("job_id") or f"JOB-{conn.execute('SELECT COUNT(*) FROM jobs').fetchone()[0] + 1:04d}"
    pid=conn.execute("INSERT INTO jobs (job_id,title,description,client_name,status) VALUES (?,?,?,?,?)",
                     (job_id,d["title"],d.get("description",""),d.get("client_name",""),d.get("status","Open"))).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":pid})

@app.route("/api/jobs/<int:jid>",methods=["PATCH"])
@login_required
def update_job(jid):
    d=request.json
    conn=get_db()
    if "title" in d: conn.execute("UPDATE jobs SET title=? WHERE id=?",(d["title"],jid))
    if "description" in d: conn.execute("UPDATE jobs SET description=? WHERE id=?",(d["description"],jid))
    if "client_name" in d: conn.execute("UPDATE jobs SET client_name=? WHERE id=?",(d["client_name"],jid))
    if "status" in d: conn.execute("UPDATE jobs SET status=? WHERE id=?",(d["status"],jid))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/jobs/<int:jid>",methods=["DELETE"])
@login_required
def del_job(jid):
    conn=get_db(); conn.execute("DELETE FROM jobs WHERE id=?",(jid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# â”€â”€ Requirements â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/requirements",methods=["GET"])
@login_required
def get_requirements():
    conn=get_db()
    status = (request.args.get("status") or "").strip()
    q = (request.args.get("q") or "").strip().lower()
    include_stats = str(request.args.get("stats", "")).lower() in {"1", "true", "yes"}
    where_parts = []
    params = []
    if status:
        where_parts.append("r.status=?")
        params.append(status)
    if q:
        where_parts.append("(lower(COALESCE(r.title,'')) LIKE ? OR lower(COALESCE(r.client_name,'')) LIKE ? OR CAST(r.id AS TEXT)=?)")
        params.extend([f"%{q}%", f"%{q}%", q])
    allowed_clients = mapped_client_names_for_current_user(conn)
    if allowed_clients is not None:
        if not allowed_clients:
            conn.close()
            return jsonify([] if not (request.args.get("page") or request.args.get("page_size")) else {
                "rows": [], "page": 1, "page_size": 50, "total": 0, "total_pages": 1,
                "stats": {"open": 0, "in_progress": 0, "closed": 0, "total_submissions": 0}
            })
        placeholders = ",".join("?" * len(allowed_clients))
        where_parts.append(f"lower(trim(COALESCE(r.client_name,''))) IN ({placeholders})")
        params.extend(sorted(allowed_clients))
    where = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""
    base_sql = f"""
        SELECT r.*, 
               ts.name as sourcer_name, 
               tr.name as recruiter_name,
               tg.name as taggd_recruiter_name,
               (SELECT COUNT(*) FROM candidates c WHERE c.requirement_id=r.id) as submissions
        FROM requirements r 
        LEFT JOIN team_members ts ON r.assigned_sourcer_id=ts.id
        LEFT JOIN team_members tr ON r.assigned_recruiter_id=tr.id
        LEFT JOIN taggd_recruiters tg ON tg.id=r.taggd_recruiter_id
        {where}
        ORDER BY r.created_at DESC
    """
    if str(request.args.get("all", "")).lower() in {"1", "true", "yes"} and not request.args.get("page"):
        rows=conn.execute(base_sql, params).fetchall()
        conn.close(); return jsonify([dict(r) for r in rows])

    if not request.args.get("page") and not request.args.get("page_size") and not status:
        rows=conn.execute(base_sql, params).fetchall()
        conn.close(); return jsonify([dict(r) for r in rows])

    page = parse_positive_int(request.args.get("page"), 1)
    page_size = parse_positive_int(request.args.get("page_size"), 50, 500)
    total = conn.execute(f"SELECT COUNT(*) FROM requirements r {where}", params).fetchone()[0]
    list_sql = f"""
        SELECT r.id,
               r.title,
               r.client_name,
               r.description,
               r.location,
               r.daily_target,
               r.status,
               r.jd_url,
               r.taggd_recruiter_id,
               COALESCE(tg.name, r.taggd_recruiter_name, '') AS taggd_recruiter_name,
               ts.name as sourcer_name,
               tr.name as recruiter_name
        FROM requirements r
        LEFT JOIN team_members ts ON r.assigned_sourcer_id=ts.id
        LEFT JOIN team_members tr ON r.assigned_recruiter_id=tr.id
        LEFT JOIN taggd_recruiters tg ON tg.id=r.taggd_recruiter_id
        {where}
        ORDER BY r.created_at DESC
        LIMIT ? OFFSET ?
    """
    rows=conn.execute(list_sql, params + [page_size, (page - 1) * page_size]).fetchall()
    payload = {
        "rows": [dict(r) for r in rows],
        "page": page,
        "page_size": page_size,
        "total": total,
        "total_pages": (total + page_size - 1) // page_size if page_size else 1,
    }
    if include_stats:
        stats = conn.execute(f"""
            SELECT
              SUM(CASE WHEN status='Open' THEN 1 ELSE 0 END) AS open_count,
              SUM(CASE WHEN status='In Progress' THEN 1 ELSE 0 END) AS in_progress_count,
              SUM(CASE WHEN status='Closed' THEN 1 ELSE 0 END) AS closed_count
            FROM requirements r
            {where}
        """, params).fetchone()
        total_submissions = conn.execute(f"""
            SELECT COUNT(*)
            FROM candidates c
            JOIN requirements r ON r.id=c.requirement_id
            {where}
        """, params).fetchone()[0]
        payload["stats"] = {
            "open": stats["open_count"] or 0,
            "in_progress": stats["in_progress_count"] or 0,
            "closed": stats["closed_count"] or 0,
            "total_submissions": total_submissions
        }
    conn.close(); return jsonify(payload)

@app.route("/api/requirements/search", methods=["GET"])
@login_required
def search_requirements():
    q = (request.args.get("q") or "").strip().lower()
    limit = parse_positive_int(request.args.get("limit"), 25, 50)
    if len(q) < 2:
        return jsonify([])

    conn = get_db()
    active_statuses = ["New", "Open", "In Progress"]
    where_parts = ["COALESCE(NULLIF(TRIM(r.status),''),'New') IN (?,?,?)"]
    params = active_statuses[:]

    allowed_clients = mapped_client_names_for_current_user(conn)
    if allowed_clients is not None:
        if not allowed_clients:
            conn.close()
            return jsonify([])
        placeholders = ",".join("?" * len(allowed_clients))
        where_parts.append(f"lower(trim(COALESCE(r.client_name,''))) IN ({placeholders})")
        params.extend(sorted(allowed_clients))

    like = f"%{q}%"
    where_parts.append("""
        (
          lower(COALESCE(r.title,'')) LIKE ?
          OR lower(COALESCE(r.client_name,'')) LIKE ?
          OR CAST(r.id AS TEXT)=?
        )
    """)
    params.extend([like, like, q])
    where = "WHERE " + " AND ".join(where_parts)
    rows = conn.execute(f"""
        SELECT r.id,
               r.title,
               r.client_name,
               COALESCE(NULLIF(TRIM(r.status),''),'New') AS status,
               r.taggd_recruiter_id,
               COALESCE(tg.name, r.taggd_recruiter_name, '') AS taggd_recruiter_name
        FROM requirements r
        LEFT JOIN taggd_recruiters tg ON tg.id = r.taggd_recruiter_id
        {where}
        ORDER BY
          CASE COALESCE(NULLIF(TRIM(r.status),''),'New')
            WHEN 'Open' THEN 1
            WHEN 'In Progress' THEN 2
            WHEN 'New' THEN 3
            ELSE 4
          END,
          r.created_at DESC
        LIMIT ?
    """, params + [limit]).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/requirements",methods=["POST"])
@login_required
def add_requirement():
    forbidden = client_viewer_write_forbidden()
    if forbidden:
        return forbidden
    d=request.json
    conn=get_db()
    title = (d.get("title") or "").strip()
    client_name = (d.get("client_name") or "").strip()
    location = (d.get("location") or "").strip()
    description = (d.get("description") or "").strip()
    taggd_recruiter = resolve_taggd_recruiter_for_client(conn, d.get("taggd_recruiter_id"), client_name)
    if not title:
        conn.close()
        return jsonify({"ok":False,"error":"Requirement title is required"}),400
    if not taggd_recruiter:
        conn.close()
        return jsonify({"ok":False,"error":"Taggd Recruiter Name is required and must belong to the selected client."}),400
    if not current_user_can_use_client(conn, client_name):
        conn.close()
        return jsonify({"ok":False,"error":"You can add requirements only for clients mapped to you by admin."}),403
    if not location or location.lower() in {"-", "na", "n/a", "none", "null"}:
        conn.close()
        return jsonify({"ok":False,"error":"Location is required."}),400
    title, title_errors = validate_requirement_title(title, client_name)
    if title_errors:
        conn.close()
        return jsonify({"ok":False,"error":"Requirement naming standard: " + " ".join(title_errors)}),400
    if not (d.get("has_jd_text") or d.get("has_jd_file")):
        conn.close()
        return jsonify({"ok":False,"error":"Add a detailed job description or attach a JD file."}),400
    duplicate = None
    title_key = normalize_requirement_title_key(title)
    client_key = normalize_requirement_client_key(client_name)
    for row in conn.execute("SELECT id, title, client_name FROM requirements").fetchall():
        if normalize_requirement_title_key(row["title"]) == title_key and normalize_requirement_client_key(row["client_name"]) == client_key:
            duplicate = row
            break
    if duplicate:
        conn.close()
        return jsonify({"ok":False,"error":"A requirement with this title and client already exists."}),409
    similar, score = find_similar_requirement(conn, title, client_name, description)
    if similar:
        conn.close()
        return jsonify({
            "ok": False,
            "error": f"Similar requirement already exists for {client_name}: #{similar['id']} {similar['title']} ({round(score * 100)}% match). Please edit/reuse the existing requirement instead of creating a duplicate.",
            "similar_requirement": {"id": similar["id"], "title": similar["title"], "client_name": similar["client_name"], "score": round(score, 3)}
        }), 409
    cursor=conn.execute("""INSERT INTO requirements 
        (title,description,client_name,location,taggd_recruiter_id,taggd_recruiter_name,assigned_sourcer_id,assigned_recruiter_id,daily_target,status,created_by)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
        (title,description,client_name,location,
         taggd_recruiter["id"], taggd_recruiter["name"],
         d.get("assigned_sourcer_id") or session.get("team_member_id"),
         d.get("assigned_recruiter_id") or session.get("team_member_id"),d.get("daily_target",3),d.get("status","New"),
         session.get("username")))
    rid=cursor.lastrowid
    
    # Auto-add default screening checks if none provided
    checks = d.get("checks",[])
    if not checks:
        default_checks = [
            "Technical Skills", "Years of Relevant Experience", "Within Given Budget",
            "Notice Period", "Location", "Non-Poachable Employee", "Updated CV"
        ]
        for i, check_name in enumerate(default_checks):
            conn.execute("""INSERT INTO requirement_checks 
                (requirement_id,check_name,check_description,check_type,pass_criteria,sort_order)
                VALUES (?,?,?,?,?,?)""",
                (rid, check_name, "", "boolean", "Yes", i))
    else:
        for i, check in enumerate(checks):
            check_name = check.get("check_name") or check.get("name")
            if not check_name:
                continue
            conn.execute("""INSERT INTO requirement_checks 
                (requirement_id,check_name,check_description,check_type,pass_criteria,sort_order)
                VALUES (?,?,?,?,?,?)""",
                (rid, check_name, check.get("description"), check.get("check_type") or check.get("type","boolean"), 
                 check.get("pass_criteria"), check.get("sort_order", i)))
    
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":rid})

@app.route("/api/requirements/<int:rid>",methods=["GET"])
@login_required
def get_requirement(rid):
    conn=get_db()
    row=conn.execute("SELECT * FROM requirements WHERE id=?",(rid,)).fetchone()
    conn.close()
    if row:
        conn_check = get_db()
        can_use = current_user_can_use_client(conn_check, row["client_name"])
        conn_check.close()
        if not can_use:
            return jsonify({"error":"Requirement is not mapped to your clients"}),403
        return jsonify(dict(row))
    return jsonify({"error":"Not found"}),404

@app.route("/api/requirements/<int:rid>",methods=["PATCH"])
@login_required
def update_requirement(rid):
    forbidden = client_viewer_write_forbidden()
    if forbidden:
        return forbidden
    d=request.json
    conn=get_db()
    current = conn.execute("SELECT title, client_name, description, jd_url, taggd_recruiter_id, taggd_recruiter_name FROM requirements WHERE id=?", (rid,)).fetchone()
    if not current:
        conn.close()
        return jsonify({"ok":False,"error":"Requirement not found"}),404
    if not current_user_can_use_client(conn, current["client_name"]):
        conn.close()
        return jsonify({"ok":False,"error":"Requirement is not mapped to your clients"}),403
    new_title = (d.get("title", current["title"]) or "").strip()
    new_client = (d.get("client_name", current["client_name"]) or "").strip()
    if "location" in d and ((not (d.get("location") or "").strip()) or (d.get("location") or "").strip().lower() in {"-", "na", "n/a", "none", "null"}):
        conn.close()
        return jsonify({"ok":False,"error":"Location is required."}),400
    if "client_name" in d and not current_user_can_use_client(conn, new_client):
        conn.close()
        return jsonify({"ok":False,"error":"You can use only clients mapped to you by admin."}),403
    if "taggd_recruiter_id" in d:
        taggd_recruiter = resolve_taggd_recruiter_for_client(conn, d.get("taggd_recruiter_id"), new_client)
        if not taggd_recruiter:
            conn.close()
            return jsonify({"ok":False,"error":"Taggd Recruiter Name is required and must belong to the selected client."}),400
    else:
        taggd_recruiter = None
    if "title" in d or "client_name" in d:
        new_title, title_errors = validate_requirement_title(new_title, new_client)
        if title_errors:
            conn.close()
            return jsonify({"ok":False,"error":"Requirement naming standard: " + " ".join(title_errors)}),400
    duplicate = None
    title_key = normalize_requirement_title_key(new_title)
    client_key = normalize_requirement_client_key(new_client)
    for row in conn.execute("SELECT id, title, client_name FROM requirements WHERE id<>?", (rid,)).fetchall():
        if normalize_requirement_title_key(row["title"]) == title_key and normalize_requirement_client_key(row["client_name"]) == client_key:
            duplicate = row
            break
    if duplicate:
        conn.close()
        return jsonify({"ok":False,"error":"A requirement with this title and client already exists."}),409
    new_description = (d.get("description", current["description"]) or "").strip()
    if "description" in d and not (d.get("has_jd_text") or d.get("has_jd_file") or current["jd_url"]):
        conn.close()
        return jsonify({"ok":False,"error":"Add a detailed job description or attach a JD file."}),400
    similar, score = find_similar_requirement(conn, new_title, new_client, new_description, exclude_id=rid)
    if similar:
        conn.close()
        return jsonify({
            "ok": False,
            "error": f"Similar requirement already exists for {new_client}: #{similar['id']} {similar['title']} ({round(score * 100)}% match). Please edit/reuse that requirement instead of creating a duplicate.",
            "similar_requirement": {"id": similar["id"], "title": similar["title"], "client_name": similar["client_name"], "score": round(score, 3)}
        }), 409
    if "title" in d: conn.execute("UPDATE requirements SET title=? WHERE id=?",(new_title,rid))
    if "description" in d: conn.execute("UPDATE requirements SET description=? WHERE id=?",(new_description,rid))
    if "client_name" in d: conn.execute("UPDATE requirements SET client_name=? WHERE id=?",(new_client,rid))
    if "location" in d: conn.execute("UPDATE requirements SET location=? WHERE id=?",(d["location"],rid))
    if "remote" in d: conn.execute("UPDATE requirements SET remote=? WHERE id=?",(1 if d["remote"] else 0,rid))
    if "assigned_sourcer_id" in d: conn.execute("UPDATE requirements SET assigned_sourcer_id=? WHERE id=?",(d["assigned_sourcer_id"],rid))
    if "assigned_recruiter_id" in d: conn.execute("UPDATE requirements SET assigned_recruiter_id=? WHERE id=?",(d["assigned_recruiter_id"],rid))
    if "daily_target" in d: conn.execute("UPDATE requirements SET daily_target=? WHERE id=?",(d["daily_target"],rid))
    if "status" in d: conn.execute("UPDATE requirements SET status=? WHERE id=?",(d["status"],rid))
    if "taggd_recruiter_id" in d:
        conn.execute(
            "UPDATE requirements SET taggd_recruiter_id=?, taggd_recruiter_name=? WHERE id=?",
            (taggd_recruiter["id"], taggd_recruiter["name"], rid)
        )
    conn.execute("UPDATE requirements SET updated_at=datetime('now','localtime') WHERE id=?",(rid,))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/requirements/<int:rid>/upload_jd", methods=["POST"])
@login_required
def upload_requirement_jd(rid):
    forbidden = client_viewer_write_forbidden()
    if forbidden:
        return forbidden
    if "file" not in request.files:
        return jsonify({"error": "No JD file uploaded"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "Empty filename"}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in [".pdf", ".doc", ".docx", ".txt"]:
        return jsonify({"error": "Invalid JD file type"}), 400
    conn = get_db()
    exists = conn.execute("SELECT id FROM requirements WHERE id=?", (rid,)).fetchone()
    if not exists:
        conn.close()
        return jsonify({"error": "Requirement not found"}), 404
    upload_folder = os.path.join(app.root_path, "uploads", "recruiters", current_upload_owner_key(), "requirements", "single", f"requirement_{rid}")
    os.makedirs(upload_folder, exist_ok=True)
    safe_name = secure_filename(f.filename)
    saved_name = f"requirement_jd_{rid}_{int(datetime.now().timestamp())}_{safe_name}"
    file_path = os.path.join(upload_folder, saved_name)
    try:
        f.save(file_path)
        file_url = f"/uploads/{os.path.relpath(file_path, os.path.join(app.root_path, 'uploads')).replace(os.sep, '/')}"
        conn.execute(
            """UPDATE requirements
               SET jd_filename=?, jd_url=?, jd_public_id=?, updated_at=datetime('now','localtime')
               WHERE id=?""",
            (safe_name, file_url, saved_name, rid)
        )
        conn.commit()
        return jsonify({"ok": True, "filename": safe_name, "url": file_url, "public_id": saved_name})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()

@app.route("/api/requirements/<int:rid>",methods=["DELETE"])
@login_required
def del_requirement(rid):
    conn=get_db()
    conn.execute("UPDATE candidates SET requirement_id=NULL WHERE requirement_id=?",(rid,))
    conn.execute("DELETE FROM requirement_submissions WHERE requirement_id=?",(rid,))
    conn.execute("DELETE FROM requirement_checks WHERE requirement_id=?",(rid,))
    conn.execute("DELETE FROM requirements WHERE id=?",(rid,))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/requirements/<int:rid>/checks",methods=["GET"])
@login_required
def get_requirement_checks(rid):
    conn=get_db()
    rows=conn.execute("SELECT * FROM requirement_checks WHERE requirement_id=? ORDER BY sort_order",(rid,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/requirements/<int:rid>/checks",methods=["POST"])
@login_required
def add_requirement_check(rid):
    d=request.json
    conn=get_db()
    cursor=conn.execute("""INSERT INTO requirement_checks 
        (requirement_id,check_name,check_description,check_type,pass_criteria,is_mandatory,sort_order)
        VALUES (?,?,?,?,?,?,?)""",
        (rid,d.get("check_name"),d.get("check_description"),d.get("check_type","boolean"),
         d.get("pass_criteria"),1 if d.get("is_mandatory",True) else 0,
         d.get("sort_order",0)))
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":cursor.lastrowid})

@app.route("/api/requirements/<int:rid>/checks/<int:cid>",methods=["DELETE"])
@login_required
def del_requirement_check(rid,cid):
    conn=get_db()
    conn.execute("DELETE FROM requirement_checks WHERE id=?",(cid,))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

# â”€â”€ Requirement Submissions â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/submissions",methods=["GET"])
@login_required
def get_submissions():
    conn=get_db()
    rows=conn.execute("""
        SELECT rs.*, r.title as requirement_title, r.client_name,
               c.candidate_name, c.email_addr, c.phone, c.current_company, c.current_role
        FROM requirement_submissions rs
        LEFT JOIN requirements r ON rs.requirement_id=r.id
        LEFT JOIN candidates c ON rs.candidate_id=c.id
        ORDER BY rs.submitted_at DESC
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/submissions",methods=["POST"])
@login_required
def add_submission():
    endpoint_started = time.perf_counter()
    d=request.json
    conn=get_db()
    
    # Create or get candidate
    candidate_id = d.get("candidate_id")
    if not candidate_id:
        # Create new candidate from data
        create_candidate_started = time.perf_counter()
        cursor = conn.execute("""INSERT INTO candidates 
            (candidate_name,email_addr,phone,current_company,current_role,experience_years,key_skills,notice_period,current_salary,expected_salary,current_location,requirement_id,recruiter_name,recruiter_email)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (d.get("candidate_name"),d.get("email_addr"),d.get("phone"),d.get("current_company"),
             d.get("current_role"),d.get("experience_years"),d.get("key_skills"),d.get("notice_period"),
             d.get("current_salary"),d.get("expected_salary"),d.get("current_location"),
             d.get("requirement_id"),session.get("recruiter_name"),session.get("recruiter_email")))
        candidate_id = cursor.lastrowid
        perf_log("submission.create_candidate", create_candidate_started, candidate_id=candidate_id)
    
    # Create submission
    insert_started = time.perf_counter()
    cursor = conn.execute("""INSERT INTO requirement_submissions
        (candidate_id,requirement_id,sourcer_id,submitted_by,status,notes)
        VALUES (?,?,?,?,?,?)""",
        (candidate_id, d.get("requirement_id"), d.get("sourcer_id"), 
         session.get("username"), 'Submitted', d.get("notes","")))
    submission_id = cursor.lastrowid
    perf_log("submission.insert", insert_started, submission_id=submission_id)
    
    # Add check results
    checks = d.get("checks",[])
    checks_started = time.perf_counter()
    for check in checks:
        conn.execute("""INSERT INTO submission_checks 
            (submission_id,check_id,check_name,passed,notes)
            VALUES (?,?,?,?,?)""",
            (submission_id, check.get("check_id"), check.get("check_name"), 
             1 if check.get("passed") else 0, check.get("notes","")))
    perf_log("submission.checks", checks_started, count=len(checks))
    
    commit_started = time.perf_counter()
    conn.commit()
    perf_log("submission.commit", commit_started, submission_id=submission_id)
    conn.close()
    
    perf_log("submission.total", endpoint_started, status=200, submission_id=submission_id)
    return jsonify({"ok":True,"id":submission_id,"candidate_id":candidate_id})

@app.route("/api/submissions/<int:sid>",methods=["PATCH"])
@login_required
def update_submission(sid):
    d=request.json
    conn=get_db()
    if "status" in d:
        conn.execute("UPDATE requirement_submissions SET status=?, updated_at=datetime('now','localtime') WHERE id=?",(d["status"],sid))
    if "recruiter_feedback" in d:
        conn.execute("UPDATE requirement_submissions SET recruiter_feedback=?, feedback_by=?, updated_at=datetime('now','localtime') WHERE id=?",
            (d["recruiter_feedback"], session.get("user_id"), sid))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/submissions/<int:sid>/checks",methods=["GET"])
@login_required
def get_submission_checks(sid):
    conn=get_db()
    rows=conn.execute("SELECT * FROM submission_checks WHERE submission_id=?",(sid,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])
    return jsonify({"ok":True})

# â”€â”€ Automation Rules â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/automation_rules",methods=["GET"])
@login_required
def get_automation_rules():
    conn=get_db()
    rows=conn.execute("SELECT * FROM automation_rules WHERE is_active=1 ORDER BY id").fetchall()
    result=[dict(r) for r in rows]
    for r in result: r["action_config"]=json.loads(r["action_config"] or "{}")
    conn.close(); return jsonify(result)

@app.route("/api/automation_rules",methods=["POST"])
@login_required
def add_automation_rule():
    d=request.json
    if not d.get("rule_name") or not d.get("trigger_type") or not d.get("action_type"):
        return jsonify({"error":"rule_name, trigger_type, action_type required"}),400
    conn=get_db()
    rid=conn.execute("INSERT INTO automation_rules (rule_name,trigger_type,trigger_value,action_type,action_config) VALUES (?,?,?,?,?)",
                     (d["rule_name"],d["trigger_type"],d.get("trigger_value",""),
                      d["action_type"],json.dumps(d.get("action_config",{})))).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":rid})

@app.route("/api/automation_rules/<int:rid>",methods=["DELETE"])
@login_required
def del_automation_rule(rid):
    conn=get_db(); conn.execute("DELETE FROM automation_rules WHERE id=?",(rid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# â”€â”€ Interviews â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/interviews",methods=["GET"])
@login_required
def get_interviews():
    conn=get_db()
    rows=conn.execute("""SELECT i.*, c.candidate_name, c.role_name
                         FROM interviews i
                         JOIN candidates c ON i.candidate_id=c.id
                         ORDER BY scheduled_at DESC""").fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/interviews",methods=["POST"])
@login_required
def add_interview():
    d=request.json
    if not d.get("candidate_id") or not d.get("scheduled_at"):
        return jsonify({"error":"candidate_id and scheduled_at required"}),400
    conn=get_db()
    iid=conn.execute("""INSERT INTO interviews
        (candidate_id,interviewer_name,interviewer_email,scheduled_at,duration_mins,location,meeting_link,notes)
        VALUES (?,?,?,?,?,?,?,?)""",
        (d["candidate_id"],d.get("interviewer_name",""),d.get("interviewer_email",""),
         d["scheduled_at"],d.get("duration_mins",60),d.get("location",""),
         d.get("meeting_link",""),d.get("notes",""))).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":iid})

@app.route("/api/interviews/<int:iid>",methods=["PATCH"])
@login_required
def update_interview(iid):
    d=request.json
    allowed=["interviewer_name","interviewer_email","scheduled_at","duration_mins","location","meeting_link","status","notes"]
    sets=", ".join(f"{k}=?" for k in d if k in allowed)
    vals=[v for k,v in d.items() if k in allowed]+[iid]
    if not sets: return jsonify({"error":"nothing to update"}),400
    conn=get_db()
    conn.execute(f"UPDATE interviews SET {sets} WHERE id=?",vals)
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/interviews/<int:iid>",methods=["DELETE"])
@login_required
def del_interview(iid):
    conn=get_db(); conn.execute("DELETE FROM interviews WHERE id=?",(iid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# â”€â”€ Approvals â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/api/approvals",methods=["GET"])
@login_required
def get_approvals():
    conn=get_db()
    rows=conn.execute("""SELECT a.*, c.candidate_name, c.role_name
                         FROM approvals a
                         JOIN candidates c ON a.candidate_id=c.id
                         ORDER BY a.created_at DESC""").fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

@app.route("/api/approvals",methods=["POST"])
@login_required
def add_approval():
    d=request.json
    if not d.get("candidate_id") or not d.get("approval_type"):
        return jsonify({"error":"candidate_id and approval_type required"}),400
    conn=get_db()
    aid=conn.execute("""INSERT INTO approvals
        (candidate_id,requester_name,requester_email,approval_type,comments)
        VALUES (?,?,?,?,?)""",
        (d["candidate_id"],d.get("requester_name",""),d.get("requester_email",""),
         d["approval_type"],d.get("comments",""))).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":aid})

@app.route("/api/approvals/<int:aid>",methods=["PATCH"])
@login_required
def update_approval(aid):
    d=request.json
    conn=get_db()
    if "status" in d:
        conn.execute("UPDATE approvals SET status=?, reviewed_by=?, reviewed_at=datetime('now','localtime') WHERE id=?",
                     (d["status"],d.get("reviewed_by",""),aid))
    if "comments" in d:
        conn.execute("UPDATE approvals SET comments=? WHERE id=?",(d["comments"],aid))
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/approvals/<int:aid>",methods=["DELETE"])
@login_required
def del_approval(aid):
    conn=get_db(); conn.execute("DELETE FROM approvals WHERE id=?",(aid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})


@app.route("/api/email_templates",methods=["POST"])
@login_required
def add_email_template():
    d=request.json
    if not d.get("name") or not d.get("subject") or not d.get("body"):
        return jsonify({"error":"name, subject, body required"}),400
    conn=get_db()
    tid=conn.execute("INSERT INTO email_templates (name,subject,body,trigger_event) VALUES (?,?,?,?)",
                     (d["name"],d["subject"],d["body"],d.get("trigger_event",""))).lastrowid
    conn.commit(); conn.close()
    return jsonify({"ok":True,"id":tid})

@app.route("/api/email_templates/<int:tid>",methods=["PATCH"])
@login_required
def update_email_template(tid):
    d=request.json
    allowed=["name","subject","body","trigger_event"]
    sets=", ".join(f"{k}=?" for k in d if k in allowed)
    vals=[v for k,v in d.items() if k in allowed]+[tid]
    if not sets: return jsonify({"error":"nothing to update"}),400
    conn=get_db()
    conn.execute(f"UPDATE email_templates SET {sets} WHERE id=?",vals)
    conn.commit(); conn.close()
    return jsonify({"ok":True})

@app.route("/api/email_templates/<int:tid>",methods=["DELETE"])
@login_required
def del_email_template(tid):
    conn=get_db(); conn.execute("DELETE FROM email_templates WHERE id=?",(tid,))
    conn.commit(); conn.close(); return jsonify({"ok":True})

# â”€â”€ Send Email â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def unwrap_template_braces(text, variables):
    output = text or ""
    for value in variables.values():
        clean = str(value or "").strip()
        if not clean:
            continue
        output = re.sub(r"\{\s*" + re.escape(clean) + r"\s*\}", clean, output)
    return output

def _attach_bytes(msg, filename, content, mimetype="application/octet-stream"):
    maintype, subtype = (mimetype.split("/", 1) + ["octet-stream"])[:2]
    part = MIMEBase(maintype, subtype)
    part.set_payload(content)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment", filename=filename)
    msg.attach(part)

def clean_email_list(value):
    if not value:
        return []
    if isinstance(value, str):
        items = re.split(r"[\n,;]+", value)
    elif isinstance(value, (list, tuple)):
        items = value
    else:
        items = []
    emails = []
    for item in items:
        email = str(item or "").strip().lower()
        if email and re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email) and email not in emails:
            emails.append(email)
    return emails

def email_list_is_valid(value, cleaned):
    if not value:
        return True
    if isinstance(value, str):
        raw_items = [item.strip() for item in re.split(r"[\n,;]+", value) if item.strip()]
    elif isinstance(value, (list, tuple)):
        raw_items = [str(item or "").strip() for item in value if str(item or "").strip()]
    else:
        raw_items = []
    return len(raw_items) == len(cleaned)

def send_google_oauth_email(to_addr, subject, body, attachments=None, html_body=None, cc_addrs=None):
    token = session.get("google_token") or {}
    access_token = token.get("access_token")
    if not access_token:
        return {"error": "Please log in with Google again before sending email."}
    creds = Credentials(
        token=access_token,
        refresh_token=token.get("refresh_token"),
        token_uri=token.get("token_uri") or "https://oauth2.googleapis.com/token",
        client_id=os.getenv("GOOGLE_CLIENT_ID"),
        client_secret=os.getenv("GOOGLE_CLIENT_SECRET"),
        scopes=GOOGLE_OAUTH_SCOPE.split()
    )
    if creds.expired and creds.refresh_token:
        try:
            creds.refresh(GoogleAuthRequest())
        except Exception as e:
            error_text = str(e).lower()
            if "invalid_scope" in error_text or "bad request" in error_text:
                return {"error": "Google email permission needs to be refreshed. Please sign out of ATS, sign in with Google again, then retry sending the email."}
            return {"error": "Google login expired. Please sign out of ATS, sign in with Google again, then retry sending the email."}
        token["access_token"] = creds.token
        session["google_token"] = token
    msg = MIMEMultipart("mixed" if attachments else "alternative")
    msg["Subject"] = subject
    msg["From"] = session.get("email") or "me"
    msg["To"] = to_addr
    if cc_addrs:
        msg["Cc"] = ", ".join(cc_addrs)
    if html_body:
        alt = MIMEMultipart("alternative")
        alt.attach(MIMEText(body or "", "plain"))
        alt.attach(MIMEText(html_body, "html"))
        msg.attach(alt)
    else:
        msg.attach(MIMEText(body, "plain"))
    for attachment in attachments or []:
        _attach_bytes(
            msg,
            attachment.get("filename") or "attachment",
            attachment.get("content") or b"",
            attachment.get("mimetype") or "application/octet-stream",
        )
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
    service = google_api_build("gmail", "v1", credentials=creds, cache_discovery=False)
    sent = service.users().messages().send(userId="me", body={"raw": raw}).execute()
    return {"ok": True, "sent_to": to_addr, "message_id": sent.get("id"), "sender": session.get("email")}

def google_oauth_credentials(scopes):
    token = session.get("google_token") or {}
    access_token = token.get("access_token")
    if not access_token:
        return None, "Please log in with Google again before using this function."
    creds = Credentials(
        token=access_token,
        refresh_token=token.get("refresh_token"),
        token_uri=token.get("token_uri") or "https://oauth2.googleapis.com/token",
        client_id=os.getenv("GOOGLE_CLIENT_ID"),
        client_secret=os.getenv("GOOGLE_CLIENT_SECRET"),
        scopes=scopes
    )
    if creds.expired and creds.refresh_token:
        creds.refresh(GoogleAuthRequest())
        token["access_token"] = creds.token
        session["google_token"] = token
    return creds, None

def central_drive_credentials(scopes):
    service_account_file = (os.getenv("GOOGLE_DRIVE_SERVICE_ACCOUNT_FILE") or "").strip()
    service_account_json = (os.getenv("GOOGLE_DRIVE_SERVICE_ACCOUNT_JSON") or "").strip()
    try:
        if service_account_file:
            return service_account.Credentials.from_service_account_file(service_account_file, scopes=scopes), None
        if service_account_json:
            return service_account.Credentials.from_service_account_info(json.loads(service_account_json), scopes=scopes), None
    except Exception as e:
        return None, f"Central Google Drive credential could not be loaded: {e}"
    return None, "Central Google Drive is not configured. Set GOOGLE_DRIVE_SERVICE_ACCOUNT_FILE and share the central Drive folder with that service account."

def central_drive_service():
    scopes = ["https://www.googleapis.com/auth/drive"]
    creds, error = central_drive_credentials(scopes)
    if error:
        raise RuntimeError(error)
    return google_api_build("drive", "v3", credentials=creds, cache_discovery=False)

def backup_database_copy(target_path):
    src = sqlite3.connect(DB_PATH)
    dst = sqlite3.connect(target_path)
    try:
        src.backup(dst)
    finally:
        dst.close()
        src.close()

def should_skip_backup_path(path):
    rel = os.path.relpath(path, BASE_DIR)
    parts = set(rel.split(os.sep))
    skip_dirs = {
        ".git", ".venv", "venv", "__pycache__", ".pytest_cache",
        "node_modules", "backups"
    }
    if parts & skip_dirs:
        return True
    name = os.path.basename(path)
    if name in {"ats.db", "ats.db-wal", "ats.db-shm"}:
        return True
    if name.endswith((".pyc", ".pyo", ".log", ".zip")):
        return True
    return False

def create_ats_backup_zip():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    tmp = tempfile.NamedTemporaryFile(prefix=f"hrguru_ats_backup_{timestamp}_", suffix=".zip", delete=False)
    tmp_path = tmp.name
    tmp.close()
    with tempfile.TemporaryDirectory() as tmp_dir:
        db_copy = os.path.join(tmp_dir, "ats.db")
        backup_database_copy(db_copy)
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.write(db_copy, "database/ats.db")
            for root, dirs, files in os.walk(BASE_DIR):
                dirs[:] = [d for d in dirs if not should_skip_backup_path(os.path.join(root, d))]
                for filename in files:
                    full_path = os.path.join(root, filename)
                    if should_skip_backup_path(full_path):
                        continue
                    arcname = os.path.join("codebase", os.path.relpath(full_path, BASE_DIR))
                    zf.write(full_path, arcname)
    return tmp_path, f"hrguru_ats_backup_{timestamp}.zip"

def get_or_create_drive_folder(service, folder_name, parent_id=None):
    safe_folder_name = folder_name.replace("'", "\\'")
    query = (
        "mimeType='application/vnd.google-apps.folder' "
        f"and name='{safe_folder_name}' "
        "and trashed=false"
    )
    if parent_id:
        query += f" and '{parent_id}' in parents"
    result = service.files().list(
        q=query,
        spaces="drive",
        fields="files(id,name)",
        pageSize=1,
        supportsAllDrives=True,
        includeItemsFromAllDrives=True
    ).execute()
    files = result.get("files", [])
    if files:
        return files[0]["id"]
    body = {"name": folder_name, "mimeType": "application/vnd.google-apps.folder"}
    if parent_id:
        body["parents"] = [parent_id]
    folder = service.files().create(
        body=body,
        fields="id",
        supportsAllDrives=True
    ).execute()
    return folder.get("id")

def make_drive_download_url(file_id):
    return f"/api/drive_file/{file_id}" if file_id else ""

def upload_file_to_google_drive(file_path, original_name, folder_name, mimetype="application/octet-stream"):
    service = central_drive_service()
    root_folder_id = (os.getenv("GOOGLE_DRIVE_ROOT_FOLDER_ID") or "").strip() or None
    folder_id = get_or_create_drive_folder(service, folder_name, root_folder_id)
    safe_original = secure_filename(original_name or os.path.basename(file_path)) or os.path.basename(file_path)
    metadata = {"name": safe_original, "parents": [folder_id]}
    with open(file_path, "rb") as fh:
        media = MediaIoBaseUpload(fh, mimetype=mimetype, resumable=False)
        uploaded = service.files().create(
            body=metadata,
            media_body=media,
            fields="id,name,webViewLink,webContentLink",
            supportsAllDrives=True
        ).execute()
    if uploaded.get("id"):
        try:
            service.permissions().create(
                fileId=uploaded["id"],
                body={"type": "anyone", "role": "reader", "allowFileDiscovery": False},
                fields="id",
                supportsAllDrives=True
            ).execute()
        except Exception as e:
            print("Drive sharing permission warning:", e)
    return {
        "id": uploaded.get("id"),
        "name": uploaded.get("name") or safe_original,
        "url": make_drive_download_url(uploaded.get("id")),
        "webViewLink": uploaded.get("webViewLink"),
        "webContentLink": uploaded.get("webContentLink"),
    }

@app.route("/api/drive_file/<file_id>")
@login_required
def download_drive_file(file_id):
    try:
        service = central_drive_service()
        meta = service.files().get(fileId=file_id, fields="name,mimeType", supportsAllDrives=True).execute()
        content = service.files().get_media(fileId=file_id, supportsAllDrives=True).execute()
        return send_file(
            io.BytesIO(content),
            download_name=meta.get("name") or "ats_file",
            as_attachment=True,
            mimetype=meta.get("mimeType") or "application/octet-stream"
        )
    except HttpError as e:
        return jsonify({"error": "Unable to download this Google Drive file. Please confirm the file was uploaded by this ATS Google account."}), 404

@app.route("/api/admin/backup/google-drive", methods=["POST"])
@login_required
def backup_ats_to_google_drive():
    if not session.get("is_admin"):
        return jsonify({"ok": False, "error": "Admin access required."}), 403
    zip_path = None
    try:
        zip_path, filename = create_ats_backup_zip()
        service = central_drive_service()
        root_folder_id = (os.getenv("GOOGLE_DRIVE_ROOT_FOLDER_ID") or "").strip() or None
        folder_id = get_or_create_drive_folder(service, "HR Guru ATS Backups", root_folder_id)
        metadata = {"name": filename, "parents": [folder_id]}
        with open(zip_path, "rb") as fh:
            media = MediaIoBaseUpload(fh, mimetype="application/zip", resumable=False)
            uploaded = service.files().create(
                body=metadata,
                media_body=media,
                fields="id,name,webViewLink,size,createdTime",
                supportsAllDrives=True
            ).execute()
        return jsonify({
            "ok": True,
            "filename": uploaded.get("name") or filename,
            "file_id": uploaded.get("id"),
            "link": uploaded.get("webViewLink"),
            "size": uploaded.get("size")
        })
    except HttpError as e:
        message = str(e)
        if "insufficient" in message.lower() or "scope" in message.lower():
            message = "Google Drive permission is missing. Please sign out and log in with Google again, then retry backup."
        return jsonify({"ok": False, "error": message}), 400
    except Exception as e:
        print("ATS backup failed:", e, flush=True)
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e) or "Backup failed."}), 500
    finally:
        if zip_path and os.path.exists(zip_path):
            try:
                os.remove(zip_path)
            except Exception:
                pass

def send_candidate_email(candidate_id, template_name, extra_vars={}):
    """Send email to candidate using a template"""
    conn=get_db()
    c=conn.execute("""SELECT c.*, r.client_name as client_name
                      FROM candidates c
                      LEFT JOIN requirements r ON c.requirement_id = r.id
                      WHERE c.id=?""",(candidate_id,)).fetchone()
    if not c: conn.close(); return {"error":"Candidate not found"}
    t=conn.execute("SELECT * FROM email_templates WHERE name=?",(template_name,)).fetchone()
    if not t: conn.close(); return {"error":"Template not found"}
    # Replace variables in subject and body
    variables={"candidate_name":c["candidate_name"] or "",
               "role_name":c["role_name"] or "",
               "role":c["role_name"] or c["current_role"] or "",
               "current_role":c["current_role"] or "",
               "status":c["status"] or "",
               "current_company":c["current_company"] or "",
               "company":c["client_name"] or c["current_company"] or "",
               "client_name":c["client_name"] or "",
               "email_addr":c["email_addr"] or "",
               "phone":c["phone"] or "",
               "recruiter_name":c["recruiter_name"] or "",
               **extra_vars}
    subject=t["subject"]
    body=t["body"]
    for k,v in variables.items():
        subject=subject.replace("{{"+k+"}}",str(v))
        body=body.replace("{{"+k+"}}",str(v))
        subject=subject.replace("{"+k+"}",str(v))
        body=body.replace("{"+k+"}",str(v))
    subject=unwrap_template_braces(subject, variables)
    body=unwrap_template_braces(body, variables)
    recipient=c["email_addr"]
    if not recipient: conn.close(); return {"error":"No email address"}
    try:
        if session.get("google_token"):
            sent_result = send_google_oauth_email(recipient, subject, body)
            if not sent_result.get("ok"):
                conn.close()
                return sent_result
        else:
            if not GMAIL_USER or not GMAIL_APP_PASS:
                conn.close()
                return {"error":"Please log in with Google again before sending email."}
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"] = GMAIL_USER
            msg["To"] = recipient
            msg.attach(MIMEText(body,"plain"))
            with smtplib.SMTP_SSL("smtp.gmail.com",465) as s:
                s.login(GMAIL_USER,GMAIL_APP_PASS)
                s.sendmail(GMAIL_USER,recipient,msg.as_string())
        # Log the email
        conn.execute("INSERT INTO email_log (candidate_id,template_name,recipient,subject,body,status) VALUES (?,?,?,?,?,'Sent')",
                     (candidate_id,template_name,recipient,subject,body))
        conn.commit(); conn.close()
        return {"ok":True,"sent_to":recipient}
    except Exception as e:
        conn.execute("INSERT INTO email_log (candidate_id,template_name,recipient,subject,body,status,error_msg) VALUES (?,?,?,?,?,'Failed',?)",
                     (candidate_id,template_name,recipient,subject,body,str(e)))
        conn.commit(); conn.close()
        return {"error":str(e)}

def send_custom_email(to_addr, subject, body, candidate_id=None, template_name="custom", attachments=None, html_body=None, cc_addrs=None):
    use_google = bool(session.get("google_token"))
    cc_addrs = clean_email_list(cc_addrs)
    try:
        if use_google:
            try:
                sent_result = send_google_oauth_email(to_addr, subject, body, attachments=attachments, html_body=html_body, cc_addrs=cc_addrs)
                if not sent_result.get("ok"):
                    return sent_result
            except HttpError as e:
                error_text = str(e)
                if "insufficient authentication scopes" not in error_text.lower() and "insufficient permission" not in error_text.lower():
                    raise
                if not GMAIL_USER or not GMAIL_APP_PASS:
                    return {"error": "Google permission is missing for Gmail send. Please sign out and log in with Google again to grant email permission."}
                use_google = False
        if not use_google:
            if not GMAIL_USER or not GMAIL_APP_PASS:
                return {"error":"Please log in with Google again before sending email."}
            msg = MIMEMultipart("mixed" if attachments else "alternative")
            msg["Subject"] = subject
            msg["From"] = GMAIL_USER
            msg["To"] = to_addr
            if cc_addrs:
                msg["Cc"] = ", ".join(cc_addrs)
            if html_body:
                alt = MIMEMultipart("alternative")
                alt.attach(MIMEText(body or "", "plain"))
                alt.attach(MIMEText(html_body, "html"))
                msg.attach(alt)
            else:
                msg.attach(MIMEText(body, "plain"))
            for attachment in attachments or []:
                _attach_bytes(
                    msg,
                    attachment.get("filename") or "attachment",
                    attachment.get("content") or b"",
                    attachment.get("mimetype") or "application/octet-stream",
                )
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
                s.login(GMAIL_USER, GMAIL_APP_PASS)
                s.sendmail(GMAIL_USER, [to_addr] + cc_addrs, msg.as_string())
        conn = get_db()
        conn.execute(
            "INSERT INTO email_log (candidate_id,template_name,recipient,subject,body,status) VALUES (?,?,?,?,?,'Sent')",
            (candidate_id, template_name, to_addr, subject, body)
        )
        conn.commit(); conn.close()
        return {"ok": True, "sent_to": to_addr, "sender": session.get("email") or GMAIL_USER}
    except Exception as e:
        raw_error = str(e)
        error_text = raw_error.lower()
        if "invalid_scope" in error_text or "bad request" in error_text:
            raw_error = "Google email permission needs to be refreshed. Please sign out of ATS, sign in with Google again, then retry sending the email."
        conn = get_db()
        conn.execute(
            "INSERT INTO email_log (candidate_id,template_name,recipient,subject,body,status,error_msg) VALUES (?,?,?,?,?,'Failed',?)",
            (candidate_id, template_name, to_addr, subject, body, raw_error)
        )
        conn.commit(); conn.close()
        return {"error": raw_error}

@app.route("/api/candidates/daily_report_email", methods=["POST"])
@login_required
def api_daily_work_report_email():
    data = request.get_json(silent=True) or {}
    to_addr = (data.get("to") or "").strip().lower()
    if not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", to_addr):
        return jsonify({"error": "Please enter a valid email address."}), 400
    cc_addrs = clean_email_list(data.get("cc") or [])
    if not email_list_is_valid(data.get("cc"), cc_addrs):
        return jsonify({"error": "Please enter valid CC email addresses."}), 400
    if not session.get("google_token"):
        return jsonify({"error": "Please log in with Google before sending the daily work report email."}), 400
    query_args = dict(data.get("filters") or {})
    if data.get("ids"):
        query_args["ids"] = ",".join(str(x) for x in data.get("ids") or [] if str(x).isdigit())
    rows = candidate_report_rows(query_args, session)
    filters_applied = bool(query_args.get("ids") or any(query_args.get(k) for k in ["q", "client", "requirement_id", "status", "skills", "exp_range"]))
    report_date = feedback_request_shared_date(rows, query_args)
    sender_name = daily_report_sender_name()
    email_columns = data.get("columns") or []
    subject, body = build_daily_work_report_email(rows, sender_name, filters_applied, to_addr, data.get("recipient_name") or "", report_date, email_columns)
    html_body = build_daily_work_report_html(rows, sender_name, filters_applied, to_addr, data.get("recipient_name") or "", report_date, email_columns)
    attachments = [{
        "filename": f"daily_work_report_{date.today().isoformat()}.xlsx",
        "content": build_candidate_export_xlsx(rows),
        "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }]
    cv_attachments, missing_cvs = candidate_cv_attachment_validation(rows)
    cv_summary = {
        "attached_cvs": len(cv_attachments),
        "cv_filenames": [a.get("filename") or f"candidate_cv_{i}" for i, a in enumerate(cv_attachments, 1)],
    }
    if missing_cvs:
        names = ", ".join(missing_cvs[:8])
        more = f" and {len(missing_cvs) - 8} more" if len(missing_cvs) > 8 else ""
        return jsonify({"error": f"Cannot send email. Attach required CVs before sending daily reporting email. Missing CV for: {names}{more}."}), 400
    attachments.extend(cv_attachments)
    result = send_custom_email(to_addr, subject, body, template_name="daily_work_report", attachments=attachments, html_body=html_body, cc_addrs=cc_addrs)
    if not result.get("ok"):
        return jsonify({"error": result.get("error") or "Unable to send daily work report email."}), 400
    return jsonify({
        "ok": True,
        "sent_to": to_addr,
        "count": len(rows),
        "attached_cvs": cv_summary["attached_cvs"],
        "cv_filenames": cv_summary["cv_filenames"],
        "attachments": 1 + len(cv_attachments),
        "message": f"Email sent successfully to {to_addr}" + (f" with CC to {', '.join(cc_addrs)}" if cc_addrs else "")
    })

@app.route("/api/candidates/daily_report_preview", methods=["POST"])
@login_required
def api_daily_work_report_preview():
    data = request.get_json(silent=True) or {}
    cc_addrs = clean_email_list(data.get("cc") or [])
    if not email_list_is_valid(data.get("cc"), cc_addrs):
        return jsonify({"error": "Please enter valid CC email addresses."}), 400
    query_args = dict(data.get("filters") or {})
    if data.get("ids"):
        query_args["ids"] = ",".join(str(x) for x in data.get("ids") or [] if str(x).isdigit())
    rows = candidate_report_rows(query_args, session)
    filters_applied = bool(query_args.get("ids") or any(query_args.get(k) for k in ["q", "client", "requirement_id", "status", "skills", "exp_range"]))
    report_date = feedback_request_shared_date(rows, query_args)
    sender_name = daily_report_sender_name()
    email_columns = data.get("columns") or []
    subject, body = build_daily_work_report_email(rows, sender_name, filters_applied, data.get("to") or "", data.get("recipient_name") or "", report_date, email_columns)
    html_body = build_daily_work_report_html(rows, sender_name, filters_applied, data.get("to") or "", data.get("recipient_name") or "", report_date, email_columns)
    cv_attachments, missing_cvs = candidate_cv_attachment_validation(rows)
    cv_summary = {
        "attached_cvs": len(cv_attachments),
        "cv_filenames": [a.get("filename") or f"candidate_cv_{i}" for i, a in enumerate(cv_attachments, 1)],
    }
    missing_message = ""
    if missing_cvs:
        names = ", ".join(missing_cvs[:8])
        more = f" and {len(missing_cvs) - 8} more" if len(missing_cvs) > 8 else ""
        missing_message = f" Missing CV for: {names}{more}."
    return jsonify({
        "ok": True,
        "subject": subject,
        "body": body,
        "html_body": html_body,
        "count": len(rows),
        "attached_cvs": cv_summary["attached_cvs"],
        "cv_filenames": cv_summary["cv_filenames"],
        "cc": cc_addrs,
        "missing_cvs": missing_cvs,
        "warning": missing_message.strip(),
        "attachments": 1 + cv_summary["attached_cvs"],
    })

@app.route("/api/candidates/feedback_request_email", methods=["POST"])
@login_required
def api_feedback_request_email():
    data = request.get_json(silent=True) or {}
    to_addr = (data.get("to") or "").strip().lower()
    if not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", to_addr):
        return jsonify({"error": "Please enter a valid email address."}), 400
    if not session.get("google_token"):
        return jsonify({"error": "Please log in with Google before sending feedback request email."}), 400
    ids = [str(x) for x in data.get("ids") or [] if str(x).isdigit()]
    if not ids:
        return jsonify({"error": "Select at least one candidate before requesting feedback."}), 400
    query_args = dict(data.get("filters") or {})
    query_args["ids"] = ",".join(ids)
    rows = candidate_report_rows(query_args, session)
    if not rows:
        return jsonify({"error": "No selected candidates were found for your access."}), 404
    report_date = feedback_request_shared_date(rows, query_args)
    subject, body, html_body = build_feedback_request_email(
        rows,
        to_addr,
        data.get("recipient_name") or "",
        session.get("recruiter_name") or session.get("username") or "",
        report_date,
        data.get("columns") or [],
    )
    cv_attachments, missing_cvs = candidate_cv_attachment_validation(rows)
    cv_summary = {
        "attached_cvs": len(cv_attachments),
        "cv_filenames": [a.get("filename") or f"candidate_cv_{i}" for i, a in enumerate(cv_attachments, 1)],
    }
    if missing_cvs:
        names = ", ".join(missing_cvs[:8])
        more = f" and {len(missing_cvs) - 8} more" if len(missing_cvs) > 8 else ""
        return jsonify({"error": f"Cannot send email. Attach required CVs before sending feedback request email. Missing CV for: {names}{more}."}), 400
    result = send_custom_email(to_addr, subject, body, template_name="feedback_request", attachments=cv_attachments, html_body=html_body)
    if not result.get("ok"):
        return jsonify({"error": result.get("error") or "Unable to send feedback request email."}), 400
    return jsonify({
        "ok": True,
        "sent_to": to_addr,
        "count": len(rows),
        "attached_cvs": cv_summary["attached_cvs"],
        "cv_filenames": cv_summary["cv_filenames"],
        "attachments": len(cv_attachments),
        "subject": subject,
        "message": f"Feedback request sent successfully to {to_addr}"
    })

@app.route("/api/candidates/feedback_request_preview", methods=["POST"])
@login_required
def api_feedback_request_preview():
    data = request.get_json(silent=True) or {}
    to_addr = (data.get("to") or "").strip().lower() or "recipient@example.com"
    ids = [str(x) for x in data.get("ids") or [] if str(x).isdigit()]
    if not ids:
        return jsonify({"error": "Select at least one candidate before previewing feedback request."}), 400
    query_args = dict(data.get("filters") or {})
    query_args["ids"] = ",".join(ids)
    rows = candidate_report_rows(query_args, session)
    if not rows:
        return jsonify({"error": "No selected candidates were found for your access."}), 404
    report_date = feedback_request_shared_date(rows, query_args)
    subject, body, html_body = build_feedback_request_email(
        rows,
        to_addr,
        data.get("recipient_name") or "",
        session.get("recruiter_name") or session.get("username") or "",
        report_date,
        data.get("columns") or [],
    )
    cv_attachments, missing_cvs = candidate_cv_attachment_validation(rows)
    cv_summary = {
        "attached_cvs": len(cv_attachments),
        "cv_filenames": [a.get("filename") or f"candidate_cv_{i}" for i, a in enumerate(cv_attachments, 1)],
    }
    missing_message = ""
    if missing_cvs:
        names = ", ".join(missing_cvs[:8])
        more = f" and {len(missing_cvs) - 8} more" if len(missing_cvs) > 8 else ""
        missing_message = f" Missing CV for: {names}{more}."
    return jsonify({
        "ok": True,
        "subject": subject,
        "body": body,
        "html_body": html_body,
        "count": len(rows),
        "attached_cvs": cv_summary["attached_cvs"],
        "cv_filenames": cv_summary["cv_filenames"],
        "missing_cvs": missing_cvs,
        "warning": missing_message.strip(),
        "attachments": cv_summary["attached_cvs"],
    })

@app.route("/api/send_email", methods=["POST"])
@login_required
def api_send_email():
    if request.is_json:
        d = request.get_json(silent=True) or {}
        uploaded_files = []
    else:
        d = request.form.to_dict()
        uploaded_files = request.files.getlist("attachments")

    try:
        attachments = []
        for uploaded in uploaded_files:
            if uploaded and uploaded.filename:
                attachments.append({
                    "filename": secure_filename(uploaded.filename) or uploaded.filename,
                    "content": uploaded.read(),
                    "mimetype": uploaded.mimetype or "application/octet-stream",
                })

        candidate_id = d.get("candidate_id")
        if candidate_id:
            try:
                candidate_id = int(candidate_id)
            except (TypeError, ValueError):
                return jsonify({"error": "Invalid candidate."}), 400
            conn = get_db()
            owner_sql, owner_params = non_admin_candidate_owner_clause(session, "c")
            allowed_candidate = conn.execute(
                f"SELECT c.id FROM candidates c WHERE c.id=?{owner_sql}",
                [candidate_id] + owner_params
            ).fetchone()
            conn.close()
            if not allowed_candidate:
                return jsonify({"error": "Candidate not found or permission denied."}), 404

        # The UI sends the final composed subject/body, so send that directly.
        # This avoids relying on an expired Google OAuth access token.
        if d.get("to") and d.get("subject") and d.get("body"):
            to_addr = (d.get("to") or "").strip().lower()
            cc_addrs = clean_email_list(d.get("cc") or d.get("cc_addrs"))
            if not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", to_addr):
                return jsonify({"error": "Please enter a valid recipient email."}), 400
            if not email_list_is_valid(d.get("cc") or d.get("cc_addrs"), cc_addrs):
                return jsonify({"error": "Please enter valid CC email addresses."}), 400
            result = send_custom_email(
                to_addr,
                d["subject"],
                d["body"],
                candidate_id,
                d.get("template_name") or "custom",
                attachments=attachments,
                cc_addrs=cc_addrs
            )
            return jsonify(result), (200 if result.get("ok") else 400)

        # Template-only email flow
        if candidate_id and d.get("template_name"):
            extra = d.get("variables", {})
            return jsonify(send_candidate_email(
                candidate_id,
                d["template_name"],
                extra
            ))

        return jsonify({"error": "Invalid payload"}), 400

    except Exception as e:
        return jsonify({"error": str(e)}), 500

LINKEDIN_CAMPAIGN_TYPE = "linkedin_page_follow"

def linkedin_schedule_steps(schedule):
    schedule = (schedule or "").lower()
    if "2 emails" in schedule:
        return [
            {"step": 1, "delay_after_days": 4},
            {"step": 2, "delay_after_days": None},
        ]
    if "manual" in schedule:
        return [
            {"step": 1, "delay_after_days": None},
            {"step": 2, "delay_after_days": None},
            {"step": 3, "delay_after_days": None},
        ]
    return [
        {"step": 1, "delay_after_days": 3},
        {"step": 2, "delay_after_days": 4},
        {"step": 3, "delay_after_days": None},
    ]

def linkedin_campaign_templates(segment, subject=None, body=None):
    segment = segment or "Active job seekers"
    role_line = "We also post openings, hiring announcements, and career opportunities that may match your profile."
    if "IT" in segment:
        role_line = "We also post technology roles and hiring updates that may match your skills."
    elif "Sales" in segment:
        role_line = "We also post sales openings, client-facing roles, and referral opportunities."
    return [
        {
            "step": 1,
            "subject": subject or "Get our latest job updates on LinkedIn",
            "body": body or (
                "Hi,\n\n"
                "We regularly share new openings, hiring updates, and career opportunities on our LinkedIn page. "
                "If you'd like to see relevant roles and hiring updates faster, you can follow us here:\n\n"
                "{{linkedin_page_url}}\n\n"
                "Regards,\nHR Guru Placement Services"
            ),
        },
        {
            "step": 2,
            "subject": "New roles and hiring updates",
            "body": (
                "Hi {{first_name}},\n\n"
                f"{role_line}\n\n"
                "By following our LinkedIn page, you can see new job openings, hiring announcements, "
                "career tips, and referral opportunities in one place:\n\n"
                "{{linkedin_page_url}}\n\n"
                "Regards,\n{{recruiter_name}}\n{{company_name}}"
            ),
        },
        {
            "step": 3,
            "subject": "Know someone looking for a new role?",
            "body": (
                "Hi {{first_name}},\n\n"
                "We regularly share new openings and hiring updates on our LinkedIn page.\n\n"
                "If you know friends or colleagues exploring new opportunities, you can follow and share the page here:\n\n"
                "{{linkedin_page_url}}\n\n"
                "Regards,\n{{recruiter_name}}\n{{company_name}}"
            ),
        },
    ]

def candidate_first_name(candidate_name):
    parts = re.findall(r"[A-Za-z]+", candidate_name or "")
    return parts[0] if parts else "there"

def render_campaign_text(text, variables):
    output = text or ""
    for key, value in variables.items():
        output = output.replace("{{" + key + "}}", str(value or ""))
        output = output.replace("{" + key + "}", str(value or ""))
    return unwrap_template_braces(output, variables)

def parse_recent_days(value, default=14):
    try:
        days = int(value)
        return max(0, min(days, 365))
    except Exception:
        return default

def build_linkedin_exclusion_rules(data):
    return {
        "exclude_unsubscribed": bool(data.get("exclude_unsubscribed", True)),
        "exclude_do_not_contact": bool(data.get("exclude_do_not_contact", True)),
        "exclude_recently_contacted": bool(data.get("exclude_recently_contacted", True)),
        "recent_days": parse_recent_days(data.get("recent_days"), 14),
        "exclude_negative_responders": bool(data.get("exclude_negative_responders", True)),
        "exclude_sensitive_active": bool(data.get("exclude_sensitive_active", True)),
    }

def linkedin_candidate_matches_segment(row, segment):
    segment_l = (segment or "").lower()
    text = " ".join(str(row.get(k) or "") for k in [
        "candidate_name", "role_name", "current_role", "key_skills", "status", "tags", "remarks"
    ]).lower()
    status = (row.get("status") or "").lower()
    if "it candidate" in segment_l:
        return any(term in text for term in [
            "python", "java", "developer", "software", "engineer", "cloud", "devops",
            "data", "network", "fullstack", "backend", "frontend", "sql", "azure", "aws"
        ])
    if "sales candidate" in segment_l:
        return any(term in text for term in ["sales", "business development", "bdm", "account manager", "client"])
    if "placed" in segment_l:
        return any(term in status for term in ["placed", "joined"])
    if "rejected" in segment_l:
        return "reject" in status or "not selected" in text
    if "previous applicant" in segment_l:
        return True
    if "passive" in segment_l:
        return not any(term in status for term in ["screening", "interview", "offered"])
    return True

def normalize_filter_terms(value):
    return [term.strip().lower() for term in re.split(r"[,;\n]+", str(value or "")) if term.strip()]

def linkedin_candidate_matches_filters(row, filters):
    filters = filters or {}
    searchable = " ".join(str(row.get(k) or "") for k in [
        "candidate_name", "email_addr", "role_name", "current_role", "key_skills",
        "status", "tags", "remarks", "current_location", "preferred_location", "client_name"
    ]).lower()
    keyword = str(filters.get("keyword") or "").strip().lower()
    if keyword and keyword not in searchable:
        return False, "outside keyword filter"
    for key, label in [("status", "outside status filter"), ("client", "outside client filter")]:
        terms = normalize_filter_terms(filters.get(key))
        if terms:
            source = str(row.get("client_name") if key == "client" else row.get("status") or "").lower()
            if not any(term in source for term in terms):
                return False, label
    skill_terms = normalize_filter_terms(filters.get("skill"))
    if skill_terms:
        source = " ".join(str(row.get(k) or "") for k in ["role_name", "current_role", "key_skills"]).lower()
        if not any(term in source for term in skill_terms):
            return False, "outside role/skill filter"
    return True, ""

def filtered_max_recipients(filters):
    try:
        value = int((filters or {}).get("max_recipients") or 0)
        return max(0, min(value, 5000))
    except Exception:
        return 0

def row_is_excluded_for_linkedin(row, rules, recent_contacted_ids):
    status = (row.get("status") or "").lower()
    tags = (row.get("tags") or "").lower()
    remarks = (row.get("remarks") or "").lower()
    feedback = (row.get("candidate_feedback") or "").lower()
    combined = " ".join([status, tags, remarks, feedback])
    if rules.get("exclude_unsubscribed") and any(term in combined for term in ["unsubscribe", "unsubscribed", "opt out", "opt-out"]):
        return "unsubscribed"
    if rules.get("exclude_do_not_contact") and any(term in combined for term in ["do not contact", "dnc", "do-not-contact"]):
        return "do not contact"
    if rules.get("exclude_negative_responders") and any(term in combined for term in ["not interested", "negative", "stop emailing", "do not email", "don't email"]):
        return "negative response"
    if rules.get("exclude_sensitive_active") and any(term in status for term in ["screening", "interview", "offered", "offer", "submitted"]):
        return "active hiring conversation"
    if rules.get("exclude_recently_contacted") and row.get("id") in recent_contacted_ids:
        return "recently contacted"
    return ""

def linkedin_campaign_candidates(conn, segment, rules, filters=None):
    filters = filters or {}
    recent_contacted_ids = set()
    if rules.get("exclude_recently_contacted"):
        since = (datetime.now() - timedelta(days=rules.get("recent_days", 14))).strftime("%Y-%m-%d %H:%M:%S")
        recent_contacted_ids = {
            row["candidate_id"]
            for row in conn.execute(
                "SELECT DISTINCT candidate_id FROM email_log WHERE candidate_id IS NOT NULL AND sent_at >= ?",
                (since,)
            ).fetchall()
        }
    rows = conn.execute("""
        SELECT c.id, c.candidate_name, c.email_addr, c.role_name, c.current_role, c.key_skills, c.status,
               c.tags, c.remarks, c.candidate_feedback, c.recruiter_name, c.recruiter_email,
               c.current_location, c.preferred_location, r.client_name
        FROM candidates c
        LEFT JOIN requirements r ON c.requirement_id = r.id
        WHERE email_addr IS NOT NULL AND trim(email_addr) != ''
        ORDER BY c.created_at DESC
    """).fetchall()
    eligible = []
    excluded = {}
    seen_emails = set()
    max_recipients = filtered_max_recipients(filters)
    for row in rows:
        row = dict(row)
        email = (row.get("email_addr") or "").strip().lower()
        if not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email):
            excluded["invalid email"] = excluded.get("invalid email", 0) + 1
            continue
        if email in seen_emails:
            excluded["duplicate email"] = excluded.get("duplicate email", 0) + 1
            continue
        if not linkedin_candidate_matches_segment(row, segment):
            excluded["outside segment"] = excluded.get("outside segment", 0) + 1
            continue
        filter_match, filter_reason = linkedin_candidate_matches_filters(row, filters)
        if not filter_match:
            excluded[filter_reason] = excluded.get(filter_reason, 0) + 1
            continue
        reason = row_is_excluded_for_linkedin(row, rules, recent_contacted_ids)
        if reason:
            excluded[reason] = excluded.get(reason, 0) + 1
            continue
        seen_emails.add(email)
        eligible.append(row)
        if max_recipients and len(eligible) >= max_recipients:
            break
    return eligible, excluded

@app.route("/api/communication/linkedin_campaign/preview", methods=["POST"])
@login_required
def preview_linkedin_campaign():
    data = request.get_json(silent=True) or {}
    rules = build_linkedin_exclusion_rules(data.get("exclusion_rules") or data)
    filters = data.get("audience_filters") or {}
    segment = data.get("target_segment") or data.get("segment") or "Active job seekers"
    conn = get_db()
    eligible, excluded = linkedin_campaign_candidates(conn, segment, rules, filters)
    conn.close()
    return jsonify({
        "ok": True,
        "eligible_count": len(eligible),
        "excluded": excluded,
        "sample": [{
            "id": row["id"],
            "candidate_name": row.get("candidate_name") or "-",
            "email_addr": row.get("email_addr") or "-",
            "role": row.get("role_name") or row.get("current_role") or "-",
            "client": row.get("client_name") or "-",
            "status": row.get("status") or "-",
        } for row in eligible[:10]]
    })

@app.route("/api/communication/linkedin_campaigns", methods=["POST"])
@login_required
def create_linkedin_campaign():
    data = request.get_json(silent=True) or {}
    linkedin_url = (data.get("linkedin_page_url") or "").strip()
    if not linkedin_url:
        return jsonify({"error": "LinkedIn company page URL is required."}), 400
    segment = data.get("target_segment") or "Active job seekers"
    schedule = data.get("send_schedule") or "3 emails over 10 days"
    steps = linkedin_schedule_steps(schedule)
    max_steps = len(steps)
    templates = linkedin_campaign_templates(segment, data.get("email_subject"), data.get("email_body"))[:max_steps]
    rules = build_linkedin_exclusion_rules(data.get("exclusion_rules") or data)
    filters = data.get("audience_filters") or {}
    conn = get_db()
    eligible, excluded = linkedin_campaign_candidates(conn, segment, rules, filters)
    campaign_name = data.get("name") or f"LinkedIn Follow - {segment}"
    tracking_link = (data.get("tracking_link") or linkedin_url).strip()
    campaign_id = conn.execute("""
        INSERT INTO communication_campaigns
            (campaign_type, name, linkedin_page_url, target_segment, send_schedule, tracking_link,
             follow_up_step, exclusion_rules_json, templates_json, analytics_json, status, created_by)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        LINKEDIN_CAMPAIGN_TYPE, campaign_name, linkedin_url, segment, schedule, tracking_link,
        data.get("follow_up_step") or "", json.dumps(rules), json.dumps(templates),
        json.dumps({"excluded": excluded, "audience_filters": filters}), "Active", session.get("username") or session.get("email") or ""
    )).lastrowid
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row in eligible:
        token = secrets.token_urlsafe(18)
        conn.execute("""
            INSERT OR IGNORE INTO communication_campaign_recipients
                (campaign_id, candidate_id, email_addr, first_name, current_step, max_steps, status, tracking_token, next_send_at)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (
            campaign_id, row["id"], (row.get("email_addr") or "").strip().lower(),
            candidate_first_name(row.get("candidate_name")), 1, max_steps, "Pending", token, now
        ))
    conn.commit()
    recipient_count = conn.execute(
        "SELECT COUNT(*) AS n FROM communication_campaign_recipients WHERE campaign_id=?",
        (campaign_id,)
    ).fetchone()["n"]
    conn.close()
    return jsonify({
        "ok": True,
        "campaign_id": campaign_id,
        "eligible_count": recipient_count,
        "excluded": excluded,
        "message": f"Campaign created with {recipient_count} eligible recipients. Email 1 is due now; later steps follow the selected day gaps."
    })

@app.route("/api/communication/linkedin_campaigns/<int:campaign_id>/send_due", methods=["POST"])
@login_required
def send_due_linkedin_campaign(campaign_id):
    data = request.get_json(silent=True) or {}
    limit = max(1, min(int(data.get("limit") or 25), 200))
    conn = get_db()
    campaign = conn.execute("SELECT * FROM communication_campaigns WHERE id=?", (campaign_id,)).fetchone()
    if not campaign:
        conn.close()
        return jsonify({"error": "Campaign not found."}), 404
    campaign = dict(campaign)
    templates = json.loads(campaign["templates_json"] or "[]")
    steps = linkedin_schedule_steps(campaign["send_schedule"])
    now = datetime.now()
    due = [dict(row) for row in conn.execute("""
        SELECT r.*, c.candidate_name, c.recruiter_name, c.recruiter_email
        FROM communication_campaign_recipients r
        JOIN candidates c ON c.id = r.candidate_id
        WHERE r.campaign_id=? AND r.status='Pending' AND datetime(r.next_send_at) <= datetime('now','localtime')
        ORDER BY datetime(r.next_send_at), r.id
        LIMIT ?
    """, (campaign_id, limit)).fetchall()]
    conn.close()
    sent = 0
    failed = 0
    errors = []
    base_url = request.host_url.rstrip("/")
    for row in due:
        step = int(row.get("current_step") or 1)
        template = next((t for t in templates if int(t.get("step") or 0) == step), templates[min(step - 1, len(templates) - 1)] if templates else None)
        if not template:
            continue
        tracking_url = f"{base_url}/api/communication/linkedin_campaigns/click/{row['tracking_token']}"
        variables = {
            "first_name": row.get("first_name") or candidate_first_name(row.get("candidate_name")),
            "candidate_name": row.get("candidate_name") or "",
            "linkedin_page_url": tracking_url,
            "recruiter_name": session.get("recruiter_name") or row.get("recruiter_name") or "Recruiter",
            "company_name": data.get("company_name") or "HR Guru",
        }
        subject = render_campaign_text(template.get("subject"), variables)
        body = render_campaign_text(template.get("body"), variables)
        try:
            result = send_custom_email(
                row["email_addr"], subject, body,
                candidate_id=row["candidate_id"],
                template_name=f"LinkedIn Follow Campaign - Email {step}"
            )
        except Exception as e:
            result = {"error": str(e)}
        update_conn = get_db()
        if result.get("ok"):
            sent += 1
            next_step = step + 1
            delay = steps[step - 1].get("delay_after_days") if step - 1 < len(steps) else None
            if next_step <= int(row.get("max_steps") or len(templates)) and delay is not None:
                next_send_at = (now + timedelta(days=delay)).strftime("%Y-%m-%d %H:%M:%S")
                status = "Pending"
            elif next_step <= int(row.get("max_steps") or len(templates)):
                next_send_at = None
                status = "Paused"
            else:
                next_send_at = None
                status = "Completed"
            update_conn.execute("""
                UPDATE communication_campaign_recipients
                SET current_step=?, status=?, sent_count=sent_count+1, last_sent_at=datetime('now','localtime'),
                    next_send_at=?, error_msg='', updated_at=datetime('now','localtime')
                WHERE id=?
            """, (next_step, status, next_send_at, row["id"]))
        else:
            failed += 1
            errors.append({"recipient": row["email_addr"], "error": result.get("error") or "Send failed"})
            update_conn.execute("""
                UPDATE communication_campaign_recipients
                SET status='Failed', error_msg=?, updated_at=datetime('now','localtime')
                WHERE id=?
            """, (result.get("error") or "Send failed", row["id"]))
        update_conn.commit()
        update_conn.close()
    conn = get_db()
    totals = conn.execute("""
        SELECT status, COUNT(*) AS n, COALESCE(SUM(sent_count),0) AS sent_total, COALESCE(SUM(clicked_count),0) AS clicked_total
        FROM communication_campaign_recipients
        WHERE campaign_id=?
        GROUP BY status
    """, (campaign_id,)).fetchall()
    analytics = {row["status"]: row["n"] for row in totals}
    analytics["emails_sent"] = sum(row["sent_total"] for row in totals)
    analytics["clicks_to_linkedin"] = sum(row["clicked_total"] for row in totals)
    conn.execute("UPDATE communication_campaigns SET analytics_json=?, updated_at=datetime('now','localtime') WHERE id=?",
                 (json.dumps(analytics), campaign_id))
    conn.commit(); conn.close()
    return jsonify({"ok": True, "sent": sent, "failed": failed, "errors": errors[:10], "analytics": analytics})

@app.route("/api/communication/linkedin_campaigns/click/<token>")
def track_linkedin_campaign_click(token):
    conn = get_db()
    row = conn.execute("""
        SELECT r.id, c.linkedin_page_url
        FROM communication_campaign_recipients r
        JOIN communication_campaigns c ON c.id = r.campaign_id
        WHERE r.tracking_token=?
    """, (token,)).fetchone()
    if not row:
        conn.close()
        return redirect("https://www.linkedin.com/")
    conn.execute("""
        UPDATE communication_campaign_recipients
        SET clicked_count=clicked_count+1, updated_at=datetime('now','localtime')
        WHERE id=?
    """, (row["id"],))
    conn.commit(); conn.close()
    return redirect(row["linkedin_page_url"] or "https://www.linkedin.com/")

@app.route("/api/communication/linkedin_campaigns/<int:campaign_id>/analytics")
@login_required
def linkedin_campaign_analytics(campaign_id):
    conn = get_db()
    rows = conn.execute("""
        SELECT status, COUNT(*) AS n, COALESCE(SUM(sent_count),0) AS sent_total, COALESCE(SUM(clicked_count),0) AS clicked_total
        FROM communication_campaign_recipients
        WHERE campaign_id=?
        GROUP BY status
    """, (campaign_id,)).fetchall()
    conn.close()
    analytics = {row["status"]: row["n"] for row in rows}
    analytics["emails_sent"] = sum(row["sent_total"] for row in rows)
    analytics["clicks_to_linkedin"] = sum(row["clicked_total"] for row in rows)
    return jsonify({"ok": True, "campaign_id": campaign_id, "analytics": analytics})

@app.route("/api/email_templates")
@login_required
def get_email_templates():
    conn = get_db()
    rows = conn.execute(
        "SELECT id,name,subject,body FROM email_templates ORDER BY name"
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/email_log",methods=["GET"])
@login_required
def get_email_log():
    conn=get_db()
    rows=conn.execute("""SELECT el.*, c.candidate_name 
                         FROM email_log el 
                         LEFT JOIN candidates c ON el.candidate_id=c.id 
                         ORDER BY sent_at DESC LIMIT 100""").fetchall()
    conn.close(); return jsonify([dict(r) for r in rows])

try:
    acquire_single_instance_lock()
    initialize_app_database()
except Exception as e:
    print("Database initialization error:", e, flush=True)

if __name__=="__main__":
    port = int(os.getenv("PORT", 5001))
    print(f"HR Guru ATS -> http://localhost:{port}", flush=True)
    app.run(debug=False, host="0.0.0.0", port=port)





